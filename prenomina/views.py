from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from proyecto.models import UserDatos, Perfil, Catorcenas, Costo, TablaFestivos, Vacaciones, Economicos, Economicos_dia_tomado, Vacaciones_dias_tomados, Empresa, Solicitud_vacaciones, Solicitud_economicos, Trabajos_encomendados
from proyecto.models import TablaVacaciones,SalarioDatos,DatosISR

from django.db import models
from django.db.models import Subquery, OuterRef, Q
from revisar.models import AutorizarPrenomina, Estado
from proyecto.filters import CostoFilter
from .models import Prenomina,PrenominaIncidencias, IncidenciaRango
from django.core.paginator import Paginator
from django.shortcuts import render, redirect
import datetime 
from dateutil import parser
from dateutil.relativedelta import relativedelta
import os

from datetime import timedelta, date
from .filters import PrenominaFilter
import math

from django.contrib import messages
from django.http import HttpResponseRedirect
from django.urls import reverse
from django.http import JsonResponse, HttpResponse
#Excel
from openpyxl import Workbook
import openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image
from openpyxl.styles import NamedStyle, Font, PatternFill
from openpyxl.utils import get_column_letter
from django.db.models.functions import Concat
from django.db.models import Value
from django.db.models import Sum
from django.db.models import Count
from django.db.models import IntegerField
from django.db.models.functions import Cast
from django.http import HttpResponseRedirect

from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.lib.pagesizes import letter,A4,landscape
import io
from reportlab.lib import colors
from reportlab.lib.colors import Color, black, blue, red, white
from reportlab.platypus import BaseDocTemplate, Frame, Paragraph, NextPageTemplate, PageBreak, PageTemplate,Table, SimpleDocTemplate,TableStyle, KeepInFrame, Spacer
import textwrap
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.utils import ImageReader
from django.http import FileResponse
from django.core.files.base import ContentFile

from decimal import Decimal 
import calendar
from esquema.models import BonoSolicitado
from django.db.models import Sum

from proyecto.models import Variables_imss_patronal
from proyecto.models import SalarioDatos

from .forms import PrenominaIncidenciasFormSet,IncidenciaRangoForm
import time

# Create your views here.

#funcion para obtener la catorcena actual
def obtener_catorcena():
    fecha_actual = datetime.date.today()
    catorcena_actual = Catorcenas.objects.filter(fecha_inicial__lte=fecha_actual, fecha_final__gte=fecha_actual).first()
    return catorcena_actual

def generar_prenomina(prenomina,catorcena,festivos):
        #crear_registro_prenomina(prenomina, prenomina.catorcena)
        economicos = Economicos_dia_tomado.objects.filter(prenomina__status=prenomina.empleado.status, fecha__range=(catorcena.fecha_inicial, catorcena.fecha_final))
        vacaciones = Vacaciones_dias_tomados.objects.filter(prenomina__status=prenomina.empleado.status,fecha_inicio__lte=catorcena.fecha_final,fecha_fin__gte=catorcena.fecha_inicial)

        incidencias = [] #es una lista de objectos que lleva los 14 dias con incidencias de la prenomina
        incidencia = 0 #se inicializa la incidencia
        fecha_incidencia = catorcena.fecha_inicial
        
        for i in range(1,15):
            fecha = fecha_incidencia
            incidencia = 16 #asistencia
            if fecha.weekday() == 6:  
                incidencia = 5 #domingo
            elif fecha in [festivo.dia_festivo for festivo in festivos]:
                incidencia = 13 #festivo:
            elif fecha in [economico.fecha for economico in economicos]:
                incidencia = 14 #economico            
            
            #Itera y saca la fecha de inicio y la fecha fin

            #se prepara el objecto PrenominaIncidencia para posterior ser almacenado en la BD
            crear_incidencia = PrenominaIncidencias(
                prenomina_id=prenomina.id,
                fecha=fecha,
                incidencia_id=incidencia  
            )
            
            #se inserta a la lista de las incidencias         
            incidencias.append(crear_incidencia)     
            #se agregar un dia para realizar el recorrido
            fecha_incidencia += timedelta(days=1)
            
        #se insertan todos los objecto en una sola consulta
        PrenominaIncidencias.objects.bulk_create(incidencias, batch_size=14)

def registrar_rango_incidencias(request,pk):    
    if request.method == 'POST':
        #catorcena
        catorcena_actual = obtener_catorcena()
        #RH
        user_filter = UserDatos.objects.get(user=request.user)
        #nombre = Perfil.objects.get(numero_de_trabajador = user_filter.numero_de_trabajador, distrito = user_filter.distrito)
        #Empleado
        costo = Costo.objects.get(id=pk)
        prenomina = Prenomina.objects.get(empleado=costo,catorcena = catorcena_actual.id)
        #Se trae el formulario de las incapacidades para ser validado
        incidencias_form = IncidenciaRangoForm(request.POST, request.FILES)
        #print(incidencias_form)
        if incidencias_form.is_valid():
            #validaciones a partir de la fecha de inicio y fecha fin
            fecha_start = incidencias_form.cleaned_data['fecha_inicio']
            fecha_end = incidencias_form.cleaned_data['fecha_fin']
            
            if fecha_start > fecha_end:
                 return JsonResponse({'poscondicion': 'La fecha de inicio debe ser menor a la fecha final'}, status=422)
            
            if fecha_start < catorcena_actual.fecha_inicial:
                return JsonResponse({'poscondicion': 'No puedes agregar una fecha anterior de la catorcena actual'}, status=422)
            
            #Busca si existe al menos una vacacion en el rango de fechas: inicio - fin
            vacaciones = Vacaciones_dias_tomados.objects.filter(
                prenomina__status=prenomina.empleado.status,
                fecha_inicio__lte=fecha_end,
                fecha_fin__gte=fecha_start
            ).values('id').exists()
            
            if vacaciones:
                return JsonResponse({'poscondicion': 'Ya existen vacaciones dentro del rango de fechas especificado'}, status=422)
                        
            festivos = TablaFestivos.objects.filter(dia_festivo__range=[fecha_start, fecha_end]).values('id').exists()  
            if festivos:
                return JsonResponse({'poscondicion': 'Ya existen dias festivos dentro del rango de fechas especificado'}, status=422) 
                
            economicos = Economicos_dia_tomado.objects.filter(fecha__range=[fecha_start, fecha_end]).values('id').exists()  
            if economicos:
                return JsonResponse({'poscondicion': 'Ya existen economicos dentro del rango de fechas especificado'}, status=422)
            
            incidencias = PrenominaIncidencias.objects.filter(fecha__range=[fecha_start, fecha_end]).values('id').exists()    
            if incidencias:
                return JsonResponse({'poscondicion': 'Ya existen incidencias dentro del rango de fechas especificado'}, status=422)
            
            
            #exit()
            #Cumple la validacion
            incidencia_rango = incidencias_form.save(commit=False)
            incidencia_rango.soporte = request.FILES['soporte'] 
            incidencia_rango.save()  
            
            fecha_actual = incidencia_rango.fecha_inicio#punto de inicio
            fecha_fin = min(incidencia_rango.fecha_fin, prenomina.catorcena.fecha_final)#toma la fecha mas chica entre las dos fechas
            
            while fecha_actual <= fecha_fin:
                incidencia = incidencia_rango.incidencia_id
                comentario = incidencia_rango.comentario
                soporte = incidencia_rango.soporte
                
                if fecha_actual.weekday() == (incidencia_rango.dia_inhabil_id - 1): 
                    #print(fecha_actual.weekday())
                    if (incidencia_rango.dia_inhabil_id - 1) == 6:# se resta 1 para obtener el dia domingo
                        incidencia = 5 #domingo
                        comentario = None
                        soporte = None
                    else:
                        incidencia = 2 #descanso
                        comentario = None
                        soporte = None
                        
                registro_prenomina, creado = PrenominaIncidencias.objects.update_or_create(
                prenomina=prenomina,
                    fecha=fecha_actual,
                    defaults={
                        'comentario': comentario, 
                        'soporte': soporte,
                        'incidencia_id': incidencia,
                        'incidencia_rango':incidencia_rango,                        
                    }
                )
                fecha_actual += timedelta(days=1)
                
                

            
            return JsonResponse({'success': 'Agregado correctamente'}, status=200)
        #error
        else:
            print("No pasa la validacion")
        
        """
        if incidencias_form.is_valid():    
            incidencia_rango = incidencias_form.save(commit=False)
            incidencia_rango.soporte = request.FILES['soporte'] 
            incidencia_rango.save()  
            
            incidencias = PrenominaIncidencias.objects.filter(prenomina = prenomina,fecha__range=[incidencia_rango.fecha_inicio,incidencia_rango.fecha_fin])
            
            for incidencia in incidencias:  
                incidencia.soporte =   incidencia_rango.soporte
                incidencia.incidencia = incidencia_rango.incidencia
                incidencia.comentario = incidencia_rango.comentario
                incidencia.incidencias_rango = incidencia_rango
                incidencia.save()
        """
        """
        context = {
            'incidencias_form':incidencias_form,
            'costo':costo,
            'prenomina':prenomina
        }  
        """  
    #return render(request, 'prenomina/Actualizar_revisar.html')
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
    #return redirect('prenomina/Actualizar_revisar.html',pk)

def actualizar_prenomina(prenominas,catorcena,festivos):
    print("aqui se debe actualizar la prenomina")
    cont = 0
    for prenomina in prenominas:
        cont = cont + 1
        print("contador ",cont)
        economicos = Economicos_dia_tomado.objects.filter(prenomina__status=prenomina.empleado.status, fecha__range=(catorcena.fecha_inicial, catorcena.fecha_final))
        vacaciones = Vacaciones_dias_tomados.objects.filter(prenomina__status=prenomina.empleado.status,fecha_inicio__lte=catorcena.fecha_final,fecha_fin__gte=catorcena.fecha_inicial)
        
        # Para los días económicos - registrados en el tiempo de la prenomina
        for economico in economicos:
            PrenominaIncidencias.objects.filter(
                prenomina=prenomina,
                fecha=economico.fecha
            ).update(
                comentario=None,
                soporte=None,
                incidencia_id=14 #economico
            )
                  
        # Para los días de vacaciones registrados en el tiempo de la prenomina
        for vacacion in vacaciones:
            # Buscar las incidencias correspondientes en la prenomina actual
            incidencias_vacaciones = PrenominaIncidencias.objects.filter(prenomina=prenomina,fecha__range=(vacacion.fecha_inicio, vacacion.fecha_fin))
            for incidencia_vacacion in incidencias_vacaciones:
                if incidencia_vacacion.incidencia_id == 16:#asistencia
                    incidencia_vacacion.incidencia_id = 15#vacacion
                elif incidencia_vacacion.fecha.weekday() == (int(vacacion.dia_inhabil_id) - 1):
                    if (int(vacacion.dia_inhabil_id) - 1) == 6:# se resta 1 para obtener el dia domingo
                        incidencia_vacacion.incidencia_id = 5 #domingo
                    else:
                        incidencia_vacacion.incidencia_id = 2 #descanso
                incidencia_vacacion.save()
        
        
       

@login_required(login_url='user-login')
def Tabla_prenomina(request):
    start_time = time.time()  # Registrar el tiempo de inicio
    user_filter = UserDatos.objects.get(user=request.user)
    revisar_perfil = Perfil.objects.get(distrito=user_filter.distrito,numero_de_trabajador=user_filter.numero_de_trabajador)
    empresa_faxton = Empresa.objects.get(empresa="Faxton")
    if user_filter.tipo.nombre == "RH":
        
        #llamar la fucion para obtener la catorcena actual
        catorcena_actual = obtener_catorcena()
        
        #para traer los empleados segun el filtro
        if revisar_perfil.empresa == empresa_faxton:
            costo = Costo.objects.filter(complete=True, status__perfil__baja=False,status__perfil__empresa=empresa_faxton).order_by("status__perfil__numero_de_trabajador")
        elif user_filter.distrito.distrito == 'Matriz':
            costo = Costo.objects.filter(complete=True, status__perfil__baja=False).order_by("status__perfil__numero_de_trabajador")
        else:
            costo = Costo.objects.filter(status__perfil__distrito=user_filter.distrito, complete=True,  status__perfil__baja=False).order_by("status__perfil__numero_de_trabajador")

        prenominas = Prenomina.objects.filter(empleado__in=costo,catorcena=catorcena_actual.id).order_by("empleado__status__perfil__numero_de_trabajador")
        #prenominas = Prenomina.objects.filter(empleado__in=costo,catorcena = catorcena_actual.id).order_by("empleado__status__perfil__numero_de_trabajador").prefetch_related('incidencias')
        festivos = TablaFestivos.objects.filter(dia_festivo__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final))
        #crear las prenominas actuales si es que ya es nueva catorcena
        nuevas_prenominas = []
        for empleado in costo:
            #checar si existe prenomina para el empleado en la catorcena actual
            prenomina_existente = prenominas.filter(empleado=empleado).exists()
            #si no existe crear una nueva prenomina
            if not prenomina_existente:
                nueva_prenomina = Prenomina(empleado=empleado, catorcena=catorcena_actual)
                nuevas_prenominas.append(nueva_prenomina)
                #generar_prenomina(nueva_prenomina,catorcena_actual,festivos)
            #else:
                #actualizar_prenomina(prenominas,catorcena_actual,festivos)  
        if nuevas_prenominas:
            Prenomina.objects.bulk_create(nuevas_prenominas)              
        #costo_filter = CostoFilter(request.GET, queryset=costo)
        #costo = costo_filter.qs
        #prenominas = Prenomina.objects.filter(empleado__in=costo,catorcena = catorcena_actual.id).order_by("empleado__status__perfil__numero_de_trabajador").prefetch_related('incidencias')
        
        prenomina_filter = PrenominaFilter(request.GET, queryset=prenominas)
        prenominas = prenomina_filter.qs

        #para verificar las autotizaciones
        for prenomina in prenominas:
            ultima_autorizacion = AutorizarPrenomina.objects.filter(prenomina=prenomina).order_by('-updated_at').first() #Ultimo modificado

            if ultima_autorizacion is not None:
                prenomina.valor = ultima_autorizacion.estado.tipo #Esta bien como agarra el dato de RH arriba que es el primero
            prenomina.estado_general = determinar_estado_general(request,ultima_autorizacion)

        if request.method =='POST' and 'Autorizar' in request.POST:
            if user_filter.tipo.nombre ==  "RH":
                prenominas_filtradas = [prenom for prenom in prenominas if prenom.estado_general == 'RH pendiente (rechazado por Controles técnicos)' or prenom.estado_general == 'RH pendiente (rechazado por Gerencia)' or prenom.estado_general == 'Sin autorizaciones']
                if prenominas_filtradas:
                    # Llamar a la función Autorizar_gerencia con las prenominas filtradas
                    return Autorizar_general(request,prenominas_filtradas, user_filter,catorcena_actual)
                else:
                    # Si no hay prenominas que cumplan la condición, manejar según sea necesario
                    messages.error(request,'Ya se han autorizado todas las prenominas pendientes')
        
        if request.method =='POST' and 'Excel' in request.POST:
            return Excel_estado_prenomina(request,prenominas, user_filter)
        if request.method =='POST' and 'Excel2' in request.POST:
            return Excel_estado_prenomina_formato(request,prenominas, user_filter)
        
        p = Paginator(prenominas, 50)
        page = request.GET.get('page')
        salidas_list = p.get_page(page)

        context = {
            'prenomina_filter':prenomina_filter,
            'salidas_list': salidas_list,
            'prenominas':prenominas
        }
        end_time = time.time()  # Registrar el tiempo de finalización
        print(f"Tiempo total de carga de la página: {end_time - start_time} segundos")
        return render(request, 'prenomina/Tabla_prenomina.html', context)
    else:
        return render(request, 'revisar/403.html')

@login_required(login_url='user-login')
def Autorizar_general(request,prenominas, user_filter, catorcena_actual):
    if request.user.is_authenticated:
        nombre = Perfil.objects.get(numero_de_trabajador=user_filter.numero_de_trabajador, distrito=user_filter.distrito)
        aprobado = Estado.objects.get(tipo="aprobado")
        fechas_domingo = [fecha for fecha in (catorcena_actual.fecha_inicial + timedelta(days=d) for d in range((catorcena_actual.fecha_final - catorcena_actual.fecha_inicial).days + 1)) if fecha.weekday() == 6]
        for prenomina in prenominas:
            prenomina.domingo = prenomina.domingo_set.filter(fecha__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final))
            prenomina.descanso = prenomina.descanso_set.filter(fecha__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final))
            revisado, created = AutorizarPrenomina.objects.get_or_create(prenomina=prenomina, tipo_perfil=user_filter.tipo) #Checa si existe autorización de su perfil y si no lo crea 
            revisado.estado = Estado.objects.get(tipo="aprobado")
            nombre = Perfil.objects.get(numero_de_trabajador=user_filter.numero_de_trabajador, distrito=user_filter.distrito)
            revisado.perfil = nombre
            revisado.comentario = 'Aprobación general'
            revisado.save()
            #Añadir dias domingo o descanso
            if prenomina.domingo.count() == 0 and prenomina.descanso.count() == 0:
                for fecha in fechas_domingo:
                    descanso, created = Domingo.objects.get_or_create(prenomina=prenomina,fecha = fecha, comentario = "generado automaticamente", editado = "sistema")
                    descanso.save()
        messages.success(request, 'Prenominas pendientes autorizadas automaticamente')
        return redirect('Prenomina')  # Cambia 'ruta_a_redirigir' por la URL a la que deseas redirigir después de autorizar las prenóminas

@login_required(login_url='user-login')
def capturarIncidencias(request, incidencia,fecha_incio,fecha_fin,dia_inhabil,prenomina,comentario,nombre,url):
            if incidencia == '2':
                evento_model = Castigos
            elif incidencia == '3':
                evento_model = Permiso_goce
            elif incidencia == '4':
                evento_model = Permiso_sin
            dias_semana = {
                            'Lunes': 0,
                            'Martes': 1,
                            'Miércoles': 2,
                            'Jueves': 3,
                            'Viernes': 4,
                            'Sábado': 5,
                            'Domingo': 6
                        }
            if evento_model:
                if url:
                    obj, created = evento_model.objects.update_or_create(fecha=fecha_incio, fecha_fin=fecha_fin, dia_inhabil = dia_inhabil,prenomina=prenomina, defaults={'comentario': comentario, 'editado': f"E:{nombre.nombres} {nombre.apellidos}"})
                    obj.url = url
                    obj.save()
                    # Obtener el día de la semana del día inhabil
                    dia_semana = dia_inhabil.nombre
                    dia_semana = dias_semana[dia_semana]

                    # Definir el tipo de modelo basado en el día de la semana
                    if dia_semana == 6:  # Domingo
                        descanso_model = Domingo
                    else:
                        descanso_model = Descanso
                    # Iterar sobre cada día correspondiente al día de la semana dentro del rango de fechas
                    for fecha_iter in (fecha_incio + timedelta(days=d) for d in range((fecha_fin - fecha_incio).days + 1)):
                        if fecha_iter.weekday() == dia_semana:
                            # Crear el objeto correspondiente
                            descanso, created = descanso_model.objects.get_or_create(prenomina=prenomina, fecha=fecha_iter, defaults={'comentario': "generado automaticamente", 'editado': "sistema"})
                            descanso.save()
                else:
                    evento_model.objects.update_or_create(fecha=fecha_incio, fecha_fin=fecha_fin, dia_inhabil = dia_inhabil, prenomina=prenomina, defaults={'comentario': comentario, 'editado': f"E:{nombre.nombres} {nombre.apellidos}"})
                    # Obtener el día de la semana del día inhabil
                    dia_semana = dia_inhabil.nombre
                    dia_semana = dias_semana[dia_semana]

                    # Definir el tipo de modelo basado en el día de la semana
                    if dia_semana == 6:  # Domingo
                        descanso_model = Domingo
                    else:
                        descanso_model = Descanso
                    # Iterar sobre cada día correspondiente al día de la semana dentro del rango de fechas
                    for fecha_iter in (fecha_incio + timedelta(days=d) for d in range((fecha_fin - fecha_incio).days + 1)):
                        if fecha_iter.weekday() == dia_semana:
                            # Crear el objeto correspondiente
                            descanso, created = descanso_model.objects.get_or_create(prenomina=prenomina, fecha=fecha_iter, defaults={'comentario': "generado automaticamente", 'editado': "sistema"})
                            descanso.save()
            else:
                #  donde nuevo_estado no tiene un mapeo en el diccionario
                print(f"Error: nuevo_estado desconocido")
            
            messages.success(request, 'Cambios guardados exitosamente')
            
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

@login_required(login_url='user-login')
def capturarIncapacidades(request, tipo,subsecuente,fecha_incio,fecha_fin,dia_inhabil,prenomina,comentario,url,nombre,incapacidades,):
    dias_semana = {
                    'Lunes': 0,
                    'Martes': 1,
                    'Miércoles': 2,
                    'Jueves': 3,
                    'Viernes': 4,
                    'Sábado': 5,
                    'Domingo': 6
                }
    if incapacidades.exists() and subsecuente == True:
        incapacidad = Incapacidades.objects.get(prenomina__empleado_id=prenomina.empleado.id,fecha__lte=fecha_fin,fecha_fin__gte=fecha_incio,)
        fecha_subsecuente = incapacidad.fecha_fin + timedelta(days=1)
        if url:
            obj, created = Incapacidades.objects.update_or_create(fecha=fecha_subsecuente, fecha_fin=fecha_fin, prenomina=prenomina, subsecuente=True, defaults={'comentario': comentario, 'editado': f"E:{nombre.nombres} {nombre.apellidos}",})
            obj.tipo = tipo
            obj.subsecuente = subsecuente
            obj.dia_inhabil = dia_inhabil
            obj.url = url
            obj.save()
            # Obtener el día de la semana del día inhabil
            dia_semana = dia_inhabil.nombre
            dia_semana = dias_semana[dia_semana]

            # Definir el tipo de modelo basado en el día de la semana
            if dia_semana == 6:  # Domingo
                descanso_model = Domingo
            else:
                descanso_model = Descanso
            # Iterar sobre cada día correspondiente al día de la semana dentro del rango de fechas
            for fecha_iter in (fecha_incio + timedelta(days=d) for d in range((fecha_fin - fecha_incio).days + 1)):
                if fecha_iter.weekday() == dia_semana:
                    # Crear el objeto correspondiente
                    descanso, created = descanso_model.objects.get_or_create(prenomina=prenomina, fecha=fecha_iter, defaults={'comentario': "generado automaticamente", 'editado': "sistema"})
                    descanso.save()
        else:
            Incapacidades.objects.update_or_create(fecha=fecha_subsecuente, fecha_fin=fecha_fin, dia_inhabil = dia_inhabil, prenomina=prenomina, subsecuente=True, defaults={'comentario': comentario, 'editado': f"E:{nombre.nombres} {nombre.apellidos}"})
            # Obtener el día de la semana del día inhabil
            dia_semana = dia_inhabil.nombre
            dia_semana = dias_semana[dia_semana]

            # Definir el tipo de modelo basado en el día de la semana
            if dia_semana == 6:  # Domingo
                descanso_model = Domingo
            else:
                descanso_model = Descanso
            # Iterar sobre cada día correspondiente al día de la semana dentro del rango de fechas
            for fecha_iter in (fecha_incio + timedelta(days=d) for d in range((fecha_fin - fecha_incio).days + 1)):
                if fecha_iter.weekday() == dia_semana:
                    # Crear el objeto correspondiente
                    descanso, created = descanso_model.objects.get_or_create(prenomina=prenomina, fecha=fecha_iter, defaults={'comentario': "generado automaticamente", 'editado': "sistema"})
                    descanso.save()
        messages.success(request, 'Se extendio la incapacidad')    
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
        
    else:
        if url:
            obj, created = Incapacidades.objects.update_or_create(fecha=fecha_incio, fecha_fin=fecha_fin, dia_inhabil = dia_inhabil, prenomina=prenomina, defaults={'comentario': comentario, 'editado': f"E:{nombre.nombres} {nombre.apellidos}"})
            obj.tipo = tipo
            obj.subsecuente = subsecuente
            obj.url = url
            obj.save()
            # Obtener el día de la semana del día inhabil
            dia_semana = dia_inhabil.nombre
            dia_semana = dias_semana[dia_semana]

            # Definir el tipo de modelo basado en el día de la semana
            if dia_semana == 6:  # Domingo
                descanso_model = Domingo
            else:
                descanso_model = Descanso
            # Iterar sobre cada día correspondiente al día de la semana dentro del rango de fechas
            for fecha_iter in (fecha_incio + timedelta(days=d) for d in range((fecha_fin - fecha_incio).days + 1)):
                if fecha_iter.weekday() == dia_semana:
                    # Crear el objeto correspondiente
                    descanso, created = descanso_model.objects.get_or_create(prenomina=prenomina, fecha=fecha_iter, defaults={'comentario': "generado automaticamente", 'editado': "sistema"})
                    descanso.save()
        else:
            Incapacidades.objects.update_or_create(fecha=fecha_incio, fecha_fin=fecha_fin, dia_inhabil = dia_inhabil, prenomina=prenomina, defaults={'comentario': comentario, 'editado': f"E:{nombre.nombres} {nombre.apellidos}"})
            # Obtener el día de la semana del día inhabil
            dia_semana = dia_inhabil.nombre
            dia_semana = dias_semana[dia_semana]

            # Definir el tipo de modelo basado en el día de la semana
            if dia_semana == 6:  # Domingo
                descanso_model = Domingo
            else:
                descanso_model = Descanso
            # Iterar sobre cada día correspondiente al día de la semana dentro del rango de fechas
            for fecha_iter in (fecha_incio + timedelta(days=d) for d in range((fecha_fin - fecha_incio).days + 1)):
                if fecha_iter.weekday() == dia_semana:
                    # Crear el objeto correspondiente
                    descanso, created = descanso_model.objects.get_or_create(prenomina=prenomina, fecha=fecha_iter, defaults={'comentario': "generado automaticamente", 'editado': "sistema"})
                    descanso.save()
        messages.success(request, 'Se guardo correctamente')    
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))


@login_required(login_url='user-login')
def programar_incidencias(request,pk):
    # crea el nuevo dato según el nuevo estado o comentario
    if request.method == 'POST' and 'btn_incidencias' in request.POST:
        
        #saber catorcena
        ahora = datetime.date.today()
        #ahora = datetime.date.today() + timedelta(days=10)
        catorcena_actual = Catorcenas.objects.filter(fecha_inicial__lte=ahora, fecha_final__gte=ahora).first()
        
        #RH
        user_filter = UserDatos.objects.get(user=request.user)
        nombre = Perfil.objects.get(numero_de_trabajador = user_filter.numero_de_trabajador, distrito = user_filter.distrito)
        
        #Empleado
        costo = Costo.objects.get(id=pk)
        prenomina = Prenomina.objects.get(empleado=costo,fecha__range=[catorcena_actual.fecha_inicial, catorcena_actual.fecha_final])
        
        #Se trae el formulario de las incapacidades para ser validado
        form_incidencias = IncapacidadesForm(request.POST, request.FILES)
        
        if form_incidencias.is_valid():
            incidencia = form_incidencias.cleaned_data.get('incidencias')
            fecha_incio = form_incidencias.cleaned_data['fecha']
            fecha_fin = form_incidencias.cleaned_data['fecha_fin']
            dia_inhabil = form_incidencias.cleaned_data['dia_inhabil']
            comentario = form_incidencias.cleaned_data['comentario']    
            url = form_incidencias.cleaned_data['url']       
            
            #VALIDACIONES
            if fecha_incio > fecha_fin:
                print("La fecha de inicio es posterior a la fecha final.")
                messages.error(request, 'La fecha de inicio debe ser menor a la fecha final')
                return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
            
            if not incidencia:
                messages.error(request, 'Debes seleccionar una incidencia')
                return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
            
            if fecha_incio < catorcena_actual.fecha_inicial:
                messages.error(request, 'No puedes agregar una fecha anterior de la catorcena actual')
                return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
            
            vacaciones = Vacaciones_dias_tomados.objects.filter(Q(prenomina__status=prenomina.empleado.status, fecha_inicio__range=[fecha_incio, fecha_fin]) | Q(prenomina__status=prenomina.empleado.status, fecha_fin__range=[fecha_incio, fecha_fin]))
            if vacaciones.exists():
                messages.error(request, 'Ya existen vacaciones dentro del rango de fechas especificado')
                return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
            
            economicos = Economicos_dia_tomado.objects.filter(prenomina__status=prenomina.empleado.status, fecha__range=[fecha_incio, fecha_fin])
            if economicos.exists():            
                messages.error(request, 'Ya existen economicos dentro del rango de fechas especificado')
                return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
            
            festivos = TablaFestivos.objects.filter(dia_festivo__range=[fecha_incio, fecha_fin])
            if festivos.exists():
                messages.error(request, 'Ya existen festivos dentro del rango de fechas especificado')
                return HttpResponseRedirect(request.META.get('HTTP_REFERER'))        
            
            
            capturarIncidencias(request, incidencia,fecha_incio,fecha_fin,dia_inhabil,prenomina,comentario,nombre,url)
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
        
        else:
            for field, errors in form_incidencias.errors.items():
                for error in errors:
                    messages.error(request,error)
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))


@login_required(login_url='user-login')
def programar_incidencias(request,pk):
    # crea el nuevo dato según el nuevo estado o comentario
    if request.method == 'POST' and 'btn_incidencias' in request.POST:
        
        #saber catorcena
        ahora = datetime.date.today()
        #ahora = datetime.date.today() + timedelta(days=10)
        catorcena_actual = Catorcenas.objects.filter(fecha_inicial__lte=ahora, fecha_final__gte=ahora).first()
        
        #RH
        user_filter = UserDatos.objects.get(user=request.user)
        nombre = Perfil.objects.get(numero_de_trabajador = user_filter.numero_de_trabajador, distrito = user_filter.distrito)
        
        #Empleado
        costo = Costo.objects.get(id=pk)
        prenomina = Prenomina.objects.get(empleado=costo,fecha__range=[catorcena_actual.fecha_inicial, catorcena_actual.fecha_final])
        
        #Se trae el formulario de las incapacidades para ser validado
        form_incidencias = IncapacidadesForm(request.POST, request.FILES)
        
        if form_incidencias.is_valid():
            incidencia = form_incidencias.cleaned_data.get('incidencias')
            fecha_incio = form_incidencias.cleaned_data['fecha']
            fecha_fin = form_incidencias.cleaned_data['fecha_fin']
            dia_inhabil = form_incidencias.cleaned_data['dia_inhabil']
            comentario = form_incidencias.cleaned_data['comentario']    
            url = form_incidencias.cleaned_data['url']       
            
            #VALIDACIONES
            if fecha_incio > fecha_fin:
                print("La fecha de inicio es posterior a la fecha final.")
                messages.error(request, 'La fecha de inicio debe ser menor a la fecha final')
                return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
            
            if not incidencia:
                messages.error(request, 'Debes seleccionar una incidencia')
                return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
            
            if fecha_incio < catorcena_actual.fecha_inicial:
                messages.error(request, 'No puedes agregar una fecha anterior de la catorcena actual')
                return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
            
            vacaciones = Vacaciones_dias_tomados.objects.filter(Q(prenomina__status=prenomina.empleado.status, fecha_inicio__range=[fecha_incio, fecha_fin]) | Q(prenomina__status=prenomina.empleado.status, fecha_fin__range=[fecha_incio, fecha_fin]))
            if vacaciones.exists():
                messages.error(request, 'Ya existen vacaciones dentro del rango de fechas especificado')
                return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
            
            economicos = Economicos_dia_tomado.objects.filter(prenomina__status=prenomina.empleado.status, fecha__range=[fecha_incio, fecha_fin])
            if economicos.exists():            
                messages.error(request, 'Ya existen economicos dentro del rango de fechas especificado')
                return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
            
            festivos = TablaFestivos.objects.filter(dia_festivo__range=[fecha_incio, fecha_fin])
            if festivos.exists():
                messages.error(request, 'Ya existen festivos dentro del rango de fechas especificado')
                return HttpResponseRedirect(request.META.get('HTTP_REFERER'))        
            
            #VALIDAR INCIDENCIAS - EDITAR UNA INCIDENCIA QUE ESTE DENTRO DEL RANGO DE LA CAT Y VERIFICAR QUE EXISTA EN LA CAT ANTERIOIR
            incapacidades = Incapacidades.objects.filter(
                prenomina__empleado_id=prenomina.empleado.id,
                fecha__lte=fecha_fin,  # La fecha de inicio de la incapacidad debe ser menor o igual a la fecha fin del rango proporcionado
                fecha_fin__gte=fecha_incio,   # La fecha fin de la incapacidad debe ser mayor o igual a la fecha inicio del rango proporcionado
            )
            
            castigos = Castigos.objects.filter(
                prenomina__empleado_id=prenomina.empleado.id,
                fecha__lte=fecha_fin,
                fecha_fin__gte=fecha_incio,
            )
            
            permisos_goce = Permiso_goce.objects.filter(
                prenomina__empleado_id=prenomina.empleado.id,
                fecha__lte=fecha_fin,
                fecha_fin__gte=fecha_incio,
            )
            
            permisos_sin = Permiso_sin.objects.filter(
                prenomina__empleado_id=prenomina.empleado.id,
                fecha__lte=fecha_fin,
                fecha_fin__gte=fecha_incio,
            )
            
            #elimina una incapacidad y es remplazada por otra en caso que sea de la misma catorcena
            if incapacidades.exists():
                if incapacidades.filter(fecha__lt=catorcena_actual.fecha_inicial).exists():
                    messages.error(request, 'Ya existen incapacidades de la catorcena anterior')
                    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))   
                else:
                    #se elimina el soporte asociado
                    soporte = incapacidades.first()
                    os.remove(soporte.url.path)
                    #se elima la incapacidad de la BD
                    incapacidades.delete()
                    
            if castigos.exists():
                if castigos.filter(fecha__lt=catorcena_actual.fecha_inicial).exists():
                    messages.error(request, 'Ya existen castigos de la catorcena anterior')
                    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))   
                else:
                    # Se elimina el soporte asociado
                    soporte = castigos.first()
                    os.remove(soporte.url.path)
                    # Se elimina el castigo de la BD
                    castigos.delete()
            
            if permisos_goce.exists():
                if permisos_goce.filter(fecha__lt=catorcena_actual.fecha_inicial).exists():
                    messages.error(request, 'Ya existen permisos de goce de sueldo de la catorcena anterior')
                    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))   
                else:
                    # Se elimina el soporte asociado
                    soporte = permisos_goce.first()
                    os.remove(soporte.url.path)
                    # Se elimina el permiso de goce de la BD
                    permisos_goce.delete()
            
            if permisos_sin.exists():
                if permisos_sin.filter(fecha__lt=catorcena_actual.fecha_inicial).exists():
                    messages.error(request, 'Ya existen permisos sin goce de sueldo de la catorcena anterior')
                    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))   
                else:
                    # Se elimina el soporte asociado
                    soporte = permisos_sin.first()
                    os.remove(soporte.url.path)
                    # Se elimina el permiso sin goce de la BD
                    permisos_sin.delete()        
            
            capturarIncidencias(request, incidencia,fecha_incio,fecha_fin,dia_inhabil,prenomina,comentario,nombre,url)
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
        
        else:
            for field, errors in form_incidencias.errors.items():
                for error in errors:
                    messages.error(request,error)
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
        
@login_required(login_url='user-login')
def programar_incapacidades(request,pk):
    # crea el nuevo dato según el nuevo estado o comentario
    if request.method == 'POST' and 'btn_incapacidades' in request.POST:
        
        #saber catorcena
        ahora = datetime.date.today()
        #ahora = datetime.date.today() + timedelta(days=10)
        catorcena_actual = Catorcenas.objects.filter(fecha_inicial__lte=ahora, fecha_final__gte=ahora).first()
        
        #RH
        user_filter = UserDatos.objects.get(user=request.user)
        nombre = Perfil.objects.get(numero_de_trabajador = user_filter.numero_de_trabajador, distrito = user_filter.distrito)
        
        #Empleado
        costo = Costo.objects.get(id=pk)
        prenomina = Prenomina.objects.get(empleado=costo,fecha__range=[catorcena_actual.fecha_inicial, catorcena_actual.fecha_final])
        
        #Se trae el formulario de las incapacidades para ser validado
        form_incapacidades = IncapacidadesTipoForm(request.POST, request.FILES)
        
        if form_incapacidades.is_valid():
            #'tipo','subsecuente','fecha','fecha_fin','comentario','url'
            tipo = form_incapacidades.cleaned_data['tipo']
            subsecuente = form_incapacidades.cleaned_data['subsecuente']
            fecha_incio = form_incapacidades.cleaned_data['fecha']
            fecha_fin = form_incapacidades.cleaned_data['fecha_fin']
            dia_inhabil = form_incapacidades.cleaned_data['dia_inhabil']
            comentario = form_incapacidades.cleaned_data['comentario']
            url = form_incapacidades.cleaned_data['url']       
                        
            #VALIDACIONES
            if fecha_incio > fecha_fin:
                print("La fecha de inicio es posterior a la fecha final.")
                messages.error(request, 'La fecha de inicio debe ser menor a la fecha final')
                return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
                        
            if fecha_incio < catorcena_actual.fecha_inicial:
                messages.error(request, 'No puedes agregar una fecha anterior de la catorcena actual')
                return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
            
            vacaciones = Vacaciones_dias_tomados.objects.filter(Q(prenomina__status=prenomina.empleado.status, fecha_inicio__range=[fecha_incio, fecha_fin]) | Q(prenomina__status=prenomina.empleado.status, fecha_fin__range=[fecha_incio, fecha_fin]))
            if vacaciones.exists():
                messages.error(request, 'Ya existen vacaciones dentro del rango de fechas especificado')
                return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
            
            economicos = Economicos_dia_tomado.objects.filter(prenomina__status=prenomina.empleado.status, fecha__range=[fecha_incio, fecha_fin])
            if economicos.exists():            
                messages.error(request, 'Ya existen economicos dentro del rango de fechas especificado')
                return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
            
            festivos = TablaFestivos.objects.filter(dia_festivo__range=[fecha_incio, fecha_fin])
            if festivos.exists():
                messages.error(request, 'Ya existen festivos dentro del rango de fechas especificado')
                return HttpResponseRedirect(request.META.get('HTTP_REFERER'))        
            
            #VALIDAR INCIDENCIAS - EDITAR UNA INCIDENCIA QUE ESTE DENTRO DEL RANGO DE LA CAT Y VERIFICAR QUE EXISTA EN LA CAT ANTERIOIR
            incapacidades = Incapacidades.objects.filter(
                prenomina__empleado_id=prenomina.empleado.id,
                fecha__lte=fecha_fin,  # La fecha de inicio de la incapacidad debe ser menor o igual a la fecha fin del rango proporcionado
                fecha_fin__gte=fecha_incio,   # La fecha fin de la incapacidad debe ser mayor o igual a la fecha inicio del rango proporcionado
            )
            
            castigos = Castigos.objects.filter(
                prenomina__empleado_id=prenomina.empleado.id,
                fecha__lte=fecha_fin,
                fecha_fin__gte=fecha_incio,
            )
            
            permisos_goce = Permiso_goce.objects.filter(
                prenomina__empleado_id=prenomina.empleado.id,
                fecha__lte=fecha_fin,
                fecha_fin__gte=fecha_incio,
            )
            
            permisos_sin = Permiso_sin.objects.filter(
                prenomina__empleado_id=prenomina.empleado.id,
                fecha__lte=fecha_fin,
                fecha_fin__gte=fecha_incio,
            )
            
            #elimina una incapacidad y es remplazada por otra en caso que sea de la misma catorcena
            if incapacidades.exists() and subsecuente is False:
                if incapacidades.filter(fecha__lt=catorcena_actual.fecha_inicial).exists():
                    messages.error(request, 'Ya existen incapacidades de la catorcena anterior')
                    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))   
                else:
                    #Traer la ultima incapacidad
                    ultima_incapacidad = Incapacidades.objects.filter(prenomina__empleado_id=prenomina.empleado.id,fecha_fin__gte=fecha_fin,).last()
                    if ultima_incapacidad is not None:
                        #Traer incidencias 
                        eliminar_subsecuentes = Incapacidades.objects.filter(prenomina__empleado_id=prenomina.empleado.id,fecha__lte=ultima_incapacidad.fecha_fin,fecha_fin__gte=fecha_incio,)
                        #se elimina el soporte asociado
                        soporte = eliminar_subsecuentes.first()
                        os.remove(soporte.url.path)
                        #se elima la incapacidad de la BD
                        eliminar_subsecuentes.delete()
                    #se elimina el soporte asociado
                    #soporte = incapacidades.first()
                    #os.remove(soporte.url.path)
                    #se elima la incapacidad de la BD
                    #incapacidades.delete()
                    
            if castigos.exists():
                if castigos.filter(fecha__lt=catorcena_actual.fecha_inicial).exists():
                    messages.error(request, 'Ya existen castigos de la catorcena anterior')
                    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))   
                else:
                    # Se elimina el soporte asociado
                    soporte = castigos.first()
                    os.remove(soporte.url.path)
                    # Se elimina el castigo de la BD
                    castigos.delete()
            
            if permisos_goce.exists():
                if permisos_goce.filter(fecha__lt=catorcena_actual.fecha_inicial).exists():
                    messages.error(request, 'Ya existen permisos de goce de sueldo de la catorcena anterior')
                    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))   
                else:
                    # Se elimina el soporte asociado
                    soporte = permisos_goce.first()
                    os.remove(soporte.url.path)
                    # Se elimina el permiso de goce de la BD
                    permisos_goce.delete()
            
            if permisos_sin.exists():
                if permisos_sin.filter(fecha__lt=catorcena_actual.fecha_inicial).exists():
                    messages.error(request, 'Ya existen permisos sin goce de sueldo de la catorcena anterior')
                    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))   
                else:
                    # Se elimina el soporte asociado
                    soporte = permisos_sin.first()
                    os.remove(soporte.url.path)
                    # Se elimina el permiso sin goce de la BD
                    permisos_sin.delete()        
            capturarIncapacidades(request, tipo,subsecuente,fecha_incio,fecha_fin,dia_inhabil,prenomina,comentario,url,nombre, incapacidades)
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
        
        else:
            for field, errors in form_incapacidades.errors.items():
                for error in errors:
                    messages.error(request,error)
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))        
    
@login_required(login_url='user-login')
def crear_rango_incidencias_dos(request,pk):
    # crea el nuevo dato según el nuevo estado o comentario
    if request.method == 'POST' and 'btn_incidencias' in request.POST:        
        #catorcena
        catorcena_actual = obtener_catorcena()
        #RH
        user_filter = UserDatos.objects.get(user=request.user)
        nombre = Perfil.objects.get(numero_de_trabajador = user_filter.numero_de_trabajador, distrito = user_filter.distrito)
        #Empleado
        costo = Costo.objects.get(id=pk)
        prenomina = Prenomina.objects.get(empleado=costo,catorcena = catorcena_actual.id)
        
        #Se trae el formulario de las incapacidades para ser validado
        incidencias_form = IncidenciasRangoForm(request.POST, request.FILES)
        
        if incidencias_form.is_valid():    
            incidencia_rango = incidencias_form.save(commit=False)  # Guarda el formulario pero no en la base de datos aún
            if incidencia_rango.fecha_inicio > incidencia_rango.fecha_fin:
                return JsonResponse({'poscondicion': 'La fecha de inicio debe ser menor a la fecha final'}, status=422)
            
            
            
            incidencia_rango.soporte = request.FILES['soporte']  # Asigna el archivo adjunto al campo 'soporte' de la incidencia
            print("es la foto ", incidencia_rango.soporte)
            incidencia_rango.save()  
            
            incidencias = PrenominaIncidencias.objects.filter(prenomina = prenomina,fecha__range=[incidencia_rango.fecha_inicio,incidencia_rango.fecha_fin])
            for incidencia in incidencias:
              
                incidencia.soporte =   incidencia_rango.soporte
                incidencia.incidencia = incidencia_rango.incidencia
                incidencia.comentario = incidencia_rango.comentario
                incidencia.incidencias_rango = incidencia_rango
                incidencia.save()
                
            print(incidencia_rango.id)
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
        else:
            for field, errors in incidencias_form.errors.items():
                for error in errors:
                    messages.error(request,error)
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
def crear_registro_prenomina(prenomina,catorcena):
    # Lista de las incidencias de la prenomina a crear
    incidencias_prenomina = []
    
    # Crear las 14 incidencias por defecto considerando domingo
    for i in range(14):
        objeto = PrenominaIncidencias(
            prenomina=prenomina,
            fecha=catorcena.fecha_inicial + timedelta(days=i),
            comentario=None,
            soporte=None,
            # 5 = domingo, 16 = asistencia, 6 y 13 == dia semana domingo
            incidencia_id = 5 if i == 6 or i == 13 else 16
        )
        incidencias_prenomina.append(objeto)
    
    PrenominaIncidencias.objects.bulk_create(incidencias_prenomina)
 
def crear_formsets():
    from django.forms import formset_factory
    from .forms import PrenominaIncidenciasForm
    # Crear el formset factory
    FormSet = formset_factory(PrenominaIncidenciasForm, extra=14)
    
    # Definir la fecha inicial
    fecha_inicial = date(2024, 6, 1)  # Fecha inicial

    # Lista para almacenar los formsets
    formsets = []

    # Generar los 14 formsets con fechas iniciales consecutivas
    for i in range(14):
        nueva_fecha = fecha_inicial + timedelta(days=i)
        formset = FormSet(initial=[{'fecha': nueva_fecha.strftime('%Y-%m-%d')}])
        formsets.append(formset)

    return formsets
    
@login_required(login_url='user-login')
def PrenominaRevisar(request, pk):
    user_filter = UserDatos.objects.get(user=request.user)
    if user_filter.tipo.id == 4: #Perfil RH
        start_time = time.time()  # Registrar el tiempo de inicio
        #llamar la fucion para obtener la catorcena actual
        catorcena_actual = obtener_catorcena()
        
        #obtener el empleado respecto a su prenomina     
        costo = Costo.objects.get(id=pk)
        prenomina = Prenomina.objects.get(empleado=costo,catorcena = catorcena_actual.id)
        
        #flujo de las autorizaciones
        autorizacion1 = prenomina.autorizarprenomina_set.filter(tipo_perfil__nombre="Control Tecnico").first()
        autorizacion2 = prenomina.autorizarprenomina_set.filter(tipo_perfil__nombre="Gerencia").first()
        
        #obtener la instancia de los formularios
        
        #Para guardar los datos en la prenomina
        if request.method == 'POST' and 'guardar_cambios' in request.POST:
            prenomina_form = PrenominaIncidenciasFormSet(request.POST)
            #print(formset)
            if prenomina_form.is_valid():
                for form in prenomina_form:
                    if form.cleaned_data.get('DELETE'):
                        # Si el formulario está marcado para eliminación, eliminar el registro
                        if form.cleaned_data['id']:
                            #ELIMINAR LO QUE SON CREADOS EN LA PRENOMINA, NO ELIMINAR SI SON DE PRENOMANIA ANTERIOR Verifica si el registro es del antes de la cat, si esta dentro o igual a la cat actual
                            print("")
                            #print("es es el id pero de id registro", form.cleaned_data['id'])
                            #print("es es el id pero del rango", form.cleaned_data['incidencia'].id)
                            
                            #PrenominaIncidencias.objects.filter(incidencia_rango_id=form.cleaned_data['id']).delete()
                            #IncidenciaRango.objects.filter(pk=2).delete()
                            #PrenominaIncidencias.objects.filter(id=form.cleaned_data['id']).delete()
                           
                    else:
                        #se extren los datos 
                        asistencia = form.cleaned_data['incidencia']
                        if asistencia.id != 16:
                            fecha = form.cleaned_data['fecha']
                            comentario = form.cleaned_data['comentario']
                            incidencia = form.cleaned_data['incidencia']
                            
                            registro_prenomina, creado = PrenominaIncidencias.objects.update_or_create(
                                prenomina=prenomina,
                                fecha=fecha,
                                defaults={
                                    'comentario': comentario, 
                                    'incidencia': incidencia
                                }
                            )
                    
            return redirect('Prenomina')
            #se obtienen todos los formularios - formset
            #prenomina_form= FormSet(request.POST)
            #print(prenomina_form)
            print('asda')
            #se obtienen todos los formularios - formset
            #prenomina_form= PrenominaIncidenciasFormSet(request.POST, request.FILES)
            #validaciones
            #if prenomina_form.is_valid():
                #se recorre cada formulario con sus respectivos datos
            #    for form in prenomina_form:
                    #se extren los datos 
            #        fecha = form.cleaned_data['fecha']
            #        comentario = form.cleaned_data['comentario']
            #        incidencia = form.cleaned_data['incidencia']
                    #soporte = form.cleaned_data['soporte']    
                    #se busca por prenomina y por fecha, se guarda o actuliza la prenomina
            #        PrenominaIncidencias.objects.filter(
            #            prenomina=prenomina,  # Criterios de búsqueda
            #            fecha = fecha
            #            ).update(
            #                comentario = comentario,
            #              soporte = soporte,
            #                incidencia = incidencia
            #            )
        else:
            #EJEUCUTA LOS QUERIES FESTIVOS, ECONOMICOS, FESTIVOS, RANGOS Y GUARDA LOS DATOS
            catorcena = obtener_catorcena()
            #obtener los queries para su posterior llenado
            festivos = TablaFestivos.objects.filter(dia_festivo__range=[catorcena_actual.fecha_inicial, catorcena_actual.fecha_final]) #festivos en la catorcena actual
            economicos = Economicos_dia_tomado.objects.filter(prenomina__status=prenomina.empleado.status, fecha__range=[catorcena_actual.fecha_inicial, catorcena_actual.fecha_final], complete = False)
            vacaciones = Vacaciones_dias_tomados.objects.filter(prenomina__status=prenomina.empleado.status,fecha_inicio__lte=catorcena.fecha_final,fecha_fin__gte=catorcena.fecha_inicial)
            
            for festivo in festivos:
                registro, created = PrenominaIncidencias.objects.update_or_create(
                    prenomina_id=prenomina.id,
                    fecha=festivo.dia_festivo,
                    defaults={
                        'incidencia_id': 13 #festivo
                    }
                )
                
            for economico in economicos:
                registro, created = PrenominaIncidencias.objects.update_or_create(
                    prenomina_id=prenomina.id,
                    fecha=economico.fecha,
                    defaults={
                        'incidencia_id': 14 # economico
                    }
                )

            for vacacion in vacaciones:
                    #se ajusta la fecha de acuerdo a la catorcena
                    fecha_inicio = max(vacacion.fecha_inicio, prenomina.catorcena.fecha_inicial)
                    fecha_fin = min(vacacion.fecha_fin, prenomina.catorcena.fecha_final)
                    #se considera el dia inhabil (descanso)       
                    dia_inhabil = vacacion.dia_inhabil_id
                    
                    fecha = fecha_inicio
                    incidencia = 0
                    #sacar las fechas
                    while fecha_inicio <= fecha_fin:
                        if fecha_inicio <= fecha <= fecha_fin:
                            incidencia = 15 # vacacion
                            #verifica si es domingo o descanso
                            if fecha_inicio.weekday() == (dia_inhabil - 1): 
                                if (dia_inhabil - 1) == 6:# se resta 1 para obtener el dia domingo
                                    incidencia = 5 #domingo
                                else:
                                    incidencia = 2 #descanso
                            elif fecha_inicio in [festivo.dia_festivo for festivo in festivos]:
                                incidencia = 13 #festivo:
                        print("fecha: ",fecha, "incidencia: ", incidencia)
                        registro, created = PrenominaIncidencias.objects.update_or_create(
                            prenomina_id=prenomina.id,
                            fecha=fecha,
                            defaults={
                             'incidencia_id': incidencia
                            }
                        )  
                           
                        #se agregar un dia para realizar el recorrido de la fecha          
                        fecha_inicio += timedelta(days=1)
                        fecha +=   timedelta(days=1)

            #PREPARA LOS FORMULARIOS PARA SER MOSTRADOS CON LAS INCIDENCIAS
            fecha_inicial = catorcena_actual.fecha_inicial  # Fecha inicial
            # 5 domingo, 16 asistencia, 6 y 13 domingo calendario
            datos_iniciales = [{'fecha': fecha_inicial + timedelta(days=i),'incidencia':5 if i == 6 or i == 13 else 16} for i in range(14)] # se preparan los 14 forms con su fecha, 12 asistencias, 2 domingos
            #se filtra las incidencias por la prenomina es decir por el empleado
            incidencias = PrenominaIncidencias.objects.filter(prenomina = prenomina)
            
            rangos = PrenominaIncidencias.objects.select_related("incidencia_rango")
            #for r in rangos:
            #    print(r.fecha)
            #    print(r.soporte)
            
            
            #for i in incidencias:
            #    print(i.id)
                #print(i.fecha)
                #print(i.soporte)
                #print(i.incidencia_rango.incidencia.id)

            # Iterar sobre las incidencias y actualizar los datos iniciales si coinciden con la fecha
            for incidencia in incidencias:
                for data in datos_iniciales:
                    if incidencia.fecha == data['fecha']:
                        data['soporte'] = incidencia.soporte
                        data['comentario'] = incidencia.comentario
                        data['incidencia'] = incidencia.incidencia_id
                        data['id'] = incidencia.id  
                        data['id_rango'] = incidencia.incidencia_rango
                        
            prenomina_incidencias_form = PrenominaIncidenciasFormSet(initial=datos_iniciales)
            incidencia_rango_form = IncidenciaRangoForm()

            #es para guardar la autorizacion
            
            revisado_rh, created = AutorizarPrenomina.objects.get_or_create(prenomina=prenomina, tipo_perfil=user_filter.tipo)
            revisado_rh.estado =  Estado.objects.get(pk=1) #aprobado
            perfil_rh = Perfil.objects.get(numero_de_trabajador = user_filter.numero_de_trabajador, distrito = user_filter.distrito)
            revisado_rh.perfil=perfil_rh
            revisado_rh.comentario="Revisado por RH"
            revisado_rh.save()
            
                     
            context = {
                'prenomina':prenomina,
                'costo':costo,
                'catorcena_actual':catorcena_actual,
                'autorizacion1':autorizacion1,
                'autorizacion2':autorizacion2,
                'catorcena_actual':catorcena_actual,
                'prenomina_incidencias_form': prenomina_incidencias_form,
                'incidencia_rango_form':incidencia_rango_form
                
            }
            end_time = time.time()  # Registrar el tiempo de finalización
            print(f"Tiempo total de carga de la página: {end_time - start_time} segundos")
            return render(request, 'prenomina/Actualizar_revisar.html',context)
    else:
        return render(request, 'revisar/403.html')

@login_required(login_url='user-login')
def determinar_estado_general(request, ultima_autorizacion):
    if ultima_autorizacion is None:
        return "Sin autorizaciones"

    tipo_perfil = ultima_autorizacion.tipo_perfil.nombre.lower()
    estado_tipo = ultima_autorizacion.estado.tipo.lower()

    if tipo_perfil == 'rh' and estado_tipo == 'aprobado': #Ultimo upd rh y fue aprobado
        return 'Controles técnicos pendiente'              #Solo puede editarlo ct

    if tipo_perfil == 'control tecnico' and estado_tipo == 'aprobado': #Ultimo upd ct y fue aprobado
        return 'Gerente pendiente'                         
    
    if tipo_perfil == 'gerencia' and estado_tipo == 'aprobado': #Ultimo upd gerencia y fue aprobado
        return 'Gerente aprobado (Prenomina aprobada)'

    if tipo_perfil == 'control tecnico' and estado_tipo == 'rechazado': #Ultimo upd ct y fue rechazado
        return 'RH pendiente (rechazado por Controles técnicos)'
    
    if tipo_perfil == 'gerencia' and estado_tipo == 'rechazado': #Ultimo upd gerencia y fue rechazado
        return 'RH pendiente (rechazado por Gerencia)'

    return 'Estado no reconocido'


@login_required(login_url='user-login')
def conteo_incidencias_aguinaldo(request,prenomina,fecha_inicio,fecha_fin):
    #incidencias es una variable que lleva el conteo de todas las incidencias para el aguinaldo y se retorna  
    incidencias = 0
    faltas = Faltas.objects.filter(prenomina__empleado = prenomina.empleado.id,fecha__range=(fecha_inicio,fecha_fin)).count()
    incidencias = faltas
    print("este es el recuento totoal de las faltas ", faltas)
    return incidencias

@login_required(login_url='user-login')
def calcular_aguinaldo_eventual(request,salario,prenomina):
    """
    se realiza el calculo del registro cuando cumpla el tiempo y se guarda en la base de datos, para guardar
    se debe estar en complete = True (pagado), mes (el ultimo mes del contrato) y se considera que se va 
    a pagar en la siguiente catorcena
    """
    #tipo contrato
    tipo_contrato = prenomina.empleado.status.tipo_de_contrato_id    
    #el mes corresponde cuando cumple 1°, 2° o 3° mes
    mes = 0
    aguinaldo = Decimal(0.00)
    if tipo_contrato == 2:#eventual
        #fecha ingreso
        fecha_ingreso =  prenomina.empleado.status.fecha_ingreso
        #calculo relacion dias laborados
        calculo_fecha = relativedelta(datetime.date.today(),fecha_ingreso)
        
        #se lleva un registro de los aguinaldos que se van registrando 
        aguinaldo_registrado = Aguinaldo_Contrato.objects.filter(empleado_id = prenomina.empleado.id).last()
        
        if aguinaldo_registrado is None:
            #se realiza el calculo por los meses y por los dias laborados correspondientes a cada condicion
            if calculo_fecha.months >= 1 and calculo_fecha.days >= 0:
                #primer mes laborado le corresponden 30 dias
                dias_aguinaldo = Decimal(30 * 15) / 365
                aguinaldo = dias_aguinaldo * salario
                mes = 1
        else:
            if aguinaldo_registrado.mes == 1 and aguinaldo_registrado.complete == True and calculo_fecha.months >= 3 and calculo_fecha.days >= 0:
                #tercer mes laborado le corresponden 60 dias
                dias_aguinaldo = Decimal(60 * 15) / 365
                aguinaldo = dias_aguinaldo * salario
                mes = 3
            elif aguinaldo_registrado.mes == 3 and aguinaldo_registrado.complete == True and calculo_fecha.months >= 6 and calculo_fecha.days >= 0:
                #sexto mes laborado le corresponden 90 dias
                dias_aguinaldo = Decimal(90 * 15) / 365
                aguinaldo = dias_aguinaldo * salario
                mes = 6
            print("este es el aguinaldo eventual: ",aguinaldo)
            #se guarda en caso que exista un valor en el aguinaldo
        
        if aguinaldo != 0:
            #saber la catorcena actual
            ahora = datetime.date.today()
            catorcena_actual = Catorcenas.objects.filter(fecha_inicial__lte=ahora, fecha_final__gte=ahora).first()
            #Guardar el aguinaldo
            aguinaldo_contrato = Aguinaldo_Contrato(
                empleado = prenomina.empleado,
                aguinaldo = aguinaldo,
                fecha = date.today(),
                catorcena = catorcena_actual.id + 1, #el aguinaldo se paga en la cat siguiente
                complete=False,
                mes = mes
            )
            aguinaldo_contrato.save()
              
@login_required(login_url='user-login')
def calcular_aguinaldo(request,salario,prenomina):
    aguinaldo = Decimal(0.00)
    #obtener la catorcena actual - se utiliza para comparar la catorcena cuando sea diciembre
    ahora = datetime.date.today()
    catorcena_actual = Catorcenas.objects.filter(fecha_inicial__lte=ahora, fecha_final__gte=ahora).first()
    print("esta es la catorcena para comparar: ",catorcena_actual.id)
    #obtener la primer cat de diciembre fecha y dia 
    catorcena_decembrina = Catorcenas.objects.filter(fecha_inicial__month=12, fecha_final__month=12,fecha_inicial__year=ahora.year,fecha_final__year=ahora.year).first()
    #print("esta es la catorcena de diciembre", catorcena_decembrina.id)
    
    #verifica si es la catorcena de diciembre para pagar el aguinaldo
    #if catorcena_actual.id == catorcena_decembrina.id:
    if catorcena_actual.id == 1000:
        
        tipo_contrato = prenomina.empleado.status.tipo_de_contrato_id
        
        if tipo_contrato in (1,3,5,6): # planta, especial, planta 1, planta 2
            fecha_planta = prenomina.empleado.status.fecha_planta
            fecha_planta_anterior = prenomina.empleado.status.fecha_planta_anterior
                    
            #Se obtiene la fecha de planta o planta anterior
            if fecha_planta is None and fecha_planta_anterior is None:
                fecha = None
            elif fecha_planta_anterior is not None and fecha_planta is not None:
                fecha = fecha_planta_anterior
            elif fecha_planta:
                fecha = fecha_planta
            elif fecha_planta_anterior:
                fecha = fecha_planta_anterior
                        
            #de acuerdo a la fecha se realiza el calculo para el aguinaldo
            if fecha is not None:
                fecha = datetime.date(2022, 12, 31)
                
                
                antiguedad = relativedelta(datetime.date.today(),fecha)
            
                if antiguedad.years >= 1:#Aguinaldo completo >= al 1 primer año de antiguedad
                    #aqui es con respecto al año 1/ENE - 31/DIC
                    año_actual = datetime.date.today().year
                    inicio_año = datetime.date(año_actual, 1, 1)
                    fin_año = datetime.date(año_actual, 12, 31)
                    
                    #llama funcion para el conteo de las incidencias necesarias para el aguinaldo
                    total_incidencias = conteo_incidencias_aguinaldo(request,prenomina,inicio_año,fin_año)
                    #se restan las incidencias con los dias laborados proporcionales
                    dias_laborados = 365 - total_incidencias
                    
                    dias_pago = Decimal((dias_laborados) * 15 / 365)
                    aguinaldo = Decimal(dias_pago * salario)
                    print("aguinaldo completo")
                    print(aguinaldo)
                    #exit()
                    return aguinaldo
                else: #aguinaldo proporcional 
                    #No cumple el año y se obtiene el proporcional
                    fecha_final = datetime.date(datetime.date.today().year, 12, 31) #obtener fin de año
                    diferencia = fecha_final - fecha 
                    dias_aguinaldo = diferencia.days #total de dias laborados
                    #fecha es la fecha de planta o anterior
                    dias_incidencias = conteo_incidencias_aguinaldo(request,prenomina,fecha,fecha_final)
                
                    #se restan las incidencias con los dias laborados proporcionales
                    dias_laborados = dias_aguinaldo - dias_incidencias
                            
                    dias_pago = Decimal((dias_laborados * 15)/365)
                    print("aguinaldo proporcional ")
                    aguinaldo = Decimal( dias_pago * salario)
                    print(aguinaldo)
                    #exit()
                    return aguinaldo
            else:
                return aguinaldo
        else:
            return aguinaldo
    else:
        print("no corresponde pago del aguinaldo")
        return aguinaldo
    
@login_required(login_url='user-login')
def calcular_prima_dominical(request,dia_extra,salario):
    dato = SalarioDatos.objects.get()
    prima_dominical = salario * Decimal(dato.prima_vacacional)
    prima_dominical = prima_dominical * dia_extra
    print("salario", salario)
    print("prima dominical", Decimal(prima_dominical))
    return Decimal(prima_dominical)

@login_required(login_url='user-login')
def calcular_cuotas_imss(request,sdi_imss):
    variables_patronal = Variables_imss_patronal.objects.get()
    salario_datos = SalarioDatos.objects.get()
    
    #multiplica el sdi * el % de cuatoas / el numero de dias de la catorcena
    invalidez_vida = sdi_imss * Decimal(variables_patronal.iv_obrero / 100) * 14
    cesantia_vejez = sdi_imss * Decimal(variables_patronal.cav_patron/100) * 14
    
    #obtener el salario cotizacion mensual
    salario_cot_men = sdi_imss * Decimal(30.4)
    gastos_medicos = sdi_imss * Decimal(variables_patronal.gmp_obrero/100) * 14
    en_dinero = sdi_imss * Decimal(variables_patronal.pd_obrero/100) * 14 
    
    #calcular cuota fija por cada trabajador hasta por 3 UMAs
    cuota_fija_umas = salario_datos.UMA * Decimal(30.4) * 3
    diferencia_sbc_umas = salario_cot_men - cuota_fija_umas
    cuota_fija = (diferencia_sbc_umas * Decimal(variables_patronal.cf_obrero / 100) / Decimal(30.4)) * 14
    enfermedades_maternidad = gastos_medicos + en_dinero + cuota_fija
    
    #La suma del calculo de cada resultado    
    calculo_imss = invalidez_vida + cesantia_vejez + enfermedades_maternidad
    print("Este es el calculo IMSS: ", calculo_imss)
    return calculo_imss

@login_required(login_url='user-login')
def calcular_isr(request,salario,prima_dominical_isr,calulo_aguinaldo_isr,calculo_aguinaldo_eventual_isr):   
    salario_datos = SalarioDatos.objects.get()
    limite_inferior = 0
    porcentaje = 0
    cuota_fija = 0
    #Salario minino queda exento
    if salario > salario_datos.Salario_minimo:
        #PRIMA DOMINICAL
        if prima_dominical_isr < salario_datos.UMA:
            #exento
            prima_dominical_isr = 0
            print("prima dominical exenta: ", prima_dominical_isr)
        else:
            #gravable
            prima_dominical_isr = prima_dominical_isr - Decimal(salario_datos.UMA)
            print("prima dominical gravable: ", prima_dominical_isr)

        #AGUINALDO
        if calulo_aguinaldo_isr < (salario_datos.UMA * 30):
            #exento
            calulo_aguinaldo_isr = 0
        else:
            #gravado
            calulo_aguinaldo_isr - Decimal(salario_datos.UMA * 30)
            
            
        #AGUINALDO EVENTUAL
        if calculo_aguinaldo_eventual_isr <  (salario_datos.UMA * 30):
            #exento
            calculo_aguinaldo_eventual_isr = 0
            print("aguinaldo envetual exento", calculo_aguinaldo_eventual_isr)
        else:
            #gravado
            calculo_aguinaldo_eventual_isr - Decimal(salario_datos.UMA * 30)
            print("aguinaldo envetual gravado", calculo_aguinaldo_eventual_isr)
        
    else:
        #exento
        prima_dominical_isr = 0
        calulo_aguinaldo_isr = 0
        calculo_aguinaldo_eventual_isr = 0
        
    #multiplicar el salario por 30.4
    salario_catorcenal = salario * Decimal(salario_datos.dias_mes) #30.4
    #se suman la prima dominical, vacacional, los aguinaldos para despues aplicar el calculo del isr
    #salario_catorcenal = salario_catorcenal + prima_dominical_isr + calulo_aguinaldo_isr + calculo_aguinaldo_eventual_isr

     
    #llamar la tabla de IRS
    tabla_irs = DatosISR.objects.all()
    
    #obtener el valor aproximado hacia abajo para obtener las variables
    for datos_irs in tabla_irs:
        if salario_catorcenal >= datos_irs.liminf:
            limite_inferior = datos_irs.liminf
            porcentaje = datos_irs.excedente
            cuota_fija = datos_irs.cuota
            
    #realizar el calculo
    isr_mensual = ((salario_catorcenal - limite_inferior) * porcentaje) + cuota_fija
    
    isr_catorcenal = (isr_mensual / salario_datos.dias_mes) * 14
    
    print("calculo ISR: ", isr_mensual)
    
    return isr_catorcenal

@login_required(login_url='user-login')
def calcular_prima_vacacional(request,salario,prenomina):
    
    tabla_vacaciones = TablaVacaciones.objects.all()
    dato = SalarioDatos.objects.get()
    
    fecha_actual = datetime.date.today()
       
    calcular_prima = True
    if prenomina.empleado.status.tipo_de_contrato_id == 4: #HONORARIOS
        calcular_prima = False
    elif prenomina.empleado.status.tipo_de_contrato_id == 2: #EVENTUAL
        fecha = fecha_actual - timedelta(days=365) # El calculo es 12 dias de vacaciones, siempre para contrato eventual
    elif prenomina.empleado.status.tipo_de_contrato_id == 7: #NR
        calcular_prima = False
    elif prenomina.empleado.status.fecha_planta is None and prenomina.empleado.status.fecha_planta_anterior is None:
        calcular_prima = False
    elif prenomina.empleado.status.fecha_planta is not None and prenomina.empleado.status.fecha_planta_anterior is not None:
        fecha = prenomina.empleado.status.fecha_planta_anterior
    elif prenomina.empleado.status.fecha_planta:
        fecha = prenomina.empleado.status.fecha_planta
    elif prenomina.empleado.status.fecha_planta_anterior:
        fecha = prenomina.empleado.status.fecha_planta_anterior

    prima_vacacional = 0
    if calcular_prima == True:#calcula la prima
        calcular_antiguedad = relativedelta(fecha_actual, fecha)
        antiguedad = calcular_antiguedad.years
        print("esta es la antiguedad ", antiguedad)

        if antiguedad > 0:
            for tabla in tabla_vacaciones:
                if antiguedad >= tabla.years:
                    dias_vacaciones = tabla.days #Se asignan los días para el calculo de la prima vacacional

            vac_reforma_actual = Decimal(dias_vacaciones) * Decimal(salario)
            print("dias vacaciones", dias_vacaciones, "salario", salario)
            print("vacaciones", vac_reforma_actual)
            
            prima_vacacional = vac_reforma_actual*Decimal(dato.prima_vacacional)
            print("esta es la prima ", prima_vacacional)
            return prima_vacacional

        else:#No calcula la prima - No tiene el año de antiguedad o más
            print("esta es la prima ", prima_vacacional)
            return prima_vacacional

    else:#No calcula la prima
        print("esta es la prima ", prima_vacacional)
        return prima_vacacional 
"""    
def calcular_dias_interseccion(fecha_inicio_rango, fecha_fin_rango, fecha_inicio_catorcena, fecha_fin_catorcena, dias_inhabiles):
    inicio = max(fecha_inicio_rango, fecha_inicio_catorcena)
    fin = min(fecha_fin_rango, fecha_fin_catorcena)
    if inicio > fin:
        return 0, 0
    else:
        total_dias = (fin - inicio).days + 1
        dias_inhabiles_count = sum(1 for day in range(total_dias)
                                   if (inicio + timedelta(days=day)).weekday() in dias_inhabiles)
        return total_dias, dias_inhabiles_count
"""    
@login_required(login_url='user-login')
def calcular_retardos(request, prenomina, catorcena_actual):
    retardos = prenomina.prenominaincidencias_set.filter(
        incidencia__id = 1,  # Filtra por tipo de incidencia
        fecha__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final)
    ).values("fecha").count()
    return retardos

@login_required(login_url='user-login')
def calcular_descanso(request, prenomina, catorcena_actual):
    descanso = prenomina.prenominaincidencias_set.filter(
        incidencia__id = 2,  # Filtra por tipo de incidencia
        fecha__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final)
    ).values("fecha").count()
    return descanso

@login_required(login_url='user-login')
def calcular_faltas(request, prenomina, catorcena_actual):
    faltas = prenomina.prenominaincidencias_set.filter(
        incidencia__id = 3,  # Filtra por tipo de incidencia
        fecha__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final)
    ).values("fecha").count()
    return faltas

@login_required(login_url='user-login')
def calcular_comision(request, prenomina, catorcena_actual):
    comision = prenomina.prenominaincidencias_set.filter(
        incidencia__id = 4,  # Filtra por tipo de incidencia
        fecha__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final)
    ).values("fecha").count()
    return comision

@login_required(login_url='user-login')
def calcular_domingo(request, prenomina, catorcena_actual):
    domingo = prenomina.prenominaincidencias_set.filter(
        incidencia__id = 5,  # Filtra por tipo de incidencia
        fecha__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final)
    ).values("fecha").count()
    return domingo

@login_required(login_url='user-login')
def calcular_dia_extra(request, prenomina, catorcena_actual):
    dia_extra = prenomina.prenominaincidencias_set.filter(
        incidencia__id = 6,  # Filtra por tipo de incidencia
        fecha__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final)
    ).values("fecha").count()
    return dia_extra
 
@login_required(login_url='user-login')    
def calcular_festivos(request,prenomina,catorcena_actual):
    festivos = prenomina.prenominaincidencias_set.filter(
        incidencia__id = 13,  # Filtra por tipo de incidencia
        fecha__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final)
    ).values("fecha").count()
    return festivos 

@login_required(login_url='user-login')    
def calcular_economicos(request,prenomina,catorcena_actual):
    economicos = prenomina.prenominaincidencias_set.filter(
        incidencia__id = 14,  # Filtra por tipo de incidencia
        fecha__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final)
    ).values("fecha").count()
    return economicos

@login_required(login_url='user-login')
def calcular_castigos(request, prenomina, catorcena_actual):
    castigos = prenomina.prenominaincidencias_set.filter(
        incidencia__id=7,
        incidencia_rango__fecha_inicio__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final)
    ).values("incidencia_rango__fecha_inicio").count()
    return castigos

@login_required(login_url='user-login')
def calcular_permisos_sin_goce(request, prenomina, catorcena_actual):
    permisos_sin_goce = prenomina.prenominaincidencias_set.filter(
        incidencia__id=8,
        incidencia_rango__fecha_inicio__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final)
    ).values("incidencia_rango__fecha_inicio").count()
    return permisos_sin_goce

@login_required(login_url='user-login')
def calcular_permisos_con_goce(request, prenomina, catorcena_actual):
    permisos_con_goce = prenomina.prenominaincidencias_set.filter(
        incidencia__id=9,
        incidencia_rango__fecha_inicio__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final)
    ).values("incidencia_rango__fecha_inicio").count()
    return permisos_con_goce

@login_required(login_url='user-login')
def calcular_vacaciones(request, prenomina, catorcena_actual):
    vacaciones = prenomina.prenominaincidencias_set.filter(
        incidencia__id=15,
        incidencia_rango__fecha_inicio__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final)
    ).values("incidencia_rango__fecha_inicio").count()
    return vacaciones

@login_required(login_url='user-login')
def calcular_incapacidad_enfermedad_general(request, prenomina, catorcena_actual):
    incapacidad_enfermedad_general = prenomina.prenominaincidencias_set.filter(
        incidencia__id=10,
        incidencia_rango__fecha_inicio__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final)
    ).values("incidencia_rango__fecha_inicio").count()
    return incapacidad_enfermedad_general

@login_required(login_url='user-login')
def calcular_incapacidad_riesgo_laboral(request, prenomina, catorcena_actual):
    incapacidad_riesgo_laboral = prenomina.prenominaincidencias_set.filter(
        incidencia__id=11,
        incidencia_rango__fecha_inicio__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final)
    ).values("incidencia_rango__fecha_inicio").count()
    return incapacidad_riesgo_laboral

@login_required(login_url='user-login')
def calcular_incapacidad_maternidad(request, prenomina, catorcena_actual):
    incapacidad_maternidad = prenomina.prenominaincidencias_set.filter(
        incidencia__id=12,
        incidencia_rango__fecha_inicio__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final)
    ).values("incidencia_rango__fecha_inicio").count()
    return incapacidad_maternidad

"""
@login_required(login_url='user-login')    
def calcular_incapacidades(request,prenomina,catorcena_actual):
    print("LLAMADA DESDE CALCULAR INCAPACIDADES")
    #LLeva el conteo de las incapacidades
    cont_maternidad = 0
    cont_riesgo = 0
    cont_enfermedad = 0
    tipo = 0 #se guarda el tipo del id de la incapacidad
    #subsecuente = False #se llevar un contador 
    cont_descanso = 0 #cuenta los dias si son domingo o dias de descanso
    fecha = None #es un contador para incrementar el ciclo en el for del rango de fechas
    dias = 0 #cuenta los dias correspondientes de incapacidades
    pagar_dias = 0 #lleva el conteo de pagar los dias de la incapacidad 1,2 ó 3 dias
    cont_subsecuente = False #se lleva un contador para controlar el pago de los dias de la incapacidad de enfermedad

    
    #Query de las incapacidades que caen en la catorcena actual
    incapacidades = Incapacidades.objects.filter(prenomina__empleado_id=prenomina.empleado.id,fecha__lte=catorcena_actual.fecha_final,fecha_fin__gte=catorcena_actual.fecha_inicial)
    if incapacidades.exists():
        for incapacidad in incapacidades:
            #solo cuenta los dias que estan en la catorcena
            dentro_de_la_catorcena = catorcena_actual.fecha_inicial <= incapacidad.fecha <= catorcena_actual.fecha_final
            
            if dentro_de_la_catorcena:
                print("La incapacidad está dentro del rango de la catorcena.")                
                #detectar el dia de descanso
                cont_descanso = 0 #puede ser domingo o un dia entre semana
                for i in range((incapacidad.fecha_fin - incapacidad.fecha).days + 1): #rango para el ciclo for
                    fecha = incapacidad.fecha + timedelta(days=i) #para incrementar el dia
                    if fecha.weekday() == (incapacidad.dia_inhabil_id - 1): #verifica si es el descanso
                        cont_descanso += 1 # cuenta los dias de descanso
                        
                #conteo de los dias en relacion dias de descanso
                dias = incapacidad.fecha_fin - incapacidad.fecha
                dias = (dias.days + 1) - cont_descanso
                
                if incapacidad.tipo_id == 1 :#riesgo
                    cont_riesgo = cont_riesgo + dias
                    tipo = incapacidad.tipo_id
                    subsecuente = incapacidad.subsecuente
                    print("dias de incapadidad riesgo ", cont_riesgo)
                
                elif incapacidad.tipo_id == 2: #enfermedad
                    cont_enfermedad = cont_enfermedad + dias
                    tipo = incapacidad.tipo_id
                    subsecuente = incapacidad.subsecuente
                    print("dias de incapadidad riesgo ", cont_enfermedad)
                    
                    if incapacidad.subsecuente == False and cont_subsecuente == False:
                        print("solo se ejecuta 1 vez")
                        cont_subsecuente == True
                        for i in range(min(3, (incapacidad.fecha_fin - incapacidad.fecha).days + 1)): #rango para el ciclo for
                            fecha_actual = incapacidad.fecha + timedelta(days=i) #para incrementar el dia
                            if catorcena_actual.fecha_inicial <= fecha_actual <= catorcena_actual.fecha_final:
                                # Verifica si el día actual es un día inhabilitado
                                if fecha_actual.weekday() == (incapacidad.dia_inhabil_id - 1):
                                    cont_descanso += 1
                                else:
                                    pagar_dias += 1
                                                                                     
                elif incapacidad.tipo_id == 3: #maternidad
                    cont_maternidad = cont_maternidad + dias
                    tipo = incapacidad.tipo_id
                    subsecuente = incapacidad.subsecuente
                    print("dias de incapadidad maternidad ", cont_maternidad)
                    
            else:
                  
                print("La incapacidad está fuera del rango de la catorcena.")
                fecha_inicio = max(incapacidad.fecha, catorcena_actual.fecha_inicial)  # Selecciona la fecha más reciente entre la fecha de la incapacidad y la fecha inicial de la catorcena
                fecha_fin = min(incapacidad.fecha_fin, catorcena_actual.fecha_final)  # Selecciona la fecha más temprana entre la fecha de fin de la incapacidad y la fecha final de la catorcena
                
                print("esta es la fecha de inicio pero de? ", fecha_inicio)
                print("esta es la fecha fin de pero de? ", fecha_fin)
                
                #detectar el dia de descanso
                cont_descanso = 0 #puede ser domingo o un dia entre semana
                for i in range((fecha_fin - fecha_inicio).days + 1): #rango para el ciclo for
                    fecha = fecha_inicio + timedelta(days=i) #para incrementar el dia
                    if fecha.weekday() == (incapacidad.dia_inhabil_id - 1): #verifica si es el descanso
                        cont_descanso += 1 # cuentas

                #conteo de los dias en relacion dias de descanso
                dias = fecha_fin - fecha_inicio
                dias = (dias.days + 1) - cont_descanso
                print("estos son los dias de la catorcena ", dias)
                
                
                if incapacidad.tipo_id == 1 :#riesgo
                    cont_riesgo = cont_riesgo + dias
                    tipo = incapacidad.tipo_id
                    print("dias de incapadidad riesgo ", cont_riesgo)
                
                elif incapacidad.tipo_id == 2: #enfermedad
                    cont_enfermedad = cont_enfermedad + dias
                    print("esta es la cont enfermedad: ", cont_enfermedad)
                    tipo = incapacidad.tipo_id
                    subsecuente = incapacidad.subsecuente
                     
                elif incapacidad.tipo_id == 3: #maternidad
                    cont_maternidad = cont_maternidad + dias
                    tipo = incapacidad.tipo_id
                    subsecuente = incapacidad.subsecuente
                    print("dias de incapadidad maternidad ", cont_enfermedad)
                        
        print("Total de días de incapacidad por riesgo: ", cont_riesgo)
        print("Total de días de incapacidad por enfermedad: ", cont_enfermedad)
        print("Total de días de incapacidad por maternidad: ", cont_maternidad) 
        print("tipo",tipo)
        print("susecuente", subsecuente)
            
    return cont_riesgo,cont_enfermedad,pagar_dias,cont_maternidad,tipo     
"""
@login_required(login_url='user-login')
def Excel_estado_prenomina(request,prenominas, user_filter):
    from datetime import datetime
    
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Reporte_prenominas_' + str(datetime.now())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Reporte')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 11)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 11)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 11)
    bold_money_style = NamedStyle(name='bold_money_style', number_format='$#,##0.00', font=Font(bold=True))
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)
    dato_style = NamedStyle(name='dato_style',number_format='DD/MM/YYYY')
    dato_style.font = Font(name="Arial Narrow", size = 11)
        
    columns = ['Empleado','#Trabajador','Distrito','#Catorcena','Fecha','Estado general','RH','CT','Gerencia','Autorizada','Retardos','Castigos','Permiso con goce de sueldo',
               'Permiso sin goce de sueldo','Descansos','Incapacidades','Faltas','Comisión','Domingo','Dia de descanso laborado','Festivos','Economicos','Vacaciones','Salario Cartocenal',
               'Previsión social', 'Total bonos','Total percepciones','Prestamo infonavit','IMSS','Fonacot','ISR Retenido','Total deducciones','Neto a pagar en nomina']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        if col_num == 1:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 50
        if col_num < 4:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30
        if col_num == 4:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30
        else:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 15


    columna_max = len(columns)+2
    
    ahora = datetime.now()
    #ahora = datetime.now() + timedelta(days=15)
    
    catorcena_actual = Catorcenas.objects.filter(fecha_inicial__lte=ahora, fecha_final__gte=ahora).first()

    (ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por Savia RH. JH}')).style = messages_style
    (ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}')).style = messages_style
    #(ws.cell(column = columna_max, row = 3, value='Algún dato')).style = messages_style
    #(ws.cell(column = columna_max +1, row=3, value = 'alguna sumatoria')).style = money_resumen_style
    (ws.cell(column = columna_max, row = 4, value=f'Catorcena: {catorcena_actual.catorcena}: {catorcena_actual.fecha_inicial.strftime("%d/%m/%Y")} - {catorcena_actual.fecha_final.strftime("%d/%m/%Y")}')).style = dato_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 50
    ws.column_dimensions[get_column_letter(columna_max + 1)].width = 50

    rows = []

    sub_salario_catorcenal_costo = Decimal(0.00) #Valor de referencia del costo
    sub_salario_catorcenal = Decimal(0.00)
    sub_apoyo_pasajes = Decimal(0.00)
    sub_total_bonos = Decimal(0.00)
    sub_total_percepciones = Decimal(0.00)
    sub_prestamo_infonavit = Decimal(0.00)
    sub_calculo_isr = Decimal(0.00)
    sub_calculo_imss = Decimal(0.00)
    sub_prestamo_fonacot = Decimal(0.00)
    sub_total_deducciones = Decimal(0.00)
    sub_pagar_nomina = Decimal(0.00)
        
    for prenomina in prenominas:        
        RH = AutorizarPrenomina.objects.filter(prenomina=prenomina, tipo_perfil__nombre="RH").first()
        CT = AutorizarPrenomina.objects.filter(prenomina=prenomina, tipo_perfil__nombre="Control Tecnico").first()
        G = AutorizarPrenomina.objects.filter(prenomina=prenomina, tipo_perfil__nombre="Gerencia").first()

        if G is not None and G.estado.tipo == 'aprobado':
            estado = 'aprobado'
        elif G is not None and G.estado == 'rechazado':
            estado = 'rechazado'
        else:
            estado = 'pendiente'

        if RH is None:
            RH ="Ninguno"   
        else:
            RH = str(RH.perfil.nombres)+(" ")+str(RH.perfil.apellidos)
        if CT is None:
            CT ="Ninguno"
        else:
            CT = str(CT.perfil.nombres)+(" ")+str(CT.perfil.apellidos)
        if G is None:
            G ="Ninguno"
        else:
            G = str(G.perfil.nombres)+(" ")+str(G.perfil.apellidos)
        
        #datos para obtener los calculos de la prenomina dependiendo el empleado
        salario = Decimal(prenomina.empleado.status.costo.sueldo_diario)
        apoyo_pasajes = prenomina.empleado.status.costo.apoyo_de_pasajes
        infonavit = prenomina.empleado.status.costo.amortizacion_infonavit
        fonacot = prenomina.empleado.status.costo.fonacot 
        sdi_imss = prenomina.empleado.status.costo.sdi_imss
                
        #realiza el calculo de las cuotas imss
        calculo_imss = calcular_cuotas_imss(request,sdi_imss)
                
        #realiza el calculo de la prima vacacional
        calulo_prima_vacacional = calcular_prima_vacacional(request,salario,prenomina)
        
        #Fecha para obtener los bonos agregando la hora y la fecha de acuerdo a la catorcena
        fecha_inicial = datetime.combine(catorcena_actual.fecha_inicial, datetime.min.time()) + timedelta(hours=00, minutes=00,seconds=00)
        fecha_final = datetime.combine(catorcena_actual.fecha_final, datetime.min.time()) + timedelta(hours=23, minutes=59,seconds=59)
        
        total_bonos = BonoSolicitado.objects.filter(
            trabajador_id=prenomina.empleado.status.perfil.id,
            solicitud__fecha_autorizacion__isnull=False,
            solicitud__fecha_autorizacion__range=(fecha_inicial, fecha_final)
        ).aggregate(total=Sum('cantidad'))['total'] or 0
           
        #calculo del infonavit
        if infonavit == 0:
            prestamo_infonavit = Decimal(0.00)
        else:
            prestamo_infonavit = Decimal((infonavit / Decimal(30.4) ) * 14 )
                                        
        #calculo del fonacot
        if fonacot == 0:
            prestamo_fonacot = Decimal(0.00)
        else:
            #Se haya la catorcena actual, y cuenta cuantas catorcenas le corresponden al mes actual
            primer_dia_mes = datetime(datetime.now().year, datetime.now().month, 1).date()
            ultimo_dia_mes = datetime(datetime.now().year, datetime.now().month,
                                    calendar.monthrange(datetime.now().year, datetime.now().month)[1]).date()
            numero_catorcenas =  Catorcenas.objects.filter(fecha_final__range=(primer_dia_mes,ultimo_dia_mes)).count()
            prestamo_fonacot = prestamo_fonacot / numero_catorcenas
            
        print("EMPLEADO: ",prenomina.empleado)
        print("infonavit", prestamo_infonavit)
        print("fonacot", prestamo_fonacot)
        print("IMSS", calculo_imss)
        print("salario: ",salario)
        
        #Contar las incidencias        
        retardos = calcular_retardos(request, prenomina, catorcena_actual)
        descanso = calcular_descanso(request, prenomina, catorcena_actual)
        faltas = calcular_faltas(request, prenomina, catorcena_actual)
        comision = calcular_comision(request, prenomina, catorcena_actual)
        domingo = calcular_domingo(request, prenomina, catorcena_actual)
        dia_extra = calcular_dia_extra(request, prenomina, catorcena_actual)
        festivos = calcular_festivos(request, prenomina, catorcena_actual)
        economicos = calcular_economicos(request, prenomina, catorcena_actual)

        castigos = calcular_castigos(request, prenomina, catorcena_actual)
        permisos_sin_goce = calcular_permisos_sin_goce(request, prenomina, catorcena_actual)
        permisos_con_goce = calcular_permisos_con_goce(request, prenomina, catorcena_actual)
        vacaciones = calcular_vacaciones(request, prenomina, catorcena_actual)

        incapacidad_enfermedad_general = calcular_incapacidad_enfermedad_general(request, prenomina, catorcena_actual)
        incapacidad_riesgo_laboral = calcular_incapacidad_riesgo_laboral(request, prenomina, catorcena_actual)
        incapacidad_maternidad = calcular_incapacidad_maternidad(request, prenomina, catorcena_actual)
        """
        print(f"Castigos: {castigos} días")
        print(f"Permisos sin goce: {permisos_sin_goce} días")
        print(f"Permisos con goce: {permisos_con_goce} días")
        print(f"Vacaciones: {vacaciones} días")
        print(f"Incapacidad enfermedad general: {incapacidad_enfermedad_general} días")
        print(f"Incapacidad riesgo laboral: {incapacidad_riesgo_laboral} días")
        print(f"Incapacidad maternidad: {incapacidad_maternidad} días")
        
        incapacidades_riesgo = 0
        incapacidades_enfermedad = 0
        incapacidades_maternidad = 0
        """

        #cont_riesgo,cont_enfermedad,pagar_dias_incapacidad,cont_maternidad,tipo = calcular_incapacidades(request,prenomina,catorcena_actual)
        """
        if tipo == 1:#riesgo de trabajo
            incapacidades_riesgo = cont_riesgo
        elif tipo == 2:#enfermedad general
            incapacidades_enfermedad = cont_enfermedad 
            print("este es el main ENFERMEDAD ", incapacidades_enfermedad)
            print("este es el man ENFERMEDAD PAGAR", pagar_dias_incapacidad)           
        elif tipo == 3:#maternidad
            incapacidades_maternidad = cont_maternidad
            
        calcular el numero de vacaciones
        cantidad_dias_vacacion = 0
        if vacaciones.exists():
            for vacacion in vacaciones:
                diferencia = vacacion.fecha_fin - vacacion.fecha_inicio
                cantidad_dias_vacacion = diferencia.days + 1        
        print("total vacaciones: ", cantidad_dias_vacacion)
        """
        #calculo de la prima se manda a llamar
        if vacaciones > 0:
            cantidad_dias_vacacion = vacaciones
            calcular_prima_vacacional(cantidad_dias_vacacion)
            
        
        #numero de catorena
        catorcena_num = catorcena_actual.catorcena 
        
        incidencias = 0
        incidencias_retardos = 0
        
        if faltas > 0:
            incidencias = incidencias + faltas
            
        if retardos > 0:
            incidencias_retardos = retardos // 3 # 3 retardos se descuenta 1 dia
            
        if castigos > 0:
            incidencias = incidencias + castigos
            
        if permisos_sin_goce  > 0:
            incidencias = incidencias + permisos_sin_goce 
            
        prima_dominical = 0
        pago_doble = 0  
        if dia_extra > 0:
            pago_doble = Decimal(dia_extra * (salario * 2))
            prima_dominical = calcular_prima_dominical(request,dia_extra,salario)
            
        #calcular aguinaldo - siempre de ejecutara al momento de generar al reporte - verifica si es diciembre
        calulo_aguinaldo = calcular_aguinaldo(request,salario,prenomina)
        
        #calcular aguinaldo eventual - siempre se ejecutara al momento de generar el reporte
        #calcular_aguinaldo_eventual(request,salario,prenomina)
        
        #se realiza el calculo del aguinaldo por el tiempo del contrato, se calcula en una catorcena y se paga en la siguiente
        """
        aguinaldo_contrato = Aguinaldo_Contrato.objects.filter(empleado_id=prenomina.empleado.id).last()
        calculo_aguinaldo_eventual = 0
        if aguinaldo_contrato is not None and catorcena_actual.id == aguinaldo_contrato.catorcena:
            calculo_aguinaldo_eventual = aguinaldo_contrato.aguinaldo # se pasa el valor del aguinaldo del contrato para ser calculado en el ISR
            aguinaldo_contrato.complete = True #se actualiza el registro para definir que se ha pagado en la catorcena correspondiente
            aguinaldo_contrato.save()
        """ 
        #realiza el calculo del ISR
        calcular_aguinaldo_eventual = 0
        calculo_isr = calcular_isr(request,salario,prima_dominical,calulo_aguinaldo,calcular_aguinaldo_eventual)
            
        #calculo de la prenomina - regla de tres   
        dias_de_pago = 12
        #print("incidencias", incidencias, "incidencias_retarods", incidencias_retardos, "incidencias_inca", incidencias_incapacidades)
        incapacidades = incapacidad_riesgo_laboral + incapacidad_maternidad  + incapacidad_enfermedad_general 
        print("incidencias", incidencias, "incidencias_retarods", incidencias_retardos, "Incapacidades_riesgo", incapacidad_riesgo_laboral + incapacidad_maternidad, + incapacidad_enfermedad_general)
        #se hace el calculo en relacion con los pagos de dias laborados
        #print("estos son los dias a pagar: ", pagar_dias_incapacidad)
        dias_laborados = dias_de_pago - (incidencias + incidencias_retardos + incapacidad_riesgo_laboral + incapacidad_maternidad + (incapacidad_enfermedad_general)) #- pagar_dias_incapacidad))
        print("estos son los dias laborados: ", dias_laborados)
        proporcion_septimos_dias = Decimal((dias_laborados * 2) / 12)
        proporcion_laborados = proporcion_septimos_dias + dias_laborados
        salario_catorcenal = (proporcion_laborados * salario) + pago_doble
        print("ESTE ES EL SALARIO CATORCENAL ", salario_catorcenal)
        
        
        print("ESTE ES EL APOYO PASAJES ahora: ", apoyo_pasajes)
        apoyo_pasajes = (apoyo_pasajes / 12 ) * (12 - (incidencias + incapacidad_enfermedad_general + incapacidad_riesgo_laboral))
          
        print("apoyos pasajes: ", apoyo_pasajes)
        print("total: ", salario_catorcenal)
        print("pagar nomina: ", apoyo_pasajes + salario_catorcenal)
        #total_percepciones = salario_catorcenal + apoyo_pasajes + total_bonos + prima_dominical
        total_percepciones = salario_catorcenal + apoyo_pasajes + total_bonos + calulo_aguinaldo + prima_dominical +prima_dominical #+ calculo_aguinaldo_eventual
        #IMSS y el ISR
        total_deducciones = prestamo_infonavit + prestamo_fonacot #+ calculo_isr + calculo_imss
        pagar_nomina = (total_percepciones - total_deducciones)
        
        if retardos == 0: 
            retardos = ''
        
        if castigos == 0:
            castigos = ''
            
        if permisos_con_goce  == 0:
            permisos_con_goce  = ''
            
        if permisos_sin_goce  == 0:
            permisos_sin_goce  = ''
            
        if descanso == 0:
            descanso = ''

        if dia_extra == 0:
            dia_extra = ''
                    
        if incapacidades == 0:
            incapacidades = ''
        
        if faltas == 0:
            faltas = ''
        
        if comision == 0:
            comision = ''
            
        if domingo == 0:
            domingo = ''
            
        if festivos == 0:
            festivos = ''
            
        if economicos == 0:
            economicos  = ''
        
        cantidad_dias_vacacion=vacaciones
        if cantidad_dias_vacacion == 0:
            cantidad_dias_vacacion = ''
            
        
            
        
        # Agregar los valores a la lista rows para cada prenomina
        row = (
            prenomina.empleado.status.perfil.nombres + ' ' + prenomina.empleado.status.perfil.apellidos,
            prenomina.empleado.status.perfil.numero_de_trabajador,
            prenomina.empleado.status.perfil.distrito.distrito,
            catorcena_num,
            str(prenomina.catorcena.fecha_inicial) + " " + str(prenomina.catorcena.fecha_final),
            prenomina.estado_general,
            str(RH),
            str(CT),
            str(G),
            estado,
            retardos,
            castigos,
            permisos_con_goce,
            permisos_sin_goce,
            descanso,
            #str("Días anteriores: ")+str(incapacidades_anterior)+str(" Días actual: ")+str(incapacidades_actual),
            incapacidades,
            faltas,
            comision,
            domingo,
            dia_extra,
            festivos,
            economicos,
            cantidad_dias_vacacion,
            salario_catorcenal,
            apoyo_pasajes,
            total_bonos,
            total_percepciones,
            prestamo_infonavit,
            calculo_imss,
            prestamo_fonacot,
            calculo_isr,
            #
            #
            total_deducciones,
            pagar_nomina,
        )
        rows.append(row)
        
        #sub_salario_catorcenal_costo = sub_salario_catorcenal_costo + salario_catorcenal_costo
        sub_salario_catorcenal = sub_salario_catorcenal + salario_catorcenal
        sub_apoyo_pasajes = sub_apoyo_pasajes + apoyo_pasajes
        sub_total_bonos = sub_total_bonos + total_bonos
        sub_total_percepciones = sub_total_percepciones + total_percepciones
        sub_prestamo_infonavit = sub_prestamo_infonavit + prestamo_infonavit
        sub_calculo_imss = sub_calculo_imss + calculo_imss
        sub_prestamo_fonacot = sub_prestamo_fonacot + prestamo_fonacot
        sub_calculo_isr = sub_calculo_isr + calculo_isr
        sub_total_deducciones = sub_total_deducciones + total_deducciones
        sub_pagar_nomina = sub_pagar_nomina + pagar_nomina
        
        
                 
    # Ahora puedes usar la lista rows como lo estás haciendo actualmente en tu código
    for row_num, row in enumerate(rows, start=2):
        for col_num, value in enumerate(row, start=1):
            if col_num < 4:
                ws.cell(row=row_num, column=col_num, value=value).style = body_style
            elif col_num == 5:
                ws.cell(row=row_num, column=col_num, value=value).style = date_style
            elif col_num > 5 and col_num < 24:
                ws.cell(row=row_num, column=col_num, value=value).style = body_style
            elif col_num >= 24:
                ws.cell(row=row_num, column=col_num, value=value).style = money_style
            else:
                ws.cell(row=row_num, column=col_num, value=value).style = body_style

    #sumar las columnas
    
    #print("Salario neto cartocelnal", salario_catorcenal)
    #sub_salario_catorcenal = sub_salario_catorcenal + salario_catorcenal
    #print("subtotal catorcenal",sub_salario_catorcenal)
    #sub_apoyo_pasajes = Decimal(sub_apoyo_pasajes) + apoyo_pasajes
    #sub_total_bonos = sub_total_bonos + total_bonos
    #sub_total_percepciones = sub_total_percepciones + total_percepciones
    #sub_prestamo_infonavit = sub_prestamo_infonavit + prestamo_infonavit
    #sub_prestamo_fonacot = sub_prestamo_fonacot + prestamo_fonacot
    #sub_total_deducciones = sub_total_deducciones + total_deducciones
    #sub_pagar_nomina = sub_pagar_nomina + pagar_nomina
    
    
    add_last_row = ['Total','','','','','','','','','','','','','','','','','','','','','','',
                    #sub_salario_catorcenal_costo,
                    sub_salario_catorcenal,
                    sub_apoyo_pasajes,
                    sub_total_bonos,
                    sub_total_percepciones,
                    sub_prestamo_infonavit,
                    sub_calculo_imss,
                    sub_prestamo_fonacot,
                    sub_calculo_isr,
                    sub_total_deducciones,
                    sub_pagar_nomina
                    ]
    ws.append(add_last_row) 
    
    # Aplicar el estilo money_style a cada celda de la fila
    for row in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row):
        for cell in row:
            cell.style = bold_money_style

    
    #referencia_celda = f'{"x"}{"24"}'
    #celda = ws[referencia_celda]
    #celda.value = 'Laravel'
    
    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)

@login_required(login_url='user-login')
def PdfFormatoEconomicos(request, solicitud):
    solicitud= Solicitud_economicos.objects.get(id=solicitud.id)
    now = date.today()
    fecha = solicitud.fecha
    periodo = str(fecha.year)
    economico = 0
    if not Economicos.objects.filter(status=solicitud.status):
        economico = 0
    else:
        last_economico = Economicos.objects.filter(status=solicitud.status).last()
        economico = last_economico.dias_disfrutados
    #Para ubicar el dia de regreso en un dia habil (Puede caer en día festivo)
    #if status.regimen.regimen == 'L-V':
    #    inhabil1 = 6
    #    inhabil2 = 7
    #elif status.regimen.regimen == 'L-S':
    #    inhabil1 = 7
    #    inhabil2 = None
    regreso = fecha + timedelta(days=1)
    #if regreso.isoweekday() == inhabil1:
    #    regreso = regreso + timedelta(days=1)
    #if regreso.isoweekday() == inhabil2:
    #    regreso = regreso + timedelta(days=1)


    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)

    #Colores utilizados
    azul = Color(0.16015625,0.5,0.72265625)
    rojo = Color(0.59375, 0.05859375, 0.05859375)

    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',16)
    c.drawCentredString(305,765,'GRUPO VORCAB SA DE CV')
    c.setFont('Helvetica-Bold',11)
    c.drawCentredString(305,750,'SOLICITUD DE DIA ECONOMICO')
    if solicitud.autorizar == False:
        c.setFillColor(rojo)
        c.setFont('Helvetica-Bold',16)
        c.drawCentredString(305,725,'SOLICITUD NO AUTORIZADA')
    c.setFillColor(black)
    c.setFont('Helvetica-Bold',11)
    c.drawString(40,690,'NOMBRE:')
    c.line(95,688,325,688)
    espacio = ' '
    nombre_completo = str(solicitud.status.perfil.nombres + espacio + solicitud.status.perfil.apellidos)
    c.drawString(100,690,nombre_completo)
    c.drawString(40,670,'PUESTO:')
    c.line(95,668,325,668)
    c.drawString(100,670,solicitud.status.puesto.puesto)
    c.drawString(335,670,'TELEFONO PARTICULAR:')
    c.line(475,668,580,668)
    c.drawString(485,670,solicitud.status.telefono)
    c.drawString(40,620,'FECHA DE PLANTA:')
    if solicitud.status.fecha_planta != None:
        dia = str(solicitud.status.fecha_planta.day)
        mes = str(solicitud.status.fecha_planta.month)
        año = str(solicitud.status.fecha_planta.year)
    else:
        dia = "NR"
        mes = "NR"
        año = "NR"
    #rect(x, y, alto, ancho, stroke=1, fill=0)
    c.rect(185,600, 150, 50)
    c.line(185,618,335,618)
    c.line(185,638,335,638)
    c.line(230,650,230,600)
    c.line(290,650,290,600)
    c.drawCentredString(210,620,dia)
    c.drawCentredString(260,620,mes)
    c.drawCentredString(310,620,año)
    c.drawString(40,600,'FECHA DE SOLICITUD:')
    c.drawCentredString(210,605,str(now.day))
    c.drawCentredString(260,605,str(now.month))
    c.drawCentredString(310,605,str(now.year))
    c.drawString(200,640,'DIA')
    c.drawString(250,640,'MES')
    c.drawString(300,640,'AÑO')
    c.drawString(400,600,'FIRMA DEL SOLICITANTE')
    c.rect(390,598, 155, 50)
    c.line(390,610,545,610)
    c.drawString(40,540,'PERIODO CORRESPONDIENTE:')
    c.drawCentredString(450,540, periodo)
    c.rect(35,538, 255, 12)
    c.rect(360,538, 190, 12)
    c.drawCentredString(385,520,'1')
    c.drawCentredString(435,520,'2')
    c.drawCentredString(485,520,'3')

    c.drawString(40,500,'NO. DE DIA ECONOMICO:')
    c.rect(35,498, 175, 12)
    c.rect(360,498, 150, 12)
    c.line(410,510,410,498)
    c.line(460,510,460,498)
    c.setFillColorRGB(0.8, 0.8, 0.8)  # Color de relleno
    if economico == 1:
        c.rect(360,498, 50, 12, stroke = 1, fill = 1)
    elif economico == 2:
        c.rect(410,498, 50, 12, stroke = 1, fill = 1)
    elif economico == 3:
        c.rect(460,498, 50, 12, stroke = 1, fill = 1)
    c.setFillColor(black)
    c.drawString(40,480,'CON GOCE DE SUELDO:')
    c.rect(35,478, 140, 12)
    c.drawString(380,480,'SI')
    c.rect(360,478, 50, 12)
    c.drawString(425,480,'NO')
    c.rect(410,478, 50, 12)
    c.drawString(40,460,'FECHA QUE DESEA SALIR DEL PERMISO:')
    c.drawCentredString(450,460,str(fecha))
    c.rect(35,458, 250, 12)
    c.rect(360,458, 190, 12)
    c.drawString(40,440,'FECHA DE REGRESO A LABORES:')
    c.drawCentredString(450,440,str(regreso))
    c.rect(35,438, 195, 12)
    c.rect(360,438, 190, 12)
    #c.drawCentredString(305,370,'OBSERVACIONES')
    text = solicitud.comentario
    x = 40
    y = 385
    c.setFillColor(black)
    c.setFont('Helvetica-Bold',12)
    c.drawCentredString(310, y - 15, 'OBSERVACIONES')
    c.setFont('Helvetica', 9)
    lines = textwrap.wrap(text, width=100)
    for line in lines:
        c.drawString(x + 10, y - 30, line)
        y -= 25
    c.rect(40,368, 530, 12)
    c.rect(40,300, 530, 68)
    c.drawCentredString(170,125,'FIRMA GERENCIA')
    c.rect(70,123, 200, 12)
    c.rect(70,135, 200, 50)
    c.drawCentredString(440,125,'FIRMA DE JEFE INMEDIATO')
    c.rect(330,123, 210, 12)
    c.rect(330,135, 210, 50)
    c.save()
    c.showPage()
    buf.seek(0)

    return FileResponse(buf, as_attachment=True, filename='Formato_Economico.pdf')

@login_required(login_url='user-login')
def PdfFormatoVacaciones(request, solicitud):
    solicitud= Solicitud_vacaciones.objects.get(id=solicitud.id)
    inicio = solicitud.fecha_inicio
    fin = solicitud.fecha_fin
    dia_inhabil = solicitud.dia_inhabil
    ######
    tabla_festivos = TablaFestivos.objects.all()
    delta = timedelta(days=1)
    day_count = (fin - inicio + delta ).days
    cuenta = day_count
    inhabil = dia_inhabil.numero
    for fecha in (inicio + timedelta(n) for n in range(day_count)):
        if fecha.isoweekday() == inhabil:
            cuenta -= 1
        else:
            for dia in tabla_festivos:
                if fecha == dia.dia_festivo:
                    cuenta -= 1
    diferencia = str(cuenta)
    #Para ubicar el dia de regreso en un dia habil (Puede caer en día festivo)
    fin = fin + timedelta(days=1)
    if fin.isoweekday() == inhabil:
        fin = fin + timedelta(days=1)
    now = date.today()
    año1 = str(inicio.year)
    año2= str(fin.year)
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)

    #Colores utilizados
    azul = Color(0.16015625,0.5,0.72265625)
    rojo = Color(0.59375, 0.05859375, 0.05859375)

    c.setFillColor(azul)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',16)
    c.drawCentredString(305,765,'GRUPO VORCAB SA DE CV')
    c.drawCentredString(305,750,'SOLICITUD DE VACACIONES')
    c.drawInlineImage('static/images/vordcab.png',50,720, 4 * cm, 2 * cm) #Imagen Savia
    if solicitud.autorizar == False:
        c.setFillColor(rojo)
        c.setFont('Helvetica-Bold',16)
        c.drawCentredString(305,725,'SOLICITUD NO AUTORIZADA')
    c.setFillColor(black)
    c.setFont('Helvetica-Bold',11)
    c.drawString(40,690,'NOMBRE:')
    c.line(95,688,325,688)
    espacio = ' '
    nombre_completo = str(solicitud.status.perfil.nombres + espacio + solicitud.status.perfil.apellidos)
    c.drawString(100,690,nombre_completo)
    c.drawString(40,670,'PUESTO:')
    c.line(95,668,325,668)
    c.drawString(100,670,solicitud.status.puesto.puesto)

    c.drawString(335,670,'TELEFONO PARTICULAR:')
    c.line(475,668,580,668)
    c.drawString(485,670,solicitud.status.telefono)

    if cuenta < 3:
        altura=200
        margen=20
        c.drawCentredString(305,502,'OBSERVACIONES')
        if solicitud.comentario_rh:
            c.drawString(55,490,solicitud.comentario_rh)
        else:
            c.drawString(55,490,'ninguna')
        c.rect(50,500, 510, 12)
        c.rect(50,435, 510, 65)
    else:
        altura=0
        margen=0
    c.drawString(40,620-margen,'FECHA DE PLANTA:')
    if solicitud.status.fecha_planta != None:
        dia = str(solicitud.status.fecha_planta.day)
        mes = str(solicitud.status.fecha_planta.month)
        año = str(solicitud.status.fecha_planta.year)
    else:
        dia = "NR"
        mes = "NR"
        año = "NR"

    c.rect(185,598-margen, 150, 55)
    c.line(185,618-margen,335,618-margen)
    c.line(185,638-margen,335,638-margen)
    c.line(230,650-margen,230,598-margen)
    c.line(290,650-margen,290,598-margen)
    c.drawCentredString(210,620-margen,dia)
    c.drawCentredString(260,620-margen,mes)
    c.drawCentredString(310,620-margen,año)
    c.drawString(40,600-margen,'FECHA DE SOLICITUD:')
    c.drawCentredString(210,600-margen,str(now.day))
    c.drawCentredString(260,600-margen,str(now.month))
    c.drawCentredString(310,600-margen,str(now.year))
    c.drawString(200,640-margen,'DIA')
    c.drawString(250,640-margen,'MES')
    c.drawString(300,640-margen,'AÑO')
    c.drawString(400,600-margen,'FIRMA DEL SOLICITANTE')
    c.rect(390,598-margen, 155, 55)
    c.line(390,610-margen,545,610-margen)

    c.drawString(40,560-altura,'PERIODO VACACIONAL CORRESPONDIENTE:')
    c.drawCentredString(425,560-altura, año1)
    c.drawCentredString(450,560-altura, '/')
    c.drawCentredString(475,560-altura, año2)
    c.rect(35,558-altura, 255, 12)
    c.rect(360,558-altura, 190, 12)
    #form = VacacionesFormato(request.POST,)
    c.drawString(40,540-altura,'NO. DE DIAS DE VACACIONES:')
    c.drawCentredString(450,540-altura,diferencia)
    c.rect(35,538-altura, 175, 12)
    c.rect(360,538-altura, 190, 12)
    c.drawString(40,520-altura,'CON GOCE DE SUELDO:')
    c.rect(35,518-altura, 140, 12)
    c.drawString(380,520-altura,'SI')
    c.rect(360,518-altura, 50, 12)
    c.drawString(425,520-altura,'NO')
    c.rect(410,518-altura, 50, 12)
    c.drawString(40,500-altura,'FECHA QUE DESEA SALIR DE VACACIONES:')
    c.drawCentredString(450,500-altura,str(inicio))
    c.rect(35,498-altura, 250, 12)
    c.rect(360,498-altura, 190, 12)
    c.drawString(40,480-altura,'FECHA DE REGRESO A LABORES:')
    c.drawCentredString(450,480-altura,str(fin))
    c.rect(35,478-altura, 195, 12)
    c.rect(360,478-altura, 190, 12)
    if cuenta >= 3: ########### Para formatos largos
        c.drawCentredString(300,440,'Entrega-Recepción')
        c.setFont('Helvetica',11)
        c.drawString(40,400,'DATOS DE QUIEN RECIBE:')
        c.drawString(40,380,'Nombre:')
        if solicitud.recibe_nombre:
            c.drawString(100,380,solicitud.recibe_nombre)
        c.line(90,378,375,378)
        c.drawString(385,380,'Area:')
        if solicitud.recibe_area:
            c.drawString(435,380,solicitud.recibe_area)
        c.line(420,378,560,378)
        c.drawString(40,360,'Puesto:')
        if solicitud.recibe_puesto:
            c.drawString(100,360,solicitud.recibe_puesto)
        c.line(90,358,375,358)
        c.drawString(40,340,'Sector:')
        if solicitud.recibe_sector:
            c.drawString(100,340,solicitud.recibe_sector)
        c.line(90,338,375,338)
        c.setFont('Helvetica-Bold',14)
        c.drawString(40,300,'SITUACIÓN DE TRABAJOS ENCOMENDADOS:')
        c.setFillColor(black)
        c.setFont('Helvetica',11)

        # Estilo de párrafo para los datos en la tabla
        styleSheet = getSampleStyleSheet()
        paragraphStyle = styleSheet['Normal']
        paragraphStyle.fontSize = 8
        #Tabla y altura guia
        high = 130
        trabajos = Trabajos_encomendados.objects.filter(solicitud_vacaciones__id=solicitud.id)
        data = []
        data.append(['No.', 'DENOMINACIÓN ASUNTO', 'ESTADO'])

        numero = 1  # Inicializar el número

        for index, trabajo in enumerate(trabajos, start=1):
            trabajo_data = []
            for i in range(1, 11):
                asunto = getattr(trabajo, f'asunto{i}', '')
                estado = getattr(trabajo, f'estado{i}', '')
                trabajo_data.append((numero, asunto, estado))
                numero += 1  # Incrementar el número
            data.extend(trabajo_data)
        high = high - 20

        #Propiedades de la tabla
        width, height = landscape(letter)
        table_style = TableStyle([ #estilos de la tabla
            #ENCABEZADO
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('TEXTCOLOR',(0,0),(-1,0), colors.black),
            ('FONTSIZE',(0,0),(-1,0), 12),
            ('BACKGROUND',(0,0),(-1,0), colors.white),
            #CUERPO
            ('TEXTCOLOR',(0,1),(-1,-1), colors.black),
            ('FONTSIZE',(0,1),(-1,-1), 12),
            ('GRID',(0,0),(-1,-1),0.5,colors.grey),
        ])

        # Definir variables para el salto de página
        rows_per_page = 7
        total_rows = len(data) - 1  # Excluye el encabezado
        current_row = 1  # Comenzar desde la primera fila (excluyendo el encabezado)

        # Generar páginas con el contenido restante
        while current_row <= total_rows:
            # Calcular el espacio disponible en la página actual
            available_height = height - 70 - 20  # Ajustar según el espaciado

            # Calcular cuántas filas caben en la página actual
            if current_row == 1:
                rows_on_page = min(rows_per_page, math.floor((available_height - 20) / 20))  # Para la primera página
            else:
                rows_on_page = min(20, math.floor((available_height - 20) / 20))  # Para las páginas restantes

            # Obtener los datos para la página actual
            end_row = int(current_row + rows_on_page) if current_row + rows_on_page <= total_rows else total_rows + 1
            page_data = data[current_row:end_row]

            # Reemplazar valores None con un espacio en blanco
            page_data = [[cell if cell is not None else " " for cell in row] for row in page_data]

            # Calcular la altura para dibujar la tabla
            table_height = len(page_data) * 20

            # Calcular la posición vertical para la tabla
            if current_row == 1:
                # Ajustar el margen superior para la primera tabla
                table_y = height - 130 - table_height - 275  # Usar la altura específica para la primera tabla
            else:
                # Calcular el margen desde la parte superior de la página
                margin_top = 40
                table_y = height - table_height - margin_top

            # Dentro del bucle para crear la tabla de cada página
            table_data = []
            for row in page_data:
                table_row = []
                for cell_data in row:
                    if isinstance(cell_data, str) and len(cell_data) > 30:
                        # Aplicar estilo CSS para dividir palabras largas
                        cell_data = cell_data.replace(' ', '<br/>')
                        cell_data = f'<font size="12">{cell_data}</font>'
                        cell = Paragraph(cell_data, paragraphStyle)
                    else:
                        cell = cell_data
                    table_row.append(cell)
                table_data.append(table_row)

            # Crear la tabla para la página actual
            table = Table([data[0]] + table_data, colWidths=[1.5 * cm, 8 * cm, 10 * cm], repeatRows=1)
            table.setStyle(table_style)

            # Dibujar la tabla en la página actual
            table.wrapOn(c, width, height)
            table.drawOn(c, 25, table_y)

            # Avanzar al siguiente conjunto de filas
            current_row += rows_on_page

            # Cambiar la cantidad de filas por página después de la primera página
            if current_row == 1:
                rows_per_page = 20

            # Agregar una nueva página si quedan más filas por dibujar
            if current_row <= total_rows:
                c.showPage()
        c.showPage()
        c.setFont('Helvetica-Bold',12)
        #Parrafo con salto de linea automatica si el texto es muy largo
        text = " "
        if solicitud.informacion_adicional:
            text = solicitud.informacion_adicional
        x = 40
        y = 750
        c.setFillColor(black)
        c.setFont('Helvetica', 12)
        c.drawString(x + 5, y - 15, 'INFORMACIÓN ADICIONAL:')
        c.setFont('Helvetica', 9)
        lines = textwrap.wrap(text, width=100)
        for line in lines:
            c.drawString(x + 10, y - 35, line)
            y -= 15

        # Estilo de párrafo para los comentarios
        styleSheet = getSampleStyleSheet()
        commentStyle = styleSheet['Normal']
        commentStyle.fontSize = 8

        def format_comment(comment):
            if comment is None:
                return ""
            return Paragraph(comment, commentStyle)

        # Datos y ajustes de la tabla
        data2 = []
        high = 465
        data2.append(['No.', 'TEMAS', 'COMENTARIO'])
        data2.append(["1", "Información sobre personal a su cargo", format_comment(solicitud.temas.comentario1)])
        data2.append(["2", "Documentos", format_comment(solicitud.temas.comentario2)])
        data2.append(["3", Paragraph("Arqueo de caja o cuenta bancaria a su cargo (cuando aplique)"), format_comment(solicitud.temas.comentario3)])
        data2.append(["4", "Proyectos pendientes", format_comment(solicitud.temas.comentario4)])
        data2.append(["5", "Estado de las operaciones a su cargo", format_comment(solicitud.temas.comentario5)])
        data2.append(["6", "Deudas con la empresa", format_comment(solicitud.temas.comentario6)])
        data2.append(["7", "Saldos por comprobar a contabilidad", format_comment(solicitud.temas.comentario7)])
        data2.append(["8", "Activos asignados", format_comment(solicitud.temas.comentario8)])
        data2.append(["9", "Otros", format_comment(solicitud.temas.comentario9)])

        table = Table(data2, colWidths=[1.5 * cm, 8 * cm, 11 * cm,], repeatRows=1)
        table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('TEXTCOLOR',(0,0),(-1,0), colors.black),
            ('FONTSIZE',(0,0),(-1,0), 13),
            ('BACKGROUND',(0,0),(-1,0), colors.white),
            ('GRID',(0,0),(-1,-1),0.5,colors.grey),
        ]))

        # Dibujar la tabla en el lienzo
        table.wrapOn(c, width, height)
        table.drawOn(c, 25, high)

        #c.drawString(40,375,'ANEXOS:')
        text = " "
        if solicitud.anexos:
            text = solicitud.anexos
        x = 40
        y = 380
        c.setFillColor(black)
        c.setFont('Helvetica', 12)
        c.drawString(x + 5, y - 15, 'Anexos:')
        c.setFont('Helvetica', 9)
        lines = textwrap.wrap(text, width=100)
        for line in lines:
            c.drawString(x + 10, y - 30, line)
            y -= 25
        c.line(40,345,570,345)
        c.line(40,320,570,320)
        c.line(40,293,570,293)
        c.line(40,270,570,270)
        c.drawCentredString(200,170,'ENTREGUE (NOMBRE Y FIRMA)')
        c.drawCentredString(200,190,nombre_completo)
        c.line(105,185,295,185)
        c.drawCentredString(400,170,'RECIBI (NOMBRE Y FIRMA)')
        c.drawCentredString(400,190,solicitud.recibe_nombre)
        c.line(320,185,480,185)

    c.drawCentredString(200,70,'FIRMA DE GERENCIA')
    c.rect(120,68, 160, 70)
    c.line(120,80,280,80)
    c.drawCentredString(400,70,'FIRMA DE JEFE INMEDIATO')
    c.rect(300,68, 195, 70)
    c.line(300,80,495,80)
    c.save()
    c.showPage()
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename='Formato_Vacaciones.pdf')

@login_required(login_url='user-login')
def obtener_fechas_con_incidencias(request, prenomina, catorcena_actual):
    # Crear lista de fechas entre fecha_inicial y fecha_final de la catorcena
    dias_entre_fechas = [(catorcena_actual.fecha_inicial + timedelta(days=i)) for i in range((catorcena_actual.fecha_final - catorcena_actual.fecha_inicial).days + 1)]

    # Obtener incidencias
    incidencias = {
        "retardos": prenomina.prenominaincidencias_set.filter(incidencia__id=1, fecha__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final)),
        "descanso": prenomina.prenominaincidencias_set.filter(incidencia__id=2, fecha__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final)),
        "faltas": prenomina.prenominaincidencias_set.filter(incidencia__id=3, fecha__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final)),
        "comision": prenomina.prenominaincidencias_set.filter(incidencia__id=4, fecha__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final)),
        "domingo": prenomina.prenominaincidencias_set.filter(incidencia__id=5, fecha__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final)),
        "dia_extra": prenomina.prenominaincidencias_set.filter(incidencia__id=6, fecha__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final)),
        "festivos": prenomina.prenominaincidencias_set.filter(incidencia__id=13, fecha__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final)),
        "economicos": prenomina.prenominaincidencias_set.filter(incidencia__id=14, fecha__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final)),
        "castigos": prenomina.prenominaincidencias_set.filter(incidencia__id=7, incidencia_rango__isnull=True, fecha__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final)),
        "permisos_sin_goce": prenomina.prenominaincidencias_set.filter(incidencia__id=8, incidencia_rango__isnull=True, fecha__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final)),
        "permisos_con_goce": prenomina.prenominaincidencias_set.filter(incidencia__id=9, incidencia_rango__isnull=True, fecha__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final)),
        "vacaciones": prenomina.prenominaincidencias_set.filter(incidencia__id=15, incidencia_rango__isnull=True, fecha__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final)),
        "incapacidad_enfermedad_general": prenomina.prenominaincidencias_set.filter(incidencia__id=10, incidencia_rango__isnull=True, fecha__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final)),
        "incapacidad_riesgo_laboral": prenomina.prenominaincidencias_set.filter(incidencia__id=11, incidencia_rango__isnull=True, fecha__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final)),
        "incapacidad_maternidad": prenomina.prenominaincidencias_set.filter(incidencia__id=12, incidencia_rango__isnull=True, fecha__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final))
    }

    # Crear una lista con fechas y etiquetas
    fechas_con_etiquetas = []
    for fecha in dias_entre_fechas:
        etiqueta = "asistencia"
        for tipo_incidencia, incidencia_set in incidencias.items():
            if fecha in incidencia_set.values_list('fecha', flat=True):
                etiqueta = tipo_incidencia
                break
        fechas_con_etiquetas.append((fecha, etiqueta))

    return fechas_con_etiquetas

@login_required(login_url='user-login')
def Excel_estado_prenomina_formato(request,prenominas, user_filter):
    from datetime import datetime
    
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Reporte_prenomina_días_' + str(datetime.now())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Reporte')
    #Comenzar en la fila 1
    row_num = 3

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 11)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 11)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 11)
    bold_money_style = NamedStyle(name='bold_money_style', number_format='$#,##0.00', font=Font(bold=True))
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)
    dato_style = NamedStyle(name='dato_style',number_format='DD/MM/YYYY')
    dato_style.font = Font(name="Arial Narrow", size = 11)
    ahora = datetime.now()
    #ahora = datetime.now() + timedelta(days=10)
    # todas las fechas de la catorcena actual
    catorcena_actual = Catorcenas.objects.filter(fecha_inicial__lte=ahora, fecha_final__gte=ahora).first()
    delta = catorcena_actual.fecha_final - catorcena_actual.fecha_inicial
    dias_entre_fechas = [catorcena_actual.fecha_inicial + timedelta(days=i) for i in range(delta.days + 1)]
    # Generar los nombres de las columnas de los días
    dias_columnas = [str(fecha.day) for fecha in dias_entre_fechas]
        
    columns = ['No.','NOMBRE DE EMPLEADO','PUESTO','PROYECTO','SUBPROYECTO'] + dias_columnas + ['Salario Catorcenal','Salario Catorcenal',
               'Previsión social', 'Total bonos','Total percepciones','Prestamo infonavit','Fonacot','Total deducciones','Neto a pagar en nomina','Salario','Salario Domingo',]
    
    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        if col_num == 1:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 50
        if col_num < 4:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30
        if col_num == 4:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30
        else:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 15


    columna_max = len(columns)+2
    
    ahora = datetime.now()
    #ahora = datetime.now() + timedelta(days=10)
    
    catorcena_actual = Catorcenas.objects.filter(fecha_inicial__lte=ahora, fecha_final__gte=ahora).first()
    (ws.cell(column = 1, row = 1, value='Reporte Prenomina SAVIA RH')).style = messages_style
    (ws.cell(column = 1, row = 2, value=f'Catorcena: {catorcena_actual.catorcena}: {catorcena_actual.fecha_inicial.strftime("%d/%m/%Y")} - {catorcena_actual.fecha_final.strftime("%d/%m/%Y")}')).style = dato_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 50
    ws.column_dimensions[get_column_letter(columna_max + 1)].width = 50

    rows = []

    sub_salario_catorcenal_costo = Decimal(0.00) #Valor de referencia del costo
    sub_salario_catorcenal = Decimal(0.00)
    sub_apoyo_pasajes = Decimal(0.00)
    sub_total_bonos = Decimal(0.00)
    sub_total_percepciones = Decimal(0.00)
    sub_prestamo_infonavit = Decimal(0.00)
    sub_prestamo_fonacot = Decimal(0.00)
    sub_total_deducciones = Decimal(0.00)
    sub_pagar_nomina = Decimal(0.00)
        
    for prenomina in prenominas:
        #datos para obtener los calculos de la prenomina dependiendo el empleado
        salario_catorcenal_costo = (prenomina.empleado.status.costo.neto_catorcenal_sin_deducciones)
        
        salario = Decimal(prenomina.empleado.status.costo.neto_catorcenal_sin_deducciones) / 14
        neto_catorcenal =  prenomina.empleado.status.costo.neto_catorcenal_sin_deducciones
        apoyo_pasajes = prenomina.empleado.status.costo.apoyo_de_pasajes
        infonavit = prenomina.empleado.status.costo.amortizacion_infonavit
        fonacot = prenomina.empleado.status.costo.fonacot 
        
        #Fecha para obtener los bonos agregando la hora y la fecha de acuerdo a la catorcena
        fecha_inicial = datetime.combine(catorcena_actual.fecha_inicial, datetime.min.time()) + timedelta(hours=00, minutes=00,seconds=00)
        fecha_final = datetime.combine(catorcena_actual.fecha_final, datetime.min.time()) + timedelta(hours=23, minutes=59,seconds=59)
        
        total_bonos = BonoSolicitado.objects.filter(
            trabajador_id=prenomina.empleado.status.perfil.id,
            solicitud__fecha_autorizacion__isnull=False,
            solicitud__fecha_autorizacion__range=(fecha_inicial, fecha_final)
        ).aggregate(total=Sum('cantidad'))['total'] or 0

        print("Total Bonos:", total_bonos)
           
        #calculo del infonavit
        if infonavit == 0:
            prestamo_infonavit = Decimal(0.00)
        else:
            prestamo_infonavit = Decimal((infonavit / Decimal(30.4) ) * 14 )
       
        #calculo del fonacot
        if fonacot == 0:
            prestamo_fonacot = Decimal(0.00)
        else:
            #Se haya la catorcena actual, y cuenta cuantas catorcenas le corresponden al mes actual
            primer_dia_mes = datetime(datetime.now().year, datetime.now().month, 1).date()
            ultimo_dia_mes = datetime(datetime.now().year, datetime.now().month,
                                    calendar.monthrange(datetime.now().year, datetime.now().month)[1]).date()
            numero_catorcenas =  Catorcenas.objects.filter(fecha_final__range=(primer_dia_mes,ultimo_dia_mes)).count()
            prestamo_fonacot = prestamo_fonacot / numero_catorcenas
            
        
        #contar no. de incidencias 
        retardos = calcular_retardos(request, prenomina, catorcena_actual)
        descanso = calcular_descanso(request, prenomina, catorcena_actual)
        faltas = calcular_faltas(request, prenomina, catorcena_actual)
        comision = calcular_comision(request, prenomina, catorcena_actual)
        domingo = calcular_domingo(request, prenomina, catorcena_actual)
        dia_extra = calcular_dia_extra(request, prenomina, catorcena_actual)
        festivos = calcular_festivos(request, prenomina, catorcena_actual)
        economicos = calcular_economicos(request, prenomina, catorcena_actual)

        castigos = calcular_castigos(request, prenomina, catorcena_actual)
        permisos_sin_goce = calcular_permisos_sin_goce(request, prenomina, catorcena_actual)
        permisos_con_goce = calcular_permisos_con_goce(request, prenomina, catorcena_actual)
        vacaciones = calcular_vacaciones(request, prenomina, catorcena_actual)

        incapacidad_enfermedad_general = calcular_incapacidad_enfermedad_general(request, prenomina, catorcena_actual)
        incapacidad_riesgo_laboral = calcular_incapacidad_riesgo_laboral(request, prenomina, catorcena_actual)
        incapacidad_maternidad = calcular_incapacidad_maternidad(request, prenomina, catorcena_actual)
        
        #numero de catorena
        catorcena_num = catorcena_actual.catorcena 
        
        incidencias = 0
        incidencias_retardos = 0
        
        if faltas > 0:
            incidencias = incidencias + faltas
            print("Faltas: ", faltas)
            
        if retardos > 0:
            incidencias_retardos = retardos // 3 #3 retardos se decuenta 1 dia
            
        if castigos > 0:
            incidencias = incidencias + castigos
            print("Castigos incidencias contadas", castigos)
        
        if permisos_sin_goce  > 0:
            incidencias = incidencias + permisos_sin_goce
        
        pago_doble = 0  
        if dia_extra > 0:
            pago_doble = Decimal(dia_extra * salario)
        
        incapacidad = str("")   
        incidencias_incapacidades = 0 #Modificar
        incidencias_incapacidades_pasajes = 0          
        incapacidades = 0         
        cantidad_dias_vacacion = 0
        #calculo de la prenomina - regla de tres   
        dias_de_pago = 12
        dias_laborados = dias_de_pago - (incidencias + incidencias_retardos + incidencias_incapacidades)
        proporcion_septimos_dias = Decimal((dias_laborados * 2) / 12)
        proporcion_laborados = proporcion_septimos_dias + dias_laborados
        salario_catorcenal = (proporcion_laborados * salario) + pago_doble
        
        print("las incidencias incapacidades", incidencias_incapacidades)
        if incidencias_incapacidades_pasajes > 0:
            apoyo_pasajes = (apoyo_pasajes / 12 * (12 - (incidencias + incidencias_incapacidades_pasajes))) #12 son los dias trabajados
            print("Aqui es donde se ejecuta el codigo")
        else:
            apoyo_pasajes = (apoyo_pasajes / 12 * (12 - (incidencias))) #12 son los dias trabajados
            print("Aqui no se deberia ejecutar el codigo")
        
        print("apoyos pasajes: ", apoyo_pasajes)
        print("total: ", salario_catorcenal)
        print("pagar nomina: ", apoyo_pasajes + salario_catorcenal)
        
        total_percepciones = salario_catorcenal + apoyo_pasajes + total_bonos
        total_deducciones = prestamo_infonavit + prestamo_fonacot
        pagar_nomina = total_percepciones - total_deducciones
        
        if retardos == 0: 
            retardos = ''
        
        if castigos == 0:
            castigos = ''
            
        if permisos_con_goce == 0:
            permisos_con_goce = ''
            
        if permisos_sin_goce == 0:
            permisos_sin_goce = ''
            
        if descanso == 0:
            descanso = ''

        if dia_extra == 0:
            dia_extra = ''
                    
        if incapacidades == 0:
            incapacidades = ''
        
        if faltas == 0:
            faltas = ''
        
        if comision == 0:
            comision = ''
            
        if domingo == 0:
            domingo = ''
            
        if festivos == 0:
            festivos = ''
            
        if economicos == 0:
            economicos = ''
            
        if cantidad_dias_vacacion == 0:
            cantidad_dias_vacacion = ''
    abreviaciones = {
        "economico": "D/E",
        "castigo": "CAS",
        "retardos": "R",
        "vacaciones": "V",
        "comision": "C",
        "faltas": "F",
        "permisos_con_goce": "PGS",
        "festivo": "FE",
        "incapacidades": "I",
        "permisos_sin_goce": "PSS",
        "descanso": "DEZ",
        "asistencia": "x",
        "domingo": "D",
    }
    abreviaciones_colores_cortas = {
        "D/E": "FF92d050",    # Verde claro
        "CAS": "FF948a54",    # Marrón
        "R": "FFe26b0a",      # Naranja
        "V": "FF00b0f0",      # Azul claro
        "C": "FF538dd5",      # Azul
        "F": "FFFF0000",      # Rojo
        "PGS": "FFfcd5b4",    # Beige
        "FE": "FFb1a0c7",     # Púrpura claro
        "I": "FF963634",      # Rojo oscuro
        "PSS": "FFc00000",    # Rojo oscuro
        "DEZ": "FF00b050",    # Verde
        "x": "FFFFFF",         # Blanco
        "D": "FFFF00"         # Amarillo
    }
    rows = []
    for prenomina in prenominas:
        fechas_con_etiquetas = obtener_fechas_con_incidencias(request, prenomina, catorcena_actual)
        estados_por_dia = [abreviaciones.get(estado, estado) for _, estado in fechas_con_etiquetas]
        row = (
            prenomina.empleado.status.perfil.numero_de_trabajador,
            prenomina.empleado.status.perfil.nombres + ' ' + prenomina.empleado.status.perfil.apellidos,
            prenomina.empleado.status.puesto.puesto,
            prenomina.empleado.status.perfil.proyecto.proyecto,
            prenomina.empleado.status.perfil.subproyecto.subproyecto,
            *estados_por_dia,  # Desempaquetar estados_por_dia aquí
            salario_catorcenal_costo,
            salario_catorcenal,
            apoyo_pasajes,  # Prevision social pasajes
            total_bonos,
            total_percepciones,
            prestamo_infonavit,
            prestamo_fonacot,
            total_deducciones,
            pagar_nomina,
            salario,
            ((proporcion_septimos_dias * salario) / 2)
        )
        rows.append(row)
        
        sub_salario_catorcenal_costo = sub_salario_catorcenal_costo + salario_catorcenal_costo
        sub_salario_catorcenal = sub_salario_catorcenal + salario_catorcenal
        sub_apoyo_pasajes = sub_apoyo_pasajes + apoyo_pasajes
        sub_total_bonos = sub_total_bonos + total_bonos
        sub_total_percepciones = sub_total_percepciones + total_percepciones
        sub_prestamo_infonavit = sub_prestamo_infonavit + prestamo_infonavit
        sub_prestamo_fonacot = sub_prestamo_fonacot + prestamo_fonacot
        sub_total_deducciones = sub_total_deducciones + total_deducciones
        sub_pagar_nomina = sub_pagar_nomina + pagar_nomina
        
        
                 
    # Ahora puedes usar la lista rows como lo estás haciendo actualmente en tu código
    for row_num, row in enumerate(rows, start=4):
        for col_num, value in enumerate(row, start=1):
            if col_num < 4:
                ws.cell(row=row_num, column=col_num, value=value).style = body_style
            elif col_num == 5:
                ws.cell(row=row_num, column=col_num, value=value).style = date_style
            elif 5 < col_num < 24:
                ws.cell(row=row_num, column=col_num, value=value).style = body_style
                # Verificar si el valor está en la lista de abreviaciones
                if value in abreviaciones_colores_cortas:
                    color_hex = abreviaciones_colores_cortas[value]
                    fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                    ws.cell(row=row_num, column=col_num).fill = fill
            elif col_num >= 24:
                ws.cell(row=row_num, column=col_num, value=value).style = money_style
            else:
                ws.cell(row=row_num, column=col_num, value=value).style = body_style

    add_last_row = ['Total','','','','','','','','','','','','','','','','','','',
                    sub_salario_catorcenal_costo,
                    sub_salario_catorcenal,
                    sub_apoyo_pasajes,
                    sub_total_bonos,
                    sub_total_percepciones,
                    sub_prestamo_infonavit,
                    sub_prestamo_fonacot,
                    sub_total_deducciones,
                    sub_pagar_nomina
                    ]
    ws.append(add_last_row) 

    # Agregar las abreviaciones como una tabla de dos columnas
    for key, value in abreviaciones.items():
        ws.append([key, value])

    # Aplicar el estilo a cada celda de la tabla de abreviaciones cortas con colores de fondo
    for row_num, row in enumerate(ws.iter_rows(min_row=ws.max_row - len(abreviaciones) + 1, max_row=ws.max_row), start=ws.max_row - len(abreviaciones) + 1):
        for col_num, cell in enumerate(row, start=1):
            cell.style = bold_money_style
            if cell.value is not None:
                # Obtener el color correspondiente para la abreviación corta actual
                color = abreviaciones_colores_cortas.get(cell.value, "FFFFFF")  # Por defecto, color blanco
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)