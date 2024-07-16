from django.shortcuts import render

from django.shortcuts import render,redirect
#verificar autenticacion del usuario
from django.contrib.auth.decorators import login_required
#se importa los modelos de la otra app
from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from django.core.serializers import serialize
import json
from datetime import datetime
from django.http import HttpResponse
from django.contrib import messages
from django.db.models import Sum
import os
import logging
from proyecto.models import Distrito,Perfil,Puesto,UserDatos,Catorcenas,DatosBancarios
from .models import Categoria,Subcategoria,Bono,Solicitud,BonoSolicitado,Requerimiento
from revisar.models import AutorizarSolicitudes
from .forms import SolicitudForm, BonoSolicitadoForm, RequerimientoForm,AutorizarSolicitudesUpdateForm,AutorizarSolicitudesGerenteUpdateForm
from django.db import connection
from django.core.paginator import Paginator
from .filters import SolicitudFilter,AutorizarSolicitudesFilter,BonoSolicitadoFilter
from django.db.models import Max
from django.db.models import Subquery, OuterRef
from django.http import Http404
import datetime
from datetime import date, timedelta
from django.db.models import Q
#Excel
from openpyxl import Workbook
import openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.db.models.functions import Concat
from django.db.models import Value
from django.db.models import Sum
from django.db.models import Count
from django.db.models import IntegerField
from django.db.models.functions import Cast
from django.http import HttpResponseRedirect
from datetime import datetime
from datetime import datetime, timedelta
from django.utils import timezone
from datetime import date
#pillow
from PIL import Image
from io import BytesIO
from django.core.files.uploadedfile import InMemoryUploadedFile

#Pagina inicial de los esquemas de los bonos
@login_required(login_url='user-login')
def inicio(request):
    
    bonos = Categoria.objects.all();
    
    context= {
        'bonos':bonos,
    }
    
    return render(request,'esquema/inicio.html',context)

#Listar las solicitudes
@login_required(login_url='user-login')
def listarBonosVarilleros(request):    
    #se obtiene el usuario logueado
    usuario = get_object_or_404(UserDatos,user_id = request.user.id)
    ids = [9,10,11]
    
    if usuario.tipo not in [1,2,3]:
        #se obtiene el perfil del usuario logueado
        solicitante = get_object_or_404(Perfil,numero_de_trabajador = usuario.numero_de_trabajador, distrito = usuario.distrito)
        
        subconsulta_ultima_fecha = AutorizarSolicitudes.objects.values('solicitud_id').annotate(
                ultima_fecha=Max('created_at')
            ).filter(solicitud_id=OuterRef('solicitud_id')).values('ultima_fecha')[:1]
        
        if usuario.tipo.id in [9,10,11]:
            #obtiene todas las ultimas autorizaciones de todos los distritos y roles independientemente en el flujo que se encuentre
            autorizaciones = AutorizarSolicitudes.objects.filter(
                created_at=Subquery(subconsulta_ultima_fecha)
            ).select_related('solicitud', 'perfil').filter(
                solicitud__complete = 1
            ).order_by("-created_at")
        elif usuario.tipo.id in [4,5,12]: #rh, supervisor, superintendente adm.
            #obtiene todas las ultimas autorizaciones de su distrito y roles
            autorizaciones = AutorizarSolicitudes.objects.filter(
                created_at=Subquery(subconsulta_ultima_fecha)
            ).select_related('solicitud', 'perfil').filter(
                solicitud__solicitante_id__distrito_id=solicitante.distrito_id ,solicitud__complete = 1
                #solicitud__solicitante_id__distrito_id=solicitante.distrito_id,tipo_perfil_id = usuario.tipo.id ,solicitud__complete = 1
            ).order_by("-created_at")
        else:
            #obtiene la ultima autorizacion independientemente en el flujo que se encuentre            
            autorizaciones = AutorizarSolicitudes.objects.filter(
                created_at=Subquery(subconsulta_ultima_fecha)
            ).select_related('solicitud', 'perfil').filter(
                solicitud__solicitante_id__distrito_id=solicitante.distrito_id ,solicitud__complete = 1,tipo_perfil_id=usuario.tipo.id
                #solicitud__solicitante_id__distrito_id=solicitante.distrito_id,tipo_perfil_id = usuario.tipo.id ,solicitud__complete = 1
            ).order_by("-created_at")
        
        autorizaciones_filter = AutorizarSolicitudesFilter(request.GET, queryset=autorizaciones)
        autorizaciones = autorizaciones_filter.qs
            
        p = Paginator(autorizaciones, 50)
        page = request.GET.get('page')
        salidas_list = p.get_page(page)
        autorizaciones= p.get_page(page)
        
        contexto = {
            'usuario':usuario,
            'autorizaciones':autorizaciones,
            'autorizaciones_filter': autorizaciones_filter,
            'salidas_list':salidas_list,
            'ids': ids
        }
        
        return render(request,'esquema/bonos_varilleros/listar.html',contexto)
    
    else:
        return render(request, 'revisar/403.html')
    
def comprimir_imagen(imagen):
    """
    Comprime una imagen y devuelve un objeto InMemoryUploadedFile.
    
    Args:
    imagen (InMemoryUploadedFile): Objeto de imagen subida.
    
    Returns:
    InMemoryUploadedFile: Objeto de imagen comprimida.
    """
    # Abre la imagen usando Pillow
    img = Image.open(imagen)
    
    if img.format != 'JPEG':
        img = img.convert('RGB')
    
    # Crea un flujo de bytes para almacenar la imagen comprimida
    img_temp_output = BytesIO()
    
    ancho_original, alto_original = img.size

    # Hace que el height y width se reduzca a la mitad
    nuevo_ancho = ancho_original // 2
    nuevo_alto = alto_original // 2

    # Redimensionar la imagen
    img = img.resize((nuevo_ancho, nuevo_alto))
    
    # Comprime la imagen y la guarda en el flujo de bytes
    img.save(img_temp_output, format='JPEG', quality=25, optimize=True)  # Puedes ajustar la calidad según tus necesidades
    
    # Restablece el puntero del flujo de bytes al principio
    img_temp_output.seek(0)
    
    # Crea un objeto InMemoryUploadedFile para la imagen comprimida
    img_comprimida = InMemoryUploadedFile(img_temp_output, None, imagen.name.split('.')[0] + 'b.jpg', 'image/jpeg', img_temp_output.getbuffer().nbytes, None)
    
    return img_comprimida

#para crear solicitudes de bonos
@login_required(login_url='user-login')
def crearSolicitudBonosVarilleros(request):
    usuario = request.user  
    
    #Todos los supervisores y RH pueden crear solicitudes
    if usuario.userdatos.tipo_id in (5,4):
        superintendente = UserDatos.objects.filter(distrito_id=usuario.userdatos.distrito, tipo_id=6).values('numero_de_trabajador').first()
        perfil_superintendente = Perfil.objects.filter(numero_de_trabajador = superintendente['numero_de_trabajador']).values('id').first() 
        #se obtiene el usuario logueado
        usuario = get_object_or_404(UserDatos,user_id = request.user.id)
        #se obtiene el perfil del usuario logueado
        solicitante = get_object_or_404(Perfil,numero_de_trabajador = usuario.numero_de_trabajador) 
        #se cargan los formularios con los valores del post
        solicitudForm = SolicitudForm()      
        bonoSolicitadoForm = BonoSolicitadoForm()
        requerimientoForm = RequerimientoForm()
        #se hace una consulta con los empleados del distrito que pertenecen
        empleados = Perfil.objects.filter(distrito_id = usuario.distrito.id).exclude(numero_de_trabajador = usuario.numero_de_trabajador).exclude(baja = 1).order_by('nombres')
        #se carga el formulario en automatico definiendo filtros
        bonoSolicitadoForm.fields["trabajador"].queryset = empleados 
        #crea el contexto
        contexto = {
                'usuario':usuario,
                'solicitante':solicitante,
                'solicitudForm':solicitudForm,
                'bonoSolicitadoForm':bonoSolicitadoForm,
                'requerimientoForm':requerimientoForm,
                'lista_archivos':None,
        }
        
        #Para guardar la solicitud
        if request.method == "POST":
            #obtiene el folio independientemente del formulario
            if request.POST.get('valor') is not None:
                folio = request.POST.get('valor')
            else:        
                folio = request.POST.get('folio')
                
            solicitud, created = Solicitud.objects.get_or_create(id=folio,defaults={'complete': False, 'id':folio,'folio':folio,'solicitante_id':solicitante.id, 'total':0.00})
                    
            #obtiene un queryset de los archivos de la solicitud
            lista_archivos = Requerimiento.objects.filter(solicitud_id = folio).values("id","url")
            #obtiene los bonos que han sido agregados a la solicitud
            datos_bonos_solicitud = BonoSolicitado.objects.filter(solicitud_id = folio)
            
            #se carga el valor del bono inicial 
            valor_bono = Solicitud.objects.filter(pk=folio).values("bono_id").first()
            if valor_bono is not None:
                solicitudForm = SolicitudForm(initial={'bono': valor_bono["bono_id"]}) 
                        
            #se obtiene la cantidad total
            buscar_solicitud = Solicitud.objects.filter(folio=folio).values_list("id","total").first()
            if buscar_solicitud is not None:
                total = buscar_solicitud[1]
            else:
                total = None  
                
            if 'btn_archivos' in request.POST:      
                #Se envian los formularios con datos                   
                requerimientoForm = RequerimientoForm(request.POST, request.FILES)  
                archivos = request.FILES.getlist('url')
                
                #creacion del contexto            
                contexto['solicitudForm'] = solicitudForm
                contexto['bonoSolicitadoForm'] = bonoSolicitadoForm
                contexto['requerimientoForm'] = requerimientoForm
                contexto['lista_archivos'] = lista_archivos
                contexto['datos_bonos_solicitud'] = datos_bonos_solicitud
                contexto['total'] = total
                contexto['folio'] = folio
                
                #validacion     
                if requerimientoForm.is_valid():
                    #Se recorren los archivos para ser almacenados
                    for archivo in archivos:
                        #Comprime imagenes
                        if archivo.content_type != 'application/pdf' and archivo.content_type != 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' and archivo.content_type != 'application/vnd.ms-excel':
                            documento = comprimir_imagen(archivo)
                        #cuando es un PDF
                        else:
                            documento = archivo
                        
                        #Guarda imagen o PDF
                        Requerimiento.objects.create(
                            solicitud_id = folio,
                             url = documento,
                        )
                        
                        
                    solicitud.complete_requerimiento = True
                    solicitud.save()
                    messages.success(request, "Los archivos se subieron correctamente")
                    return render(request,'esquema/bonos_varilleros/crear_solicitud.html',contexto)
                else:
                    return render(request,'esquema/bonos_varilleros/crear_solicitud.html',contexto)
            
            elif 'btn_agregar' in request.POST: 
                    
                solicitudForm = SolicitudForm(request.POST)      
                bonoSolicitadoForm = BonoSolicitadoForm(request.POST)  
                bonoSolicitadoForm.fields["trabajador"].queryset = empleados 
                
                contexto['solicitudForm'] = solicitudForm
                contexto['bonoSolicitadoForm'] = bonoSolicitadoForm
                contexto['requerimientoForm'] = requerimientoForm
                contexto['lista_archivos'] = lista_archivos
                contexto['datos_bonos_solicitud'] = datos_bonos_solicitud
                contexto['total'] = total
                contexto['folio'] = folio
                            
                #validación de los formularios
                if solicitudForm.is_valid() and bonoSolicitadoForm.is_valid():
                    #obtener los datos de los formularios
                    bono = solicitudForm.cleaned_data['bono']
                    trabajador = bonoSolicitadoForm.cleaned_data['trabajador']
                    puesto = bonoSolicitadoForm.cleaned_data['puesto']
                    cantidad = bonoSolicitadoForm.cleaned_data['cantidad']
                    
                    #se agrega el bono a la solicitud
                    solicitud.bono_id = bono
                    solicitud.save()
                    solicitud.complete_bono = True
                    solicitud.save()
                    
                    BonoSolicitado.objects.create(
                        solicitud_id = solicitud.id,
                        trabajador_id = trabajador.id,
                        puesto_id = puesto.id,
                        distrito_id = usuario.distrito.id,
                        cantidad = cantidad,
                    )
                            
                    #Actuliza la cantidad del total de la solicitud 
                    total = BonoSolicitado.objects.filter(solicitud_id = solicitud.id).values("cantidad").aggregate(total=Sum('cantidad'))['total']                 
                    Solicitud.objects.filter(pk=solicitud.id).values("total").update(total=total)
                        
                    messages.success(request, "El bono se ha agregado a la solicitud correctamente")
                            
                    #se llama el formulario vacio para que pueda agregar mas bonos
                    bonoSolicitadoForm = BonoSolicitadoForm()
                            
                    contexto['bonoSolicitadoForm'] = bonoSolicitadoForm
                    contexto['total'] = total
                                            
                    contexto["bonoSolicitadoForm"].fields["trabajador"].queryset = empleados
                        
                    return render(request,'esquema/bonos_varilleros/crear_solicitud.html',contexto)
                #se muestran los errores de validaciones      
                else:
                                                    
                    return render(request,'esquema/bonos_varilleros/crear_solicitud.html',contexto)
                            
        #Es metodo GET - Carga los formularios
        else:        
            #Genera el número de folio automaticamente
            ultimo_registro = Solicitud.objects.values('id').last()
            
            if ultimo_registro is not None:
                folio = ultimo_registro['id'] + 1 
            else:
                folio = 1
                        
            contexto['folio'] = folio
             
            return render(request,'esquema/bonos_varilleros/crear_solicitud.html',contexto)
    else:
        return render(request, 'revisar/403.html')

@login_required(login_url='user-login')  
def verificarSolicitudBonosVarilleros(request,solicitud):
    
    usuario = get_object_or_404(UserDatos, user_id=request.user.id)
    print(usuario)
    
    #solamente RH y supervisores
    if usuario.tipo.id in [4,5]:
    
        perfil = Perfil.objects.filter(numero_de_trabajador=usuario.numero_de_trabajador).values('id')
        permiso = Solicitud.objects.filter(solicitante_id=perfil[0]['id'], folio=solicitud)
        
        #checa el perfil correspondiente para cambiar la solicitud - policy
        if not permiso:
            return render(request, 'revisar/403.html')
        
        autorizacion = AutorizarSolicitudes.objects.select_related('solicitud').filter(solicitud=solicitud).last()
        requerimientoForm = RequerimientoForm()
        solicitudForm = SolicitudForm(initial={'bono': autorizacion.solicitud.bono.id})
        bonoSolicitadoForm = BonoSolicitadoForm()
        datos_bonos_solicitud = BonoSolicitado.objects.filter(solicitud_id=solicitud)
        #empleados = Perfil.objects.filter(distrito_id=usuario.distrito.id).exclude(numero_de_trabajador=usuario.numero_de_trabajador).exclude(baja=1).order_by('nombres')
        empleados = Perfil.objects.filter(empresa_id = 5).exclude(numero_de_trabajador = usuario.numero_de_trabajador).exclude(baja = 1).order_by('nombres')
        bonoSolicitadoForm.fields["trabajador"].queryset = empleados
        lista_archivos = Requerimiento.objects.filter(solicitud_id=solicitud).values("id", "url")

        if request.method == 'POST':
            if 'btn_archivos' in request.POST:
                requerimientoForm = RequerimientoForm(request.POST, request.FILES)
                archivos = request.FILES.getlist('url')

                if requerimientoForm.is_valid():
                    for archivo in archivos:
                        
                        #Comprime imagenes
                        if archivo.content_type != 'application/pdf':
                            documento = comprimir_imagen(archivo)
                        #cuando es un PDF
                        else:
                            documento = archivo
                        
                        Requerimiento.objects.create(
                            solicitud_id=solicitud,
                            url=documento,
                        )
                    messages.success(request, "Los archivos se han subido correctamente")
                
                

            elif 'btn_agregar' in request.POST:
                solicitudForm = SolicitudForm(request.POST)
                bonoSolicitadoForm = BonoSolicitadoForm(request.POST)
                bonoSolicitadoForm.fields["trabajador"].queryset = empleados

                if solicitudForm.is_valid() and bonoSolicitadoForm.is_valid():
                    bono = solicitudForm.cleaned_data['bono']
                    trabajador = bonoSolicitadoForm.cleaned_data['trabajador']
                    puesto = bonoSolicitadoForm.cleaned_data['puesto']
                    cantidad = bonoSolicitadoForm.cleaned_data['cantidad']

                    Solicitud.objects.filter(pk=solicitud).update(bono=bono)
                    
                    BonoSolicitado.objects.create(
                            solicitud_id=solicitud,
                            trabajador_id=trabajador.id,
                            puesto_id=puesto.id,
                            distrito_id=usuario.distrito.id,
                            cantidad=cantidad,
                        )
                    total = BonoSolicitado.objects.filter(solicitud_id=solicitud).values("cantidad").aggregate(total=Sum('cantidad'))['total']
                    Solicitud.objects.filter(pk=solicitud).values("total").update(total=total)
                    messages.success(request, "El bono se ha agregado a la solicitud correctamente")
                    bonoSolicitadoForm = BonoSolicitadoForm()
                    bonoSolicitadoForm.fields["trabajador"].queryset = empleados
                    

            elif 'btn_actualizar' in request.POST:
                autorizar = AutorizarSolicitudes.objects.get(solicitud_id=solicitud, tipo_perfil_id=6)
                autorizar.estado_id = 3  # pendiente
                autorizar.comentario = autorizacion.comentario
                autorizar.revisar = True
                autorizar.save()
                messages.success(request, "Se ha actualizado la solicitud y se envía al Superintendente")
                return redirect('listarBonosVarilleros')

        contexto = {
            'requerimientoForm': requerimientoForm,
            'solicitudForm': solicitudForm,
            'bonoSolicitadoForm': bonoSolicitadoForm,
            'solicitud': solicitud,
            'solicitante': autorizacion.solicitud,
            'datos_bonos_solicitud': datos_bonos_solicitud,
            'total': autorizacion.solicitud.total,
            'lista_archivos': lista_archivos,
            'estado': autorizacion
        }

        return render(request, 'esquema/bonos_varilleros/verificar_solicitud.html', contexto)
    
    else:
        return render(request, 'revisar/403.html')
        
#para ver detalles de la solicitud
@login_required(login_url='user-login')
def verDetallesSolicitud(request,solicitud_id):    
    usuario = get_object_or_404(UserDatos,user_id = request.user.id)
    if usuario.tipo not in [1,2,3]:
        #obtener_bono = Solicitud.objects.filter(pk=solicitud_id).values('bono_id').first()
        obtener_bono = Solicitud.objects.filter(pk=solicitud_id).first()
        
        if obtener_bono is None:
            raise Http404("No encontrado") 
        
        #Permisos - Los usuarios pueden ver de todos los distritos - Los demas solo pueden ver de su distrito
        if usuario.tipo.id not in [9,10,11] and usuario.distrito.id != obtener_bono.solicitante.distrito.id:
            return render(request, 'revisar/403.html')
        
        soporte_detalles = Subcategoria.objects.filter(pk=obtener_bono.bono_id).values('soporte').first()
        soporte = soporte_detalles['soporte']
        
        #los bonos solicitados
        bonos = BonoSolicitado.objects.filter(solicitud_id = solicitud_id)
        #los archivos
        requerimientos = Requerimiento.objects.filter(solicitud_id = solicitud_id)
        
        #busca la ultima solicitud con relacion a sus modelos     
        autorizaciones = AutorizarSolicitudes.objects.filter(
            solicitud__folio=solicitud_id
        ).annotate(
            ultima_fecha=Max('created_at')
        ).order_by('-ultima_fecha').first()
                
        #Para obtener el rol del solicitante
        no_trabajador = autorizaciones.solicitud.solicitante.numero_de_trabajador
        distrito = autorizaciones.solicitud.solicitante.distrito.id
        rol = UserDatos.objects.get(distrito_id = distrito, numero_de_trabajador = no_trabajador)
        rol = rol.tipo
        
        #se carga el formulario con datos iniciales
        autorizarSolicitudesUpdateForm = AutorizarSolicitudesUpdateForm(initial={'estado':autorizaciones.estado.id,'comentario':autorizaciones.comentario})
        autorizarSolicitudesGerenteUpdateForm = AutorizarSolicitudesGerenteUpdateForm(initial={'estado':autorizaciones.estado.id,'comentario':autorizaciones.comentario})
            
        contexto = {
            "usuario":usuario,
            "autorizaciones":autorizaciones,
            "bonos":bonos,
            "requerimientos": requerimientos,
            "autorizarSolicitudesUpdateForm":autorizarSolicitudesUpdateForm,
            "autorizarSolicitudesGerenteUpdateForm":autorizarSolicitudesGerenteUpdateForm,
            "soporte":soporte,
            "rol":rol
        }
        
        return render(request,'esquema/bonos_varilleros/detalles_solicitud.html',contexto)
    else:
        return render(request, 'revisar/403.html')

#lista bonos aprobados
@login_required(login_url='user-login')
def listarBonosVarillerosAprobados(request):
     
    #se obtiene el usuario logueado
    usuario = get_object_or_404(UserDatos,user_id = request.user.id)
    ids = [9,10,11]
    
    #Se muestran por catorcenas
    fecha_actual = datetime.now()
    
    catorcena_actual = Catorcenas.objects.filter(fecha_inicial__lte=fecha_actual, fecha_final__gte=fecha_actual).first()
    fecha_inicial = datetime.combine(catorcena_actual.fecha_inicial, datetime.min.time()) + timedelta(hours=00, minutes=00,seconds=00)
    fecha_final = datetime.combine(catorcena_actual.fecha_final, datetime.min.time()) + timedelta(hours=23, minutes=59,seconds=59)

    #flujo permisos y autorizaciones
    autorizaciones = None
    #Si es usuario RH de distrito matriz
    if usuario.tipo.id in (9,10,11):
        #obtiene todos los bonos aprobados de todos los distritos
        autorizaciones = AutorizarSolicitudes.objects.filter(
            solicitud__complete = 1,
            estado_id = 1,
            tipo_perfil_id = 8,
            updated_at__range=(fecha_inicial,fecha_final)
           
        ).order_by("-created_at").values('solicitud_id')
    
    elif usuario.tipo.id in (12,8):
        #obtiene todos los bonos aprobados de un solo distrito al que pertenece
        autorizaciones = AutorizarSolicitudes.objects.filter(
            solicitud__complete = 1,
            estado_id = 1,
            tipo_perfil_id = 8,
            perfil__distrito_id = usuario.distrito.id,
            updated_at__range=(fecha_inicial,fecha_final)
           
        ).order_by("-created_at").values('solicitud_id')
    
    else:
        return render(request, 'revisar/403.html')
        
    #Permisos
    if autorizaciones is None:
        return render(request, 'revisar/403.html')
    
    #se buscan los perfiles acredores al bono
    solicitudes = []
    for item in autorizaciones:
        solicitud_id = item['solicitud_id']
        solicitudes.append(solicitud_id)
        
    bonos = BonoSolicitado.objects.filter(solicitud_id__in = solicitudes).order_by('trabajador_id')
    bonosolicitado_filter = BonoSolicitadoFilter(request.GET, queryset=bonos) 
    bonos = bonosolicitado_filter.qs
    
    total_monto = bonos.aggregate(total_monto=Sum('cantidad'))['total_monto']
    cantidad_bonos_aprobados = bonos.count()
    
    if request.method =='POST' and 'excel' in request.POST:
        return convert_excel_bonos_aprobados(bonos,catorcena_actual,total_monto,cantidad_bonos_aprobados)
    
    p = Paginator(bonos, 50)
    page = request.GET.get('page')
    salidas_list = p.get_page(page)
    bonos = p.get_page(page)
    
    contexto = {
        'bonos':bonos,
        'salidas_list':salidas_list,
        'bonosolicitado_filter':bonosolicitado_filter,
        'cantidad_bonos_aprobados':cantidad_bonos_aprobados,
        'total_monto':total_monto,
        'catorcena':catorcena_actual,
        'usuario':usuario,
        'ids':ids
    }
    
    return render(request,'esquema/bonos_varilleros/listar_bonos_aprobados.html',contexto)

#generar reportes bonos aprobados
@login_required(login_url='user-login')
def generarReporteBonosVarillerosAprobados(request):
    
    #se obtiene el usuario logueado
    usuario = get_object_or_404(UserDatos,user_id = request.user.id)
    ids = [9,10,11]
    
    #Flujo de las autorizaciones y permisos
    if usuario.tipo.id in (9,10,11): #
        #se buscan los perfiles acredores al bono
        folios = Solicitud.objects.filter(fecha_autorizacion__isnull=False).values('folio')
    elif usuario.tipo.id in (4,12,8): #RH, SA, GE
        folios = Solicitud.objects.filter(fecha_autorizacion__isnull=False, solicitante__distrito_id = usuario.distrito.id).values('folio')
    else:
        return render(request, 'revisar/403.html')
        
    if not folios:
        return render(request, 'revisar/403.html')
    
    #se prepara un 
    solicitudes = []
    for item in folios:
        solicitud_id = item['folio']
        solicitudes.append(solicitud_id)
        
    bonos = BonoSolicitado.objects.filter(solicitud_id__in = solicitudes).order_by('trabajador_id')
            
    bonosolicitado_filter = BonoSolicitadoFilter(request.GET, queryset=bonos) 
    bonos = bonosolicitado_filter.qs
    
    for b in bonos:
        print(b)
            
    bono = bonos.last()
    
    if bono is not None:
        catorcena = Catorcenas.objects.filter(fecha_inicial__lte=bono.solicitud.fecha_autorizacion, fecha_final__gte=bono.solicitud.fecha_autorizacion).first()
    else:
        catorcena = None
        
    total_monto = bonos.aggregate(total_monto=Sum('cantidad'))['total_monto']
    cantidad_bonos_aprobados = bonos.count()
        
    if request.method =='POST' and 'excel' in request.POST:
        return convert_excel_bonos_aprobados(bonos,catorcena,total_monto,cantidad_bonos_aprobados)
    
    p = Paginator(bonos, 50)
    page = request.GET.get('page')
    salidas_list = p.get_page(page)
    bonos = p.get_page(page)
        
    contexto = {
        'bonos':bonos,
        'bonosolicitado_filter':bonosolicitado_filter,
        'cantidad_bonos_aprobados':cantidad_bonos_aprobados,
        'catorcena':catorcena,
        'total_monto':total_monto,
        'salidas_list':salidas_list,
        'usuario':usuario,
        'ids':ids
    }
            
    return render(request,'esquema/bonos_varilleros/generar_reporte_bonos.html',contexto)

#para remover bonos agregados
@login_required(login_url='user-login')
def removerBono(request,bono_id):
    #hacer el complete requerimiento a 0 - contar el numero de archivos cuando es 0
    if request.method == "POST":
        
        usuario = get_object_or_404(UserDatos,user_id = request.user.id)
        
        if usuario.tipo.id in (4,5):
            try:
                bono = BonoSolicitado.objects.get(pk=bono_id)
                solicitud = Solicitud.objects.get(pk=bono.solicitud_id)
                solicitud.total -= bono.cantidad
                solicitud.save()
                bono.delete()
                return JsonResponse({'bono_id': bono_id,'total':solicitud.total} ,status=200, safe=True)
            except:
                return JsonResponse({'mensaje': 'No encontrado'}, status=404,safe=True)
        else:
            return JsonResponse({'mensaje': 'Prohibido'}, status=403,safe=True)
        
#para eliminar los bonos al editar una solicitud
@login_required(login_url='user-login')
def removerBonosEditar(request, solicitud_id):
    if request.method == 'POST':
        
        usuario = get_object_or_404(UserDatos,user_id = request.user.id)
        
        if usuario.tipo.id in (4,5):
            try:
                get_object_or_404(Solicitud,pk=solicitud_id)
                BonoSolicitado.objects.filter(solicitud_id = solicitud_id).delete()
                return JsonResponse({'mensaje':'eliminados'},status=200,safe=True)
                
            except:
                return JsonResponse({'mensaje':'error del servidor'},status=500,safe=True)    
        else:
            return JsonResponse({'mensaje': 'Prohibido'}, status=403,safe=True)
        
#para remover archivos agregados
@login_required(login_url='user-login')
def removerArchivo(request,archivo_id):
    if request.method == "POST":
        
        usuario = get_object_or_404(UserDatos,user_id = request.user.id)
        if usuario.tipo.id in (4,5):
            try:
                archivo = get_object_or_404(Requerimiento,pk=archivo_id)
                
                if os.path.isfile(archivo.url.path):
                    os.remove(archivo.url.path)
                    
                archivo.delete()
                            
                return JsonResponse({"archivo_id":archivo_id},status=200,safe=False)
            except:
                return JsonResponse({'mensaje':'objecto no encontrado'},status=404,safe=True)
            
        else:
            return JsonResponse({'mensaje': 'Prohibido'}, status=403,safe=True)        

@login_required(login_url='user-login')
def EnviarSolicitudEsquemaBono(request):
    usuario = get_object_or_404(UserDatos,user_id = request.user.id)
    if usuario.tipo.id in (4,5):
        try:
            #se obtiene la solicitud desde el request 
            data = json.loads(request.body)
            #se busca la solicitud en la BD
            solicitud = Solicitud.objects.get(pk=data['solicitud'])
            #se verifica que la solicitud este complete para crear la autorizacion
            if solicitud.complete_bono == True and solicitud.complete_requerimiento == True:
                solicitud.complete = True
                solicitud.save()    
                
                usuario = request.user  
                superintendente = UserDatos.objects.filter(distrito_id=usuario.userdatos.distrito.id, tipo_id=6).values('numero_de_trabajador').first()
                perfil_superintendente = Perfil.objects.filter(numero_de_trabajador = superintendente['numero_de_trabajador']).values('id').first() 
                
                #se crea la autorizacion
                AutorizarSolicitudes.objects.create(
                    solicitud_id = solicitud.id,
                    perfil_id =  perfil_superintendente['id'],
                    tipo_perfil_id = 6, # superintendente
                    estado_id = 3, # pendiente
                )
                
                return JsonResponse({'mensaje':1},status=200,safe=False)
            else:
                #falta subir los requerimientos
                return JsonResponse({"mensaje":0},status=422,safe=False)
            
        except:
            return JsonResponse({'mensaje':'no encontrado'},status=404,safe=False) 
    else:
            return JsonResponse({'mensaje': 'Prohibido'}, status=403,safe=True)
    
#solicita la cantidad de un bono en especifico de la tabla de esquema de bonos definidos
@login_required(login_url='user-login')
def solicitarEsquemaBono(request):
    if request.method == "POST":
        #se obtiene el usuario logueado
        usuario = get_object_or_404(UserDatos,user_id = request.user.id)
        #se obtienen los datos enviados del servidor            
        data = json.loads(request.body)
        
        #response_data = {'message': data}
        #return JsonResponse(response_data,safe=False)
            
        esquema_bono = Bono.objects.filter(esquema_subcategoria_id = data['bono'], distrito_id = usuario.distrito.id, puesto_id = data['puesto'])
        serialized_data = serialize("json", esquema_bono)
        serialized_data = json.loads(serialized_data)
        return JsonResponse(serialized_data, safe=False, status=200)

        #datos = {'mensaje': 'comunicacion con el back'}
        #return JsonResponse(datos)

@login_required(login_url='user-login')
def solicitarSoporteBono(request):
     if request.method == "POST":
        #se obtienen los datos enviados del servidor            
        data = json.loads(request.body)
        subcategoria = Subcategoria.objects.get(pk=data['bono'])
        return JsonResponse({'soporte':subcategoria.soporte},status=200,safe=False)
         
         
         
#GENERACION DE REPORTES EN EXCEL
def convert_excel_bonos_aprobados(bonos,catorcena,total_monto,cantidad_bonos_aprobados):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Reporte_bonos_varilleros_aprobados_' + str(date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Reporte')
    #Comenzar en la fila 1
    row_num = 1
    
    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    number_style = NamedStyle(name='number_style', number_format='#,##0')
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY HH:MM')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)
    dato_style = NamedStyle(name='dato_style',number_format='DD/MM/YYYY')
    dato_style.font = Font(name="Arial Narrow", size = 11)
    
    
    
    #se crea el encabezado de la tabla en excel 
    columns = ['Folio','Fecha emisión','Fecha aprobación','Nombre','No. de cuenta','No. de tarjeta','Banco','Distrito','Bono','Puesto','Cantidad']
    
    #se añade el ancho de cada columna
    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        if col_num < 4:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 10
        if col_num == 4:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30
        else:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 15
            
    columna_max = len(columns)+2
    
    (ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por Savia RH. UH}')).style = messages_style
    (ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}')).style = messages_style
    (ws.cell(column = columna_max, row = 3, value='')).style = messages_style
    (ws.cell(column = columna_max, row = 5, value=f'Catorcena: {catorcena.catorcena}: {catorcena.fecha_inicial} - {catorcena.fecha_final}')).style = dato_style
    (ws.cell(column = columna_max, row = 6, value=f'Bonos aprobados: {cantidad_bonos_aprobados}')).style = messages_style
    (ws.cell(column = columna_max, row = 7, value=f'Total $: {total_monto}')).style = messages_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 45
    ws.column_dimensions[get_column_letter(columna_max + 1)].width = 45
    ws.column_dimensions[get_column_letter(columna_max + 1)].width = 45
    ws.column_dimensions[get_column_letter(columna_max + 1)].width = 45
    ws.column_dimensions[get_column_letter(columna_max + 1)].width = 45
    
    rows = []
    
    #aqui se recorre el query de los bonos y se debe formatear los objectos a un tipo de dato
    for bono in bonos:
        row = (
            bono.solicitud.folio,
            bono.fecha.strftime('%Y-%m-%d %H:%M'),
            bono.solicitud.fecha_autorizacion.strftime('%Y-%m-%d %H:%M'),
            str(bono.trabajador),
            bono.trabajador.status.datosbancarios.no_de_cuenta,
            bono.trabajador.status.datosbancarios.numero_de_tarjeta,
            str(bono.trabajador.status.datosbancarios.banco),
            str(bono.distrito),
            str(bono.solicitud.bono),
            str(bono.puesto),
            bono.cantidad
        )
        rows.append(row)
        
        #aqui se empieza el recorrido para el llenado de datos de acuerdo a su tipo
        for row_num, row in enumerate(rows, start=2):
            for col_num, value in enumerate(row, start=1):
                if col_num == 1:
                    ws.cell(row=row_num, column=col_num, value=value).style = number_style
                elif col_num == 2 or col_num == 3:
                    ws.cell(row=row_num, column=col_num, value=value).style = date_style
                elif col_num <= 9:
                    ws.cell(row=row_num, column=col_num, value=value).style = body_style
                else:
                    ws.cell(row=row_num, column=col_num, value=value).style = money_style
    
    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)
    
    return(response)

@login_required(login_url='user-login')
def tabuladorBonos(request):
    
     return render(request, 'esquema/crear_bonos/tabulador_bonos.html')
