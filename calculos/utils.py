from datetime import datetime
#Excel
from openpyxl import Workbook
import openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image
from openpyxl.styles import NamedStyle, Font, PatternFill
from openpyxl.utils import get_column_letter
from django.db.models.functions import Concat
from django.db.models import Value
from django.db.models import Sum
from django.db.models import Count
from django.db.models import IntegerField
from django.db.models.functions import Cast
from django.http import HttpResponseRedirect
from django.http import HttpResponse 

from proyecto.models import Catorcenas, Variables_imss_patronal, SalarioDatos, TablaVacaciones, DatosISR, TablaSubsidio
from revisar.models import AutorizarPrenomina
from esquema.models import BonoSolicitado
from prenomina.models import PrenominaIncidencias, TipoAguinaldo, Aguinaldo

from datetime import timedelta, datetime
from decimal import Decimal
import calendar
from dateutil.relativedelta import relativedelta
from django.contrib.auth.decorators import login_required
from django.db.models import Q

#CUOTAS IMSS
def calcular_cuotas_imss(request,sdi_imss):
    variables_patronal = Variables_imss_patronal.objects.get()
    salario_datos = SalarioDatos.objects.get()
    
    #multiplica el sdi * el % de cuatoas / el numero de dias de la catorcena
    invalidez_vida = sdi_imss * Decimal(variables_patronal.iv_obrero / 100) * 14
    cesantia_vejez = sdi_imss * Decimal(variables_patronal.cav_patron/100) * 14
    
    #obtener el salario cotizacion mensual
    salario_cot_men = sdi_imss * Decimal(30.4)
    gastos_medicos = sdi_imss * Decimal(variables_patronal.gmp_obrero/100) * 14
    en_dinero = sdi_imss * Decimal(variables_patronal.pd_obrero/100) * 14 
    
    #calcular cuota fija por cada trabajador hasta por 3 UMAs
    cuota_fija_umas = salario_datos.UMA * Decimal(30.4) * 3
    diferencia_sbc_umas = salario_cot_men - cuota_fija_umas
    cuota_fija = (diferencia_sbc_umas * Decimal(variables_patronal.cf_obrero / 100) / Decimal(30.4)) * 14
    enfermedades_maternidad = gastos_medicos + en_dinero + cuota_fija
    
    #La suma del calculo de cada resultado    
    calculo_imss = invalidez_vida + cesantia_vejez + enfermedades_maternidad
    return calculo_imss

#CALULAR ISR
def calcular_isr(salario,prima_dominical_isr,prima_vacacional_isr,aguinaldo_isr,prenomina,catorcena):           
    salario_datos = SalarioDatos.objects.get()
    limite_inferior = 0
    porcentaje = 0
    cuota_fija = 0
    
    #PRIMA DOMINICAL
    if prima_dominical_isr < salario_datos.UMA:
        #exento
        prima_dominical_isr = 0
    else:
        #gravable
        prima_dominical_isr = Decimal(prima_dominical_isr - salario_datos.UMA)
    
    #PRIMA VACACIONAL
    if prima_vacacional_isr < (salario_datos.UMA * 15):
        #exento
        prima_vacacional_isr = 0
    else:
        #gravado
        prima_vacacional_isr = Decimal(prima_vacacional_isr - (salario_datos.UMA * 15))
        
    #AGUINALDO
    if aguinaldo_isr is not None:
        if aguinaldo_isr.monto < (salario_datos.UMA * 30):
            #exento
            aguinaldo_isr = 0
        else:
            #gravado
            aguinaldo_isr = Decimal(aguinaldo_isr.monto - (salario_datos.UMA *30))            
    else:
        aguinaldo_isr = 0
    
    #Verifica que sea salario minimo y no tenga percepciones de lo contrario realiza el pago de ISR porque genera parte gravada  
    if salario == Decimal(salario_datos.Salario_minimo):
        realizar_caclulo_isr = False
        comprobar_isr = prima_dominical_isr + prima_vacacional_isr + aguinaldo_isr #comprueba el calculo de isr
        if comprobar_isr > 0:
            realizar_caclulo_isr = True
    else:
        realizar_caclulo_isr = True
        
    #Dependiendo del salario y las partes gravadas entra al flujo para el calculo de isr
    if realizar_caclulo_isr == True:
        #multiplicar el salario por 30.4
        salario_catorcenal = salario * Decimal(salario_datos.dias_mes) #30.4
        #se suman la prima dominical, vacacional, los aguinaldos para despues aplicar el calculo del isr
        salario_catorcenal = salario_catorcenal + prima_dominical_isr + prima_vacacional_isr + aguinaldo_isr
        
        #llamar la tabla de IRS
        tabla_irs = DatosISR.objects.all() #extraer para no caer en el ciclo for
        
        #obtener el valor aproximado hacia abajo para obtener las variables
        for datos_irs in tabla_irs:
            if salario_catorcenal >= datos_irs.liminf:
                limite_inferior = datos_irs.liminf
                porcentaje = datos_irs.excedente
                cuota_fija = datos_irs.cuota
                
        #realizar el calculo
        isr_mensual = ((salario_catorcenal - limite_inferior) * porcentaje) + cuota_fija
        
        isr_catorcenal = (isr_mensual / salario_datos.dias_mes) * 14
        return isr_catorcenal
    
    else:
        #error
        isr_catorcenal = 0
        return isr_catorcenal
    
#OBTENER EL TOTAL DE BONOS POR EMPLEADO Y CATORCENA
def obtener_total_bonos(request,prenomina, catorcena):
    #Fecha para obtener los bonos agregando la hora y la fecha de acuerdo a la catorcena
    fecha_inicial = datetime.combine(catorcena.fecha_inicial, datetime.min.time()) + timedelta(hours=00, minutes=00,seconds=00)
    fecha_final = datetime.combine(catorcena.fecha_final, datetime.min.time()) + timedelta(hours=23, minutes=59,seconds=59)
    
    total_bonos = BonoSolicitado.objects.filter(
        trabajador_id=prenomina.empleado.status.perfil.id,
        solicitud__fecha_autorizacion__isnull=False,
        solicitud__fecha_autorizacion__range=(fecha_inicial, fecha_final)
    ).aggregate(total=Sum('cantidad'))['total'] or 0
        
    return total_bonos

#CALCULAR INFONAVIT
def calcular_infonavit(request, infonavit):
    if infonavit == 0:
        prestamo_infonavit = Decimal(0.00)
    else:
        prestamo_infonavit = Decimal((infonavit / Decimal(30.4) ) * 14 )
    
    return prestamo_infonavit 

#CALCULAR FONACOT
def calcular_fonacot(fonacot,catorcena_actual):
    if fonacot == 0:
        prestamo_fonacot = Decimal(0.00)
    else:
        primer_dia_mes = datetime(catorcena_actual.fecha_inicial.year, catorcena_actual.fecha_inicial.month, 1).date()
        ultimo_dia_mes = datetime(catorcena_actual.fecha_inicial.year, catorcena_actual.fecha_inicial.month,
                                calendar.monthrange(catorcena_actual.fecha_inicial.year, catorcena_actual.fecha_inicial.month)[1]).date()
        numero_catorcenas =  Catorcenas.objects.filter(fecha_final__range=(primer_dia_mes,ultimo_dia_mes)).count()
        prestamo_fonacot = fonacot / numero_catorcenas
        
    return prestamo_fonacot

#PRIMA DOMINICAL
def calcular_prima_dominical(prenomina,salario):
    
    #Se obtiene las fechas de los domingos calendario
    fecha1 = prenomina.catorcena.fecha_inicial + timedelta(days=6) #primer domingo
    fecha2 = fecha1 + timedelta(days=7) #segundo domingo
        
    #se hacen dos consultas para obtener si exiten dias laborados 6, dias festivos 17 y asistencia en domingos
    dominical1 = PrenominaIncidencias.objects.filter(prenomina__empleado_id=prenomina.empleado_id,fecha = fecha1).values_list('incidencia', flat="True")
    dominical2 = PrenominaIncidencias.objects.filter(prenomina__empleado_id=prenomina.empleado_id,fecha = fecha2).values_list('incidencia', flat="True")
    
    #Se empieza el conteo de los domingos para prima
    cont = 0
    if not dominical1:
        cont+=1
    elif dominical1[0] in (6,17): #incidencia descanso laborado y festivo
        cont+=1
        
    if not dominical2: 
        cont+=1
    elif dominical2[0] in (6,17): #incidencia descanso laborado y festivo
        cont+=1
    
    #calculo de la prima dominical
    prima_dominical = Decimal(salario * Decimal(0.25)) * cont
    return prima_dominical
    
#PRIMA VACACIONAL
def calcular_prima_vacacional(prenomina):
    tipo_contrato = prenomina.empleado.status.tipo_de_contrato_id
    prima_vacacional = 0
    
    if tipo_contrato in (1,3,5,6): # planta, especial, planta 1, planta 2
        fecha_planta = prenomina.empleado.status.fecha_planta
        fecha_planta_anterior = prenomina.empleado.status.fecha_planta_anterior
                 
        #Se obtiene la fecha de planta o planta anterior
        if fecha_planta is None and fecha_planta_anterior is None: 
            fecha = None
        elif fecha_planta_anterior is not None and fecha_planta is not None:
            fecha = fecha_planta_anterior
        elif fecha_planta:
            fecha = fecha_planta
        elif fecha_planta_anterior:
            fecha = fecha_planta_anterior    
            
        if fecha is not None:
            #extraer mes y dia de la fecha para comprobar si cae en la catorcena 
            fecha_inicial = (prenomina.catorcena.fecha_inicial.month, prenomina.catorcena.fecha_inicial.day)
            fecha_final = (prenomina.catorcena.fecha_final.month, prenomina.catorcena.fecha_final.day)
            fecha_aniversario = (fecha.month, fecha.day)
                        
            #Solo se verifica el año y el mes si cae dentro de la catorcena el dia y mes
            if fecha_inicial <= fecha_aniversario <= fecha_final:
                calcular_antiguedad = relativedelta(prenomina.catorcena.fecha_final,fecha)
                antiguedad = calcular_antiguedad.years  
                if antiguedad > 0:
                    salario = prenomina.empleado.status.costo.sueldo_diario
                    dias_vacaciones = 0
                    tabla_vacaciones = TablaVacaciones.objects.all()
                    for tabla in tabla_vacaciones:
                        if antiguedad >= tabla.years:
                            dias_vacaciones = tabla.days   
                    prima_vacacional = Decimal(dias_vacaciones * salario) * Decimal(0.25)
                    return prima_vacacional
                else:
                    return prima_vacacional   
        else:
            return prima_vacacional
    #En caso que no tenga el año de antiguedad es 0
    return prima_vacacional     

def calcular_incidencias_aguinaldo(prenomina, fecha_inicio, fecha_fin):
    faltas = 0
    permiso_sin_goce = 0
    castigos = 0
    incapacidad_enfermedad = 0
    incapacidad_dias_pago = 0 #lleva el conteo de la paga de los 3 dias de la incapacidad
    
    incidencias = PrenominaIncidencias.objects.filter(
        prenomina__empleado_id=prenomina.empleado_id,
        fecha__range=(fecha_inicio, fecha_fin)
    ).select_related('incidencia_rango').values(
        'incidencia', 
        'complete',
        'fecha',
        'incidencia_rango__dia_inhabil_id'
    )

    for incidencia in incidencias:
        if incidencia['incidencia'] == 3:  # falta
            faltas += 1
        elif incidencia['incidencia'] == 8:  # permiso sin goce de sueldo
            permiso_sin_goce += 1
        elif incidencia['incidencia'] == 7:  # castigos
            castigos += 1
        elif incidencia['incidencia'] == 10:  # incapacidad enfermedad
            if incidencia['incidencia_rango__dia_inhabil_id'] == 7:
                if incidencia['fecha'].weekday() != 6:
                    if incidencia['complete']:
                        incapacidad_dias_pago +=1
                    incapacidad_enfermedad +=1
            elif incidencia['fecha'].weekend() != (incidencia['incidencia_rango__dia_inhabil_id'] - 1):
                if incidencia['complete']:
                        incapacidad_dias_pago +=1
                incapacidad_enfermedad +=1

    incidencias = faltas + permiso_sin_goce + castigos + (incapacidad_enfermedad - incapacidad_dias_pago)
    
    return incidencias

#Calcular aguinaldo enventual - Solo registra el aguinaldo en la BD - Solo se ejecuta en RH revisar, enviar no en el reporte
def calcular_aguinaldo_eventual(prenomina):
    from prenomina.views import obtener_catorcena
    """
    se realiza el calculo del registro cuando cumpla el tiempo, ademas caiga en la catorcena y se guarda en la base de datos, para guardar 
    y se considera que se va a pagar en la siguiente catorcena. Se llama en la funcion de revisar y creara o se actaulizara segun sea el caso
    """ 
    tipo_contrato = prenomina.empleado.status.tipo_de_contrato_id
    salario = prenomina.empleado.status.costo.sueldo_diario
    mes = 0 #corresponde al mes 1er ,3er y 6to
    aguinaldo = Decimal(0.00)
    
    if tipo_contrato == 2: #eventual
        fecha_ingreso =  prenomina.empleado.status.fecha_ingreso #fecha ingreso
        #se lleva un registro de los aguinaldos que se van registrando - tipo eventual
        aguinaldo_registrado = Aguinaldo.objects.filter(empleado_id = prenomina.empleado.id, tipo_id = 3).last()
        catorcena = obtener_catorcena()
        
        if aguinaldo_registrado is None or aguinaldo_registrado.catorcena.id == catorcena.id: #primer contrato
            #se realiza el calculo por los meses y por los dias laborados correspondientes a cada condicion
            fecha = fecha_ingreso + relativedelta(months=1)
            if catorcena.fecha_inicial <= fecha <= catorcena.fecha_final:
                #verifica si cae en la catorcena 
                #La fecha de ingreso se le suma un mes y se le pasa como parametros los rangos de incidencias para realizar el conteo
                fecha_fin = fecha_ingreso + relativedelta(months=1)
                #llama la funcion para conteo de las incidencias para el aguinaldo
                incidencias = calcular_incidencias_aguinaldo(prenomina,fecha_ingreso,fecha_fin)
                #calculo aguinaldo - primer mes laborado le corresponden 30 dias
                dias_laborados = 30 - incidencias
                dias_aguinaldo = Decimal(dias_laborados * 15) / 365
                aguinaldo = dias_aguinaldo * salario
                mes = 1
        else:
            if aguinaldo_registrado.mes == 1 or aguinaldo_registrado.catorcena.id == catorcena.id: #catorcena.id: #segundo contrato
                fecha = fecha_ingreso + relativedelta(months=3)
                if catorcena.fecha_inicial <= fecha <= catorcena.fecha_final:
                    #La fecha de ingreso se le suman 3 meses y se le pasa como parametros los rangos de incidencias para realizar el conteo por los tres meses
                    fecha_ingreso = fecha_ingreso + relativedelta(months=1) #rango de fechas para el segundo contrado 
                    fecha_fin = fecha_ingreso + relativedelta(months=2)
                    #llama la funcion para conteo de las incidencias para el aguinaldos
                    incidencias = calcular_incidencias_aguinaldo(prenomina,fecha_ingreso,fecha_fin)
                    #tercer mes laborado le corresponden 60 dias
                    dias_laborados = 60 - incidencias
                    dias_aguinaldo = Decimal(dias_laborados * 15) / 365
                    aguinaldo = dias_aguinaldo * salario
                    mes = 3
            elif aguinaldo_registrado.mes == 3 or aguinaldo_registrado.catorcena.id == catorcena.id: #Tercer contrato
                fecha = fecha_ingreso + relativedelta(months=6)
                if catorcena.fecha_inicial <= fecha <= catorcena.fecha_final:
                    fecha_ingreso = fecha_ingreso + relativedelta(months=3)
                    fecha_fin = fecha_ingreso + relativedelta(months=3)
                    #llama la funcion para conteo de las incidencias para el aguinaldos               
                    incidencias = calcular_incidencias_aguinaldo(prenomina,fecha_ingreso,fecha_fin)
                    #sexto mes laborado le corresponden 90 dias
                    dias_laborados = 90 - incidencias
                    dias_aguinaldo = Decimal(dias_laborados * 15) / 365
                    aguinaldo = dias_aguinaldo * salario
                    mes = 6
                    
        if aguinaldo != 0:
            #obtiene la catorcena actual y la fecha
            fecha_actual = datetime.now().date()
            catorcena = obtener_catorcena()
            #Guardar el aguinaldo
            
            #Actuliza la infomacion cada que ve le de enviar en revisar la infomación siempre y cuando este en la misma catorcena
            aguinaldo_contrato, created = Aguinaldo.objects.update_or_create(
                empleado_id = prenomina.empleado_id,
                catorcena_id = catorcena.id,
                defaults={
                    'monto':aguinaldo,
                    'fecha': fecha_actual,
                    'complete': False,
                    'tipo_id':3, #eventual
                    'mes':mes
                                        
                }
            )

#CALCULAR AGUINALDO ANUAL O PROPORCIONAL - Solo registra el aguinaldo en la BD - Solo se ejecuta en RH revisar, enviar no en el reporte
def calcular_aguinaldo(prenomina):
    """
    se realiza el calculo del registro cuando cumpla el tiempo y se guarda en la base de datos, para guardar 
    y se considera que se va a pagar en la siguiente catorcena. Sera llamada esta funcion de otra app, se creara o actualizara
    """
    #obtener la catorcena actual - se utiliza para comparar la catorcena cuando sea diciembre
    ahora = datetime.today()
    #ahora = datetime(2024,11,29)
    catorcena_actual = Catorcenas.objects.filter(fecha_inicial__lte=ahora, fecha_final__gte=ahora).first()    
    #obtener catorcena para generar y guardar el calculo del aguinaldo | posteriormente falta pagarlo, solo se registra
    cat_registro_aguinaldo = Catorcenas.objects.filter(fecha_inicial__month=11,fecha_final__month=12).last()
    
    
    #verifica si es la primera catorcena de diciembre  para realizar el calculo y registro del aguinaldo - la siguiente cat se paga en la prenomina
    if catorcena_actual.id == cat_registro_aguinaldo.id:
        tipo_contrato = prenomina.empleado.status.tipo_de_contrato_id
        
        if tipo_contrato in (1,3,5,6):
            
            #Verifica si el aguinaldo ya fue registrado tipo_id: 1 anual, 2 proporcional, 3 eventual | solo se registrar un aguinaldo por empleado al año
            aguinaldo = Aguinaldo.objects.filter(empleado_id=prenomina.empleado.id, fecha__year=ahora.year).exclude(tipo_id = 3).last()
            #Para que no no haga el flujo cuando el aguinaldo ya existe
            if aguinaldo is None or aguinaldo.catorcena.id == catorcena_actual.id:
                #se obtiene el tipo de contrato que aplica para el aguinaldo
                tipo_contrato = prenomina.empleado.status.tipo_de_contrato_id       
                salario = prenomina.empleado.status.costo.sueldo_diario
                
                if tipo_contrato in (1,3,5,6): # planta, especial, planta 1, planta 2
                    fecha_planta = prenomina.empleado.status.fecha_planta
                    fecha_planta_anterior = prenomina.empleado.status.fecha_planta_anterior
                            
                    #Se obtiene la fecha de planta o planta anterior
                    if fecha_planta is None and fecha_planta_anterior is None:
                        fecha = None
                    elif fecha_planta_anterior is not None and fecha_planta is not None:
                        fecha = fecha_planta_anterior
                    elif fecha_planta:
                        fecha = fecha_planta
                    elif fecha_planta_anterior:
                        fecha = fecha_planta_anterior
                        
                    #de acuerdo a la fecha se realiza el calculo para el aguinaldo
                    if fecha is not None:                    
                        #se obtiene o no los años de antiguedad del empleado
                        antiguedad = relativedelta(datetime.today(),fecha)                
                        if antiguedad.years >= 1:#Aguinaldo completo >= al 1 primer año de antiguedad
                            #aqui es con respecto al año 1/ENE/Y - 31/DIC/Y
                            año_actual = datetime.today().year
                            inicio_año = datetime(año_actual, 1, 1).date()
                            fin_año = datetime(año_actual, 12, 31).date()
                            
                            #llama funcion para el conteo de las incidencias necesarias para el aguinaldo
                            total_incidencias = calcular_incidencias_aguinaldo(prenomina, inicio_año,fin_año)
                            #se restan las incidencias con los dias laborados proporcionales
                            dias_laborados = 365 - total_incidencias                   
                            dias_pago = Decimal((dias_laborados) * 15 / 365)
                            aguinaldo = Decimal(dias_pago * salario)
                            
                            #Se guarda o actualiza el aguinaldo
                            aguinaldo_contrato, created = Aguinaldo.objects.update_or_create(
                                empleado_id = prenomina.empleado_id,
                                catorcena_id = catorcena_actual.id,
                                defaults={
                                    'monto':aguinaldo,
                                    'fecha': ahora,
                                    'complete': False,
                                    'tipo_id':1, #Anual
                                }
                            )
                                        
                        else: #aguinaldo proporcional 
                            #No cumple el año y se obtiene el proporcional
                            fecha_final = datetime(datetime.today().year, 12, 31).date() #obtener fin de año
                            diferencia = fecha_final - fecha
                            dias_aguinaldo = diferencia.days #total de dias laborados
                            #llama funcion para el conteo de las incidencias necesarias para el aguinaldo
                            total_incidencias = calcular_incidencias_aguinaldo(prenomina, fecha,fecha_final)
                            #se restan las incidencias con los dias laborados proporcionales
                            dias_laborados = dias_aguinaldo - total_incidencias   
                            dias_pago = Decimal((dias_laborados * 15)/365)
                            aguinaldo = Decimal( dias_pago * salario)
                            
                            aguinaldo_contrato, created = Aguinaldo.objects.update_or_create(
                                empleado_id = prenomina.empleado_id,
                                catorcena_id = catorcena_actual.id,
                                defaults={
                                    'monto':aguinaldo,
                                    'fecha': ahora,
                                    'complete': False,
                                    'tipo_id':2, #Proporcional
                                }
                            )
                            
def calcular_subsidio(salario, isr):
    #PENDIENTE DESARROLLO
    #el salario del empleado * dias de la catorcena
    salario_catorcenal = Decimal(salario * 14) 
    
    subsidio = 0
    
    #se obtiene el subsidio en relacion a la catorcena 14
    tabla_subsidio = TablaSubsidio.objects.all()
    for dato in tabla_subsidio:
        if salario_catorcenal >= ((float(dato.liminf)/30.4)*14):
            subsidio=dato.cuota
            
    #Depende del salario si se aplica el subsidio o no 
    isr = (isr - subsidio)
    
    return isr, subsidio

#Obtener el aguinaldo - Se ejecuta para elaborar el reporte
def obtener_aguinaldo(prenomina, catorcena):
    #Se obtiene el aguinaldo desde la base de datos
    aguinaldo = Aguinaldo.objects.filter(empleado_id=prenomina.empleado.id).last()
    
    #se verifica si existe el aguinaldo | se verifica la catorcena y se paga en la siguiente
    if aguinaldo:        
        if (aguinaldo.catorcena.id + 1) == catorcena.id:
            return aguinaldo
        else:
            return None        
    else:
        return aguinaldo
            
def calcular_incidencias(prenomina, catorcena_actual):
    # Contadores para cada tipo de incidencia
    retardos = 0
    descansos = 0
    faltas = 0
    comisiones = 0
    domingos = 0
    dia_extra = 0 #dia de descanso laborado
    festivos = 0
    economicos = 0
    castigos = 0
    permisos_sin_goce = 0
    permisos_con_goce = 0
    vacaciones = 0
    incapacidad_riesgo_laboral = 0
    incapacidad_maternidad = 0
    cont_incapacidad_enfermedad = 0
    cont_incapacidad_maternidad = 0
    cont_incapacidad_riesgo = 0
    incapacidad_enfermedad = 0
    festivo_laborado = 0    
    #ayudan a realizar los calculos de las incidencias - variables auxiliares
    incapacidad_dias_pago = 0 #incapacidad enfermedad - 3 dias pago
    
    # Obtener todas las incidencias en una sola consulta
    incidencias = PrenominaIncidencias.objects.filter(
        prenomina__empleado_id=prenomina.empleado_id,
        fecha__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final)
    ).select_related('incidencia_rango')

    # Contar las incidencias por tipo
    for incidencia in incidencias:
        
        checar = incidencia # checar se utiliza para verificar el complete de los dias pagados de incapacidad enfermedad, dia fecha de la incidencia, inhabil
        dia = incidencia #
        inhabil = incidencia.incidencia_rango
        
        incidencia = incidencia.incidencia.id
        if incidencia == 1:
            retardos += 1
        elif incidencia == 2:
            descansos += 1
        elif incidencia == 3:
            faltas += 1
        elif incidencia == 4:
            comisiones += 1
        elif incidencia == 5:
            domingos += 1
        elif incidencia == 6: # dia de descanso laborado
            dia_extra += 1 
        elif incidencia == 7:
            castigos += 1
        elif incidencia == 8:
            permisos_sin_goce += 1
        elif incidencia == 9:
            permisos_con_goce += 1
        #Incapacidad enfermedad general
        elif incidencia == 10:
            cont_incapacidad_enfermedad +=1 #lleva el numero de incapacidades    
            if inhabil.dia_inhabil == 7:
                if dia.fecha.weekday() != 6:
                    if checar.complete:
                        incapacidad_dias_pago +=1 #lleva los dias de pago
                    incapacidad_enfermedad +=1 #lleva el total incapacidad
            elif dia.fecha.weekday() != (inhabil.dia_inhabil.id - 1 ):
                if checar.complete:
                        incapacidad_dias_pago +=1 #lleva los dias de pago
                incapacidad_enfermedad +=1 #lleva el total incapacidad
        #Incapacidad riesgo laboral
        elif incidencia == 11:
            cont_incapacidad_riesgo += 1
            if inhabil.dia_inhabil == 7:
                if dia.fecha.weekday() != 6:
                    incapacidad_riesgo_laboral += 1
            elif dia.fecha.weekday() != (inhabil.dia_inhabil.id - 1 ):
                    incapacidad_riesgo_laboral += 1 #lleva el total incapacidad
        # Incapacidad maternidad
        elif incidencia == 12:
            cont_incapacidad_maternidad += 1
            if inhabil.dia_inhabil == 7:
                if dia.fecha.weekday() != 6:
                    incapacidad_maternidad += 1
            elif dia.fecha.weekday() != (inhabil.dia_inhabil.id - 1 ):
                    incapacidad_maternidad += 1 #lleva el total incapacidad
        elif incidencia == 13:
            festivos += 1
        elif incidencia == 14:
            economicos += 1
        elif incidencia == 15:
            vacaciones += 1
        elif incidencia == 17:
            festivo_laborado += 1 #festivo laborado cualquier dia de la semana
                            
    return (retardos, descansos, faltas, comisiones, domingos, dia_extra, castigos, 
            permisos_sin_goce, permisos_con_goce, incapacidad_riesgo_laboral,cont_incapacidad_riesgo, 
            incapacidad_maternidad,cont_incapacidad_maternidad,incapacidad_enfermedad,cont_incapacidad_enfermedad,incapacidad_dias_pago,
            festivos, economicos, vacaciones,festivo_laborado)
    
#GENERAR REPORTE PRENOMINA ACTUAL
@login_required(login_url='user-login')
def excel_estado_prenomina(request,prenominas,filtro,user_filter):        
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Reporte_prenominas_' + str(datetime.now())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Reporte')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 11)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 11)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 11)
    bold_money_style = NamedStyle(name='bold_money_style', number_format='$#,##0.00', font=Font(bold=True))
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)
    dato_style = NamedStyle(name='dato_style',number_format='DD/MM/YYYY')
    dato_style.font = Font(name="Arial Narrow", size = 11)
        
    columns = ['Empleado','#Trabajador','Distrito','#Catorcena','Fecha','Estado general','RH','CT','Gerencia','Autorizada','Retardos','Castigos','Permiso con goce de sueldo',
               'Permiso sin goce de sueldo','Descansos','Incapacidad Enfermedad','Incapacidad Riesgo Laboral','Incapacidad Maternidad','Faltas','Comisión','Domingo','Dia de descanso laborado','Festivos','Festivos laborados','Economicos','Vacaciones','Salario Cartocenal',
               'Previsión social', 'Total bonos','Prima Vacacional','Prima dominical','Aguinaldo','Total percepciones','Prestamo infonavit','IMSS','Fonacot','ISR Retenido','Total deducciones','Neto a pagar en nomina']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        if col_num == 1:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 50
        if col_num < 4:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30
        if col_num == 4:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30
        else:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 15


    columna_max = len(columns)+2
    
    if request.method == 'POST':
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
    (ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por Savia RH. JH}')).style = messages_style
    (ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}')).style = messages_style
    
    if filtro == False:
        #No se aplica el filtro de busquedas - Peticion enviada desde Prenomina app
        prenomina = prenominas.first()
        (ws.cell(column = columna_max, row = 4, value='{Cartorcena }')).style = messages_style
        (ws.cell(column = columna_max, row = 5, value= str(prenomina.catorcena) )).style = messages_style
    else:
        #No se aplica el filtro de busquedas - Peticion enviada desde Proyecto app
        date_range = f"{start_date} - {end_date}"
        (ws.cell(column = columna_max, row = 4, value='{Rango de Fechas }')).style = messages_style
        (ws.cell(column = columna_max, row = 5, value=date_range )).style = messages_style
            
    ws.column_dimensions[get_column_letter(columna_max)].width = 50
    ws.column_dimensions[get_column_letter(columna_max + 1)].width = 50

    rows = []

    sub_salario_catorcenal_costo = Decimal(0.00) #Valor de referencia del costo
    sub_salario_catorcenal = Decimal(0.00)
    sub_apoyo_pasajes = Decimal(0.00)
    sub_total_bonos = Decimal(0.00)
    prima_vacacional_total = Decimal(0.00)
    prima_dominical_total = Decimal(0.00)
    aguinaldo_total = Decimal(0.00)
    sub_total_percepciones = Decimal(0.00)
    sub_prestamo_infonavit = Decimal(0.00)
    sub_calculo_isr = Decimal(0.00)
    sub_calculo_imss = Decimal(0.00)
    sub_prestamo_fonacot = Decimal(0.00)
    #subsidio_total = Decimal(0.00)
    sub_total_deducciones = Decimal(0.00)
    sub_pagar_nomina = Decimal(0.00)
     
    for prenomina in prenominas:
        catorcena_actual = prenomina.catorcena
                  
        RH = AutorizarPrenomina.objects.filter(prenomina=prenomina, tipo_perfil__nombre="RH").first()
        CT = AutorizarPrenomina.objects.filter(prenomina=prenomina, tipo_perfil__nombre="Control Tecnico").first()
        G = AutorizarPrenomina.objects.filter(prenomina=prenomina, tipo_perfil__nombre="Gerencia").first()

        if G is not None and G.estado.tipo == 'aprobado':
            estado = 'aprobado'
        elif G is not None and G.estado == 'rechazado':
            estado = 'rechazado'
        else:
            estado = 'pendiente'

        if RH is None:
            RH ="Ninguno"   
        else:
            RH = str(RH.perfil.nombres)+(" ")+str(RH.perfil.apellidos)
        if CT is None:
            CT ="Ninguno"
        else:
            CT = str(CT.perfil.nombres)+(" ")+str(CT.perfil.apellidos)
        if G is None:
            G ="Ninguno"
        else:
            G = str(G.perfil.nombres)+(" ")+str(G.perfil.apellidos)
        
        #datos para obtener los calculos de la prenomina dependiendo el empleado
        salario = Decimal(prenomina.empleado.status.costo.sueldo_diario)
        apoyo_pasajes = prenomina.empleado.status.costo.apoyo_de_pasajes
        infonavit = prenomina.empleado.status.costo.amortizacion_infonavit
        fonacot = prenomina.empleado.status.costo.fonacot 
        sdi_imss = prenomina.empleado.status.costo.sdi_imss
                
        #realiza el calculo de las cuotas imss
        calculo_imss = calcular_cuotas_imss(request,sdi_imss)
        
        #obtener el total de los bonos
        total_bonos = obtener_total_bonos(request,prenomina,catorcena_actual)
                           
        #calculo del infonavit
        prestamo_infonavit = calcular_infonavit(request,infonavit)
                                        
        #calculo del fonacot
        prestamo_fonacot = calcular_fonacot(fonacot,catorcena_actual)
                
        #Extrar el conteo de las incidencias de la funcion      
        retardos, descansos, faltas, comisiones, domingos, dia_extra, castigos, permisos_sin_goce, permisos_con_goce, incapacidad_riesgo_laboral,cont_incapacidad_riesgo,incapacidad_maternidad,cont_incapacidad_maternidad,incapacidad_enfermedad,cont_incapacidad_enfermedad,incapacidad_dias_pago,festivos, economicos, vacaciones, festivo_laborado = calcular_incidencias(prenomina, catorcena_actual)           
        #numero de catorena
        catorcena_num = catorcena_actual.catorcena 
        
        incidencias = 0 #contador de incidencias
        descuento_pasajes = 0 #contador de las incidencias para descontar pasajes
        
        if faltas > 0:
            incidencias += faltas
            descuento_pasajes += faltas
            
        if retardos > 0:
            incidencias_retardos = retardos // 3 # Tres retardos se descuenta 1 dia
            incidencias += incidencias_retardos
            
        if castigos > 0:
            incidencias += castigos
            descuento_pasajes += castigos
            
        if permisos_sin_goce  > 0:
            incidencias += permisos_sin_goce
            descuento_pasajes += permisos_sin_goce 
            
        #dia de descanso laborado 
        pago_doble = 0  
        if dia_extra > 0:
            pago_doble = Decimal(dia_extra * (salario * 2))
                    
        #festivo laborado
        pago_doble_festivo = 0
        if festivo_laborado > 0:
            pago_doble_festivo = Decimal(festivo_laborado * (salario * 2))

        
        #incapacidades 
        if incapacidad_maternidad > 0:
            incidencias += incapacidad_maternidad
            descuento_pasajes += incapacidad_maternidad
            
        if incapacidad_riesgo_laboral > 0:
            incidencias += incapacidad_riesgo_laboral
            descuento_pasajes += incapacidad_riesgo_laboral
            
        if incapacidad_enfermedad > 0:
            incidencias += (incapacidad_enfermedad - incapacidad_dias_pago) #solo se consideran los primeros 3 dias que caigan en la cat y no sean subsecuentes
            descuento_pasajes += incapacidad_enfermedad #pasajes se descuentan por el total de numero de dias de la incapacidad
                
        #calcular la prima dominical
        prima_dominical = calcular_prima_dominical(prenomina,salario)
        
        #calcular la prima vacacional
        prima_vacacional = calcular_prima_vacacional(prenomina)
        
        #calcular el aguinaldo
        aguinaldo = obtener_aguinaldo(prenomina,catorcena_actual)
                    
        #realiza el calculo del ISR 
        calculo_isr = calcular_isr(salario,prima_dominical,prima_vacacional,aguinaldo,prenomina,catorcena_actual)
        
        #Como es un objeto, se sobreescribe aguinaldo para que tome un valor entero
        if aguinaldo is None:
            aguinaldo = 0
        else:
            aguinaldo = aguinaldo.monto
            
        #DESARROLLO PENDENTE - calcular el subsidio
        #total_isr, subsidio = calcular_subsidio(salario, calculo_isr)
    
        #Pagos dobles dias laborados (No caen en domingo)
        pagos_dobles = pago_doble + pago_doble_festivo
        
        #calculo de la prenomina - regla de tres   
        dias_de_pago = 12
        
        dias_laborados = dias_de_pago - incidencias

        proporcion_septimos_dias = Decimal((dias_laborados * 2) / 12)
        proporcion_laborados = proporcion_septimos_dias + dias_laborados
        salario_catorcenal = (proporcion_laborados * salario) + pagos_dobles

        apoyo_pasajes = (apoyo_pasajes / 12 ) * (12 - (descuento_pasajes))
                  
        total_percepciones = salario_catorcenal + apoyo_pasajes + total_bonos + prima_dominical + prima_vacacional + aguinaldo
        #IMSS y el ISR
        total_deducciones = prestamo_infonavit + prestamo_fonacot + calculo_isr + calculo_imss
        pagar_nomina = (total_percepciones - total_deducciones)
        
        #Mostrar el conteo de las incidencias del empleado    
        if retardos == 0: 
            retardos = ''
        
        if castigos == 0:
            castigos = ''
            
        if permisos_con_goce  == 0:
            permisos_con_goce  = ''
            
        if permisos_sin_goce  == 0:
            permisos_sin_goce  = ''
            
        if descansos == 0:
            descansos = ''

        if dia_extra == 0:
            dia_extra = ''
            
        if cont_incapacidad_enfermedad == 0:
            cont_incapacidad_enfermedad = ''
            
        if cont_incapacidad_riesgo == 0:
            cont_incapacidad_riesgo = ''
            
        if cont_incapacidad_maternidad == 0:
            cont_incapacidad_maternidad = ''
            
        if faltas == 0:
            faltas = ''
        
        if comisiones == 0:
            comisiones = ''
            
        if domingos == 0:
            domingos = ''
            
        if festivos == 0:
            festivos = ''
            
        if festivo_laborado == 0:
            festivo_laborado = ''
            
        if economicos == 0:
            economicos  = ''
        
        if vacaciones == 0:
            vacaciones = ''
        
        # Agregar los valores a la lista rows para cada prenomina
        row = (
            prenomina.empleado.status.perfil.nombres + ' ' + prenomina.empleado.status.perfil.apellidos,
            prenomina.empleado.status.perfil.numero_de_trabajador,
            prenomina.empleado.status.perfil.distrito.distrito,
            catorcena_num,
            str(prenomina.catorcena.fecha_inicial) + " " + str(prenomina.catorcena.fecha_final),
            prenomina.estado_general,
            str(RH),
            str(CT),
            str(G),
            estado,
            retardos,
            castigos,
            permisos_con_goce,
            permisos_sin_goce,
            descansos,
            cont_incapacidad_enfermedad,
            cont_incapacidad_riesgo,
            cont_incapacidad_maternidad,
            faltas,
            comisiones,
            domingos,
            dia_extra,
            festivos,
            festivo_laborado,
            economicos,
            vacaciones,
            salario_catorcenal,
            apoyo_pasajes,
            total_bonos,
            prima_vacacional,
            prima_dominical,
            aguinaldo,
            total_percepciones,
            prestamo_infonavit,
            calculo_imss,
            prestamo_fonacot,
            #subsidio,
            calculo_isr,
            total_deducciones,
            pagar_nomina,
        )
        rows.append(row)
        
        #es la suma del total de cada columna
        sub_salario_catorcenal = sub_salario_catorcenal + salario_catorcenal
        sub_apoyo_pasajes = sub_apoyo_pasajes + apoyo_pasajes
        sub_total_bonos = sub_total_bonos + total_bonos            
        prima_vacacional_total = prima_vacacional_total + prima_vacacional
        prima_dominical_total = prima_dominical_total + prima_dominical
        aguinaldo_total   = aguinaldo_total + aguinaldo
        sub_total_percepciones = sub_total_percepciones + total_percepciones
        sub_prestamo_infonavit = sub_prestamo_infonavit + prestamo_infonavit
        sub_calculo_imss = sub_calculo_imss + calculo_imss
        sub_prestamo_fonacot = sub_prestamo_fonacot + prestamo_fonacot
        #subsidio_total = subsidio_total + subsidio
        sub_calculo_isr = sub_calculo_isr + calculo_isr
        sub_total_deducciones = sub_total_deducciones + total_deducciones
        sub_pagar_nomina = sub_pagar_nomina + pagar_nomina
        
    # Ahora puedes usar la lista rows como lo estás haciendo actualmente en tu código
    for row_num, row in enumerate(rows, start=2):
        for col_num, value in enumerate(row, start=1):
            if col_num < 4:
                ws.cell(row=row_num, column=col_num, value=value).style = body_style
            elif col_num == 5: #fecha
                ws.cell(row=row_num, column=col_num, value=value).style = date_style
            elif col_num > 5 and col_num < 27: #Salario catorcenal
                ws.cell(row=row_num, column=col_num, value=value).style = body_style
            elif col_num >= 24:
                ws.cell(row=row_num, column=col_num, value=value).style = money_style
            else:
                ws.cell(row=row_num, column=col_num, value=value).style = body_style
       
    #Muestra la suma total de cada columna             
    add_last_row = ['Total','','','','','','','','','','','','','','','','','','','','','','','','','',
                    #sub_salario_catorcenal_costo,
                    sub_salario_catorcenal,
                    sub_apoyo_pasajes,
                    sub_total_bonos,
                    prima_vacacional_total,
                    prima_dominical_total,
                    aguinaldo_total,
                    sub_total_percepciones,
                    sub_prestamo_infonavit,
                    sub_calculo_imss,
                    sub_prestamo_fonacot,
                    #subsidio_total,
                    sub_calculo_isr,
                    sub_total_deducciones,
                    sub_pagar_nomina
                    ]
    ws.append(add_last_row) 
    
    # Aplicar el estilo money_style a cada celda de la fila
    for row in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row):
        for cell in row:
            cell.style = bold_money_style
            
    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)
        
    return(response)

#SE UTILIZA PARA MOSTRAR EL REPORTE CON LOS DIAS
@login_required(login_url='user-login')
def obtener_fechas_con_incidencias(request, prenomina, catorcena_actual):
    # Crear lista de fechas entre fecha_inicial y fecha_final de la catorcena
    dias_entre_fechas = [(catorcena_actual.fecha_inicial + timedelta(days=i)) for i in range((catorcena_actual.fecha_final - catorcena_actual.fecha_inicial).days + 1)]

    # Obtener todas las incidencias en una sola consulta
    todas_incidencias = PrenominaIncidencias.objects.filter(
        prenomina__empleado_id=prenomina.empleado_id,
        fecha__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final)
    ).values('fecha', 'incidencia__id')

    # Mapeo de id de incidencia a su etiqueta
    id_to_etiqueta = {
        1: "retardos",
        2: "descanso",
        3: "faltas",
        4: "comision",
        5: "domingo",
        6: "dia_extra",
        7: "castigos",
        8: "permisos_sin_goce",
        9: "permisos_con_goce",
        10: "incapacidad_enfermedad_general",
        11: "incapacidad_riesgo_laboral",
        12: "incapacidad_maternidad",
        13: "festivo",
        14: "economicos",
        15: "vacaciones",
        17: "festivo_laborado"
    }

    # Crear un diccionario para almacenar las fechas con su incidencia
    fecha_a_etiqueta = {fecha: "asistencia" for fecha in dias_entre_fechas}

    # Asignar etiquetas a las fechas según las incidencias
    for incidencia in todas_incidencias:
        fecha = incidencia['fecha']
        incidencia_id = incidencia['incidencia__id']
        etiqueta = id_to_etiqueta.get(incidencia_id, "asistencia")
        fecha_a_etiqueta[fecha] = etiqueta

    # Crear la lista final de fechas con sus etiquetas
    fechas_con_etiquetas = [(fecha, etiqueta) for fecha, etiqueta in fecha_a_etiqueta.items()]

    return fechas_con_etiquetas

#GENERAR REPORTE PRENOMINA VISTA DIAS
@login_required(login_url='user-login')
def excel_estado_prenomina_formato(request,prenominas, user_filter, reporte):
    from datetime import datetime
    
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Reporte_prenomina_días_' + str(datetime.now())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Reporte')
    #Comenzar en la fila 1
    row_num = 3

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 11)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 11)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 11)
    bold_money_style = NamedStyle(name='bold_money_style', number_format='$#,##0.00', font=Font(bold=True))
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)
    dato_style = NamedStyle(name='dato_style',number_format='DD/MM/YYYY')
    dato_style.font = Font(name="Arial Narrow", size = 11)
    ahora = datetime.now()
    #ahora = datetime.now() + timedelta(days=10)
    # todas las fechas de la catorcena actual
    catorcena_actual = Catorcenas.objects.filter(fecha_inicial__lte=ahora, fecha_final__gte=ahora).first()
    delta = catorcena_actual.fecha_final - catorcena_actual.fecha_inicial
    if reporte == False:
        dias_entre_fechas = [catorcena_actual.fecha_inicial + timedelta(days=i) for i in range(delta.days + 1)]
        # Generar los nombres de las columnas de los días
        dias_columnas = [str(fecha.day) for fecha in dias_entre_fechas]
    else:
        dias_entre_fechas = [f"Día {i+1}" for i in range(14)]
        # Generar los nombres de las columnas de los días
        dias_columnas = [f"Día {i+1}" for i in range(14)]
        
    columns = ['No.','NOMBRE DE EMPLEADO','PUESTO','PROYECTO','SUBPROYECTO','CATORCENA','FECHA','ESTADO PRENOMINA','RH','CT','GERENCIA'] + dias_columnas + ['Salario Catorcenal','Salario Catorcenal',
               'Previsión social', 'Total bonos','Total percepciones','Prestamo infonavit','IMSS','Fonacot','Total deducciones','Neto a pagar en nomina','Salario','Salario Domingo',]
    
    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        if col_num == 1:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 50
        if col_num < 4:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30
        if col_num == 4:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30
        else:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 15


    columna_max = len(columns)+2
    
    #ahora = datetime.now()
    #ahora = datetime.now() + timedelta(days=10)
    
    #catorcena_actual = Catorcenas.objects.filter(fecha_inicial__lte=ahora, fecha_final__gte=ahora).first()
    (ws.cell(column = 1, row = 1, value='Reporte Prenomina SAVIA RH')).style = messages_style
    (ws.cell(column = 1, row = 2, value=f'Catorcena: {catorcena_actual.catorcena}: {catorcena_actual.fecha_inicial.strftime("%d/%m/%Y")} - {catorcena_actual.fecha_final.strftime("%d/%m/%Y")}')).style = dato_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 50
    ws.column_dimensions[get_column_letter(columna_max + 1)].width = 50

    rows = []

    sub_salario_catorcenal_costo = Decimal(0.00) #Valor de referencia del costo
    sub_salario_catorcenal = Decimal(0.00)
    sub_apoyo_pasajes = Decimal(0.00)
    sub_total_bonos = Decimal(0.00)
    sub_total_percepciones = Decimal(0.00)
    sub_prestamo_infonavit = Decimal(0.00)
    sub_calculo_isr = Decimal(0.00)
    sub_calculo_imss = Decimal(0.00)
    sub_prestamo_fonacot = Decimal(0.00)
    sub_total_deducciones = Decimal(0.00)
    sub_pagar_nomina = Decimal(0.00)
    abreviaciones = {
        "economicos": "D/E",
        "castigos": "CAS",
        "retardos": "R",
        "vacaciones": "V",
        "comision": "C",
        "faltas": "F",
        "permisos_con_goce": "PGS",
        "festivo": "FE",
        "incapacidades": "I",
        "permisos_sin_goce": "PSS",
        "descanso": "DEZ",
        "asistencia": "x",
        "domingo": "D",
        "festivo_laborado": "FL",
    }
    abreviaciones_colores_cortas = {
        "D/E": "FF92d050",    # Verde claro
        "CAS": "FF948a54",    # Marrón
        "R": "FFe26b0a",      # Naranja
        "V": "FF00b0f0",      # Azul claro
        "C": "FF538dd5",      # Azul
        "F": "FFFF0000",      # Rojo
        "PGS": "FFfcd5b4",    # Beige
        "FE": "FFb1a0c7",     # Púrpura claro
        "I": "FF963634",      # Rojo oscuro
        "PSS": "FFc00000",    # Rojo oscuro
        "DEZ": "FF00b050",    # Verde
        "x": "FFFFFF",         # Blanco
        "D": "FFFF00",         # Amarillo
        "FL": "FFFF00",     
    }
    for prenomina in prenominas:        
        RH = AutorizarPrenomina.objects.filter(prenomina=prenomina, tipo_perfil__nombre="RH").first()
        CT = AutorizarPrenomina.objects.filter(prenomina=prenomina, tipo_perfil__nombre="Control Tecnico").first()
        G = AutorizarPrenomina.objects.filter(prenomina=prenomina, tipo_perfil__nombre="Gerencia").first()

        if G is not None and G.estado.tipo == 'aprobado':
            estado = 'aprobado'
        elif G is not None and G.estado == 'rechazado':
            estado = 'rechazado'
        else:
            estado = 'pendiente'

        if RH is None:
            RH ="Ninguno"   
        else:
            RH = str(RH.perfil.nombres)+(" ")+str(RH.perfil.apellidos)
        if CT is None:
            CT ="Ninguno"
        else:
            CT = str(CT.perfil.nombres)+(" ")+str(CT.perfil.apellidos)
        if G is None:
            G ="Ninguno"
        else:
            G = str(G.perfil.nombres)+(" ")+str(G.perfil.apellidos)
        
        #datos para obtener los calculos de la prenomina dependiendo el empleado
        salario = Decimal(prenomina.empleado.status.costo.sueldo_diario)
        apoyo_pasajes = prenomina.empleado.status.costo.apoyo_de_pasajes
        infonavit = prenomina.empleado.status.costo.amortizacion_infonavit
        fonacot = prenomina.empleado.status.costo.fonacot 
        sdi_imss = prenomina.empleado.status.costo.sdi_imss
                
        #realiza el calculo de las cuotas imss
        calculo_imss = calcular_cuotas_imss(request,sdi_imss)
        
        #obtener el total de los bonos
        total_bonos = obtener_total_bonos(request,prenomina,catorcena_actual)
                           
        #calculo del infonavit
        prestamo_infonavit = calcular_infonavit(request,infonavit)
                                        
        #calculo del fonacot
        prestamo_fonacot = calcular_fonacot(fonacot,catorcena_actual)
        
        print("EMPLEADO: ",prenomina.empleado)
        print("infonavit", prestamo_infonavit)
        print("fonacot", prestamo_fonacot)
        print("IMSS", calculo_imss)
        print("salario: ",salario)
        
        #Extrar el conteo de las incidencias de la funcion      
        retardos, descansos, faltas, comisiones, domingos, dia_extra, castigos, permisos_sin_goce, permisos_con_goce, incapacidad_riesgo_laboral, incapacidad_maternidad, incapacidad_enfermedad,incapacidad_dias_pago,festivos, economicos, vacaciones, domingo_laborado, festivo_laborado,festivo_domingo_laborado = calcular_incidencias(prenomina, catorcena_actual)
        
        print("Es el numero de incapacidades: ", cont_incapacidad_enfermedad)
        
        #numero de catorena
        catorcena_num = catorcena_actual.catorcena 
        
        incidencias = 0 #contador de incidencias
        descuento_pasajes = 0 #contador de las incidencias para descontar pasajes
        
        if faltas > 0:
            incidencias += faltas
            descuento_pasajes += faltas
            
        if retardos > 0:
            incidencias_retardos = retardos // 3 # Tres retardos se descuenta 1 dia
            incidencias += incidencias_retardos
            
        if castigos > 0:
            incidencias += castigos
            descuento_pasajes += castigos
            
        if permisos_sin_goce  > 0:
            incidencias += permisos_sin_goce
            descuento_pasajes += permisos_sin_goce 
            
        #dia de descanso laborado 
        pago_doble = 0  
        if dia_extra > 0:
            pago_doble = Decimal(dia_extra * (salario * 2))
        
        #dia de descando laborado - domingo
        prima_dominical = 0
        pago_doble_domingo = 0  
        if domingo_laborado > 0:
            pago_doble_domingo = Decimal(domingo_laborado * (salario * 2))
            dia_extra = dia_extra + domingo_laborado
            
        #festivo laborado
        pago_doble_festivo = 0
        if festivo_laborado > 0:
            pago_doble_festivo = Decimal(festivo_laborado * (salario * 2))
            
        #festivo laborado - domingo
        pago_festivo_domingo = 0
        if festivo_domingo_laborado > 0:
            pago_festivo_domingo = Decimal(festivo_domingo_laborado * (salario * 2 ))
            festivo_laborado = festivo_laborado + festivo_domingo_laborado
        
        #incapacidades 
        if incapacidad_maternidad > 0:
            descuento_pasajes += incapacidad_maternidad
            
        if incapacidad_riesgo_laboral > 0:
            descuento_pasajes += incapacidad_riesgo_laboral
            
        if incapacidad_enfermedad > 0:
            incidencias += (incapacidad_enfermedad - incapacidad_dias_pago) #solo se consideran los primeros 3 dias que caigan en la cat y no sean subsecuentes
            descuento_pasajes += incapacidad_enfermedad #pasajes se descuentan por el total de numero de dias de la incapacidad
                 
        #calcular la prima dominical
        prima_dominical = calcular_prima_dominical(festivo_domingo_laborado,domingo_laborado,salario)
        
        #calcular la prima vacacional
        prima_vacacional = calcular_prima_vacacional(vacaciones,salario)
        
        #calcular el aguinaldo
        aguinaldo = obtener_aguinaldo(prenomina,catorcena_actual)
                    
        #realiza el calculo del ISR
        calculo_isr = calcular_isr(salario,prima_dominical,aguinaldo,prenomina,catorcena_actual)
        
        #Como es un objeto, se sobreescribe aguinaldo para que tome un valor entero
        if aguinaldo is None:
            aguinaldo = 0
        else:
            aguinaldo = aguinaldo.monto
            
        #calcular el subsidio
        total_isr, subsidio = calcular_subsidio(salario, calculo_isr)
            
        #pagos dobles de los dias laborados - domingos, domingos-laborados, festivos, festivos-domingos
        pagos_dobles = pago_doble + pago_doble_domingo + pago_doble_festivo + pago_festivo_domingo
        
        #calculo de la prenomina - regla de tres   
        dias_de_pago = 12
        print("total incidencias: ", incidencias)
        dias_laborados = dias_de_pago - incidencias
        print("estos son los dias laborados: ", dias_laborados)
        proporcion_septimos_dias = Decimal((dias_laborados * 2) / 12)
        proporcion_laborados = proporcion_septimos_dias + dias_laborados
        salario_catorcenal = (proporcion_laborados * salario) + pagos_dobles
        print("ESTE ES EL SALARIO CATORCENAL ", salario_catorcenal)
                
        print("ESTE ES EL APOYO PASAJES ahora: ", apoyo_pasajes)
        apoyo_pasajes = (apoyo_pasajes / 12 ) * (12 - (descuento_pasajes))
          
        print("apoyos pasajes: ", apoyo_pasajes)
        print("total: ", salario_catorcenal)
        print("pagar nomina: ", apoyo_pasajes + salario_catorcenal)
        
        total_percepciones = salario_catorcenal + apoyo_pasajes + total_bonos + prima_dominical + prima_vacacional + aguinaldo
        #IMSS y el ISR
        total_deducciones = prestamo_infonavit + prestamo_fonacot + total_isr + calculo_imss
        pagar_nomina = (total_percepciones - total_deducciones)
        
        #Mostrar el conteo de las incidencias del empleado    
        if retardos == 0: 
            retardos = ''
        
        if castigos == 0:
            castigos = ''
            
        if permisos_con_goce  == 0:
            permisos_con_goce  = ''
            
        if permisos_sin_goce  == 0:
            permisos_sin_goce  = ''
            
        if descansos == 0:
            descansos = ''

        if dia_extra == 0:
            dia_extra = ''
            
        if incapacidad_enfermedad == 0:
            incapacidad_enfermedad = ''
            
        if incapacidad_riesgo_laboral == 0:
            incapacidad_riesgo_laboral = ''
            
        if incapacidad_maternidad == 0:
            incapacidad_maternidad = ''
            
        if faltas == 0:
            faltas = ''
        
        if comisiones == 0:
            comisiones = ''
            
        if domingos == 0:
            domingos = ''
            
        if festivos == 0:
            festivos = ''
            
        if festivo_laborado == 0:
            festivo_laborado = ''
            
        if economicos == 0:
            economicos  = ''
        
        if vacaciones == 0:
            vacaciones = ''
        fechas_con_etiquetas = obtener_fechas_con_incidencias(request, prenomina, catorcena_actual)
        estados_por_dia = [abreviaciones.get(estado, estado) for _, estado in fechas_con_etiquetas]
        row = (
            prenomina.empleado.status.perfil.numero_de_trabajador,
            prenomina.empleado.status.perfil.nombres + ' ' + prenomina.empleado.status.perfil.apellidos,
            prenomina.empleado.status.puesto.puesto,
            prenomina.empleado.status.perfil.proyecto.proyecto,
            prenomina.empleado.status.perfil.subproyecto.subproyecto,
            catorcena_num,
            str(prenomina.catorcena.fecha_inicial) + " " + str(prenomina.catorcena.fecha_final),
            prenomina.estado_general,
            str(RH),
            str(CT),
            str(G),
            *estados_por_dia,  # Desempaquetar estados_por_dia aquí
            salario*14,
            salario_catorcenal,
            apoyo_pasajes,  # Prevision social pasajes
            total_bonos,
            total_percepciones,
            prestamo_infonavit,
            calculo_imss,
            prestamo_fonacot,
            total_deducciones,
            pagar_nomina,
            salario,
            ((proporcion_septimos_dias * salario) / 2)
        )
        rows.append(row)
        
        sub_salario_catorcenal_costo = salario*14 + sub_salario_catorcenal_costo
        sub_salario_catorcenal = sub_salario_catorcenal + salario_catorcenal
        sub_apoyo_pasajes = sub_apoyo_pasajes + apoyo_pasajes
        sub_total_bonos = sub_total_bonos + total_bonos
        sub_total_percepciones = sub_total_percepciones + total_percepciones
        sub_prestamo_infonavit = sub_prestamo_infonavit + prestamo_infonavit
        sub_calculo_imss = sub_calculo_imss + calculo_imss
        sub_prestamo_fonacot = sub_prestamo_fonacot + prestamo_fonacot
        sub_total_deducciones = sub_total_deducciones + total_deducciones
        sub_pagar_nomina = sub_pagar_nomina + pagar_nomina
        
        
                 
    # Ahora puedes usar la lista rows como lo estás haciendo actualmente en tu código
    for row_num, row in enumerate(rows, start=4):
        for col_num, value in enumerate(row, start=1):
            if col_num < 4:
                ws.cell(row=row_num, column=col_num, value=value).style = body_style
            elif col_num == 5:
                ws.cell(row=row_num, column=col_num, value=value).style = date_style
            elif 11 < col_num <= 25:
                ws.cell(row=row_num, column=col_num, value=value).style = body_style
                # Verificar si el valor está en la lista de abreviaciones
                if value in abreviaciones_colores_cortas:
                    color_hex = abreviaciones_colores_cortas[value]
                    fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                    ws.cell(row=row_num, column=col_num).fill = fill
            elif col_num > 25:
                ws.cell(row=row_num, column=col_num, value=value).style = money_style
            else:
                ws.cell(row=row_num, column=col_num, value=value).style = body_style

    add_last_row = ['Total','','','','','','','','','','','','','','','','','','','','','','','','',
                    sub_salario_catorcenal_costo,
                    sub_salario_catorcenal,
                    sub_apoyo_pasajes,
                    sub_total_bonos,
                    sub_total_percepciones,
                    sub_prestamo_infonavit,
                    sub_calculo_imss,
                    sub_prestamo_fonacot,
                    sub_total_deducciones,
                    sub_pagar_nomina
                    ]
    ws.append(add_last_row) 

    # Agregar las abreviaciones como una tabla de dos columnas
    for key, value in abreviaciones.items():
        ws.append([key, value])

    # Aplicar el estilo a cada celda de la tabla de abreviaciones cortas con colores de fondo
    for row_num, row in enumerate(ws.iter_rows(min_row=ws.max_row - len(abreviaciones) + 1, max_row=ws.max_row), start=ws.max_row - len(abreviaciones) + 1):
        for col_num, cell in enumerate(row, start=1):
            cell.style = money_style
            if cell.value is not None:
                # Obtener el color correspondiente para la abreviación corta actual
                color = abreviaciones_colores_cortas.get(cell.value, "FFFFFF")  # Por defecto, color blanco
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)