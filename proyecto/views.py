from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect
from django.http import FileResponse
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Sum, F
from django.contrib import messages
from django.db.models import Max
import locale
import math

locale.setlocale( locale.LC_ALL, '' )

from .models import DatosISR, Costo, TablaVacaciones, Perfil, Status, Uniformes, DatosBancarios, Bonos, Vacaciones, Economicos, Puesto, Empleados_Batch, RegistroPatronal, Banco, TablaFestivos
from .models import Status_Batch, Empresa, Distrito, Nivel, Contrato, Sangre, Sexo, Civil, UserDatos, Catorcenas, Uniforme, Tallas, Ropa, SubProyecto, Proyecto,Costos_Batch, Bancarios_Batch, Tallas
from .models import Seleccion, SalarioDatos, FactorIntegracion, TablaCesantia, Solicitud_economicos, Solicitud_vacaciones, Empleado_cv
from .models import Temas_comentario_solicitud_vacaciones, Trabajos_encomendados, Vacaciones_anteriores_Batch, Dia_vacacion, Datos_baja
import csv
import json

from django.http import HttpResponse
import datetime
from datetime import timedelta, date
from dateutil.relativedelta import relativedelta #Años entre 2 fechas con años bisiestos
from django.db.models.functions import Concat
#PDF generator
from django.db.models import Q, Max
from .forms import CostoForm, BonosForm, VacacionesForm, EconomicosForm, UniformesForm, DatosBancariosForm, PerfilForm, StatusForm, IsrForm,PerfilUpdateForm
from .forms import CostoUpdateForm, BancariosUpdateForm, BonosUpdateForm, VacacionesUpdateForm, EconomicosUpdateForm, StatusUpdateForm, CatorcenasForm, BajaEmpleadoUpdate
from .forms import Dias_VacacionesForm, Empleados_BatchForm, Status_BatchForm, PerfilDistritoForm, UniformeForm, Costos_BatchForm, Bancarios_BatchForm, BajaEmpleadoForm
from .forms import SolicitudEconomicosForm, SolicitudEconomicosUpdateForm, SolicitudVacacionesForm, SolicitudVacacionesUpdateForm, Vacaciones_anteriores_BatchForm, CvAgregar
from .filters import BonosFilter, Costo_historicFilter, PerfilFilter, StatusFilter, BancariosFilter, CostoFilter, VacacionesFilter, EconomicosFilter
from .filters import CatorcenasFilter, DistritoFilter, SolicitudesVacacionesFilter, SolicitudesEconomicosFilter
from decimal import Decimal
#Excel
from openpyxl import Workbook
import openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image
from openpyxl.styles import NamedStyle, Font, PatternFill
from openpyxl.utils import get_column_letter
from django.db.models.functions import Concat
from django.db.models import Value
from django.db.models import Sum
from django.db.models import Count
from django.db.models import IntegerField
from django.db.models.functions import Cast
from django.http import HttpResponseRedirect




from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.lib.pagesizes import letter,A4,landscape
import io
from reportlab.lib import colors
from reportlab.lib.colors import Color, black, blue, red, white
from reportlab.platypus import BaseDocTemplate, Frame, Paragraph, NextPageTemplate, PageBreak, PageTemplate,Table, SimpleDocTemplate,TableStyle, KeepInFrame, Spacer
import textwrap
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.utils import ImageReader
import os

@login_required(login_url='user-login')
def Principal(request):
    return render(request, 'proyecto/Principal.html')

@login_required(login_url='user-login')
def Index(request):
    return render(request, 'proyecto/Inicio.html')

@login_required(login_url='user-login')
def Tabla_isr(request):

    isrs= DatosISR.objects.all()

    context= {
        'isrs':isrs,
        }

    return render(request, 'proyecto/Tabla_isr.html',context)

@login_required(login_url='user-login')
def Tabla_catorcenas(request):

    catorcenas = Catorcenas.objects.filter(complete=True).order_by("-fecha_final")

    catorcena_filter = CatorcenasFilter(request.GET, queryset=catorcenas)
    catorcenas = catorcena_filter.qs
    context= {
        'catorcenas':catorcenas,
        'catorcena_filter':catorcena_filter,
        }

    return render(request, 'proyecto/Tabla_catorcenas.html',context)

@login_required(login_url='user-login')
def FormularioCatorcenas(request):
    catorcena,created=Catorcenas.objects.get_or_create(complete=False)
    form = CatorcenasForm()
    if request.method == 'POST' and 'btnSend' in request.POST:
        form = CatorcenasForm(request.POST,instance=catorcena)
        form.save(commit=False)

        if form.is_valid():
            messages.success(request, 'Catorcena capturada con éxito')
            catorcena.complete=True
            form.save()
            return redirect('Tabla_catorcenas')
    context = {'form':form,}

    return render(request, 'proyecto/CatorcenasForm.html',context)

@login_required(login_url='user-login')
def CatorcenasUpdate(request, pk):

    item = Catorcenas.objects.get(id=pk)

    if request.method == 'POST':
        form = CatorcenasForm(request.POST, instance=item)

        if form.is_valid():
            messages.success(request, 'Cambios guardados con éxito en la catorcena')
            item = form.save(commit=False)
            item.save()

            catorcenas = Catorcenas.objects.filter(complete=True)
            bonos = Bonos.objects.filter(complete=True)
            for bono in bonos:
                for catorcena in catorcenas:
                    if bono.fecha_bono >= catorcena.fecha_inicial:
                        bono.mes_bono = catorcena.fecha_final
                        bono.save()

            return redirect('Tabla_catorcenas')
    else:
        form = CatorcenasForm(instance=item)

    context = {'form':form,'item':item}

    return render(request, 'proyecto/Catorcenas_update.html',context)

@login_required(login_url='user-login')
def IsrUpdate(request, pk):

    item = DatosISR.objects.get(id=pk)

    if request.method == 'POST':
        form = IsrForm(request.POST, instance=item)

        messages.success(request, 'Cambios guardados con éxito en la tabla ISR')
        if form.is_valid():
            item = form.save(commit=False)
            item.save()
            return redirect('Tabla_isr')
    else:
        form = IsrForm(instance=item)

    context = {'form':form,'item':item}

    return render(request, 'proyecto/Isr_update.html',context)

@login_required(login_url='user-login')
def Dias_VacacionesUpdate(request, pk):

    item = TablaVacaciones.objects.get(id=pk)

    if request.method == 'POST':
        form = Dias_VacacionesForm(request.POST, instance=item)

        messages.success(request, 'Cambios guardados con éxito en la tabla días de vacaciones')
        if form.is_valid():
            item = form.save(commit=False)
            item.save()
            return redirect('Tabla_isr')
    else:
        form = Dias_VacacionesForm(instance=item)

    context = {
        'form':form,
        'item':item,
        }

    return render(request, 'proyecto/Dias_vacaciones_update.html',context)

@login_required(login_url='user-login')
def Tabla_dias_vacaciones(request):
    año = datetime.date.today().year
    descansos= TablaVacaciones.objects.all()

    context= {
        'descansos':descansos,
        }

    return render(request, 'proyecto/Tabla_dias_vacaciones.html',context)

@login_required(login_url='user-login')
def Perfil_vista(request):
    user_filter = UserDatos.objects.get(user=request.user)

    if user_filter.distrito.distrito == 'Matriz':
        perfiles= Perfil.objects.filter(complete=True, baja=False).order_by("numero_de_trabajador")
    else:
        perfiles= Perfil.objects.filter(distrito=user_filter.distrito,complete=True,baja=False).order_by("numero_de_trabajador")
    perfil_filter = PerfilFilter(request.GET, queryset=perfiles)
    perfiles = perfil_filter.qs

    if request.method =='POST' and 'Excel' in request.POST:
        return convert_excel_perfil(perfiles)

    p = Paginator(perfiles, 50)
    page = request.GET.get('page')
    salidas_list = p.get_page(page)

    context= {
        'perfiles':perfiles,
        'perfil_filter':perfil_filter,
        'salidas_list':salidas_list,
        }

    return render(request, 'proyecto/Perfil.html',context)

@login_required(login_url='user-login')
def Perfil_vista_baja(request):
    user_filter = UserDatos.objects.get(user=request.user)

    if user_filter.distrito.distrito == 'Matriz':
        #perfiles= Perfil.objects.filter(complete=True, baja=True).order_by("numero_de_trabajador")
        #perfiles= Datos_baja.objects.filter(perfil__in=perfiles, complete=True).order_by("fecha")
        perfiles_con_ultima_fecha = Datos_baja.objects.filter(perfil__complete=True,perfil__baja=True).values('perfil__numero_de_trabajador').annotate(max_fecha=Max('fecha'))
        perfiles = Datos_baja.objects.filter(perfil__numero_de_trabajador__in=perfiles_con_ultima_fecha.values('perfil__numero_de_trabajador'),
                                             fecha__in=perfiles_con_ultima_fecha.values('max_fecha')).order_by('perfil__numero_de_trabajador', '-fecha')
    else:
        #perfiles= Perfil.objects.filter(distrito=user_filter.distrito,complete=True,baja=True).order_by("numero_de_trabajador")
        #perfiles= Datos_baja.objects.filter(perfil__in=perfiles,complete=True).order_by("perfil__numero_de_trabajador")
        perfiles_con_ultima_fecha = Datos_baja.objects.filter(perfil__complete_status=True,perfil__distrito=user_filter.distrito,perfil__complete=True,perfil__baja=True).values('perfil__numero_de_trabajador').annotate(max_fecha=Max('fecha'))
        perfiles = Datos_baja.objects.filter(perfil__numero_de_trabajador__in=perfiles_con_ultima_fecha.values('perfil__numero_de_trabajador'),
                                             fecha__in=perfiles_con_ultima_fecha.values('max_fecha')).order_by('perfil__numero_de_trabajador', '-fecha')
    
    perfil_filter = PerfilFilter(request.GET, queryset=perfiles)
    perfiles = perfil_filter.qs

    if request.method =='POST' and 'Excel' in request.POST:
        return convert_excel_perfil_baja(perfiles)

    p = Paginator(perfiles, 50)
    page = request.GET.get('page')
    salidas_list = p.get_page(page)

    context= {
        'perfiles':perfiles,
        'perfil_filter':perfil_filter,
        'salidas_list':salidas_list,
        }

    return render(request, 'proyecto/Perfiles_baja.html',context)

@login_required(login_url='user-login')
def FormularioPerfil(request):
    user_filter = UserDatos.objects.get(user=request.user)
    empleado,created=Perfil.objects.get_or_create(complete=False)
    subproyectos = SubProyecto.objects.all()

    if user_filter.distrito.distrito == 'Matriz':
        form = PerfilForm()
    else:
        form = PerfilDistritoForm()
    ahora = datetime.date.today()

    if request.method == 'POST' and 'btnSend' in request.POST:
        if user_filter.distrito.distrito == 'Matriz':
            form = PerfilForm(request.POST, request.FILES, instance=empleado)
        else:
            form = PerfilDistritoForm(request.POST, request.FILES, instance=empleado)
        form.save(commit=False)

        if empleado.foto and empleado.foto.size > 2097152:
            messages.error(request,'El tamaño del archivo es mayor de 2 MB')
        elif empleado.numero_de_trabajador < 0:
            messages.error(request, '(Número empleado) El numero de empleado debe ser mayor o igual a 0')
        elif empleado.fecha_nacimiento >= ahora:
            messages.error(request, 'La fecha de nacimiento no puede ser mayor o igual a hoy')
        elif Perfil.objects.filter(numero_de_trabajador=empleado.numero_de_trabajador):
            messages.error(request, '(Número empleado) El numero de empleado se repite con otro')
        else:
            messages.success(request, 'Información capturada con éxito')
            if form.is_valid():
                if user_filter.distrito.distrito != 'Matriz':
                    empleado.distrito = user_filter.distrito
                nombre = Perfil.objects.get(numero_de_trabajador = user_filter.numero_de_trabajador, distrito = user_filter.distrito)
                empleado.editado = str("C:"+nombre.nombres+" "+nombre.apellidos)
                empleado.complete=True
                form.save()
                return redirect('Perfil')

    context = {
        'form':form,
        'subproyectos':subproyectos
        }

    return render(request, 'proyecto/PerfilForm.html',context)

@login_required(login_url='user-login')
def PerfilUpdate(request, pk):
    empleado = Perfil.objects.get(id=pk)
    ahora = datetime.date.today()
    subproyectos = SubProyecto.objects.all()
    registros = empleado.history.filter(complete=True)

    if request.method == 'POST' and 'btnSend' in request.POST:
        #request.FILES permite subir imagenes en el form
        form = PerfilUpdateForm(request.POST, request.FILES, instance=empleado)
        empleado = form.save(commit=False)
        if empleado.foto and empleado.foto.size > 2097152:
            messages.error(request,'El tamaño del archivo es mayor de 2 MB')
        elif empleado.fecha_nacimiento >= ahora:
            messages.error(request, 'La fecha de nacimiento no puede ser mayor o igual a hoy')
        elif form.is_valid():
            user_filter = UserDatos.objects.get(user=request.user)
            nombre = Perfil.objects.get(numero_de_trabajador = user_filter.numero_de_trabajador, distrito = user_filter.distrito)
            empleado.editado = str("U:"+nombre.nombres+" "+nombre.apellidos)
            messages.success(request, f'Cambios guardados con éxito en el perfil de {empleado.nombres} {empleado.apellidos}')
            empleado = form.save(commit=False)
            empleado.save()
            return redirect('Perfil')
    else:
        form = PerfilUpdateForm(instance=empleado)

    context = {
        'form':form,
        'empleado':empleado,
        'subproyectos':subproyectos,
        'registros':registros,
        }


    return render(request, 'proyecto/Perfil_update.html',context)

@login_required(login_url='user-login')
def Perfil_revisar(request, pk):

    empleado = Perfil.objects.get(id=pk)


    context = {
        'empleado':empleado,

        }

    return render(request, 'proyecto/Perfil_revisar.html',context)

@login_required(login_url='user-login')
def Status_vista(request):
    user_filter = UserDatos.objects.get(user=request.user)

    if user_filter.distrito.distrito == 'Matriz':
        status= Status.objects.filter(complete=True).order_by("perfil__numero_de_trabajador")
    else:
        status = Status.objects.filter(perfil__distrito = user_filter.distrito, complete=True).order_by("perfil__numero_de_trabajador")
    status_filter = StatusFilter(request.GET, queryset=status)

    status = status_filter.qs

    if request.method =='POST' and 'Excel' in request.POST:
        return convert_excel_status(status)

                    #Set up pagination
    p = Paginator(status, 50)
    page = request.GET.get('page')
    salidas_list = p.get_page(page)

    context= {
        'status':status,
        'status_filter':status_filter,
        'salidas_list':salidas_list,
        'baja': request.GET.get('baja', False)
        }
    
    return render(request, 'proyecto/Status.html',context)

@login_required(login_url='user-login')
def FormularioStatus(request):
    user_filter = UserDatos.objects.get(user=request.user)

    if user_filter.distrito.distrito == 'Matriz':
        empleados = Perfil.objects.filter(complete=True, complete_status=False, baja=False)
    else:
        empleados = Perfil.objects.filter(distrito=user_filter.distrito,complete=True, complete_status=False, baja=False)

    estado,created=Status.objects.get_or_create(complete=False)
    form = StatusForm()
    ahora = datetime.date.today()
    registro_patronal = RegistroPatronal.objects.all()
    puestos = Puesto.objects.all()
    valido = False
    if request.method == 'POST' and 'btnSend' in request.POST:
        form = StatusForm(request.POST,instance=estado)
        form.save(commit=False)
        if estado.fecha_planta_anterior == None and estado.fecha_planta == None:
            valido=True
        elif estado.fecha_planta_anterior == None:
            if estado.fecha_planta > ahora:
                messages.error(request, '(Fecha planta) La fecha no puede ser posterior a hoy')
            else:
                valido=True
        elif estado.fecha_planta == None:
            if estado.fecha_planta_anterior > ahora:
                messages.error(request, '(Fecha planta anterior) La fecha no puede ser posterior a hoy')
            else:
                valido=True
        else:
            if estado.fecha_planta_anterior > ahora:
                messages.error(request, '(Fecha planta anterior) La fecha no puede ser posterior a hoy')
            else:
                if estado.fecha_planta > ahora:
                    messages.error(request, '(Fecha planta) La fecha no puede ser posterior a hoy')
                else:
                    if estado.fecha_planta < estado.fecha_planta_anterior:
                        messages.error(request, '(Fechas) La fecha de planta anterior no puede ser posterior a la fecha de planta')
                    else:
                        valido=True
        empleado = Perfil.objects.get(id = estado.perfil.id)
        if form.is_valid() and valido  == True:
            nombre = Perfil.objects.get(numero_de_trabajador = user_filter.numero_de_trabajador, distrito = user_filter.distrito)
            estado.editado = str("C:"+nombre.nombres+" "+nombre.apellidos)
            messages.success(request, 'Información capturada con éxito')
            estado.complete=True
            form.save()
            estado.save()
            empleado.complete_status=True
            empleado.save()
            return redirect('Status')

    context = {
        'form':form,
        'empleados':empleados,
        'registro_patronal': registro_patronal,
        'puestos':puestos,
        }

    return render(request, 'proyecto/StatusForm.html',context)

@login_required(login_url='user-login')
def StatusUpdate(request, pk):
    puestos = Puesto.objects.all()
    estado = Status.objects.get(id=pk)
    ahora = datetime.date.today()
    registros = estado.history.filter(complete=True)
    valido = False
    if request.method == 'POST' and 'btnSend' in request.POST:
        form = StatusUpdateForm(request.POST, instance=estado)
        estado = form.save(commit=False)
        if estado.fecha_planta_anterior != None and estado.fecha_planta != None:
            if estado.fecha_planta_anterior > ahora:
                messages.error(request, '(Fecha planta anterior) La fecha no puede ser posterior a hoy')
            if estado.fecha_planta > ahora:
                messages.error(request, '(Fecha planta) La fecha no puede ser posterior a hoy')
            if estado.fecha_planta_anterior > estado.fecha_planta:
                messages.error(request, '(Fecha planta anterior) La fecha no puede ser posterior a fecha de planta')
            else:
                valido=True
        elif estado.fecha_planta_anterior == None and estado.fecha_planta == None:
            valido=True
        elif estado.fecha_planta == None:
            if estado.fecha_planta_anterior > ahora:
                messages.error(request, '(Fecha planta anterior) La fecha no puede ser posterior a hoy')
            else:
                valido=True
        elif estado.fecha_planta_anterior == None:
            if estado.fecha_planta > ahora:
                messages.error(request, '(Fecha planta) La fecha no puede ser posterior a hoy')
            else:
                valido=True
        if form.is_valid() and valido == True:
            user_filter = UserDatos.objects.get(user=request.user)
            nombre = Perfil.objects.get(numero_de_trabajador = user_filter.numero_de_trabajador, distrito = user_filter.distrito)
            estado.editado = str("U:"+nombre.nombres+" "+nombre.apellidos)
            messages.success(request, f'Cambios guardados con éxito en el Status de {estado.perfil.nombres} {estado.perfil.apellidos}')
            estado = form.save(commit=False)
            estado.save()
            return redirect('Status')
    else:
        form = StatusUpdateForm(instance=estado)

    context = {'form':form,'estado':estado,'puestos':puestos,'registros':registros,}

    return render(request, 'proyecto/Status_update.html',context)

@login_required(login_url='user-login')
def Status_revisar(request, pk):

    estado = Status.objects.get(id=pk)
    if estado.ultimo_contrato_vence == datetime.date(6000, 1, 1): #Esta es la manera correcta de la fecha
        estado.ultimo_contrato_vence = 'ESPECIAL'
    elif estado.ultimo_contrato_vence == datetime.date(6001, 1, 1): #Esta es la manera correcta de la fecha
        estado.ultimo_contrato_vence = 'INDETERMINADO'
    elif estado.ultimo_contrato_vence == datetime.date(6002, 1, 1): #Esta es la manera correcta de la fecha
        estado.ultimo_contrato_vence = 'HONORARIOS'
    elif estado.ultimo_contrato_vence == datetime.date(6003, 1, 1): #Esta es la manera correcta de la fecha
        estado.ultimo_contrato_vence = 'NR'
    context = {'estado':estado,}

    return render(request, 'proyecto/Status_revisar.html',context)

@login_required(login_url='user-login')
def Administrar_tablas(request):
    puestos = Puesto.objects.all()
    salario = SalarioDatos.objects.get()
    distritos = Distrito.objects.filter(complete = True)
    perfil = Perfil.objects.filter(complete = True)
    status = Status.objects.filter(complete = True)
    bancarios = DatosBancarios.objects.filter(complete = True)
    costo = Costo.objects.filter(complete = True)
    bonos = Bonos.objects.filter(complete = True)
    vacaciones = Vacaciones.objects.filter(complete = True)
    economicos = Economicos.objects.filter(complete = True)
    distrito_seleccionado = request.POST.get('distrito_seleccionado', None)
    if distrito_seleccionado != '':
        perfill = Perfil.objects.filter(distrito__distrito = distrito_seleccionado)
        statuss = Status.objects.filter(perfil__distrito__distrito = distrito_seleccionado)
        bancarioss = DatosBancarios.objects.filter(status__perfil__distrito__distrito = distrito_seleccionado)
        costoo = Costo.objects.filter(status__perfil__distrito__distrito = distrito_seleccionado)
        bonoss = Bonos.objects.filter(costo__status__perfil__distrito__distrito = distrito_seleccionado)
        vacacioness = Vacaciones.objects.filter(status__perfil__distrito__distrito = distrito_seleccionado)
        economicoss = Economicos.objects.filter(status__perfil__distrito__distrito = distrito_seleccionado)
        if request.method =='POST' and 'Excel' in request.POST:
            return excel_reporte_especifico(distrito_seleccionado,perfill,statuss,bancarioss,costoo,bonoss,vacacioness,economicoss,)
        if request.method =='POST' and 'Pdf' in request.POST:
            return reporte_pdf_especifico(distrito_seleccionado,perfill,statuss,bancarioss,costoo,bonoss,vacacioness,economicoss,)
    #    if request.method =='POST' and 'Excel2' in request.POST:
    #        return excel_reporte_puestos()
    else:
        if request.method =='POST' and 'Excel' in request.POST:
            return excel_reporte_general(perfil,status,bancarios,costo,bonos,vacaciones,economicos,)
        if request.method =='POST' and 'Pdf' in request.POST:
            return reporte_pdf_general(perfil,status,bancarios,costo,bonos,vacaciones,economicos,)
    context= {
        'distritos':distritos,
        'distrito_seleccionado':distrito_seleccionado,
        'salario':salario,
        }
    return render(request, 'proyecto/Administrar_tablas.html', context)

@login_required(login_url='user-login')
def FormularioBonos(request):
    user_filter = UserDatos.objects.get(user=request.user)

    if user_filter.distrito.distrito == 'Matriz':
        empleados= Costo.objects.filter(complete = True, status__perfil__baja=False)
    else:
        perfil = Perfil.objects.filter(distrito = user_filter.distrito, baja=False)
        empleados= Costo.objects.filter(status__perfil__id__in=perfil.all(),complete = True)

    bonos= Bonos.objects.filter(complete=True)
    catorcenas = Catorcenas.objects.filter(complete=True)
    bono,created=Bonos.objects.get_or_create(complete=False)
    form = BonosForm()
    form.fields["costo"].queryset = empleados
    if request.method == 'POST':
        form = BonosForm(request.POST,instance=bono)
        form.save(commit=False)

        trabajador = bono.costo.status
        user = DatosBancarios.objects.filter(status=trabajador).last()
        if user is None:
            messages.error(request, '(Empleado) El empleado no tiene datos bancarios')
        else:
            bono.datosbancarios = user
            if bono.monto < 0:
                messages.error(request, '(Monto) La cantidad capturada debe ser mayor o igual 0')
            else:
                for catorcena in catorcenas:
                    if bono.fecha_bono >= catorcena.fecha_inicial:
                        bono.mes_bono = catorcena.fecha_final
                if form.is_valid():
                    nombre = Perfil.objects.get(numero_de_trabajador = user_filter.numero_de_trabajador, distrito = user_filter.distrito)
                    bono.editado = str("C:"+nombre.nombres+" "+nombre.apellidos)
                    messages.success(request, 'Información capturada con éxito')
                    bono.complete=True
                    form.save()
                    return redirect('Tabla_bonos')
    context = {'form':form,'bonos':bonos,'empleados':empleados,}

    return render(request, 'proyecto/BonosForm.html',context)

@login_required(login_url='user-login')
def BonosUpdate(request, pk):
    bono = Bonos.objects.get(id=pk)
    registros = bono.history.filter(complete=True)
    catorcenas = Catorcenas.objects.filter(complete=True)
    if request.method == 'POST':
        form =BonosUpdateForm(request.POST, instance=bono)
        bono = form.save(commit=False)
        if bono.monto < 0:
            messages.error(request, '(Monto) La cantidad capturada debe ser mayor o igual 0')
        else:
            for catorcena in catorcenas:
                if bono.fecha_bono >= catorcena.fecha_inicial:
                    bono.mes_bono = catorcena.fecha_final
        if form.is_valid():
            user_filter = UserDatos.objects.get(user=request.user)
            nombre = Perfil.objects.get(numero_de_trabajador = user_filter.numero_de_trabajador, distrito = user_filter.distrito)
            bono.editado = str("C:"+nombre.nombres+" "+nombre.apellidos)
            messages.success(request, f'Cambios guardados con éxito en los bonos de {bono.costo.status.perfil.nombres} {bono.costo.status.perfil.apellidos}')
            bono = form.save(commit=False)
            bono.save()
            return redirect('Tabla_bonos')
    else:
        form = BonosUpdateForm(instance=bono)

    context = {'form':form,'bono':bono,'registros':registros,}

    return render(request, 'proyecto/Bonos_update.html',context)

@login_required(login_url='user-login')
def Tabla_uniformes(request):
    user_filter = UserDatos.objects.get(user=request.user)

    if user_filter.distrito.distrito == 'Matriz':
        status= Status.objects.filter(complete=True).order_by("perfil__numero_de_trabajador")
    else:
        perfil = Perfil.objects.filter(distrito = user_filter.distrito)
        status= Status.objects.filter(complete=True,perfil__id__in=perfil.all()).order_by("perfil__numero_de_trabajador")

    status_filter = StatusFilter(request.GET, queryset=status)
    status = status_filter.qs
    if request.method =='POST' and 'Excel' in request.POST:
        return convert_excel_uniformes(status)
                #Set up pagination
    p = Paginator(status, 50)
    page = request.GET.get('page')
    salidas_list = p.get_page(page)

    context= {
        'status':status,
        'status_filter':status_filter,
        'salidas_list':salidas_list,
        'baja': request.GET.get('baja', False)
        }

    return render(request, 'proyecto/Tabla_uniformes.html',context)

@login_required(login_url='user-login')
def Orden_uniformes(request, pk):
    status = Status.objects.get(id=pk)
    orden, created=Uniformes.objects.get_or_create(complete=False, status=status)
    if Seleccion.objects.filter(status = status,seleccionado=True,):
        seleccion = Seleccion.objects.filter(status = status, seleccionado=True,)
        ropas = Ropa.objects.filter(complete=True)
        for dato in seleccion:
            ropas = ropas.exclude(ropa=dato.ropa.ropa)
    else:
        ropas = Ropa.objects.filter(complete=True,)
    tallas = Tallas.objects.all()
    form = UniformesForm(instance=orden)
    form_uniforme = UniformeForm()
    uniformes_pedidos = Uniforme.objects.filter(orden=orden)

    if request.method == 'POST' and  "crear" in request.POST:
        form = UniformesForm(request.POST, instance=orden)
        if uniformes_pedidos.count() == 0:
            messages.error(request, 'Debe añadir al menos un elemento a la Orden')
        else:
            empleado = Status.objects.get(id=pk)
            if form.is_valid():
                messages.success(request, 'Orden realizada con exito')
                empleado.complete_uniformes = True
                orden.complete = True
                for dato in seleccion:
                    dato.seleccionado = False
                    dato.save()
                form.save()
                orden.save()
                empleado.save()

                return redirect('Tabla_uniformes')

    context= {
        'form':form,
        'status':status,
        'tallas':tallas,
        'orden':orden,
        'form_uniforme':form_uniforme,
        'uniformes_pedidos':uniformes_pedidos,
        'ropas':ropas,
        }

    return render(request, 'proyecto/Uniformes_ordenes.html',context)

@login_required(login_url='user-login')
def update_uniformes(request, pk):
    data= json.loads(request.body)
    action = data['action']
    orden_id = int(data['orden_id'])
    #ropa_id = int(data['uniforme']) Se DESFAZA AL FILTRARLO INSERVIBLE
    talla_id = int(data['talla'])
    cantidad = int(data['cantidad'])
    orden = Uniformes.objects.get(id = orden_id)

    talla = Tallas.objects.get(id = talla_id) #talla
    prenda = Ropa.objects.get(id = talla.ropa.id) #prenda

    seleccionado, created = Seleccion.objects.get_or_create(status = orden.status, ropa = prenda) #Seleccionado
    ropa = Ropa.objects.get(id = talla.ropa.id)
    talla = Tallas.objects.get(id=talla_id)
    producto, created = Uniforme.objects.get_or_create(orden = orden, ropa = ropa, talla = talla)
    if action == "add":
        producto.cantidad = cantidad
        seleccionado.seleccionado = True
        producto.complete = True
        producto.save()
        seleccionado.save()
        messages.success(request,f'Se agregan {producto.cantidad} {producto.ropa} a la orden')
    if action == "remove":
        seleccionado.seleccionado = False
        producto.delete()
        seleccionado.save()

    return JsonResponse('Item updated, action executed: '+data["action"], safe=False)

@login_required(login_url='user-login')
def Uniformes_revisar_completados(request, pk):

    ropas = Uniformes.objects.filter(status_id=pk)
    perfil = Status.objects.get(id=pk)

    context = {'ropas':ropas,'perfil':perfil,}

    return render(request, 'proyecto/Uniformes_revisar_completados.html',context)

def Solicitudes_revisar_empleado(request):
    user_filter = UserDatos.objects.get(user=request.user)
    perfil = Perfil.objects.get(distrito=user_filter.distrito.id, numero_de_trabajador=user_filter.numero_de_trabajador)
    solicitudes_vacaciones = Solicitud_vacaciones.objects.filter(status__perfil=perfil, complete=True).order_by("-id")
    solicitudes_economicos = Solicitud_economicos.objects.filter(status__perfil=perfil, complete=True).order_by("-id")


    context = {'solicitudes_vacaciones':solicitudes_vacaciones,'solicitudes_economicos':solicitudes_economicos,}

    return render(request, 'proyecto/Solicitudes_revisar_empleado.html',context)

@login_required(login_url='user-login')
def Uniformes_revisar_ordenes(request, pk):

    ordenes = Uniforme.objects.filter(orden_id=pk)
    pedido = Uniformes.objects.get(id=pk)

    context = {'ordenes':ordenes,'pedido':pedido,}

    return render(request, 'proyecto/Uniformes_revisar_ordenes.html',context)

@login_required(login_url='user-login')
def FormularioDatosBancarios(request):
    user_filter = UserDatos.objects.get(user=request.user)

    if user_filter.distrito.distrito == 'Matriz':
        empleados= Status.objects.filter(complete = True, complete_bancarios=False, perfil__baja=False)
    else:
        perfil = Perfil.objects.filter(distrito = user_filter.distrito, baja=False)
        empleados= Status.objects.filter(perfil__id__in=perfil.all(),complete = True, complete_bancarios=False)

    bancario,created=DatosBancarios.objects.get_or_create(complete=False)
    form = DatosBancariosForm()
    form.fields["status"].queryset = empleados

    if request.method == 'POST' and 'btnSend' in request.POST:
        form = DatosBancariosForm(request.POST,instance=bancario)
        form.save(commit=False)

        if form.is_valid():
            nombre = Perfil.objects.get(numero_de_trabajador = user_filter.numero_de_trabajador, distrito = user_filter.distrito)
            bancario.editado = str("C:"+nombre.nombres+" "+nombre.apellidos)
            empleado = Status.objects.get(id = bancario.status.id)
            messages.success(request, 'Información capturada con éxito')
            bancario.complete=True
            empleado.complete_bancarios = True
            form.save()
            empleado.save()
            return redirect('Tabla_datosbancarios')


    context = {'form':form,'empleados':empleados,}

    return render(request, 'proyecto/DatosBancariosForm.html',context)

@login_required(login_url='user-login')
def BancariosUpdate(request, pk):

    item = DatosBancarios.objects.get(id=pk)
    registros = item.history.filter(complete=True)

    if request.method == 'POST':
        form = BancariosUpdateForm(request.POST, instance=item)

        if form.is_valid():
            user_filter = UserDatos.objects.get(user=request.user)
            nombre = Perfil.objects.get(numero_de_trabajador = user_filter.numero_de_trabajador, distrito = user_filter.distrito)
            item.editado = str("U:"+nombre.nombres+" "+nombre.apellidos)
            messages.success(request, f'Cambios guardados con éxito los datos bancarios de {item.status.perfil.nombres} {item.status.perfil.apellidos}')
            item = form.save(commit=False)
            item.save()
            return redirect('Tabla_datosbancarios')
    else:
        form = BancariosUpdateForm(instance=item)

    context = {'form':form,'item':item,'registros':registros,}

    return render(request, 'proyecto/Bancario_update.html',context)

@login_required(login_url='user-login')
def FormularioCosto(request):
    user_filter = UserDatos.objects.get(user=request.user)

    if user_filter.distrito.distrito == 'Matriz':
        empleados= Status.objects.filter(~Q(fecha_ingreso=None), complete = True, complete_costo = False, perfil__baja=False)
        #empleados= empleados.filter(~Q(fecha_ingreso=None))
    else:
        perfil = Perfil.objects.filter(distrito = user_filter.distrito, baja=False)
        empleados= Status.objects.filter(~Q(fecha_ingreso=None), perfil__id__in=perfil.all(),complete = True, complete_costo = False)

    tablas = DatosISR.objects.all()
    dato = SalarioDatos.objects.get()
    factores = FactorIntegracion.objects.all()
    tcesantias= TablaCesantia.objects.all() ###

    costo,created=Costo.objects.get_or_create(complete=False)
    form = CostoForm()
    form.fields["status"].queryset = empleados

    #Constantes
    quincena=Decimal(14.00)
    mes=Decimal(30.40)
    impuesto_est=Decimal(0.0315)
    sar=Decimal(0.02)
    #cesantia=Decimal(0.04625)
    cesantia=Decimal(0.0135) ###
    infonavit=Decimal(0.05)
    comision=Decimal(0.09)


    if request.method == 'POST' and 'btnSend' in request.POST:
        form = CostoForm(request.POST,instance=costo)
        form.save(commit=False)

        if costo.amortizacion_infonavit < 0:
            messages.error(request, '(Amortización) La cantidad capturada debe ser mayor a 0')
        else:
            if costo.fonacot < 0:
                messages.error(request, '(Fonacot) La cantidad capturada debe ser mayor o igual 0')
            else:
                if costo.neto_catorcenal_sin_deducciones <= 0:
                    messages.error(request, '(Neto catorcenal) La cantidad capturada debe ser mayor a 0')
                else:
                    if costo.complemento_salario_catorcenal < 0:
                        messages.error(request, '(Complemento salario) La cantidad capturada debe ser mayor o igual 0')
                    else:
                        if costo.sueldo_diario <= 0:
                            messages.error(request, '(Sueldo diario) La cantidad capturada debe ser mayor a 0')
                        else:
                            if costo.laborados <= 0:
                                messages.error(request, '(Días laborados) La cantidad capturada debe ser mayor a 0')
                            else:
                                if costo.apoyo_de_pasajes < 0:
                                    messages.error(request, '(Apoyo pasajes) La cantidad capturada debe ser mayor o igual 0')
                                else:
                                    if costo.laborados > 31:
                                        messages.error(request, '(Días laborados) La cantidad capturada debe ser menor a 31')
                                    else:
                                        if costo.apoyo_vist_familiar < 0:
                                            messages.error(request, '(Visita familiar) La cantidad capturada debe ser mayor o igual 0')
                                        else:
                                            if costo.estancia < 0:
                                                messages.error(request, '(Estancia) La cantidad capturada debe ser mayor o igual 0')
                                            else:
                                                if costo.renta < 0:
                                                    messages.error(request, '(Renta) La cantidad capturada debe ser mayor o igual 0')
                                                else:
                                                    if costo.apoyo_estudios < 0:
                                                        messages.error(request, '(Estudios) La cantidad capturada debe ser mayor o igual 0')
                                                    else:
                                                        if costo.amv < 0:
                                                            messages.error(request, '(AMV) La cantidad capturada debe ser mayor o igual 0')
                                                        else:
                                                            if costo.gasolina < 0:
                                                                messages.error(request, '(Gasolina) La cantidad capturada debe ser mayor o igual 0')
                                                            else:
                                                                if costo.campamento < 0:
                                                                    messages.error(request, '(Campamento) La cantidad capturada debe ser mayor o igual 0')
                                                                else:
                                                                    #SDI Calculo
                                                                    prima_riesgo = costo.status.registro_patronal.prima
                                                                    excedente = dato.UMA*3
                                                                    cuotafija = (dato.UMA*Decimal(0.204))*costo.laborados
                                                                    excedente_patronal = (costo.sueldo_diario-excedente)*Decimal(0.011)*costo.laborados
                                                                    excedente_obrero = (costo.sueldo_diario-excedente)*Decimal(0.004)*costo.laborados
                                                                    if excedente_patronal < 0:
                                                                        excedente_patronal = 0
                                                                    if excedente_obrero < 0:
                                                                        excedente_obrero = 0
                                                                    prestaciones_patronal = (costo.sueldo_diario*Decimal(0.007))*costo.laborados
                                                                    prestaciones_obrero = (costo.sueldo_diario*Decimal(0.0025))*costo.laborados
                                                                    gastosmp_patronal = (costo.sueldo_diario*Decimal(0.0105))*costo.laborados
                                                                    gastosmp_obrero = (costo.sueldo_diario*Decimal(0.00375))*costo.laborados
                                                                    riesgo_trabajo = (costo.sueldo_diario*(prima_riesgo/100))*costo.laborados
                                                                    invalidezvida_patronal = (costo.sueldo_diario*Decimal(0.0175))*costo.laborados
                                                                    invalidezvida_obrero = (costo.sueldo_diario*Decimal(0.00625))*costo.laborados
                                                                    guarderias_prestsociales = (costo.sueldo_diario*Decimal(0.01))*costo.laborados
                                                                    costo.imms_obrero_patronal = (cuotafija+excedente_patronal+excedente_obrero+prestaciones_patronal
                                                                                    +prestaciones_obrero+gastosmp_patronal+gastosmp_obrero+riesgo_trabajo+invalidezvida_patronal
                                                                                    +invalidezvida_obrero+guarderias_prestsociales)
                                                                    totall = costo.imms_obrero_patronal
                                                                    #Calculo de la antiguedad para el factor de integracion
                                                                    actual = date.today()
                                                                    años_ingreso = relativedelta(actual, costo.status.fecha_ingreso)
                                                                    años_ingreso = años_ingreso.years
                                                                    if años_ingreso == 0:
                                                                        años_ingreso=1
                                                                    for factor in factores:
                                                                        if años_ingreso >= factor.years:
                                                                            factor_integracion = factor.factor
                                                                    costo.sdi = factor_integracion*costo.sueldo_diario
                                                                    sdi = costo.sdi
                                                                    #Costo calculo
                                                                    costo.total_deduccion = costo.amortizacion_infonavit + costo.fonacot
                                                                    costo.neto_pagar = costo.neto_catorcenal_sin_deducciones - costo.total_deduccion
                                                                    costo.sueldo_mensual_neto = (costo.neto_catorcenal_sin_deducciones/quincena)*mes
                                                                    costo.complemento_salario_mensual = (costo.complemento_salario_catorcenal/quincena)*mes
                                                                    costo.sueldo_mensual = costo.sueldo_diario*mes
                                                                    costo.sueldo_mensual_sdi = costo.sdi*mes
                                                                    costo.total_percepciones_mensual = costo.apoyo_de_pasajes + costo.sueldo_mensual
                                                                    for tabla in tablas:
                                                                        if costo.total_percepciones_mensual >= tabla.liminf:
                                                                            costo.lim_inferior = tabla.liminf
                                                                            costo.tasa=tabla.excedente
                                                                            costo.cuota_fija=tabla.cuota
                                                                        if costo.lim_inferior >= tabla.p_ingresos:
                                                                            costo.subsidio=tabla.subsidio
                                                                    costo.impuesto_estatal= costo.total_percepciones_mensual*impuesto_est
                                                                    costo.sar= costo.sueldo_mensual_sdi*sar
                                                                    #Parte de cesantia
                                                                    busqueda_cesantia= sdi/dato.UMA ###
                                                                    for tcesantia in tcesantias:   ####
                                                                        if  busqueda_cesantia >= tcesantia.sbc:
                                                                            cesantia_valor = tcesantia.cuota_patronal
                                                                    cesantia_ley= costo.sueldo_mensual_sdi*(cesantia_valor/100)                        ###
                                                                    costo.cesantia= (costo.sueldo_mensual_sdi*cesantia)+cesantia_ley  ####
                                                                    #Parte de vacaciones
                                                                    vac_reforma_actual = Decimal((12/365)*365)*Decimal(costo.sueldo_diario)
                                                                    prima_vacacional = vac_reforma_actual*Decimal(0.25)
                                                                    aguinaldo = Decimal((15/365)*365)*Decimal(costo.sueldo_diario)
                                                                    total_vacaciones = (vac_reforma_actual+prima_vacacional+aguinaldo)/12
                                                                    #costo.cesantia= costo.sueldo_mensual_sdi*cesantia
                                                                    costo.infonavit= costo.sueldo_mensual_sdi*infonavit
                                                                    costo.excedente= costo.total_percepciones_mensual - costo.lim_inferior
                                                                    costo.impuesto_marginal= costo.excedente * costo.tasa
                                                                    costo.impuesto= costo.impuesto_marginal + costo.cuota_fija
                                                                    costo.isr= costo.impuesto
                                                                    costo.total_apoyosbonos_empleadocomp= costo.apoyo_vist_familiar + costo.estancia + costo.renta + costo.apoyo_estudios + costo.amv + costo.campamento + costo.gasolina
                                                                    costo.total_apoyosbonos_agregcomis = costo.campamento #Modificar falta suma
                                                                    costo.comision_complemeto_salario_bonos= (costo.complemento_salario_mensual + costo.campamento)*comision #Falta suma dentro de la multiplicacion
                                                                    costo.total_costo_empresa = costo.sueldo_mensual_neto + costo.complemento_salario_mensual + costo.apoyo_de_pasajes + costo.impuesto_estatal + costo.imms_obrero_patronal + costo.sar + costo.cesantia + costo.infonavit + costo.isr + costo.total_apoyosbonos_empleadocomp
                                                                    costo.total_costo_empresa = costo.total_costo_empresa + total_vacaciones
                                                                    costo.ingreso_mensual_neto_empleado= costo.sueldo_mensual_neto + costo.complemento_salario_mensual + costo.apoyo_de_pasajes + costo.total_apoyosbonos_empleadocomp # + costo.total_apoyosbonos_agregcomis
                                                                    empleado = Status.objects.get(id = costo.status.id)
                                                                    #Debes dejar lo que este entre '' para que aparezca
                                                                    if form.is_valid():
                                                                        nombre = Perfil.objects.get(numero_de_trabajador = user_filter.numero_de_trabajador, distrito = user_filter.distrito)
                                                                        costo.editado = str("C:"+nombre.nombres+" "+nombre.apellidos)
                                                                        messages.success(request, 'Datos guardados con éxito')
                                                                        costo.complete=True
                                                                        empleado.complete_costo = True
                                                                        form.save()
                                                                        empleado.save()
                                                                        return redirect('Tabla_costo')


    context = {
        'form':form,
        'empleados':empleados,
        'tablas':tablas,
        }

    return render(request, 'proyecto/CostoForm.html',context)

@login_required(login_url='user-login')
def CostoUpdate(request, pk):
    tablas= DatosISR.objects.all()
    tcesantias= TablaCesantia.objects.all() ###
    dato = SalarioDatos.objects.get()
    factores = FactorIntegracion.objects.all()
    costo = Costo.objects.get(id=pk)
    registros = costo.history.filter(~Q(amortizacion_infonavit = None))
    myfilter = Costo_historicFilter(request.GET, queryset=registros)
    registros=myfilter.qs

    comision=Decimal(0.09)
    quincena=Decimal(14.00)
    mes=Decimal(30.40)
    impuesto_est=Decimal(0.0315)
    sar=Decimal(0.02)
    #cesantia=Decimal(0.04625)
    cesantia=Decimal(0.0135) ###
    infonavit=Decimal(0.05)
    if request.method == 'POST' and 'btnSend' in request.POST:
        form = CostoUpdateForm(request.POST, instance=costo)
        form.save(commit=False)

        if costo.amortizacion_infonavit < 0:
            messages.error(request, '(Amortización) La cantidad capturada debe ser mayor o igual 0')
        else:
            if costo.fonacot < 0:
                messages.error(request, '(Fonacot) La cantidad capturada debe ser mayor o igual 0')
            else:
                if costo.neto_catorcenal_sin_deducciones <= 0:
                    messages.error(request, '(Neto catorcenal) La cantidad capturada debe ser mayor a 0')
                else:
                    if costo.complemento_salario_catorcenal < 0:
                        messages.error(request, '(Complemento salario) La cantidad capturada debe ser mayor o igual 0')
                    else:
                        if costo.sueldo_diario <= 0:
                            messages.error(request, '(Sueldo diario) La cantidad capturada debe ser mayor a 0')
                        else:
                            if costo.laborados <= 0:
                                messages.error(request, '(Días laborados) La cantidad capturada debe ser mayor a 0')
                            else:
                                if costo.apoyo_de_pasajes < 0:
                                    messages.error(request, '(Apoyo pasajes) La cantidad capturada debe ser mayor o igual 0')
                                else:
                                    if costo.laborados > 31:
                                        messages.error(request, '(Días laborados) La cantidad capturada debe ser menor a 31')
                                    else:
                                        if costo.apoyo_vist_familiar < 0:
                                            messages.error(request, '(Visita familiar) La cantidad capturada debe ser mayor o igual 0')
                                        else:
                                            if costo.estancia < 0:
                                                messages.error(request, '(Estancia) La cantidad capturada debe ser mayor o igual 0')
                                            else:
                                                if costo.renta < 0:
                                                    messages.error(request, '(Renta) La cantidad capturada debe ser mayor o igual 0')
                                                else:
                                                    if costo.apoyo_estudios < 0:
                                                        messages.error(request, '(Estudios) La cantidad capturada debe ser mayor o igual 0')
                                                    else:
                                                        if costo.amv < 0:
                                                            messages.error(request, '(AMV) La cantidad capturada debe ser mayor o igual 0')
                                                        else:
                                                            if costo.gasolina < 0:
                                                                messages.error(request, '(Gasolina) La cantidad capturada debe ser mayor o igual 0')
                                                            else:
                                                                if costo.campamento < 0:
                                                                    messages.error(request, '(Campamento) La cantidad capturada debe ser mayor o igual 0')
                                                                else:
                                                                                #SDI Calculo
                                                                    prima_riesgo = costo.status.registro_patronal.prima
                                                                    excedente = dato.UMA*3
                                                                    cuotafija = (dato.UMA*Decimal(0.204))*costo.laborados
                                                                    excedente_patronal = (costo.sueldo_diario-excedente)*Decimal(0.011)*costo.laborados
                                                                    excedente_obrero = (costo.sueldo_diario-excedente)*Decimal(0.004)*costo.laborados
                                                                    if excedente_patronal < 0:
                                                                        excedente_patronal = 0
                                                                    if excedente_obrero < 0:
                                                                        excedente_obrero = 0
                                                                    prestaciones_patronal = (costo.sueldo_diario*Decimal(0.007))*costo.laborados
                                                                    prestaciones_obrero = (costo.sueldo_diario*Decimal(0.0025))*costo.laborados
                                                                    gastosmp_patronal = (costo.sueldo_diario*Decimal(0.0105))*costo.laborados
                                                                    gastosmp_obrero = (costo.sueldo_diario*Decimal(0.00375))*costo.laborados
                                                                    riesgo_trabajo = (costo.sueldo_diario*(prima_riesgo/100))*costo.laborados
                                                                    invalidezvida_patronal = (costo.sueldo_diario*Decimal(0.0175))*costo.laborados
                                                                    invalidezvida_obrero = (costo.sueldo_diario*Decimal(0.00625))*costo.laborados
                                                                    guarderias_prestsociales = (costo.sueldo_diario*Decimal(0.01))*costo.laborados
                                                                    costo.imms_obrero_patronal = (cuotafija+excedente_patronal+excedente_obrero+prestaciones_patronal
                                                                                    +prestaciones_obrero+gastosmp_patronal+gastosmp_obrero+riesgo_trabajo+invalidezvida_patronal
                                                                                    +invalidezvida_obrero+guarderias_prestsociales)
                                                                    totall = costo.imms_obrero_patronal
                                                                    actual = date.today()
                                                                    años_ingreso = actual.year-costo.status.fecha_ingreso.year
                                                                    if años_ingreso == 0:
                                                                        años_ingreso=1
                                                                    for factor in factores:
                                                                        if años_ingreso >= factor.years:
                                                                            factor_integracion = factor.factor
                                                                    costo.sdi = factor_integracion*costo.sueldo_diario
                                                                    sdi = costo.sdi
                                                                    #Costo calculo
                                                                    costo.total_deduccion = costo.amortizacion_infonavit + costo.fonacot
                                                                    costo.neto_pagar = costo.neto_catorcenal_sin_deducciones - costo.total_deduccion
                                                                    costo.sueldo_mensual_neto = (costo.neto_catorcenal_sin_deducciones/quincena)*mes
                                                                    costo.complemento_salario_mensual = (costo.complemento_salario_catorcenal/quincena)*mes
                                                                    costo.sueldo_mensual = costo.sueldo_diario*mes
                                                                    costo.sueldo_mensual_sdi = costo.sdi*mes
                                                                    costo.total_percepciones_mensual = costo.apoyo_de_pasajes + costo.sueldo_mensual
                                                                    for tabla in tablas:
                                                                        if costo.total_percepciones_mensual >= tabla.liminf:
                                                                            costo.lim_inferior = tabla.liminf
                                                                            costo.tasa=tabla.excedente
                                                                            costo.cuota_fija=tabla.cuota
                                                                        if costo.lim_inferior >= tabla.p_ingresos:
                                                                            costo.subsidio=tabla.subsidio
                                                                    costo.impuesto_estatal= costo.total_percepciones_mensual*impuesto_est
                                                                    costo.sar= costo.sueldo_mensual_sdi*sar
                                                                    #Parte de cesantia
                                                                    busqueda_cesantia= sdi/dato.UMA ###
                                                                    for tcesantia in tcesantias:   ####
                                                                        if  busqueda_cesantia >= tcesantia.sbc:
                                                                            cesantia_valor = tcesantia.cuota_patronal
                                                                    cesantia_ley= costo.sueldo_mensual_sdi*(cesantia_valor/100)                        ###
                                                                    costo.cesantia= (costo.sueldo_mensual_sdi*cesantia)+cesantia_ley  ####
                                                                    #Parte de vacaciones
                                                                    vac_reforma_actual = Decimal((12/365)*365)*Decimal(costo.sueldo_diario)
                                                                    prima_vacacional = vac_reforma_actual*Decimal(0.25)
                                                                    aguinaldo = Decimal((15/365)*365)*Decimal(costo.sueldo_diario)
                                                                    total_vacaciones = (vac_reforma_actual+prima_vacacional+aguinaldo)/12
                                                                    #costo.cesantia= costo.sueldo_mensual_sdi*cesantia
                                                                    costo.infonavit= costo.sueldo_mensual_sdi*infonavit
                                                                    costo.excedente= costo.total_percepciones_mensual - costo.lim_inferior
                                                                    costo.impuesto_marginal= costo.excedente * costo.tasa
                                                                    costo.impuesto= costo.impuesto_marginal + costo.cuota_fija
                                                                    costo.isr= costo.impuesto
                                                                    costo.total_apoyosbonos_empleadocomp= costo.apoyo_vist_familiar + costo.estancia + costo.renta + costo.apoyo_estudios + costo.amv + costo.campamento + costo.gasolina
                                                                    costo.total_apoyosbonos_agregcomis = costo.campamento #Modificar falta suma
                                                                    costo.comision_complemeto_salario_bonos= (costo.complemento_salario_mensual + costo.campamento)*comision #Falta suma dentro de la multiplicacion
                                                                    costo.total_costo_empresa = costo.sueldo_mensual_neto + costo.complemento_salario_mensual + costo.apoyo_de_pasajes + costo.impuesto_estatal + costo.imms_obrero_patronal + costo.sar + costo.cesantia + costo.infonavit + costo.isr + costo.total_apoyosbonos_empleadocomp
                                                                    costo.total_costo_empresa = costo.total_costo_empresa + total_vacaciones
                                                                    costo.ingreso_mensual_neto_empleado= costo.sueldo_mensual_neto + costo.complemento_salario_mensual + costo.apoyo_de_pasajes + costo.total_apoyosbonos_empleadocomp # + costo.total_apoyosbonos_agregcomis
                                                                    if form.is_valid():
                                                                        user_filter = UserDatos.objects.get(user=request.user)
                                                                        nombre = Perfil.objects.get(numero_de_trabajador = user_filter.numero_de_trabajador, distrito = user_filter.distrito)
                                                                        costo.editado = str("U:"+nombre.nombres+" "+nombre.apellidos)
                                                                        messages.success(request, f'Cambios guardados con éxito los costos de {costo.status.perfil.nombres} {costo.status.perfil.apellidos}')
                                                                        costo = form.save(commit=False)
                                                                        costo.save()
                                                                        return redirect('Tabla_costo')
    else:
        form = CostoUpdateForm(instance=costo)

    context = {'form':form,'costo':costo, 'registros':registros,'comision':comision,'myfilter':myfilter,}

    return render(request, 'proyecto/Costo_update.html',context)

@login_required(login_url='user-login')
def Costo_revisar(request, pk):

    costo = Costo.objects.get(id=pk)
    ahora = datetime.date.today()
    catorcena = Catorcenas.objects.filter(fecha_inicial__lte=ahora, fecha_final__gte=ahora).first()
    bonos_dato = Bonos.objects.filter(costo=costo, fecha_bono__range=[catorcena.fecha_inicial, catorcena.fecha_final])
    sum_bonos = bonos_dato.aggregate(Sum('monto'))
    bonototal = sum_bonos['monto__sum']
    if bonototal == None:
        bonototal = 0
    comision=Decimal(0.09)

    vac_reforma_actual = Decimal((12/365)*365)*Decimal(costo.sueldo_diario)
    prima_vacacional = vac_reforma_actual*Decimal(0.25)
    aguinaldo = Decimal((15/365)*365)*Decimal(costo.sueldo_diario)
    total_vacaciones = (vac_reforma_actual+prima_vacacional+aguinaldo)/12
    costo.total_apoyosbonos_agregcomis = costo.campamento + bonototal #Modificar falta suma
    costo.comision_complemeto_salario_bonos= (costo.complemento_salario_mensual + costo.campamento + bonototal)*comision #Falta suma dentro de la multiplicacion
    costo.total_costo_empresa = costo.sueldo_mensual_neto + costo.complemento_salario_mensual + Decimal(costo.apoyo_de_pasajes) + costo.impuesto_estatal + costo.imms_obrero_patronal + costo.sar + costo.cesantia + costo.infonavit + costo.isr + costo.total_apoyosbonos_empleadocomp
    costo.total_costo_empresa = costo.total_costo_empresa + total_vacaciones
    costo.ingreso_mensual_neto_empleado= costo.sueldo_mensual_neto + costo.complemento_salario_mensual + Decimal(costo.apoyo_de_pasajes) + costo.total_apoyosbonos_empleadocomp # + costo.total_apoyosbonos_agregcomis

    costo.numero_de_trabajador=costo.status.perfil.numero_de_trabajador
    costo.empresa=costo.status.perfil.empresa
    costo.distrito=costo.status.perfil.distrito
    costo.proyecto=costo.status.perfil.proyecto
    costo.nombres=costo.status.perfil.nombres
    costo.apellidos=costo.status.perfil.apellidos
    costo.tipo_de_contrato=costo.status.tipo_de_contrato

    costo.amortizacion_infonavit=locale.currency(costo.amortizacion_infonavit, grouping=True)
    costo.fonacot=locale.currency(costo.fonacot, grouping=True)
    costo.neto_catorcenal_sin_deducciones=locale.currency(costo.neto_catorcenal_sin_deducciones, grouping=True)
    costo.complemento_salario_catorcenal=locale.currency(costo.complemento_salario_catorcenal, grouping=True)
    costo.sueldo_diario=locale.currency(costo.sueldo_diario, grouping=True)
    costo.sdi=locale.currency(costo.sdi, grouping=True)
    costo.apoyo_de_pasajes=locale.currency(costo.apoyo_de_pasajes, grouping=True)
    costo.imms_obrero_patronal=locale.currency(costo.imms_obrero_patronal, grouping=True)
    costo.apoyo_vist_familiar=locale.currency(costo.apoyo_vist_familiar, grouping=True)
    costo.estancia=locale.currency(costo.estancia, grouping=True)
    costo.renta=locale.currency(costo.renta, grouping=True)
    costo.apoyo_estudios=locale.currency(costo.apoyo_estudios, grouping=True)
    costo.amv=locale.currency(costo.amv, grouping=True)
    costo.gasolina=locale.currency(costo.gasolina, grouping=True)
    costo.campamento=locale.currency(costo.campamento, grouping=True)
    costo.total_deduccion=locale.currency(costo.total_deduccion, grouping=True)
    costo.neto_pagar=locale.currency(costo.neto_pagar, grouping=True)
    costo.sueldo_mensual_neto=locale.currency(costo.sueldo_mensual_neto, grouping=True)
    costo.complemento_salario_mensual=locale.currency(costo.complemento_salario_mensual, grouping=True)
    costo.sueldo_mensual=locale.currency(costo.sueldo_mensual, grouping=True)
    costo.sueldo_mensual_sdi=locale.currency(costo.sueldo_mensual_sdi, grouping=True)
    costo.total_percepciones_mensual=locale.currency(costo.total_percepciones_mensual, grouping=True)
    costo.impuesto_estatal=locale.currency(costo.impuesto_estatal, grouping=True)
    costo.sar=locale.currency(costo.sar, grouping=True)
    costo.cesantia=locale.currency(costo.cesantia, grouping=True)
    costo.infonavit=locale.currency(costo.infonavit, grouping=True)
    costo.isr=locale.currency(costo.isr, grouping=True)
    costo.lim_inferior=locale.currency(costo.lim_inferior, grouping=True)
    costo.excedente=locale.currency(costo.excedente, grouping=True)
    costo.tasa=locale.currency(costo.tasa, grouping=True)
    costo.impuesto_marginal=locale.currency(costo.impuesto_marginal, grouping=True)
    costo.cuota_fija=locale.currency(costo.cuota_fija, grouping=True)
    costo.impuesto=locale.currency(costo.impuesto, grouping=True)
    costo.subsidio=locale.currency(costo.subsidio, grouping=True)
    costo.total_apoyosbonos_empleadocomp=locale.currency(costo.total_apoyosbonos_empleadocomp, grouping=True)
    costo.total_apoyosbonos_agregcomis=locale.currency(costo.total_apoyosbonos_agregcomis, grouping=True)
    costo.comision_complemeto_salario_bonos=locale.currency(costo.comision_complemeto_salario_bonos, grouping=True)
    costo.total_costo_empresa=locale.currency(costo.total_costo_empresa, grouping=True)
    costo.ingreso_mensual_neto_empleado=locale.currency(costo.ingreso_mensual_neto_empleado, grouping=True)
    bonototal = locale.currency(bonototal, grouping=True)
    if request.method =='POST' and 'Pdf' in request.POST:
        return reporte_pdf_costo_detalles(costo,bonototal)

    context = {'costo':costo,
               'bonototal':bonototal,}

    return render(request, 'proyecto/Costo_revisar.html',context)

@login_required(login_url='user-login')
def Empleado_Costo(request, pk):

    empleado = Status.objects.get(id=pk)
    costo = Costo.objects.get(status__id=empleado.id)

    mes = datetime.date.today().month
    comision=Decimal(0.09)

    bonos_dato = Bonos.objects.filter(costo = costo).filter(fecha_bono__month = mes)
    sum_bonos = bonos_dato.aggregate(Sum('monto'))
    bonototal = sum_bonos['monto__sum']
    if bonototal == None:
        bonototal = 0
    costo.total_apoyosbonos_agregcomis = costo.campamento + bonototal
    costo.comision_complemeto_salario_bonos= (costo.complemento_salario_mensual + costo.campamento + bonototal)*comision #Falta suma dentro de la multiplicacion
    costo.total_costo_empresa = costo.sueldo_mensual_neto + costo.complemento_salario_mensual + costo.apoyo_de_pasajes + costo.impuesto_estatal + costo.imms_obrero_patronal + costo.sar + costo.cesantia + costo.infonavit + costo.isr + costo.total_apoyosbonos_empleadocomp + costo.total_apoyosbonos_agregcomis + costo.comision_complemeto_salario_bonos
    costo.ingreso_mensual_neto_empleado= costo.sueldo_mensual_neto + costo.complemento_salario_mensual + costo.apoyo_de_pasajes + costo.total_apoyosbonos_empleadocomp + costo.total_apoyosbonos_agregcomis

    costo.amortizacion_infonavit=locale.currency(costo.amortizacion_infonavit, grouping=True)
    costo.fonacot=locale.currency(costo.fonacot, grouping=True)
    costo.neto_catorcenal_sin_deducciones=locale.currency(costo.neto_catorcenal_sin_deducciones, grouping=True)
    costo.complemento_salario_catorcenal=locale.currency(costo.complemento_salario_catorcenal, grouping=True)
    costo.sueldo_diario=locale.currency(costo.sueldo_diario, grouping=True)
    costo.sdi=locale.currency(costo.sdi, grouping=True)
    costo.apoyo_de_pasajes=locale.currency(costo.apoyo_de_pasajes, grouping=True)
    costo.imms_obrero_patronal=locale.currency(costo.imms_obrero_patronal, grouping=True)
    costo.apoyo_vist_familiar=locale.currency(costo.apoyo_vist_familiar, grouping=True)
    costo.estancia=locale.currency(costo.estancia, grouping=True)
    costo.renta=locale.currency(costo.renta, grouping=True)
    costo.apoyo_estudios=locale.currency(costo.apoyo_estudios, grouping=True)
    costo.amv=locale.currency(costo.amv, grouping=True)
    costo.gasolina=locale.currency(costo.gasolina, grouping=True)
    costo.campamento=locale.currency(costo.campamento, grouping=True)
    costo.total_deduccion=locale.currency(costo.total_deduccion, grouping=True)
    costo.neto_pagar=locale.currency(costo.neto_pagar, grouping=True)
    costo.sueldo_mensual_neto=locale.currency(costo.sueldo_mensual_neto, grouping=True)
    costo.complemento_salario_mensual=locale.currency(costo.complemento_salario_mensual, grouping=True)
    costo.sueldo_mensual=locale.currency(costo.sueldo_mensual, grouping=True)
    costo.sueldo_mensual_sdi=locale.currency(costo.sueldo_mensual_sdi, grouping=True)
    costo.total_percepciones_mensual=locale.currency(costo.total_percepciones_mensual, grouping=True)
    costo.impuesto_estatal=locale.currency(costo.impuesto_estatal, grouping=True)
    costo.sar=locale.currency(costo.sar, grouping=True)
    costo.cesantia=locale.currency(costo.cesantia, grouping=True)
    costo.infonavit=locale.currency(costo.infonavit, grouping=True)
    costo.isr=locale.currency(costo.isr, grouping=True)
    costo.lim_inferior=locale.currency(costo.lim_inferior, grouping=True)
    costo.excedente =locale.currency(costo.excedente, grouping=True)
    costo.tasa=locale.currency(costo.tasa, grouping=True)
    costo.impuesto_marginal=locale.currency(costo.impuesto_marginal, grouping=True)
    costo.cuota_fija=locale.currency(costo.cuota_fija, grouping=True)
    costo.impuesto=locale.currency(costo.impuesto, grouping=True)
    costo.subsidio=locale.currency(costo.subsidio, grouping=True)
    costo.total_apoyosbonos_empleadocomp=locale.currency(costo.total_apoyosbonos_empleadocomp, grouping=True)
    if bonototal == None:
        costo.bonototal =locale.currency(0, grouping=True)
    else:
        costo.bonototal=locale.currency(bonototal, grouping=True)
    costo.total_apoyosbonos_agregcomis=locale.currency(costo.total_apoyosbonos_agregcomis, grouping=True)
    costo.comision_complemeto_salario_bonos=locale.currency(costo.comision_complemeto_salario_bonos, grouping=True)
    costo.total_costo_empresa=locale.currency(costo.total_costo_empresa, grouping=True)
    costo.ingreso_mensual_neto_empleado=locale.currency(costo.ingreso_mensual_neto_empleado, grouping=True)
    if request.method =='POST' and 'Pdf' in request.POST:
        return reporte_pdf_costo_detalles(costo,bonototal)

    context = {
        'costo':costo,
        }

    return render(request, 'proyecto/Costo_revisar.html',context)

@login_required(login_url='user-login')
def TablaCosto(request):
    año = datetime.date.today().year

    user_filter = UserDatos.objects.get(user=request.user)

    if user_filter.distrito.distrito == 'Matriz':
        costos= Costo.objects.filter(complete=True).order_by("status__perfil__numero_de_trabajador")
    else:
        perfil = Perfil.objects.filter(distrito = user_filter.distrito,complete=True)
        costos = Costo.objects.filter(status__perfil__id__in=perfil.all(), complete=True).order_by("status__perfil__numero_de_trabajador")

    costo_filter = CostoFilter(request.GET, queryset=costos)
    costos = costo_filter.qs

    comision=Decimal(0.09)

    if request.method =='POST' and 'Excel' in request.POST:
        return convert_excel_costo(costos)

                #Set up pagination
    p = Paginator(costos, 50)
    page = request.GET.get('page')
    salidas_list = p.get_page(page)

    ahora = datetime.date.today()
    catorcena = Catorcenas.objects.filter(fecha_inicial__lte=ahora, fecha_final__gte=ahora).first()
    for costo in salidas_list:
        bonos_dato = Bonos.objects.filter(costo=costo, fecha_bono__range=[catorcena.fecha_inicial, catorcena.fecha_final])
        sum_bonos = bonos_dato.aggregate(Sum('monto'))
        bonototal = sum_bonos['monto__sum']
        if bonototal == None:
            bonototal = 0
        vac_reforma_actual = Decimal((12/365)*365)*Decimal(costo.sueldo_diario)
        prima_vacacional = vac_reforma_actual*Decimal(0.25)
        aguinaldo = Decimal((15/365)*365)*Decimal(costo.sueldo_diario)
        total_vacaciones = (vac_reforma_actual+prima_vacacional+aguinaldo)/12
        costo.total_apoyosbonos_agregcomis = costo.campamento + bonototal #Modificar falta suma
        costo.comision_complemeto_salario_bonos= (costo.complemento_salario_mensual + costo.campamento + bonototal)*comision #Falta suma dentro de la multiplicacion
        costo.total_costo_empresa = costo.sueldo_mensual_neto + costo.complemento_salario_mensual + Decimal(costo.apoyo_de_pasajes) + costo.impuesto_estatal + costo.imms_obrero_patronal + costo.sar + costo.cesantia + costo.infonavit + costo.isr + costo.total_apoyosbonos_empleadocomp
        costo.total_costo_empresa = costo.total_costo_empresa + total_vacaciones
        costo.ingreso_mensual_neto_empleado= costo.sueldo_mensual_neto + costo.complemento_salario_mensual + Decimal(costo.apoyo_de_pasajes) + costo.total_apoyosbonos_empleadocomp # + costo.total_apoyosbonos_agregcomis

        costo.numero_de_trabajador=costo.status.perfil.numero_de_trabajador
        costo.empresa=costo.status.perfil.empresa
        costo.distrito=costo.status.perfil.distrito
        costo.proyecto=costo.status.perfil.proyecto
        costo.nombres=costo.status.perfil.nombres
        costo.apellidos=costo.status.perfil.apellidos
        costo.tipo_de_contrato=costo.status.tipo_de_contrato

        costo.amortizacion_infonavit=locale.currency(costo.amortizacion_infonavit, grouping=True)
        costo.fonacot=locale.currency(costo.fonacot, grouping=True)
        costo.neto_catorcenal_sin_deducciones=locale.currency(costo.neto_catorcenal_sin_deducciones, grouping=True)
        costo.complemento_salario_catorcenal=locale.currency(costo.complemento_salario_catorcenal, grouping=True)
        costo.sueldo_diario=locale.currency(costo.sueldo_diario, grouping=True)
        costo.sdi=locale.currency(costo.sdi, grouping=True)
        costo.apoyo_de_pasajes=locale.currency(costo.apoyo_de_pasajes, grouping=True)
        costo.imms_obrero_patronal=locale.currency(costo.imms_obrero_patronal, grouping=True)
        costo.apoyo_vist_familiar=locale.currency(costo.apoyo_vist_familiar, grouping=True)
        costo.estancia=locale.currency(costo.estancia, grouping=True)
        costo.renta=locale.currency(costo.renta, grouping=True)
        costo.apoyo_estudios=locale.currency(costo.apoyo_estudios, grouping=True)
        costo.amv=locale.currency(costo.amv, grouping=True)
        costo.gasolina=locale.currency(costo.gasolina, grouping=True)
        costo.campamento=locale.currency(costo.campamento, grouping=True)
        costo.total_deduccion=locale.currency(costo.total_deduccion, grouping=True)
        costo.neto_pagar=locale.currency(costo.neto_pagar, grouping=True)
        costo.sueldo_mensual_neto=locale.currency(costo.sueldo_mensual_neto, grouping=True)
        costo.complemento_salario_mensual=locale.currency(costo.complemento_salario_mensual, grouping=True)
        costo.sueldo_mensual=locale.currency(costo.sueldo_mensual, grouping=True)
        costo.sueldo_mensual_sdi=locale.currency(costo.sueldo_mensual_sdi, grouping=True)
        costo.total_percepciones_mensual=locale.currency(costo.total_percepciones_mensual, grouping=True)
        costo.impuesto_estatal=locale.currency(costo.impuesto_estatal, grouping=True)
        costo.sar=locale.currency(costo.sar, grouping=True)
        costo.cesantia=locale.currency(costo.cesantia, grouping=True)
        costo.infonavit=locale.currency(costo.infonavit, grouping=True)
        costo.isr=locale.currency(costo.isr, grouping=True)
        costo.lim_inferior=locale.currency(costo.lim_inferior, grouping=True)
        costo.excedente=locale.currency(costo.excedente, grouping=True)
        costo.tasa=locale.currency(costo.tasa, grouping=True)
        costo.impuesto_marginal=locale.currency(costo.impuesto_marginal, grouping=True)
        costo.cuota_fija=locale.currency(costo.cuota_fija, grouping=True)
        costo.impuesto=locale.currency(costo.impuesto, grouping=True)
        costo.subsidio=locale.currency(costo.subsidio, grouping=True)
        costo.total_apoyosbonos_empleadocomp=locale.currency(costo.total_apoyosbonos_empleadocomp, grouping=True)
        costo.total_apoyosbonos_agregcomis=locale.currency(costo.total_apoyosbonos_agregcomis, grouping=True)
        costo.comision_complemeto_salario_bonos=locale.currency(costo.comision_complemeto_salario_bonos, grouping=True)
        costo.total_costo_empresa=locale.currency(costo.total_costo_empresa, grouping=True)
        costo.ingreso_mensual_neto_empleado=locale.currency(costo.ingreso_mensual_neto_empleado, grouping=True)

    context = {'costos':costos,'costo_filter':costo_filter,'salidas_list':salidas_list, 'baja': request.GET.get('baja', False)}

    return render(request, 'proyecto/Tabla_costo.html',context)

@login_required(login_url='user-login')
def TablaBonos(request):
    año = datetime.date.today().year
    user_filter = UserDatos.objects.get(user=request.user)

    if user_filter.distrito.distrito == 'Matriz':
        bonos= Bonos.objects.filter(complete=True,mes_bono__year=año).order_by("costo__status__perfil__numero_de_trabajador")
    else:
        perfil = Perfil.objects.filter(distrito = user_filter.distrito,complete=True)
        bonos = Bonos.objects.filter(costo__status__perfil__id__in=perfil.all(),mes_bono__year=año, complete=True).order_by("costo__status__perfil__numero_de_trabajador")

    bono_filter = BonosFilter(request.GET, queryset=bonos)
    bonos = bono_filter.qs

    if request.method =='POST' and 'Excel' in request.POST:
        return convert_excel_bonos(bonos)

    for bono in bonos:
        if bono.monto == None:
            bono.monto = 0
        else:
            bono.monto=locale.currency(bono.monto, grouping=True)
                #Set up pagination
    p = Paginator(bonos, 50)
    page = request.GET.get('page')
    salidas_list = p.get_page(page)

    context= {
        'bonos':bonos,
        'bono_filter':bono_filter,
        'salidas_list':salidas_list,
        'baja': request.GET.get('baja', False)
        }

    return render(request, 'proyecto/BonosTabla.html',context)

@login_required(login_url='user-login')
def Empleado_Bonos(request, pk):
    #año = datetime.date.today().year
    bonos= Bonos.objects.filter(costo__status__id=pk,complete=True).order_by("fecha_bono")
    perfil = bonos.last()

    bono_filter = BonosFilter(request.GET, queryset=bonos)
    bonos = bono_filter.qs

    if request.method =='POST' and 'Excel' in request.POST:
        return convert_excel_bonos(bonos)

    for bono in bonos:
        if bono.monto == None:
            bono.monto = 0
        else:
            bono.monto=locale.currency(bono.monto, grouping=True)

    context= {
        'bonos':bonos,
        'bono_filter':bono_filter,
        'perfil':perfil,
        }

    return render(request, 'proyecto/Empleado_bonos.html',context)


@login_required(login_url='user-login')
def VacacionesUpdate(request, pk):
    currentFieldCount = 10
    descanso = Vacaciones.objects.get(id=pk)
    registros = descanso.history.filter(~Q(dias_disfrutados = None))

    if request.method == 'POST' and 'btnSend' in request.POST:
        form = VacacionesUpdateForm(request.POST, instance=descanso)
        descanso = form.save(commit=False)


        tabla_festivos = TablaFestivos.objects.all()
        delta = timedelta(days=1)
        day_count = (descanso.fecha_fin - descanso.fecha_inicio + delta ).days
        cuenta = day_count #Dias entre las dos fechas
        inhabil = descanso.dia_inhabil.numero
        for fecha in (descanso.fecha_inicio + timedelta(n) for n in range(day_count)):
            if fecha.isoweekday() == inhabil:
                cuenta -= 1 #Se le restan a la cuenta los días inhabiles para sacar los dias reales
            else:
                for dia in tabla_festivos:
                    if fecha == dia.dia_festivo:
                        cuenta -= 1 # Se le restan tambien los días festivos para sacar los días reales que va a tomar (Cantida de días)
        dias_vacacion = cuenta #Para escribir en el comentario los días
        if cuenta > 0: #Aqui salgo bien con los 2 dias--------
            #Aqui se buscan las vacaciones anteriores y se van modificando los datos para poder llevar la toma de dias pendientes de años anteriores
            if Vacaciones.objects.filter(status=descanso.status.id).last().total_pendiente > 0 or Vacaciones.objects.filter(status=descanso.status.id).first().total_pendiente > 0:
                datos = Vacaciones.objects.filter(status=descanso.status.id, total_pendiente__gt=0,).order_by(Cast('periodo', output_field=IntegerField()))#Trae todas las vacaciones del mas antiguo al actual 2019-2022
                if datos.exclude(id=datos.last().id) != None:
                    datos = datos.exclude(id=datos.last().id)
                    for dato in datos: #Se pasa por los datos del mas antiguo al mas actual de los que se tenia
                        if cuenta <= dato.total_pendiente and cuenta > 0:
                            if dato.dias_disfrutados == None:
                                dato.dias_disfrutados = 0
                            dato.total_pendiente -= cuenta
                            dato.dias_disfrutados += cuenta
                            cuenta = 0
                            break
                        elif cuenta > dato.total_pendiente and cuenta > 0:
                            if dato.dias_disfrutados == None:
                                dato.dias_disfrutados = 0
                            dato.dias_disfrutados += dato.total_pendiente
                            cuenta -=dato.total_pendiente
                            dato.total_pendiente = 0
                descanso.dias_disfrutados += cuenta #Días que disfrutara son los que vienen de la cuenta
                descanso.fecha_planta_anterior = descanso.status.fecha_planta_anterior
                descanso.fecha_planta = descanso.status.fecha_planta

                if descanso.dias_disfrutados > descanso.dias_de_vacaciones:
                    messages.error(request, f'(Dias disfrutados) La cantidad total capturada debe ser menor o igual a {descanso.total_pendiente}, cantidad actual: {cuenta}')
                else:
                    periodofecha = descanso.created_at.year
                    descanso.periodo = periodofecha
                    descanso.total_pendiente = descanso.dias_de_vacaciones - descanso.dias_disfrutados
                    if form.is_valid():
                        solicitud, created = Solicitud_vacaciones.objects.get_or_create(complete=False)
                        num_campos = int(request.POST.get('num_campos', 0))

                        for i in range(1, num_campos + 1):
                            asunto = request.POST.get(f'asunto{i}')
                            estado = request.POST.get(f'estado{i}')

                            # Crear un nuevo grupo de Trabajos_encomendados para cada conjunto de 10 campos
                            if (i - 1) % 10 == 0:
                                if i > 1:
                                    trabajos_grupo.complete = True
                                    trabajos_grupo.save()

                                trabajos_grupo = Trabajos_encomendados()
                                trabajos_grupo.save()
                                solicitud.asunto.add(trabajos_grupo)

                            setattr(trabajos_grupo, f'asunto{(i - 1) % 10 + 1}', asunto)
                            setattr(trabajos_grupo, f'estado{(i - 1) % 10 + 1}', estado)
                            trabajos_grupo.save()

                        # Marcar el último grupo como completo
                        trabajos_grupo.complete = True
                        trabajos_grupo.save()

                        temas, created = Temas_comentario_solicitud_vacaciones.objects.get_or_create(complete=False,)
                        for i in range(1, 10):
                            comentario = request.POST.get(f'comentario{i}')
                            setattr(temas, f'comentario{i}', comentario)
                        temas.complete=True
                        temas.save()

                        solicitud.status = descanso.status
                        solicitud.periodo = descanso.periodo
                        solicitud.fecha_inicio = descanso.fecha_inicio
                        solicitud.fecha_fin = descanso.fecha_fin
                        solicitud.dia_inhabil = descanso.dia_inhabil
                        solicitud.recibe_nombre = request.POST.get('recibe')
                        solicitud.recibe_area = request.POST.get('area')
                        solicitud.recibe_puesto = request.POST.get('puesto')
                        solicitud.recibe_sector = request.POST.get('sector')
                        solicitud.informacion_adicional = request.POST.get('adicional')
                        solicitud.temas = temas
                        solicitud.anexos = request.POST.get('anexos')
                        solicitud.autorizar = True
                        solicitud.comentario_rh = descanso.comentario
                        solicitud.complete=True
                        solicitud.save()
                        for dato in datos:
                            historial = dato.history.first()  # Obtener la primera versión histórica del objeto
                            if historial and historial.total_pendiente != dato.total_pendiente:
                                # El campo 'total_pendiente' ha cambiado
                                dato._meta.get_field('created_at').auto_now = False
                                dato.comentario ="Solicitado periodo: " + str(descanso.periodo)
                                dato.fecha_inicio = descanso.fecha_inicio
                                dato.fecha_fin =  descanso.fecha_fin
                                dato.save()
                                dato._meta.get_field('created_at').auto_now = True
                        if cuenta > 0:
                            messages.success(request, f'Cambios guardados con éxito los días de vacaciones de {descanso.status.perfil.nombres} {descanso.status.perfil.apellidos}')
                        else:
                            messages.success(request, f'Datos capturados con éxito empleado {descanso.status.perfil.nombres} {descanso.status.perfil.apellidos} y descontados a sus días pendientes')
                        descanso.comentario +=" " + "Dias tomados:" + str(dias_vacacion)
                        form.save()
                        return redirect('Tabla_vacaciones_empleados')
            else:
                messages.error(request, 'Ya a tomado todos sus días de vacaciones')
        else:
            messages.error(request, 'La cantidad de días que disfrutara debe ser mayor a 0')
    else:
        form = VacacionesUpdateForm()
    context = {
        'form':form,
        'descanso':descanso,
        'registros':registros,
        'currentFieldCount': currentFieldCount,
        }
    return render(request, 'proyecto/Vacaciones_update.html',context)

@login_required(login_url='user-login')
def VacacionesRevisar(request, pk):
    if Vacaciones.objects.filter(id=pk):
        usuario = Vacaciones.objects.get(id=pk)
        usuario = usuario.status
    elif Status.objects.filter(id=pk):
        usuario = Status.objects.get(id=pk)
    datos = Vacaciones.objects.filter(status=usuario).order_by(Cast('periodo', output_field=IntegerField()).desc()) #Identifico las vacaciones del usuario de la mas antigua a la mas actual
    actual = Vacaciones.objects.filter(status=usuario).order_by(Cast('periodo', output_field=IntegerField())).last()
    resultado = 0
    for dato in datos:
        resultado += dato.total_pendiente

    context = {
        'actual':actual, #vacaciones del empleado
        'datos':datos, #vacaciones pendientes por año
        'resultado':resultado, #total vacaciones pendientes
        }

    return render(request, 'proyecto/Vacaciones_revisar.html',context)

@login_required(login_url='user-login')
def Tabla_Vacaciones(request): #Ya esta
    user_filter = UserDatos.objects.get(user=request.user)
    
    #Se reinician las vacaciones para los empleados que ya cumplan otro año de antiguedad con su planta anterior o actual
    fecha_actual = date.today()
    año_actual = str(fecha_actual.year)
    #Busca todos los status que no tengan vacaciones del año actual (periodo)
    status_filtrados = Status.objects.exclude(Q(fecha_planta_anterior__isnull=True, fecha_planta__isnull=True) |Q(vacaciones__periodo=año_actual)) 
    fecha_hace_un_año = fecha_actual - relativedelta(years=1)
    #Filtra todos aquellos con un año o mas de dias con respecto a la fecha actual
    #reinicio = status_filtrados.filter(complete=True,perfil__baja=False,fecha_planta_anterior__lte=fecha_hace_un_año)
    reinicio = status_filtrados.filter(complete=True,perfil__baja=False,fecha_planta_anterior__lte=fecha_hace_un_año)
    reinicio = reinicio.filter(Q(fecha_planta_anterior__month__lte=fecha_actual.month) & Q(fecha_planta_anterior__day__lte=fecha_actual.day))
    #Busco el fecha de planta en los que no tengan fecha de planta anterior
    reinicio2 = status_filtrados.filter(complete=True, perfil__baja=False, fecha_planta_anterior=None, fecha_planta__lte=fecha_hace_un_año)
    reinicio2 = reinicio2.filter(Q(fecha_planta__month__lte=fecha_actual.month) & Q(fecha_planta__day__lte=fecha_actual.day))
    reinicio = reinicio | reinicio2 #Junto los datos de los empleados que ya tienen mas de 1 año de antiguedad
    
    
    #if reinicio:
    for empleado in reinicio:
            #Se calculan los días para la vacación actual 
            ahora = datetime.date.today()
            tablas= TablaVacaciones.objects.all()
            if empleado.fecha_planta_anterior:
                days = empleado.fecha_planta_anterior
            else:
                days = empleado.fecha_planta
                
            calcular_antiguedad = relativedelta(ahora, days)
            antiguedad = calcular_antiguedad.years
            
            for tabla in tablas:
                if antiguedad >= tabla.years:
                    dias_vacaciones = tabla.days #Se asignan los días que le tocan en esta vacación
                        
            vacacion = Vacaciones.objects.create(
                complete=True, 
                status=empleado, 
                periodo=año_actual, 
                dias_de_vacaciones=dias_vacaciones,
                fecha_inicio=None,
                fecha_fin=None,
                dias_disfrutados=0,
                dia_inhabil=None,
                total_pendiente=dias_vacaciones,
                comentario="Generado autom. al cumplir otro año de antigüedad",
                editado="Sistema")
            empleado.complete_vacaciones = True #Para confirmar que ya tiene vacacion actual
            empleado.save()
  
    if user_filter.distrito.distrito == 'Matriz':
        perfil = Perfil.objects.all();
        
        #descansos= Vacaciones.objects.filter(complete=True,periodo=año_actual).annotate(Sum('dias_disfrutados')).order_by("status__perfil__numero_de_trabajador")
        periodo = Vacaciones.objects.filter(
            Q(periodo=año_actual) | Q(periodo=str(fecha_hace_un_año.year)),
            status__perfil__id__in=perfil.all(),
            complete=True
        ).annotate(Sum('dias_disfrutados')).order_by("status__perfil__numero_de_trabajador")
        
        periodo1 = periodo.filter(periodo = año_actual) #traingo los de 2024
        periodo2 = periodo.filter(periodo = fecha_hace_un_año.year) #traigo los del 2023
        
        #como el perfil se repite dos veces 2023 y 2024 elimina 1 y se queda con el 2024 
        periodo3 = periodo2.exclude(status_id__in=periodo1.values('status_id')) 
        
        descansos = periodo1 | periodo3
        
        
        print(descansos.count())
    
    #elif user_filter.distrito.distrito == 'Poza Rica':
    #    perfil = Perfil.objects.filter(distrito = user_filter.distrito,complete=True, baja=False)
    #    descansos = Vacaciones.objects.filter(status__perfil__id__in=perfil.all(), complete=True, periodo__in=["2022", "2023"]).annotate(Sum('dias_disfrutados')).order_by("status__perfil__numero_de_trabajador")
    else:
        
        perfil = Perfil.objects.filter(distrito = user_filter.distrito,complete=True)
        #descansos = Vacaciones.objects.filter(status__perfil__id__in=perfil.all(), complete=True, periodo=año_actual).annotate(Sum('dias_disfrutados')).order_by("status__perfil__numero_de_trabajador")
        periodo = Vacaciones.objects.filter(
            Q(periodo=año_actual) | Q(periodo=str(fecha_hace_un_año.year)),
            status__perfil__id__in=perfil.all(),
            complete=True
        ).annotate(Sum('dias_disfrutados')).order_by("status__perfil__numero_de_trabajador")
        
        periodo1 = periodo.filter(periodo = año_actual) #traingo los de 2024
        periodo2 = periodo.filter(periodo = fecha_hace_un_año.year) #traigo los del 2023
        
        #como el perfil se repite dos veces 2023 y 2024 elimina 1 y se queda con el 2024 
        periodo3 = periodo2.exclude(status_id__in=periodo1.values('status_id')) 
        
        descansos = periodo1 | periodo3
        
        print('proyecto: ',descansos.count())
        
        
    
    vacaciones_filter = VacacionesFilter(request.GET, queryset=descansos)
    descansos = vacaciones_filter.qs
    if request.method =='POST' and 'Excel' in request.POST:
        return convert_excel_vacaciones(descansos)
                #Set up pagination
    p = Paginator(descansos, 50)
    page = request.GET.get('page')
    salidas_list = p.get_page(page)

    context= {
        'descansos':descansos,
        'vacaciones_filter':vacaciones_filter,
        'salidas_list':salidas_list,
        'baja': request.GET.get('baja', False)
        }

    return render(request, 'proyecto/TablaVacaciones.html',context)

@login_required(login_url='user-login')
def FormularioEconomicos(request):
    user_filter = UserDatos.objects.get(user=request.user)

    año_actual = str(date.today().year) #Quitar el complete_economicos a todos aquellos que ya cumplan el año de planta
    mes_actual = datetime.date.today().month
    dia_actual = datetime.date.today().day
    #Filtra todos aquellos con un mes y dia menor a la fecha actual, con esto ya se que cumplen más del año
    status_filtrados = Status.objects.exclude(Q(fecha_planta_anterior__isnull=True, fecha_planta__isnull=True) |Q(economicos__periodo=año_actual))
    reinicio = status_filtrados.filter(complete=True,complete_vacaciones=True,perfil__baja=False,fecha_planta_anterior__month__lte=mes_actual,fecha_planta_anterior__day__lte=dia_actual,)
    reinicio2 = status_filtrados.filter(complete=True, complete_vacaciones=True, perfil__baja=False, fecha_planta_anterior=None, fecha_planta__month__lte=mes_actual,fecha_planta__day__lte=dia_actual,)
    reinicio = reinicio | reinicio2
    for empleado in reinicio:
        empleado.complete_economicos = False
        empleado.save()

    if user_filter.distrito.distrito == 'Matriz':
        empleados= Status.objects.filter(complete= True, complete_economicos = False, perfil__baja=False)
    else:
        perfil = Perfil.objects.filter(distrito = user_filter.distrito, baja=False)
        empleados= Status.objects.filter(perfil__id__in=perfil.all(),complete= True, complete_economicos = False)

    economico,created=Economicos.objects.get_or_create(complete=False)
    form = EconomicosForm()
    form.fields["status"].queryset = empleados
    total_dias_economicos=3
    dias_disfrutados=1

    if request.method == 'POST' and 'btnSend' in request.POST:
        form = EconomicosForm(request.POST,instance=economico)
        form.save(commit=False)

        #Se verifica si el empleado tiene un año de antigüedad
        if economico.status.fecha_planta_anterior:
            days = economico.status.fecha_planta_anterior
        else:
            days = economico.status.fecha_planta
        ahora = datetime.date.today()
        año = 1
        if ahora.month > days.month or (ahora.month == days.month and ahora.day >= days.day):
            antiguedad = ahora.year - days.year
        else:
            antiguedad = ahora.year - days.year - 1
        if antiguedad >= año:

            economico.dias_disfrutados = dias_disfrutados
            economico.dias_pendientes = total_dias_economicos - economico.dias_disfrutados
            periodo = economico.created_at.year
            str(periodo)
            economico.periodo=periodo
            empleado = Status.objects.get(id = economico.status.id)
            if form.is_valid():
                solicitud, created = Solicitud_economicos.objects.get_or_create(complete=False)
                solicitud.status = economico.status
                solicitud.periodo = economico.periodo
                solicitud.fecha = economico.fecha
                solicitud.comentario = economico.comentario
                solicitud.autorizar = True
                solicitud.complete=True
                solicitud.save()
                nombre = Perfil.objects.get(numero_de_trabajador = user_filter.numero_de_trabajador, distrito = user_filter.distrito)
                economico.editado = str("C:"+nombre.nombres+" "+nombre.apellidos)
                messages.success(request, 'Datos capturados con exíto')
                economico.complete=True
                economico.complete_dias=False
                empleado.complete_economicos = True
                form.save()
                empleado.save()
                return redirect('Tabla_economicos')
        else:
            messages.error(request, f'El empleado aún no cumple su año de antigüedad por lo que no puede solicitar vacaciones')

    context = {'form':form,'empleados':empleados,}

    return render(request, 'proyecto/EconomicosForm.html',context)

@login_required(login_url='user-login')
def EconomicosUpdate(request, pk):
    user_filter = UserDatos.objects.get(user=request.user)
    status = Status.objects.get(id=pk)
    economico = Economicos.objects.filter(complete=True,status=status).last()
    form = EconomicosForm()
    total_dias_economicos=3
    dias_disfrutados=1
    anterior_fecha=economico.fecha
    anterior_dia=economico.dias_disfrutados
    if request.method == 'POST' and 'btnSend' in request.POST:
        form = EconomicosUpdateForm(request.POST,instance=economico)
        fecha_form = request.POST.get('fecha')
        year, month, day = map(int, fecha_form.split('-'))
        fecha_form = date(year, month, day)
        if anterior_fecha == fecha_form - timedelta(days=1):
            messages.error(request,'Los días económicos no pueden ser seguidos')
        else:
            economico.status=status
            economico.dias_disfrutados = dias_disfrutados + anterior_dia
            economico.dias_pendientes = total_dias_economicos - economico.dias_disfrutados
            periodo = economico.created_at.year
            str(periodo)
            economico.periodo=periodo
            orden = Economicos.objects.filter(status = status.id).last()
            orden.complete_dias=True
            if economico.dias_disfrutados == total_dias_economicos:
                economico.complete_dias=True
            else:
                economico.complete_dias=False
            if form.is_valid():
                solicitud, created = Solicitud_economicos.objects.get_or_create(complete=False)
                solicitud.status = economico.status
                solicitud.periodo = economico.periodo
                solicitud.fecha = economico.fecha
                solicitud.comentario = economico.comentario
                solicitud.autorizar = True
                solicitud.complete=True
                solicitud.save()
                nombre = Perfil.objects.get(numero_de_trabajador = user_filter.numero_de_trabajador, distrito = user_filter.distrito)
                economico.editado = str("A:"+nombre.nombres+" "+nombre.apellidos)
                messages.success(request, 'Se capturaron con exíto los datos')
                economico.complete=True
                orden.save()
                form.save()
                return redirect('Tabla_economicos')

    context = {'form':form,'economico':economico,'status':status,}

    return render(request, 'proyecto/Economicos_update.html',context)

@login_required(login_url='user-login')
def Tabla_Economicos(request): #Ya esta
    user_filter = UserDatos.objects.get(user=request.user)

    #Se reinician las vacaciones para los empleados que ya cumplan otro año de antiguedad con su planta anterior o actual
    fecha_actual = date.today()
    año_actual = str(fecha_actual.year)
    fecha_hace_un_año = fecha_actual - relativedelta(years=1)
    #Busca todos los status que no tengan vacaciones del año actual (periodo)
    status_filtrados = Status.objects.exclude(Q(fecha_planta_anterior__isnull=True, fecha_planta__isnull=True) |Q(economicos__periodo=año_actual)) 
    fecha_hace_un_año = fecha_actual - relativedelta(years=1)
    #Filtra todos aquellos con un año o mas de dias con respecto a la fecha actual
    reinicio = status_filtrados.filter(complete=True,perfil__baja=False,fecha_planta_anterior__lte=fecha_hace_un_año)
    #Busco el fecha de planta en los que no tengan fecha de planta anterior
    reinicio2 = status_filtrados.filter(complete=True, perfil__baja=False, fecha_planta_anterior=None, fecha_planta__lte=fecha_hace_un_año,)
    reinicio = reinicio | reinicio2 #Junto los datos de los empleados que ya tienen mas de 1 año de antiguedad
    if reinicio:
        for empleado in reinicio:                        
            economicos = Economicos.objects.create(complete=True, status=empleado, periodo=año_actual, dias_disfrutados=0, dias_pendientes=3, fecha=None, 
                                                comentario="Generado autom. al cumplir otro año de antigüedad", editado="Sistema")
            empleado.complete_economicos = True #Para confirmar que ya tiene economico actual
            empleado.save()

    if user_filter.distrito.distrito == 'Matriz':
        economicos= Economicos.objects.filter(complete=True,complete_dias=False,created_at__year=año_actual).order_by("status__perfil__numero_de_trabajador")
        #economicost= economicos.last()
        economicoss= Economicos.objects.filter(dias_pendientes=0,complete=True,complete_dias=True,created_at__year=año_actual).order_by("status__perfil__numero_de_trabajador")
    else:
        perfil = Perfil.objects.filter(distrito = user_filter.distrito,complete=True)
        economicos = Economicos.objects.filter(status__perfil__id__in=perfil.all(),complete=True,complete_dias=False,created_at__year=año_actual).order_by("status__perfil__numero_de_trabajador")
        #economicost = economicos.last()
        economicoss = Economicos.objects.filter(status__perfil__id__in=perfil.all(),complete=True,complete_dias=True,created_at__year=año_actual).order_by("status__perfil__numero_de_trabajador")
    #economicos= Economicos.objects.filter(complete=True).annotate(Sum('dias_disfrutados'))
    economico_filter = EconomicosFilter(request.GET, queryset=economicos)
    economicos = economico_filter.qs
    economico_filters = EconomicosFilter(request.GET, queryset=economicoss)
    economicoss = economico_filters.qs
    if request.method =='POST' and 'Excel' in request.POST:
        return convert_excel_economicos(economicos,economicoss)

                #Set up pagination
    p = Paginator(economicos, 50)
    page = request.GET.get('page')
    salidas_list = p.get_page(page)
                #Set up pagination
    p = Paginator(economicoss, 50)
    page = request.GET.get('page')
    salidas_listt = p.get_page(page)
    context= {
        'economicos':economicos,
        'economico_filter':economico_filter,
        'economicoss':economicoss,
        'economico_filters':economico_filters,
        'salidas_list':salidas_list,
        'salidas_listt':salidas_listt,
        'baja': request.GET.get('baja', False)
        }

    return render(request, 'proyecto/Tabla_economicos.html',context)

@login_required(login_url='user-login')
def EconomicosRevisar(request, pk):
    economicos = Economicos.objects.filter(status__id=pk)
    #registros = economicos.history.filter(complete=True)
    empleado = economicos.last()
    if request.method =='POST' and 'Pdf' in request.POST:
        return reporte_pdf_economico_detalles(economicos,empleado)

    context = {
        'empleado':empleado,
        'economicos':economicos,
        }

    return render(request, 'proyecto/Economicos_revisar.html',context)

@login_required(login_url='user-login')
def Tabla_Datosbancarios(request):
    user_filter = UserDatos.objects.get(user=request.user)

    if user_filter.distrito.distrito == 'Matriz':
        bancarios= DatosBancarios.objects.filter(complete=True).order_by("status__perfil__numero_de_trabajador")
    else:
        perfil = Perfil.objects.filter(distrito = user_filter.distrito,complete=True)
        bancarios = DatosBancarios.objects.filter(status__perfil__id__in=perfil.all(), complete=True).order_by("status__perfil__numero_de_trabajador")
    bancario_filter = BancariosFilter(request.GET, queryset=bancarios)
    bancarios = bancario_filter.qs
    for bancario in bancarios:
        bancario.numero_de_tarjeta = str(bancario.numero_de_tarjeta)
        bancario.clabe_interbancaria = str(bancario.clabe_interbancaria)
    if request.method =='POST' and 'Excel' in request.POST:
        return convert_excel_bancarios(bancarios)

                #Set up pagination
    p = Paginator(bancarios, 50)
    page = request.GET.get('page')
    salidas_list = p.get_page(page)

    context= {
        'bancarios':bancarios,
        'bancario_filter':bancario_filter,
        'salidas_list':salidas_list,
        'baja': request.GET.get('baja', False)
        }

    return render(request, 'proyecto/Tabla_Datosbancarios.html',context)

@login_required(login_url='user-login')
def Empleado_Datosbancarios(request, pk):

    empleado = Status.objects.get(id=pk)
    datos = DatosBancarios.objects.get(status__id=empleado.id)

    context = {
        'datos':datos,
        }

    return render(request, 'proyecto/Empleado_Datosbancarios.html',context)

@login_required(login_url='user-login')
def HistoryCosto(request, pk):
    costos = Costo.objects.get(id=pk)
    registros = costos.history.filter(~Q(amortizacion_infonavit = None))
    myfilter = Costo_historicFilter(request.GET, queryset=registros)
    registros=myfilter.qs
    if request.method == 'POST':

        costo.impuesto_estatal=locale.currency(costo.impuesto_estatal, grouping=True)
        costo.sar=locale.currency(costo.sar, grouping=True)
        costo.cesantia=locale.currency(costo.cesantia, grouping=True)
        costo.infonavit=locale.currency(costo.infonavit, grouping=True)
        costo.isr=locale.currency(costo.isr, grouping=True)



    context = {'costos':costos, 'registros':registros,'myfilter':myfilter,}

    return render(request, 'proyecto/Costo_history.html',context)


def convert_excel_costo(costos):
    response = HttpResponse(content_type="application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Reporte_costos_' + str(datetime.date.today()) + '.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Reporte')
    # Comenzar en la fila 1
    row_num = 1

    # Create heading style and add it to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name="head_style")
    head_style.font = Font(name='Arial', color='00FFFFFF', bold=True, size=11)
    head_style.fill = PatternFill("solid", fgColor='00003366')
    wb.add_named_style(head_style)
    # Create body style and add it to workbook
    body_style = NamedStyle(name="body_style")
    body_style.font = Font(name='Calibri', size=10)
    wb.add_named_style(body_style)
    # Create messages style and add it to workbook
    messages_style = NamedStyle(name="mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size=11)
    wb.add_named_style(messages_style)
    # Create date style and add it to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name='Calibri', size=10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name='Calibri', size=10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name='Calibri', size=14, bold=True)
    wb.add_named_style(money_resumen_style)

    columns = ['Empresa', 'Distrito', 'Proyecto', 'Subproyecto', '#Empleado', 'Nombre', 'Puesto','Nivel','Neto Catorcenal',
               'Complemento Salario Catorcenal', 'Apoyo de Pasajes', 'Total percepciones mensual',
               'Impuesto Estatal', 'IMSS obrero patronal', 'SAR 2%', 'Cesantía', 'Infonavit', 'ISR',
               'Apoyo Visita Familiar', 'Apoyo Estancia', 'Apoyo Renta', 'Apoyo de Estudios',
               'Apoyo de Mantto Vehicular', 'Gasolina', 'Total apoyos', 'Total costo mensual para la empresa',
               'Ingreso mensual neto del empleado', 'Bonos']

    for col_num in range(len(columns)):
        (ws.cell(row=row_num, column=col_num + 1, value=columns[col_num])).style = head_style
        if col_num < 4:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 15
        if col_num == 5 or col_num == 6:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 35
        else:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 15

    columna_max = len(columns) + 2

    (ws.cell(column=columna_max, row=1, value='{Reporte Creado Automáticamente por Savia RH. UH}')).style = messages_style
    (ws.cell(column=columna_max, row=2, value='{Software desarrollado por Vordcab S.A. de C.V.}')).style = messages_style
    #(ws.cell(column=columna_max, row=3, value='Algún dato')).style = messages_style
    #(ws.cell(column=columna_max + 1, row=3, value='alguna sumatoria')).style = money_resumen_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 20
    ws.column_dimensions[get_column_letter(columna_max + 1)].width = 20

    # Se busca el bono actual
    ahora = datetime.date.today()
    catorcena = Catorcenas.objects.filter(fecha_inicial__lte=ahora, fecha_final__gte=ahora).first()

    rows = costos.annotate(bonototal=Sum('bonos__monto', filter=Q(bonos__fecha_bono__range=[catorcena.fecha_inicial, catorcena.fecha_final]))).values_list(
        'status__perfil__empresa__empresa','status__perfil__distrito__distrito','status__perfil__proyecto__proyecto','status__perfil__subproyecto__subproyecto','status__perfil__numero_de_trabajador',
        Concat('status__perfil__nombres', Value(' '), 'status__perfil__apellidos'),'status__puesto__puesto','status__nivel__nivel','neto_catorcenal_sin_deducciones','complemento_salario_catorcenal','apoyo_de_pasajes','total_percepciones_mensual',
        'impuesto_estatal','imms_obrero_patronal','sar','cesantia','infonavit','isr','apoyo_vist_familiar','estancia','renta','apoyo_estudios','amv','gasolina','total_apoyosbonos_agregcomis','total_costo_empresa',
        'ingreso_mensual_neto_empleado','bonototal')

    for id_costo, row in enumerate(rows):
        row_num += 1
        for col_num in range(len(row)):
            if col_num <= 5:
                (ws.cell(row=row_num, column=col_num + 1, value=row[col_num])).style = body_style
            if col_num > 7 and col_num < 27:
                (ws.cell(row=row_num, column=col_num + 1, value=row[col_num])).style = money_style
            else:
                (ws.cell(row=row_num, column=col_num + 1, value=str(row[col_num]))).style = body_style

    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return response


def convert_excel_bancarios(bancarios):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Reporte_datos_bancarios_' + str(datetime.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Reporte')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)

    columns = ['Empresa','Distrito','Nombre','No. de cuenta','No. de tarjeta','Clabe interbancaria','Banco','Bono de la catorcena']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        if col_num < 4:
            ws.column_dimensions[get_column_letter(col_num + 2)].width = 10
        if col_num == 4:
            ws.column_dimensions[get_column_letter(col_num + 2)].width = 30
        else:
            ws.column_dimensions[get_column_letter(col_num + 2)].width = 15


    columna_max = len(columns)+2

    (ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por Savia RH. UH}')).style = messages_style
    (ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}')).style = messages_style
    (ws.cell(column = columna_max, row = 3, value='Algún dato')).style = messages_style
    (ws.cell(column = columna_max +1, row=3, value = 'alguna sumatoria')).style = money_resumen_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 20
    ws.column_dimensions[get_column_letter(columna_max + 1)].width = 20

    rows = bancarios.values_list('status__perfil__empresa__empresa','status__perfil__distrito__distrito',Concat('status__perfil__nombres',Value(' '),
                                'status__perfil__apellidos'),'no_de_cuenta','numero_de_tarjeta','clabe_interbancaria','banco__banco')


    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            if col_num <= 8:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = body_style

    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)

def convert_excel_bonos(bonos):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Reporte_bonos_' + str(datetime.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Reporte')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)

    columns = ['Empresa','Distrito','Nombre','No. de cuenta','No. de tarjeta','Clabe interbancaria',
                'Banco','Fecha del bono','Bono total',]

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        if col_num < 4:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 10
        if col_num == 4:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30
        else:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 15


    columna_max = len(columns)+2

    (ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por Savia RH. UH}')).style = messages_style
    (ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}')).style = messages_style
    (ws.cell(column = columna_max, row = 3, value='Algún dato')).style = messages_style
    (ws.cell(column = columna_max +1, row=3, value = 'alguna sumatoria')).style = money_resumen_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 20
    ws.column_dimensions[get_column_letter(columna_max + 1)].width = 20

    rows = bonos.values_list('costo__status__perfil__empresa__empresa','costo__status__perfil__distrito__distrito',Concat('costo__status__perfil__nombres',Value(' '),
                            'costo__status__perfil__apellidos'),'datosbancarios__no_de_cuenta','datosbancarios__numero_de_tarjeta',
                            'datosbancarios__clabe_interbancaria','datosbancarios__banco','fecha_bono','monto',)


    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            if col_num < 7:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = body_style
            if col_num == 7:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = date_style
            if col_num == 8:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = money_style

    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)

def convert_excel_vacaciones(descansos):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Reporte_vacaciones_' + str(datetime.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Reporte')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)

    columns = ['Empresa','Distrito','Nombre','Fecha de planta anterior','Fecha de planta','Periodo vacacional','Días disponibles año actual',
                'Días disfrutados año actual','Días pendientes año actual','Comentario','Total pendiente',]

    rows = []
    for descanso in descansos:
        # Consulta para obtener vacaciones con total_pendiente != 0
        vacaciones_pendientes = Vacaciones.objects.filter(status=descanso.status, total_pendiente__gt=0,).annotate(Sum('total_pendiente'))
        
        total_pendiente_status = descanso.total_pendiente_status  # Calcula el valor aquí total pendiente total
        row = (descanso.status.perfil.empresa.empresa,descanso.status.perfil.distrito.distrito,f"{descanso.status.perfil.nombres} {descanso.status.perfil.apellidos}",
            descanso.status.fecha_planta_anterior,descanso.status.fecha_planta,descanso.periodo,descanso.dias_de_vacaciones,descanso.dias_disfrutados,
            descanso.total_pendiente,descanso.comentario,total_pendiente_status)
        if vacaciones_pendientes.exists():
            for vacacion_pendiente in vacaciones_pendientes:
                # Añade los valores de total_pendiente y periodo al row original
                row += (vacacion_pendiente.total_pendiente, vacacion_pendiente.periodo)
        rows.append(row)
    max_col = 0
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            if col_num < 3:
                (ws.cell(row=row_num, column=col_num+1, value=row[col_num])).style = body_style
            if col_num > 3:
                (ws.cell(row=row_num, column=col_num+1, value=row[col_num])).style = date_style
            if col_num > 5:
                (ws.cell(row=row_num, column=col_num+1, value=row[col_num])).style = body_style
        if col_num > max_col:
            max_col = col_num

    dynamic_columns = []
    for i in range(14, max_col):
        dynamic_columns.extend([f'Pendiente', f'Periodo'])
    columns = columns + dynamic_columns

    for col_num in range(len(columns)):
        (ws.cell(row = 1, column = col_num+1, value=columns[col_num])).style = head_style
        if col_num < 4:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 10
        if col_num == 4:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30
        else:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 15


    (ws.cell(column = 2, row = row_num + 2, value='{Reporte Creado Automáticamente por Savia RH. UH}')).style = messages_style
    (ws.cell(column = 2, row = row_num + 3, value='{Software desarrollado por Vordcab S.A. de C.V.}')).style = messages_style
    (ws.cell(column = 2, row = row_num + 4, value='Algún dato')).style = messages_style
    (ws.cell(column = 2 + 1, row = row_num + 5, value = 'alguna sumatoria')).style = money_resumen_style

    
    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)

def convert_excel_uniformes(ropas):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Reporte_uniformes_' + str(datetime.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Reporte')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)

    columns = ['Empresa','Distrito','Nombre','Fecha de ultima entrega','Uniformes entregados','Camisola','Pantalon','Camisa administrativa blanca',
                'Camisa administrativa azul','Camisa administrativa beige','Playera polo','Overol','Botas',]

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        if col_num < 4:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 10
        if col_num == 4:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30
        else:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 15


    columna_max = len(columns)+2

    (ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por Savia RH. UH}')).style = messages_style
    (ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}')).style = messages_style
    (ws.cell(column = columna_max, row = 3, value='Algún dato')).style = messages_style
    (ws.cell(column = columna_max +1, row=3, value = 'alguna sumatoria')).style = money_resumen_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 20
    ws.column_dimensions[get_column_letter(columna_max + 1)].width = 20

    rows = ropas.values_list('status__perfil__empresa__empresa','status__perfil__distrito__distrito',Concat('status__perfil__nombres',Value(' '),
                            'status__perfil__apellidos'),'fecha_ultima_entrega','uniformes_entregados','camisola','pantalon','camisa_blanca','camisa_azul',
                            'camisa_beige','playera_polo','overol','botas',)


    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            if col_num < 3:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = body_style
            if col_num == 3:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = date_style
            if col_num > 3:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = body_style


    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)

def convert_excel_economicos(economicos,economicoss):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Reporte_días_economicos_' + str(datetime.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Reporte')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)

    columns = ['Empresa','Distrito','Nombre','Días económicos disfrutados','Días económicos pendientes',]

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        if col_num < 4:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 10
        if col_num == 4:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30
        else:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 15


    columna_max = len(columns)+2

    (ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por Savia RH. UH}')).style = messages_style
    (ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}')).style = messages_style
    (ws.cell(column = columna_max, row = 3, value='Algún dato')).style = messages_style
    (ws.cell(column = columna_max +1, row=3, value = 'alguna sumatoria')).style = money_resumen_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 20
    ws.column_dimensions[get_column_letter(columna_max + 1)].width = 20

    rows = economicos.values_list('status__perfil__empresa__empresa','status__perfil__distrito__distrito',Concat('status__perfil__nombres',Value(' '),
                            'status__perfil__apellidos'),'dias_disfrutados','dias_pendientes',)
    rows2 = economicoss.values_list('status__perfil__empresa__empresa','status__perfil__distrito__distrito',Concat('status__perfil__nombres',Value(' '),
                            'status__perfil__apellidos'),'dias_disfrutados','dias_pendientes',)

    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            if col_num <= 5:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = body_style
    for row in rows2:
        row_num += 1
        for col_num in range(len(row)):
            if col_num <= 5:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = body_style

    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)

def convert_excel_perfil(perfiles):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Reporte_empleados_' + str(datetime.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Reporte')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)

    columns = ['Nombre','Número de trabajador','Empresa','Distrito','Fecha de nacimiento','Correo electrónico','Proyecto','Subproyecto',]

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        if col_num == 0:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30
        elif col_num > 0 and col_num < 5:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 12
        elif col_num > 4 and col_num < 7:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 25
        else:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 10


    columna_max = len(columns)+2

    (ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por Savia RH. UH}')).style = messages_style
    (ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}')).style = messages_style
    (ws.cell(column = columna_max, row = 3, value='Empleados:')).style = messages_style

    ws.column_dimensions[get_column_letter(columna_max)].width = 20
    ws.column_dimensions[get_column_letter(columna_max + 1)].width = 20

    rows = perfiles.values_list(Concat('nombres',Value(' '),'apellidos'),'numero_de_trabajador','empresa__empresa','distrito__distrito',
                                        'fecha_nacimiento','correo_electronico','proyecto__proyecto','subproyecto__subproyecto',)


    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            if col_num < 4:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = body_style
            if col_num == 4:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = date_style
            if col_num > 4:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = body_style

    (ws.cell(column = columna_max +1, row=3, value = row_num - 1)).style = messages_style

    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)

def convert_excel_perfil_baja(perfiles):
    #datos_baja = Datos_baja.objects.filter(perfil__in=perfiles).order_by("perfil__numero_de_trabajador")
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Reporte_empleados_bajas_' + str(datetime.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Reporte')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)

    columns = ['Nombre','Número de trabajador','Empresa','Distrito','Fecha de nacimiento','Correo electrónico','Proyecto','Subproyecto','Fecha baja','Finiquito','Liquidación','Motivo','Exitosa']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        if col_num == 0:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30
        elif col_num > 0 and col_num < 5:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 12
        elif col_num > 4 and col_num < 7:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 25
        else:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 10


    columna_max = len(columns)+2

    (ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por Savia RH. UH}')).style = messages_style
    (ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}')).style = messages_style
    (ws.cell(column = columna_max, row = 3, value='Empleados:')).style = messages_style

    ws.column_dimensions[get_column_letter(columna_max)].width = 20
    ws.column_dimensions[get_column_letter(columna_max + 1)].width = 20

    rows = perfiles.values_list(Concat('perfil__nombres',Value(' '),'perfil__apellidos'),'perfil__numero_de_trabajador','perfil__empresa__empresa','perfil__distrito__distrito',
                                        'perfil__fecha_nacimiento','perfil__correo_electronico','perfil__proyecto__proyecto','perfil__subproyecto__subproyecto','fecha','finiquito',
                                        'liquidacion','motivo','exitosa')


    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            if col_num < 4:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = body_style
            if col_num == 4:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = date_style
            if col_num > 4:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = body_style
            if col_num == 8:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = date_style
            if col_num > 8:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = body_style               

    (ws.cell(column = columna_max +1, row=3, value = row_num - 1)).style = messages_style

    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)

def convert_excel_status(status):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Reporte_empleados_status_' + str(datetime.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Reporte')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)

    columns = ['Nombre','# Trabajador','Distrito','Puesto','Registro patronal','NSS','CURP','RFC','Profesión','No. de cédula','Fecha expedición de cedula','Nivel del empleado','Tipo de contrato','Último contrato vence',
                'Tipo de sangre','Género','Teléfono','Domicilio','Estado civil','Fecha de planta anterior','Fecha de planta actual','Fecha de ingreso','Antigüedad años']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        if col_num == 0:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30
        elif col_num>0 and col_num < 3 and col_num > 3 and col_num < 13:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 15
        elif col_num == 3:
             ws.column_dimensions[get_column_letter(col_num + 1)].width = 20
        elif col_num == 13:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30
        else:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 15

    fecha_actual = date.today()
    columna_max = len(columns)+2
    
    (ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por Savia RH. UH}')).style = messages_style
    (ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}')).style = messages_style
    (ws.cell(column = columna_max, row = 3, value='Empleados:')).style = messages_style
    (ws.cell(column = columna_max, row = 4, value='Generado: ' + str(fecha_actual))).style = messages_style

    ws.column_dimensions[get_column_letter(columna_max)].width = 20
    ws.column_dimensions[get_column_letter(columna_max + 1)].width = 20

    rows = status.values_list(Concat('perfil__nombres',Value(' '),'perfil__apellidos'),'perfil__numero_de_trabajador','perfil__distrito__distrito','puesto__puesto','registro_patronal__patronal','nss','curp','rfc','profesion','no_cedula',
                                        'fecha_cedula','nivel','tipo_de_contrato__contrato','ultimo_contrato_vence','tipo_sangre__sangre','sexo__sexo','telefono','domicilio','estado_civil__estado_civil',
                                        'fecha_planta_anterior','fecha_planta','fecha_ingreso')
    #for row in rows:
    #    if row == datetime.date(6000, 1, 1):
    #        row = "Especial"
    
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            if col_num < 10:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = body_style
            if col_num == 10:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = date_style
            if col_num < 13 and col_num > 10:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = body_style
            if col_num == 13:
                value=row[col_num]
                if value == datetime.date(6000, 1, 1):
                    (ws.cell(row = row_num, column = col_num+1, value= "Especial")).style = body_style
                elif value == datetime.date(6001, 1, 1):
                    (ws.cell(row = row_num, column = col_num+1, value= "INDETERMINADO")).style = body_style
                elif value == datetime.date(6002, 1, 1):
                    (ws.cell(row = row_num, column = col_num+1, value= "HONORARIOS")).style = body_style
                elif value == datetime.date(6003, 1, 1):
                    (ws.cell(row = row_num, column = col_num+1, value= "Sin fecha")).style = body_style
                else:
                    (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = date_style
            if col_num > 13:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = body_style
            if col_num >= 19:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = date_style
        if row[col_num] != None:
            antiguedad = (fecha_actual.year - row[col_num].year)
            (ws.cell(row = row_num, column = col_num+2, value=antiguedad)).style = body_style
                
    (ws.cell(column = columna_max +1, row=3, value = row_num-1)).style = messages_style

    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)

@login_required(login_url='user-login')
def upload_batch_empleados(request):

    form = Empleados_BatchForm(request.POST or None, request.FILES or None)

    if form.is_valid():
        form.save()
        form = Empleados_BatchForm()
        empleados_list = Empleados_Batch.objects.get(activated = False)
        f = open(empleados_list.file_name.path, 'r', encoding='latin1')
        reader = csv.reader(f)
        next(reader) #Advance past the reader

        for row in reader:
            fecha=datetime.datetime.strptime(row[5], "%d/%m/%Y").date()
            #if Perfil.objects.get(numero_de_trabajador != row[0]):
            if Perfil.objects.filter(numero_de_trabajador = row[0], distrito__distrito = row[2]):
                messages.error(request,f'El perfil del empleado #{row[0]} ya existe dentro de la base de datos')
            else:
                if Empresa.objects.filter(empresa = row[1]):
                    empresa = Empresa.objects.get(empresa = row[1])
                    if row[2] == "MOTOS":
                        distrito = Distrito.objects.get(distrito = "Planta Veracruz")
                        division = "MOTOS"
                    elif row[2] == "PETROLERA":
                        distrito = Distrito.objects.get(distrito = "Planta Veracruz")
                        division = "PETROLERA"
                    elif Distrito.objects.filter(distrito = row[2]):
                        distrito = Distrito.objects.get(distrito = row[2])
                        division = ""
                    else:
                        messages.error(request,f'El distrito no existe dentro de la base de datos, empleado #{row[0]}')
                    if Proyecto.objects.filter(proyecto = row[7]):
                        proyecto = Proyecto.objects.get(proyecto = row[7])
                        if SubProyecto.objects.get(subproyecto = row[8], proyecto__proyecto = row[7]):
                            subproyecto = SubProyecto.objects.get(subproyecto = row[8], proyecto__proyecto = row[7])
                            empleado = Perfil(numero_de_trabajador=row[0], empresa=empresa, distrito=distrito, nombres=row[3],
                                apellidos=row[4],fecha_nacimiento=fecha,correo_electronico=row[6],proyecto=proyecto,subproyecto=subproyecto,
                                complete=True, complete_status=False,)
                            empleado.division = division
                            empleado.save()
                        else:
                            messages.error(request,f'El subproyecto no existe dentro de la base de datos, empleado #{row[0]}')
                    else:
                        messages.error(request,f'El proyecto no existe dentro de la base de datos, empleado #{row[0]}')
                else:
                    messages.error(request,f'La empresa no existe dentro de la base de datos, empleado #{row[0]}')
        empleados_list.activated = True
        empleados_list.save()

    context = {
        'form': form,
        }

    return render(request,'proyecto/upload_batch_empleados.html', context)

@login_required(login_url='user-login')
def upload_batch_status(request):

    form = Status_BatchForm(request.POST or None, request.FILES or None)

    if form.is_valid():
        form.save()
        form = Status_BatchForm()
        status_list = Status_Batch.objects.get(activated = False)
        f = open(status_list.file_name.path, 'r', encoding='latin1')
        reader = csv.reader(f)
        next(reader) #Advance past the reader

        for row in reader:
            ultimo_contrato = datetime.datetime.strptime(row[11], "%d/%m/%Y").date()
            if row[18] == '':
                ingreso = None
            else:
                ingreso = datetime.datetime.strptime(row[18], "%d/%m/%Y").date()
            if row[17] == '':
                planta = None
            else:
                planta = datetime.datetime.strptime(row[17], "%d/%m/%Y").date()
            if row[16] == '':
                planta_anterior = None
            else:
                planta_anterior = datetime.datetime.strptime(row[16], "%d/%m/%Y").date()
            if row[1] == "MOTOS":
                distrito = Distrito.objects.get(distrito = "Planta Veracruz")
            elif row[1] == "PETROLERA":
                distrito = Distrito.objects.get(distrito = "Planta Veracruz")
            elif Distrito.objects.filter(distrito = row[1]):
                distrito = Distrito.objects.get(distrito = row[1])
            else:
                messages.error(request,f'El distrito no existe dentro de la base de datos, empleado #{row[0]}')
            if Perfil.objects.filter(numero_de_trabajador = row[0], distrito__distrito = distrito):
                perfil = Perfil.objects.get(numero_de_trabajador = row[0], distrito__distrito = distrito)
                if Status.objects.filter(perfil__numero_de_trabajador = row[0], perfil__distrito__distrito = distrito):
                    messages.error(request,f'El status del empleado #{row[0]} ya existe dentro de la base de datos')
                else:
                    if RegistroPatronal.objects.filter(patronal = row[2]):
                        registro_patronal = RegistroPatronal.objects.get(patronal = row[2])
                        if Nivel.objects.filter(nivel = row[9]):
                            nivel = Nivel.objects.get(nivel = row[9])
                            if Contrato.objects.filter(contrato = row[10]):
                                tipo_de_contrato = Contrato.objects.get(contrato = row[10])
                                if Sangre.objects.filter(sangre = row[12]):
                                    sangre = Sangre.objects.get(sangre = row[12])
                                    if Sexo.objects.filter(sexo = row[13]):
                                        genero = Sexo.objects.get(sexo = row[13])
                                        if Civil.objects.filter(estado_civil = row[15]):
                                            civil = Civil.objects.get(estado_civil = row[15])
                                            if Puesto.objects.filter(puesto = row[19]):
                                                puesto = Puesto.objects.get(puesto = row[19])
                                                perfil.complete_status = True
                                                perfil.save()
                                                status = Status(perfil=perfil,registro_patronal= registro_patronal,nss=row[3],curp=row[4],rfc=row[5],telefono=row[6],profesion=row[7],
                                                        no_cedula=row[8],nivel=nivel,tipo_de_contrato=tipo_de_contrato,ultimo_contrato_vence=ultimo_contrato,tipo_sangre=sangre,sexo=genero,domicilio=row[14],
                                                        estado_civil=civil,fecha_planta_anterior=planta_anterior,fecha_planta=planta,fecha_ingreso=ingreso,puesto=puesto,complete=True,)

                                                status.save()
                                            else:
                                                messages.error(request,f'El puesto no existe dentro de la base de datos, empleado #{row[0]}')
                                        else:
                                            messages.error(request,f'El estado civil no existe dentro de la base de datos, empleado #{row[0]}')
                                    else:
                                        messages.error(request,f'El genero no existe dentro de la base de datos, empleado #{row[0]}')
                                else:
                                    messages.error(request,f'El tipo de sangre no existe dentro de la base de datos, empleado #{row[0]}')
                            else:
                                messages.error(request,f'El tipo de contrato no existe dentro de la base de datos, empleado #{row[0]}')
                        else:
                            messages.error(request,f'El nivel no existe dentro de la base de datos, empleado #{row[0]}')

                    else:
                        messages.error(request,f'El registro patronal no existe dentro de la base de datos, empleado #{row[0]}')
            else:
                messages.error(request,f'El perfil del empleado #{row[0]} no existe dentro de la base de datos')
        status_list.activated = True
        status_list.save()



    context = {
        'form': form,
        }

    return render(request,'proyecto/upload_batch_status.html', context)

@login_required(login_url='user-login')
def upload_batch_costos(request):
    dato = SalarioDatos.objects.get()
    factores = FactorIntegracion.objects.all()
    tablas= DatosISR.objects.all()
    tcesantias= TablaCesantia.objects.all() ###
    quincena=Decimal(14.00)
    mes=Decimal(30.40)
    impuesto_est=Decimal(0.0315)
    sar=Decimal(0.02)
    cesantia=Decimal(0.0135) ###
    infonavit=Decimal(0.05)
    comision=Decimal(0.09)
    form = Costos_BatchForm(request.POST or None, request.FILES or None)

    if form.is_valid():
        form.save()
        form = Costos_BatchForm()
        costo_list = Costos_Batch.objects.get(activated = False)
        f = open(costo_list.file_name.path, 'r', encoding='latin1')
        reader = csv.reader(f)
        next(reader) #Advance past the reader

        for row in reader:
            #planta = datetime.datetime.strptime(row[16], "%d/%m/%Y").date()
            #planta_anterior = datetime.datetime.strptime(row[15], "%d/%m/%Y").date()
            if row[1] == "MOTOS":
                distrito = Distrito.objects.get(distrito = "Planta Veracruz")
            elif row[1] == "PETROLERA":
                distrito = Distrito.objects.get(distrito = "Planta Veracruz")
            elif Distrito.objects.filter(distrito = row[1]):
                distrito = Distrito.objects.get(distrito = row[1])
            else:
                messages.error(request,f'El distrito no existe dentro de la base de datos, empleado #{row[0]}')
            if Status.objects.filter(perfil__numero_de_trabajador = row[0], perfil__distrito__distrito = distrito):
                status = Status.objects.get(perfil__numero_de_trabajador = row[0], perfil__distrito__distrito = distrito)
                if status.fecha_ingreso != None:
                    if Costo.objects.filter(status__perfil__numero_de_trabajador = row[0], status__perfil__distrito__distrito = distrito):
                        messages.error(request,f'El empleado #{row[0]} ya se encuentra en la base de datos')
                    else:
                        status.complete_costo = True
                        costo = Costo(status=status,neto_catorcenal_sin_deducciones=row[2],complemento_salario_catorcenal=row[3],sueldo_diario=row[4],
                                amortizacion_infonavit=row[5],fonacot=row[6],apoyo_de_pasajes=row[7],apoyo_vist_familiar=row[8],estancia=row[9],renta=row[10],
                                campamento=row[11],apoyo_estudios=row[12],gasolina=row[13],amv=row[14],laborados=row[15],complete=True,)

                        neto_catorcenal_sin_deducciones = Decimal(row[2])
                        complemento_salario_catorcenal = Decimal(row[3])
                        sueldo_diario = Decimal(row[4])
                        #sdi = Decimal(row[5])
                        #imms_obrero_patronal = Decimal(row[6])
                        amortizacion_infonavit = Decimal(row[5])
                        fonacot = Decimal(row[6])
                        apoyo_de_pasajes = Decimal(row[7])
                        apoyo_vist_familiar = Decimal(row[8])
                        estancia = Decimal(row[9])
                        renta = Decimal(row[10])
                        campamento= Decimal(row[11])
                        apoyo_estudios= Decimal(row[12])
                        gasolina= Decimal(row[13])
                        amv= Decimal(row[14])
                        #SDI Calculo
                        prima_riesgo = costo.status.registro_patronal.prima
                        excedente = dato.UMA*3
                        cuotafija = (dato.UMA*Decimal(0.204))*Decimal(costo.laborados)
                        excedente_patronal = (Decimal(costo.sueldo_diario)-excedente)*Decimal(0.011)*Decimal(costo.laborados)
                        excedente_obrero = (Decimal(costo.sueldo_diario)-excedente)*Decimal(0.004)*Decimal(costo.laborados)
                        if excedente_patronal < 0:
                            excedente_patronal = 0
                        if excedente_obrero < 0:
                            excedente_obrero = 0
                        prestaciones_patronal = (Decimal(costo.sueldo_diario)*Decimal(0.007))*Decimal(costo.laborados)
                        prestaciones_obrero = (Decimal(costo.sueldo_diario)*Decimal(0.0025))*Decimal(costo.laborados)
                        gastosmp_patronal = (Decimal(costo.sueldo_diario)*Decimal(0.0105))*Decimal(costo.laborados)
                        gastosmp_obrero = (Decimal(costo.sueldo_diario)*Decimal(0.00375))*Decimal(costo.laborados)
                        riesgo_trabajo = (Decimal(costo.sueldo_diario)*(prima_riesgo/100))*Decimal(costo.laborados)
                        invalidezvida_patronal = (Decimal(costo.sueldo_diario)*Decimal(0.0175))*Decimal(costo.laborados)
                        invalidezvida_obrero = (Decimal(costo.sueldo_diario)*Decimal(0.00625))*Decimal(costo.laborados)
                        guarderias_prestsociales = (Decimal(costo.sueldo_diario)*Decimal(0.01))*Decimal(costo.laborados)
                        costo.imms_obrero_patronal = (cuotafija+excedente_patronal+excedente_obrero+prestaciones_patronal
                                                    +prestaciones_obrero+gastosmp_patronal+gastosmp_obrero+riesgo_trabajo+invalidezvida_patronal
                                                    +invalidezvida_obrero+guarderias_prestsociales)
                        totall = costo.imms_obrero_patronal
                        #Calculo de la antiguedad para el factor de integracion
                        actual = date.today()
                        años_ingreso = relativedelta(actual, costo.status.fecha_ingreso)
                        años_ingreso = años_ingreso.years
                        if años_ingreso == 0:
                            años_ingreso=1
                        for factor in factores:
                            if años_ingreso >= factor.years:
                                factor_integracion = factor.factor
                        costo.sdi = factor_integracion*Decimal(costo.sueldo_diario) ###
                        sdi = costo.sdi
                            #Costo calculo
                        costo.total_deduccion = amortizacion_infonavit + fonacot
                        costo.neto_pagar = neto_catorcenal_sin_deducciones - costo.total_deduccion
                        costo.sueldo_mensual_neto = (neto_catorcenal_sin_deducciones/quincena)*mes
                        costo.complemento_salario_mensual = (complemento_salario_catorcenal/quincena)*mes
                        costo.sueldo_mensual = sueldo_diario*mes
                        costo.sueldo_mensual_sdi = sdi*mes
                        costo.total_percepciones_mensual = apoyo_de_pasajes + costo.sueldo_mensual
                        for tabla in tablas:
                            if costo.total_percepciones_mensual >= tabla.liminf:
                                costo.lim_inferior = tabla.liminf
                                costo.tasa=tabla.excedente
                                costo.cuota_fija=tabla.cuota
                            if costo.lim_inferior >= tabla.p_ingresos:
                                costo.subsidio=tabla.subsidio
                        costo.impuesto_estatal= costo.total_percepciones_mensual*impuesto_est
                        costo.sar= costo.sueldo_mensual_sdi*sar
                        #Parte de cesantia
                        busqueda_cesantia= sdi/dato.UMA ###
                        for tcesantia in tcesantias:   ####
                            if  busqueda_cesantia >= tcesantia.sbc:
                                cesantia_valor = tcesantia.cuota_patronal
                        cesantia_ley= costo.sueldo_mensual_sdi*(cesantia_valor/100)                        ###
                        costo.cesantia= (costo.sueldo_mensual_sdi*cesantia)+cesantia_ley  ####
                        #Parte de vacaciones
                        vac_reforma_actual = Decimal((12/365)*365)*Decimal(costo.sueldo_diario)
                        prima_vacacional = vac_reforma_actual*Decimal(0.25)
                        aguinaldo = Decimal((15/365)*365)*Decimal(costo.sueldo_diario)
                        total_vacaciones = (vac_reforma_actual+prima_vacacional+aguinaldo)/12

                        costo.infonavit= costo.sueldo_mensual_sdi*infonavit
                        costo.excedente= costo.total_percepciones_mensual - costo.lim_inferior
                        costo.impuesto_marginal= costo.excedente * costo.tasa
                        costo.impuesto= costo.impuesto_marginal + costo.cuota_fija
                        costo.isr= costo.impuesto
                        #dato.otros_bonos= dato.bonos.bonos_ct_ocho + dato.bonos.bonos_ct_nueve
                        costo.total_apoyosbonos_empleadocomp= apoyo_vist_familiar + estancia + renta + apoyo_estudios + amv + campamento + gasolina

                        costo.total_apoyosbonos_agregcomis = campamento #Modificar falta suma
                        costo.comision_complemeto_salario_bonos= (costo.complemento_salario_mensual + campamento)*comision #Falta suma dentro de la multiplicacion
                        costo.total_costo_empresa = costo.sueldo_mensual_neto + costo.complemento_salario_mensual + Decimal(costo.apoyo_de_pasajes) + costo.impuesto_estatal + costo.imms_obrero_patronal + costo.sar + costo.cesantia + costo.infonavit + costo.isr + costo.total_apoyosbonos_empleadocomp
                        costo.total_costo_empresa = costo.total_costo_empresa + total_vacaciones
                        costo.ingreso_mensual_neto_empleado= costo.sueldo_mensual_neto + costo.complemento_salario_mensual + Decimal(costo.apoyo_de_pasajes) + costo.total_apoyosbonos_empleadocomp # + costo.total_apoyosbonos_agregcomis

                        costo.save()
            else:
                messages.error(request,f'El status del empleado #{row[0]} no existe dentro de la base de datos')

        costo_list.activated = True
        costo_list.save()
        status.save()


    context = {
        'form': form,
        }

    return render(request,'proyecto/upload_batch_costos.html', context)
"""
#Bancarios actualizar datos
@login_required(login_url='user-login')
def upload_batch_bancarios(request):

    form = Bancarios_BatchForm(request.POST or None, request.FILES or None)

    if form.is_valid():
        form.save()
        form = Bancarios_BatchForm()
        bancarios_list = Bancarios_Batch.objects.get(activated = False)
        f = open(bancarios_list.file_name.path, 'r')
        reader = csv.reader(f)
        next(reader) #Advance past the reader

        for row in reader:
            if Status.objects.filter(perfil__numero_de_trabajador = row[0], perfil__distrito__distrito = row[1]):
                status = Status.objects.get(perfil__numero_de_trabajador = row[0], perfil__distrito__distrito = row[1])
                bancario = DatosBancarios.objects.get(status=status)
                if Banco.objects.filter(banco = row[5]):
                    banco = Banco.objects.get(banco = row[5])
                    status.complete_bancarios = True
                    bancario.no_de_cuenta = row[2]
                    bancario.numero_de_tarjeta = row[3]
                    bancario.clabe_interbancaria = row[4]
                    bancario.banco = banco
                    bancario.complete = True
                    bancario.save()
                else:
                    messages.error(request,f'El banco no existe dentro de la base de datos, empleado #{row[0]}')
            else:
                messages.error(request,f'El empleado no existe dentro de la base de datos, empleado #{row[0]}')
        bancarios_list.activated = True
        bancarios_list.save()
        status.save()

    context = {
        'form': form,
        }

    return render(request,'proyecto/upload_batch_bancarios.html', context)

    """
 #Bancarios agregar datos
@login_required(login_url='user-login')
def upload_batch_bancarios(request):

    form = Bancarios_BatchForm(request.POST or None, request.FILES or None)

    if form.is_valid():
        form.save()
        form = Bancarios_BatchForm()
        bancarios_list = Bancarios_Batch.objects.get(activated = False)
        f = open(bancarios_list.file_name.path, 'r')
        reader = csv.reader(f)
        next(reader) #Advance past the reader

        for row in reader:
            if Status.objects.filter(perfil__numero_de_trabajador = row[0], perfil__distrito__distrito = row[1]):
                status = Status.objects.get(perfil__numero_de_trabajador = row[0], perfil__distrito__distrito = row[1])
                if DatosBancarios.objects.filter(status__perfil__numero_de_trabajador=row[0], status__perfil__distrito__distrito = row[1]):
                    messages.error(request,f'El empleado #{row[0]} ya existe dentro de la base de datos')
                else:
                    if Banco.objects.filter(banco = row[5]):
                        banco = Banco.objects.get(banco = row[5])
                        status.complete_bancarios = True
                        bancarios = DatosBancarios(status=status, no_de_cuenta=row[2], numero_de_tarjeta=row[3], clabe_interbancaria=row[4],banco=banco,complete=True,)
                        bancarios.save()
                    else:
                        messages.error(request,f'El banco no existe dentro de la base de datos, empleado #{row[0]}')
            else:
                messages.error(request,f'El empleado no existe dentro de la base de datos, empleado #{row[0]}')
        bancarios_list.activated = True
        bancarios_list.save()
        status.save()

    context = {
        'form': form,
        }

    return render(request,'proyecto/upload_batch_bancarios.html', context)

@login_required(login_url='user-login')
def upload_batch_bonos(request): #Bonos en la tabla de bancarios

    form = Bancarios_BatchForm(request.POST or None, request.FILES or None)

    if form.is_valid():
        form.save()
        form = Bancarios_BatchForm()
        """bancarios_list = Bancarios_Batch.objects.get(activated = False)
        f = open(bancarios_list.file_name.path, 'r')
        reader = csv.reader(f)
        next(reader) #Advance past the reader

        for row in reader:
            if Status.objects.filter(perfil__numero_de_trabajador = row[0], perfil__distrito__distrito = row[1]):
                status = Status.objects.get(perfil__numero_de_trabajador = row[0], perfil__distrito__distrito = row[1])
                if DatosBancarios.objects.filter(status__perfil__numero_de_trabajador=row[0], status__perfil__distrito__distrito = row[1]):
                    messages.error(request,f'El empleado #{row[0]} ya existe dentro de la base de datos')
                else:
                    if Banco.objects.filter(banco = row[5]):
                        banco = Banco.objects.get(banco = row[5])
                        status.complete_bancarios = True
                        bancarios = DatosBancarios(status=status, no_de_cuenta=row[2], numero_de_tarjeta=row[3], clabe_interbancaria=row[4],banco=banco,complete=True,)
                        bancarios.save()
                    else:
                        messages.error(request,f'El banco no existe dentro de la base de datos, empleado #{row[0]}')
            else:
                messages.error(request,f'El empleado no existe dentro de la base de datos, empleado #{row[0]}')
        bancarios_list.activated = True
        bancarios_list.save()
        status.save()"""

    context = {
        'form': form,
        }

    return render(request,'proyecto/upload_batch_bonos.html', context) #tabla de bancarios esta el view del batch

def reporte_pdf_uniformes(uniformes, pk):
    orden = Uniformes.objects.get(id=pk)
    uniformes = Uniforme.objects.filter(orden__id=pk)
    ropas = uniformes.aggregate(Sum('cantidad'))
    suma_ropas = ropas['cantidad__sum']
    cantidad = str(suma_ropas)
    #Configuration of the PDF object
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)

    #Colores utilizados
    azul = Color(0.16015625,0.5,0.72265625)
    rojo = Color(0.59375, 0.05859375, 0.05859375)
    #Encabezado
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',10)
    c.drawString(440,735,'Folio:')

    c.setFillColor(azul)
    c.setStrokeColor(azul)
    c.setLineWidth(20)
    c.line(20,760,585,760) #Linea azul superior
    c.setLineWidth(0.2)
    c.line(20,727.5,585,727.5) #Linea posterior horizontal
    c.line(250,727.5,250,590) #Linea vertical

    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica',10)
    c.drawCentredString(295,755,'ORDEN DE UNIFORMES')

    c.drawInlineImage('static/images/Logo-Vordtec.png',50,620, 6 * cm, 3 * cm) #Imagen VORDCAB
    #Primera columna
    c.setFillColor(black)
    c.setFont('Helvetica',10)
    c.drawString(260,710,'Empleado #:')
    numero_trabajador = str(orden.status.perfil.numero_de_trabajador)
    c.drawString(325,710, numero_trabajador)
    c.drawString(260,690,'Nombre:',)
    c.drawString(325,690, orden.status.perfil.nombres)
    c.drawString(390,690, orden.status.perfil.apellidos)
    c.drawString(260,670,'Empresa:')
    empresa = str(orden.status.perfil.empresa)
    c.drawString(325,670, empresa)
    c.drawString(260,650,'Distrito')
    distrito = str(orden.status.perfil.distrito)
    c.drawString(325,650, distrito)
    c.drawString(260,630,'Fecha de pedido:')
    fecha = str(orden.fecha_pedido)
    c.drawString(345,630, fecha)
    c.drawString(260,610,'Cantidad total de piezas')
    c.drawString(380,610, cantidad)
    #Segunda columna
    #c.drawString(420,710,'Activo:')
    #c.drawString(420,690, 'NA')
    #c.drawString(420,670, 'Sector:')
    #c.drawString(420,650, 'NA')
    #c.drawString(420,630, 'Fecha Emisión:')
    #c.drawString(420,610,'28-06-2022 11:16:21')
    c.setFillColor(rojo) ## NUMERO DEL FOLIO
    id = str(orden.id)
    c.drawString(475,735, id)

    #Tabla y altura guia
    data =[]
    high = 550
    data.append(['''Orden #''','''Producto''','''Cantidad''', '''Talla''',])
    for uniforme in uniformes: #Salen todos los datos
        data.append([uniforme.id,uniforme.ropa,uniforme.cantidad,uniforme.talla,])
        high = high - 18

    #Observaciones
    c.setFillColor(azul)
    c.setLineWidth(20)
    c.line(20,high-35,585,high-35) #Linea posterior horizontal
    c.setFillColor(white)
    c.setLineWidth(.1)
    c.setFont('Helvetica-Bold',10)
    c.drawCentredString(295,high-40,'Observaciones')
    c.setFillColor(black)
    c.setFont('Helvetica',8)
    c.drawCentredString(295,high-60,'                                                                                                                ')
    c.drawCentredString(295,high-70,'                                                                                                                ')

    #Autorizacion parte
    c.setFillColor(azul)
    c.setFont('Helvetica',8)
    c.setLineWidth(1)
    c.line(150,high-150,275,high-150) #Linea posterior horizontal
    c.line(350,high-150,475,high-150) #Linea posterior horizontal
    c.setFillColor(black)
    c.drawCentredString(212.5,high-160,'Empleado')
    c.drawCentredString(412.5,high-160,'Aprobación')
    c.drawCentredString(180,high-145,orden.status.perfil.nombres)
    c.drawCentredString(240,high-145,orden.status.perfil.apellidos)

    c.drawCentredString(412.5,high-145,'Nombre aprobador')

    #Pie de pagina
    c.setFillColor(azul)
    c.setLineWidth(40)
    c.line(20,50,585,50) #Linea posterior horizontal
    c.setFillColor(white)
    #c.drawCentredString(70,53,'Clasificación:')
    #c.drawCentredString(140,53,'Nivel:')
    #c.drawCentredString(240,53,'Preparado por:')
    #c.drawCentredString(350,53,'Aprobado:')
    #c.drawCentredString(450,53,'Fecha emisión:')
    #c.drawCentredString(550,53,'Rev:')
    #Parte de abajo
    #c.drawCentredString(70,39,'Controlado')
    #c.drawCentredString(140,39,'N5')
    #c.drawCentredString(240,39,'SEOV-ALM-N4-01-01')
    #c.drawCentredString(350,39,'SUB ADM')
    #c.drawCentredString(450,39,'24/Oct/2018')
    #c.drawCentredString(550,39,'001')

    #Propiedades de la tabla
    width, height = letter
    table = Table(data, colWidths=[2.6 * cm, 2.6 * cm, 11.8 * cm, 2.6 * cm], repeatRows=1)
    table.setStyle(TableStyle([ #estilos de la tabla
        #ENCABEZADO
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('TEXTCOLOR',(0,0),(-1,0), white),
        ('FONTSIZE',(0,0),(-1,0), 13),
        ('BACKGROUND',(0,0),(-1,0), azul),
        #CUERPO
        ('TEXTCOLOR',(0,1),(-1,-1), colors.black),
        ('FONTSIZE',(0,1),(-1,-1), 10),
        ]))
    table.wrapOn(c, width, height)
    table.drawOn(c, 25, high)
    c.save()
    c.showPage()
    buf.seek(0)

    return FileResponse(buf, as_attachment=True, filename='Uniforme_reporte.pdf')

def reporte_pdf_costo_detalles(costo,bonototal):
    now = datetime.date.today()
    fecha = str(now)
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)

    #Colores utilizados
    azul = Color(0.16015625,0.5,0.72265625)
    rojo = Color(0.59375, 0.05859375, 0.05859375)
    #Encabezado
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',10)
    c.drawString(240,735,'Fecha del costo:')
    created_at = str(costo.created_at)
    c.drawString(330,735,created_at)
    c.drawString(440,735,'Fecha reporte:')
    c.drawString(510,735,fecha)

    c.setFillColor(azul)
    c.setStrokeColor(azul)
    c.setLineWidth(20)
    c.line(20,760,585,760) #Linea azul superior
    c.setLineWidth(0.2)
    c.line(20,727.5,585,727.5) #Linea posterior horizontal
    c.line(250,727.5,250,590) #Linea vertical

    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica',10)
    c.drawCentredString(295,755,'Detalles de Costo')

    #c.drawInlineImage(costo.status.perfil.empresa.logo,50,620, 6 * cm, 3 * cm) #Imagen VORDCAB
    logo = ImageReader(costo.status.perfil.empresa.logo)
    c.drawImage(logo,70,640, 4 * cm, 2 * cm)
    #Primera columna
    c.setFillColor(black)
    c.setFont('Helvetica-Bold',10)
    c.drawString(260,710,'Empleado #:')
    c.drawString(260,690,'Nombre:',)
    c.drawString(260,670,'Empresa:')
    c.drawString(260,650,'Distrito')
    c.line(20,590,585,590) #Linea posterior horizontal
    #Primera columna
    c.drawString(40,575,'Puesto:')
    c.drawString(40,550,'Amortización infonavit:')
    c.drawString(40,525,'Fonacot:')
    c.drawString(40,500,'Neto catorcenal sin deducciones:')
    c.drawString(40,475,'Complemento salario catorcenal:')
    c.drawString(40,450,'Sueldo diario:')
    c.drawString(40,425,'SDI:')
    c.drawString(40,400,'Apoyo de pasajes:')
    c.drawString(40,375,'IMSS obrero patronal:')
    c.drawString(40,350,'Apoyo visita familiar:')
    c.drawString(40,325,'Estancia:')
    c.drawString(40,300,'Renta:')
    c.drawString(40,275,'Apoyo de estudios:')
    c.drawString(40,250,'Apoyo mantenimiento vehicular:')
    c.drawString(40,225,'Gasolina:')
    c.drawString(40,200,'Campamento:')
    c.drawString(40,175,'Total deducción:')
    c.drawString(40,150,'Neto a pagar:')
    c.drawString(40,125,'Sueldo mensual neto:')
    #Segunda columna
    c.drawString(300,575,'Sueldo mensual:')
    c.drawString(300,550,'Sueldo mensual SDI:')
    c.drawString(300,525,'Total percepcion mensual:')
    c.drawString(300,500,'Impuesto estatal:')
    c.drawString(300,475,'SAR:')
    c.drawString(300,450,'Cesantia:')
    c.drawString(300,425,'Infonavit:')
    c.drawString(300,400,'ISR:')
    c.drawString(300,375,'Limite inferior:')
    c.drawString(300,350,'Excedente:')
    c.drawString(300,325,'Tasa:')
    c.drawString(300,300,'Impuesto marginal:')
    c.drawString(300,275,'Cuota fija:')
    c.drawString(300,250,'Impuesto:')
    c.drawString(300,225,'Subsidio:')
    c.drawString(300,200,'Total apoyos y bonos empleado comprueba:')
    c.drawString(300,175,'Bono total:')
    c.drawString(300,150,'Comision complemento de salario bonos:')
    c.drawString(300,125,'Total costo para la empresa:')
    c.drawString(300,100,'Ingreso mensual neto del empleado:')
    c.setFont('Helvetica',10)
    #Parte superior
    numero_trabajador = str(costo.status.perfil.numero_de_trabajador)
    c.drawString(325,710, numero_trabajador)
    c.drawString(325,690, costo.status.perfil.nombres+' '+costo.status.perfil.apellidos)
    empresa = str(costo.status.perfil.empresa)
    c.drawString(325,670, empresa)
    distrito = str(costo.status.perfil.distrito)
    c.drawString(325,650, distrito)
    #Primera columna
    puesto = str(costo.status.puesto)
    c.drawString(90,575,puesto)
    c.drawString(170,550,costo.amortizacion_infonavit)
    c.drawString(100,525,costo.fonacot)
    c.drawString(220,500,costo.neto_catorcenal_sin_deducciones)
    c.drawString(220,475,costo.complemento_salario_catorcenal)
    c.drawString(130,450,costo.sueldo_diario)
    c.drawString(80,425,costo.sdi)
    c.drawString(150,400,costo.apoyo_de_pasajes)
    c.drawString(170,375,costo.imms_obrero_patronal)
    c.drawString(170,350,costo.apoyo_vist_familiar)
    c.drawString(100,325,costo.estancia)
    c.drawString(100,300,costo.impuesto_marginal)
    c.drawString(150,275,costo.apoyo_estudios)
    c.drawString(220,250,costo.amv)
    c.drawString(120,225,costo.gasolina)
    c.drawString(130,200,costo.campamento)
    c.drawString(140,175,costo.total_deduccion)
    c.drawString(120,150,costo.neto_pagar)
    c.drawString(170,125,costo.sueldo_mensual_neto)
    #Segunda columna
    c.drawString(405,575,costo.sueldo_mensual)
    c.drawString(425,550,costo.sueldo_mensual_sdi)
    c.drawString(445,525,costo.total_percepciones_mensual)
    c.drawString(405,500,costo.impuesto_estatal)
    c.drawString(365,475,costo.sar)
    c.drawString(365,450,costo.cesantia)
    c.drawString(365,425,costo.infonavit)
    c.drawString(365,400,costo.isr)
    c.drawString(395,375,costo.lim_inferior)
    c.drawString(385,350,costo.excedente)
    c.drawString(355,325,costo.tasa)
    c.drawString(415,300,costo.impuesto_marginal)
    c.drawString(365,275,costo.cuota_fija)
    c.drawString(365,250,costo.impuesto)
    c.drawString(365,225,costo.subsidio)
    c.drawString(525,200,costo.total_apoyosbonos_empleadocomp)
    c.drawString(375,175,bonototal)
    c.drawString(515,150,costo.comision_complemeto_salario_bonos)
    c.drawString(465,125,costo.total_costo_empresa)
    c.drawString(485,100,costo.ingreso_mensual_neto_empleado)

    #Pie de pagina
    c.setFillColor(azul)
    c.setLineWidth(40)
    c.line(20,50,585,50) #Linea posterior horizontal
    c.setFillColor(white)
    #c.drawCentredString(70,53,'Clasificación:')
    #c.drawCentredString(140,53,'Nivel:')
    #c.drawCentredString(240,53,'Preparado por:')
    #c.drawCentredString(350,53,'Aprobado:')
    #c.drawCentredString(450,53,'Fecha emisión:')
    #c.drawCentredString(550,53,'Rev:')
    #Parte de abajo
    #c.drawCentredString(70,39,'Controlado')
    #c.drawCentredString(140,39,'N5')
    #c.drawCentredString(240,39,'SEOV-ALM-N4-01-01')
    #c.drawCentredString(350,39,'SUB ADM')
    #c.drawCentredString(450,39,'24/Oct/2018')
    #c.drawCentredString(550,39,'001')
    c.save()
    c.showPage()
    buf.seek(0)

    return FileResponse(buf, as_attachment=True, filename='CostoDetalle.pdf')

def reporte_pdf_economico_detalles(economicos,empleado):
    now = datetime.date.today()
    fecha = str(now)
    #Colores utilizados
    azul = Color(0.16015625,0.5,0.72265625)
    rojo = Color(0.59375, 0.05859375, 0.05859375)
    if empleado.complete_dias == True:
        estado = "Complete"
        color = azul
    else:
        estado = "Incomplete"
        color = rojo
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)

    #Encabezado
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',10)
    c.drawString(440,735,'Fecha reporte:')
    c.drawString(510,735,fecha)

    c.setFillColor(azul)
    c.setStrokeColor(azul)
    c.setLineWidth(20)
    c.line(20,760,585,760) #Linea azul superior
    c.setLineWidth(0.2)
    c.line(20,727.5,585,727.5) #Linea posterior horizontal
    c.line(250,727.5,250,590) #Linea vertical

    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica',10)
    c.drawCentredString(295,755,'Detalle de días economicos')

    #c.drawInlineImage(costo.status.perfil.empresa.logo,50,620, 6 * cm, 3 * cm) #Imagen VORDCAB
    logo = ImageReader(empleado.status.perfil.empresa.logo)
    c.drawImage(logo,70,640, 4 * cm, 2 * cm)
    #Primera columna
    c.setFillColor(black)
    c.setFont('Helvetica-Bold',10)
    c.drawString(260,710,'Empleado #:')
    c.drawString(260,690,'Nombre:',)
    c.drawString(260,670,'Empresa:')
    c.drawString(260,650,'Distrito:')
    c.drawString(260,630,'Estado:')
    c.line(20,590,585,590) #Linea posterior horizontal

    c.setFont('Helvetica',10)
    #Parte superior
    numero_trabajador = str(empleado.status.perfil.numero_de_trabajador)
    c.drawString(325,710, numero_trabajador)
    c.drawString(325,690, empleado.status.perfil.nombres)
    c.drawString(390,690, empleado.status.perfil.apellidos)
    empresa = str(empleado.status.perfil.empresa)
    c.drawString(325,670, empresa)
    distrito = str(empleado.status.perfil.distrito)
    c.drawString(325,650, distrito)
    c.setFillColor(color)
    c.drawString(325,630, estado)
    c.setFillColor(black)
    #Tabla y altura guia
    data =[]
    high = 570
    data.append(['''Periodo''','''Fecha''','''Días disfrutados''', '''Días pendientes''','''Creado''',])
    for economico in economicos: #Salen todos los datos
        creado = economico.created_at.date()
        data.append([economico.periodo,economico.fecha,economico.dias_disfrutados,economico.dias_pendientes,creado,])
        high = high - 18

    #Pie de pagina
    c.setFillColor(azul)
    c.setLineWidth(40)
    c.line(20,50,585,50) #Linea posterior horizontal
    c.setFillColor(white)
    #c.drawCentredString(70,53,'Clasificación:')
    #c.drawCentredString(140,53,'Nivel:')
    #c.drawCentredString(240,53,'Preparado por:')
    #c.drawCentredString(350,53,'Aprobado:')
    #c.drawCentredString(450,53,'Fecha emisión:')
    #c.drawCentredString(550,53,'Rev:')
    #Parte de abajo
    #c.drawCentredString(70,39,'Controlado')
    #c.drawCentredString(140,39,'N5')
    #c.drawCentredString(240,39,'SEOV-ALM-N4-01-01')
    #c.drawCentredString(350,39,'SUB ADM')
    #c.drawCentredString(450,39,'24/Oct/2018')
    #c.drawCentredString(550,39,'001')
        #Propiedades de la tabla
    width, height = letter
    table = Table(data, colWidths=[3.2 * cm, 3.2 * cm, 4.2 * cm, 4.5 * cm, 4.5 * cm], repeatRows=1)
    table.setStyle(TableStyle([ #estilos de la tabla
        #ENCABEZADO
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('TEXTCOLOR',(0,0),(-1,0), white),
        ('FONTSIZE',(0,0),(-1,0), 13),
        ('BACKGROUND',(0,0),(-1,0), azul),
        #CUERPO
        ('TEXTCOLOR',(0,1),(-1,-1), colors.black),
        ('FONTSIZE',(0,1),(-1,-1), 10),
        ]))
    table.wrapOn(c, width, height)
    table.drawOn(c, 25, high)
    c.save()
    c.showPage()
    buf.seek(0)

    return FileResponse(buf, as_attachment=True, filename='EconomicoDetalle.pdf')

@login_required(login_url='user-login')
def FormatoVacaciones(request):
    usuario = UserDatos.objects.get(user__id=request.user.id)
    dato = Vacaciones.objects.filter(status__perfil__numero_de_trabajador=usuario.numero_de_trabajador).last()
    datos = Vacaciones.objects.filter(status__perfil__numero_de_trabajador=usuario.numero_de_trabajador)
    context= {
        'usuario':usuario,
        'dato':dato,
        'datos':datos,
        }

    return render(request, 'proyecto/Formato_vacaciones.html',context)

@login_required(login_url='user-login')
def SolicitudVacaciones(request):
    currentFieldCount = 10
    usuario = UserDatos.objects.get(user__id=request.user.id)
    status = Status.objects.get(perfil__numero_de_trabajador=usuario.numero_de_trabajador, perfil__distrito=usuario.distrito)

    #Se reinician las vacaciones para los empleados que ya cumplan otro año de antiguedad con su planta anterior o actual
    fecha_actual = date.today()
    año_actual = str(fecha_actual.year)
    #Busca todos los status que no tengan vacaciones del año actual (periodo)
    status_filtrados = Status.objects.exclude(Q(fecha_planta_anterior__isnull=True, fecha_planta__isnull=True) |Q(vacaciones__periodo=año_actual)) 
    fecha_hace_un_año = fecha_actual - relativedelta(years=1)
    #Filtra todos aquellos con un año o mas de dias con respecto a la fecha actual
    reinicio = status_filtrados.filter(complete=True,perfil__baja=False,fecha_planta_anterior__lte=fecha_hace_un_año)
    #Busco el fecha de planta en los que no tengan fecha de planta anterior
    reinicio2 = status_filtrados.filter(complete=True, perfil__baja=False, fecha_planta_anterior=None, fecha_planta__lte=fecha_hace_un_año,)
    reinicio = reinicio | reinicio2 #Junto los datos de los empleados que ya tienen mas de 1 año de antiguedad
    for empleado in reinicio:
            #Se calculan los días para la vacación actual 
            ahora = datetime.date.today()
            tablas= TablaVacaciones.objects.all()
            if empleado.fecha_planta_anterior:
                days = empleado.fecha_planta_anterior
            else:
                days = empleado.fecha_planta
                
            calcular_antiguedad = relativedelta(ahora, days)
            antiguedad = calcular_antiguedad.years
            
            for tabla in tablas:
                if antiguedad >= tabla.years:
                    dias_vacaciones = tabla.days #Se asignan los días que le tocan en esta vacación
                        
            vacacion = Vacaciones.objects.create(
                complete=True, 
                status=empleado, 
                periodo=año_actual, 
                dias_de_vacaciones=dias_vacaciones,
                fecha_inicio=None,
                fecha_fin=None,
                dias_disfrutados=0,
                dia_inhabil=None,
                total_pendiente=dias_vacaciones,
                comentario="Generado autom. al cumplir otro año de antigüedad",
                editado="Sistema")
            empleado.complete_vacaciones = True #Para confirmar que ya tiene vacacion actual
            empleado.save()

    datos = Vacaciones.objects.filter(status=status).order_by(Cast('periodo', output_field=IntegerField())) #Identifico las vacaciones del usuario de la mas antigua a la mas actual
    
    pendiente=0
    for dato in datos:
        pendiente += dato.total_pendiente #Para sacar el total de días pendientes
    solicitud, created = Solicitud_vacaciones.objects.get_or_create(complete=False)
    form = SolicitudVacacionesForm()
    valido = False

    now = date.today()
    periodo = str(now.year)
    datos= Vacaciones.objects.get(complete=True,status=status,periodo=periodo) #Para sacar el dato actual


    if request.method == 'POST' and 'btnSend' in request.POST:
        form = SolicitudVacacionesForm(request.POST, instance=solicitud)
        form.save(commit=False)

        #Se quita a la cantidad de días de vacaciones el día inhabil y los días festivos para sacar la cuenta de días que tomara
        tabla_festivos = TablaFestivos.objects.all()
        delta = timedelta(days=1)
        day_count = (solicitud.fecha_fin - solicitud.fecha_inicio + delta ).days
        cuenta = day_count
        inhabil = solicitud.dia_inhabil.numero
        for fecha in (solicitud.fecha_inicio + timedelta(n) for n in range(day_count)):
            if fecha.isoweekday() == inhabil:
                cuenta -= 1
            else:
                for dia in tabla_festivos:
                    if fecha == dia.dia_festivo:
                        cuenta -= 1  #Días que va a tomar con esta solicitud

        if cuenta <= pendiente:
            if solicitud.fecha_fin > solicitud.fecha_inicio:
                verificar = Solicitud_vacaciones.objects.filter(status=status,periodo=periodo).last()
                if verificar is None:
                    valido = True
                elif verificar.autorizar is not None:
                    valido = True
                else:
                    messages.error(request, 'Tiene una solicitud generada sin revisar')
            else:
                messages.error(request,'La fecha de inicio no puede ser posterior a la final')
        else:
            messages.error(request,f'Tiene {pendiente} días pendientes, y esta solicitando {cuenta} días de vacaciones')
        if valido and form.is_valid():
            num_campos = int(request.POST.get('num_campos', 0))

            for i in range(1, num_campos + 1):
                asunto = request.POST.get(f'asunto{i}')
                estado = request.POST.get(f'estado{i}')

                # Crear un nuevo grupo de Trabajos_encomendados para cada conjunto de 10 campos
                if (i - 1) % 10 == 0:
                    if i > 1:
                        trabajos_grupo.complete = True
                        trabajos_grupo.save()

                    trabajos_grupo = Trabajos_encomendados()
                    trabajos_grupo.save()
                    solicitud.asunto.add(trabajos_grupo)

                setattr(trabajos_grupo, f'asunto{(i - 1) % 10 + 1}', asunto)
                setattr(trabajos_grupo, f'estado{(i - 1) % 10 + 1}', estado)
                trabajos_grupo.save()

            # Marcar el último grupo como completo
            trabajos_grupo.complete = True
            trabajos_grupo.save()

            solicitud.recibe_nombre = request.POST.get('recibe')
            solicitud.recibe_area = request.POST.get('area')
            solicitud.recibe_puesto = request.POST.get('puesto')
            solicitud.recibe_sector = request.POST.get('sector')
            solicitud.informacion_adicional = request.POST.get('adicional')
            solicitud.anexos = request.POST.get('anexos')
            #trabajos_encomendados, created = Trabajos_encomendados.objects.get_or_create(complete=False,)
            temas, created = Temas_comentario_solicitud_vacaciones.objects.get_or_create(complete=False,)
            for i in range(1, 10):
                comentario = request.POST.get(f'comentario{i}')
                setattr(temas, f'comentario{i}', comentario)

            temas.save()
            #trabajos_encomendados.complete=True
            #trabajos_encomendados.save()
            temas.complete=True
            temas.save()
            messages.success(request, 'Solicitud enviada a RH')
            now = date.today()
            solicitud.periodo = str(now.year)
            solicitud.status = status
            #solicitud.asunto =  trabajos_encomendados
            solicitud.temas = temas
            solicitud.complete=True
            form.save()
            return redirect('index')
    context= {
        'usuario':usuario,
        'form':form,
        'status':status,
        'datos':datos,
        'pendiente':pendiente,
        'currentFieldCount': currentFieldCount,
        }

    return render(request, 'proyecto/Formato_VacacionesForm.html',context)

@login_required(login_url='user-login')
def solicitud_vacacion_verificar(request, pk):
    user_filter = UserDatos.objects.get(user=request.user)
    solicitud = Solicitud_vacaciones.objects.get(id=pk)
    trabajos = Trabajos_encomendados.objects.filter(solicitud_vacaciones__id=solicitud.id)
    #Para mostrar en el html la ennumeracion correcta de los datos
    trabajos_data = []
    for index, trabajo in enumerate(trabajos):
        trabajo_data = []
        start_number = index * 10 + 1  # Calcular el número inicial del conjunto de datos
        for i in range(1, 11):
            asunto = getattr(trabajo, f'asunto{i}', '')
            estado = getattr(trabajo, f'estado{i}', '')
            trabajo_data.append((start_number + i - 1, asunto, estado))  # Agregar el número al conjunto de datos
        trabajos_data.append(trabajo_data)


    temas = Temas_comentario_solicitud_vacaciones.objects.get(solicitud_vacaciones__id=solicitud.id)
    tabla_festivos = TablaFestivos.objects.all()
    delta = timedelta(days=1)
    valido = True

    if request.method == 'POST' and 'btnSend' in request.POST:
        form =SolicitudVacacionesUpdateForm(request.POST, instance=solicitud)
        solicitud = form.save(commit=False)
            #Para las condicionales
        if solicitud.fecha_fin < solicitud.fecha_inicio:
            messages.error(request,'La fecha de inicio no puede ser posterior a la final')
            valido=False

        #Se quita a la cantidad de días de vacaciones el día inhabil y los días festivos para sacar la cuenta de días que tomara
        tabla_festivos = TablaFestivos.objects.all()
        delta = timedelta(days=1)
        day_count = (solicitud.fecha_fin - solicitud.fecha_inicio + delta ).days
        cuenta = day_count
        inhabil = solicitud.dia_inhabil.numero
        for fecha in (solicitud.fecha_inicio + timedelta(n) for n in range(day_count)):
            if fecha.isoweekday() == inhabil:
                cuenta -= 1
            else:
                for dia in tabla_festivos:
                    if fecha == dia.dia_festivo:
                        cuenta -= 1  #Días que va a tomar con esta solicitud
        dias_vacacion = cuenta
        if cuenta < 0:
            messages.error(request, 'La cantidad de días que disfrutara debe ser mayor a 0')
            valido=False

        #Aqui se buscan las vacaciones anteriores y se van modificando los datos para poder llevar la toma de dias pendientes de años anteriores
        ultima_vacacion = Vacaciones.objects.filter(status=solicitud.status.id).last()
        if ultima_vacacion is not None and ultima_vacacion.total_pendiente > 0:
            datos = Vacaciones.objects.filter(status=solicitud.status.id, total_pendiente__gt=0,).order_by(Cast('periodo', output_field=IntegerField()))#Trae todas las vacaciones del mas antiguo al actual 2019-2022
            suma_total = datos.aggregate(total_suma=Sum('total_pendiente'))['total_suma']
            if suma_total < cuenta:
                messages.error(request, f'Esta pidiendo {cuenta} días cuando solo tiene {suma_total}')
                valido=False
            if datos.exclude(id=datos.last().id) != None:
                datos = datos.exclude(id=datos.last().id) #Hasta aqui bien
                for dato in datos: #Se pasa por los datos del mas antiguo al mas actual de los que se tenia
                    if cuenta <= dato.total_pendiente and cuenta > 0:
                        if dato.dias_disfrutados == None:
                            dato.dias_disfrutados = 0
                        dato.total_pendiente -= cuenta
                        dato.dias_disfrutados += cuenta
                        cuenta = 0
                        break
                    elif cuenta > dato.total_pendiente and cuenta > 0:
                        if dato.dias_disfrutados == None:
                            dato.dias_disfrutados = 0
                        dato.dias_disfrutados += dato.total_pendiente
                        cuenta -=dato.total_pendiente
                        dato.total_pendiente = 0
        #else:
        #    datos = Vacaciones.objects.filter(status=solicitud.status.id, total_pendiente__gt=0,).order_by(Cast('periodo', output_field=IntegerField()))#Trae todas las vacaciones del mas antiguo al actual 2019-2022
        ############################
        if valido and form.is_valid():
            solicitud = form.save(commit=False)
            solicitud.comentario_rh= request.POST.get('comentario')
            solicitud.save()
            coment = request.POST.get('comentario')
            if solicitud.autorizar == True:
                vacacion = Vacaciones.objects.get(complete=True, status=solicitud.status, periodo=solicitud.periodo)
                vacacion.dias_disfrutados += cuenta
                vacacion.total_pendiente = vacacion.dias_de_vacaciones - vacacion.dias_disfrutados
                vacacion.dia_inhabil = solicitud.dia_inhabil
                vacacion.fecha_fin = solicitud.fecha_fin
                vacacion.fecha_inicio = solicitud.fecha_inicio
                vacacion.comentario = coment
                # Actualizamos el objeto status
                status = Status.objects.get(id=vacacion.status.id)
                status.complete_vacaciones = True
                #Guardamos las vacaciones anteriores
                for dato in datos:
                    historial = dato.history.first()  # Obtener la primera versión histórica del objeto
                    if historial and historial.total_pendiente != dato.total_pendiente:
                        # El campo 'total_pendiente' ha cambiado
                        dato._meta.get_field('created_at').auto_now = False
                        dato.comentario ="Tomado periodo:" + str(solicitud.periodo)
                        dato.fecha_inicio = solicitud.fecha_inicio
                        dato.fecha_fin =  solicitud.fecha_fin
                        dato.save()
                        dato._meta.get_field('created_at').auto_now = True
                # Guardamos los cambios en la base de datos
                vacacion.comentario +=" "+"Dias tomados:" + str(dias_vacacion)
                nombre = Perfil.objects.get(numero_de_trabajador = user_filter.numero_de_trabajador, distrito = user_filter.distrito)
                vacacion.editado = str("A:"+nombre.nombres+" "+nombre.apellidos)
                vacacion.save()
                status.save()
                messages.success(request, 'Solicitud autorizada y días de vacaciones agregados')
            else:
                messages.success(request, 'Solicitud guardada como no autorizado')

            return redirect('Solicitudes_vacaciones')
    else:
        form = SolicitudVacacionesUpdateForm(instance=solicitud)

    context = {'form':form,'solicitud':solicitud, 'temas':temas, 'trabajos_data': trabajos_data}

    return render(request,'proyecto/solicitud_vacaciones_update.html',context)

def PdfFormatoVacaciones(request, pk):
    solicitud= Solicitud_vacaciones.objects.get(id=pk)
    inicio = solicitud.fecha_inicio
    fin = solicitud.fecha_fin
    dia_inhabil = solicitud.dia_inhabil
    ######
    tabla_festivos = TablaFestivos.objects.all()
    delta = timedelta(days=1)
    day_count = (fin - inicio + delta ).days
    cuenta = day_count
    inhabil = dia_inhabil.numero
    for fecha in (inicio + timedelta(n) for n in range(day_count)):
        if fecha.isoweekday() == inhabil:
            cuenta -= 1
        else:
            for dia in tabla_festivos:
                if fecha == dia.dia_festivo:
                    cuenta -= 1
    diferencia = str(cuenta)
    #Para ubicar el dia de regreso en un dia habil (Puede caer en día festivo)
    fin = fin + timedelta(days=1)
    if fin.isoweekday() == inhabil:
        fin = fin + timedelta(days=1)
    now = date.today()
    año1 = str(inicio.year)
    año2= str(fin.year)
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)

    #Colores utilizados
    azul = Color(0.16015625,0.5,0.72265625)
    rojo = Color(0.59375, 0.05859375, 0.05859375)

    c.setFillColor(azul)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',16)
    c.drawCentredString(305,765,'GRUPO VORCAB SA DE CV')
    c.drawCentredString(305,750,'SOLICITUD DE VACACIONES')
    c.drawInlineImage('static/images/vordcab.png',50,720, 4 * cm, 2 * cm) #Imagen Savia
    if solicitud.autorizar == False:
        c.setFillColor(rojo)
        c.setFont('Helvetica-Bold',16)
        c.drawCentredString(305,725,'SOLICITUD NO AUTORIZADA')
    c.setFillColor(black)
    c.setFont('Helvetica-Bold',11)
    c.drawString(40,690,'NOMBRE:')
    c.line(95,688,325,688)
    espacio = ' '
    nombre_completo = str(solicitud.status.perfil.nombres + espacio + solicitud.status.perfil.apellidos)
    c.drawString(100,690,nombre_completo)
    c.drawString(40,670,'PUESTO:')
    c.line(95,668,325,668)
    c.drawString(100,670,solicitud.status.puesto.puesto)

    c.drawString(335,670,'TELEFONO PARTICULAR:')
    c.line(475,668,580,668)
    c.drawString(485,670,solicitud.status.telefono)

    if cuenta < 3:
        altura=200
        margen=20
        c.drawCentredString(305,502,'OBSERVACIONES')
        if solicitud.comentario_rh:
            c.drawString(55,490,solicitud.comentario_rh)
        else:
            c.drawString(55,490,'ninguna')
        c.rect(50,500, 510, 12)
        c.rect(50,435, 510, 65)
    else:
        altura=0
        margen=0
    c.drawString(40,620-margen,'FECHA DE PLANTA:')
    if solicitud.status.fecha_planta != None:
        dia = str(solicitud.status.fecha_planta.day)
        mes = str(solicitud.status.fecha_planta.month)
        año = str(solicitud.status.fecha_planta.year)
    else:
        dia = "NR"
        mes = "NR"
        año = "NR"

    c.rect(185,598-margen, 150, 55)
    c.line(185,618-margen,335,618-margen)
    c.line(185,638-margen,335,638-margen)
    c.line(230,650-margen,230,598-margen)
    c.line(290,650-margen,290,598-margen)
    c.drawCentredString(210,620-margen,dia)
    c.drawCentredString(260,620-margen,mes)
    c.drawCentredString(310,620-margen,año)
    c.drawString(40,600-margen,'FECHA DE SOLICITUD:')
    c.drawCentredString(210,600-margen,str(now.day))
    c.drawCentredString(260,600-margen,str(now.month))
    c.drawCentredString(310,600-margen,str(now.year))
    c.drawString(200,640-margen,'DIA')
    c.drawString(250,640-margen,'MES')
    c.drawString(300,640-margen,'AÑO')
    c.drawString(400,600-margen,'FIRMA DEL SOLICITANTE')
    c.rect(390,598-margen, 155, 55)
    c.line(390,610-margen,545,610-margen)

    c.drawString(40,560-altura,'PERIODO VACACIONAL CORRESPONDIENTE:')
    c.drawCentredString(425,560-altura, año1)
    c.drawCentredString(450,560-altura, '/')
    c.drawCentredString(475,560-altura, año2)
    c.rect(35,558-altura, 255, 12)
    c.rect(360,558-altura, 190, 12)
    #form = VacacionesFormato(request.POST,)
    c.drawString(40,540-altura,'NO. DE DIAS DE VACACIONES:')
    c.drawCentredString(450,540-altura,diferencia)
    c.rect(35,538-altura, 175, 12)
    c.rect(360,538-altura, 190, 12)
    c.drawString(40,520-altura,'CON GOCE DE SUELDO:')
    c.rect(35,518-altura, 140, 12)
    c.drawString(380,520-altura,'SI')
    c.rect(360,518-altura, 50, 12)
    c.drawString(425,520-altura,'NO')
    c.rect(410,518-altura, 50, 12)
    c.drawString(40,500-altura,'FECHA QUE DESEA SALIR DE VACACIONES:')
    c.drawCentredString(450,500-altura,str(inicio))
    c.rect(35,498-altura, 250, 12)
    c.rect(360,498-altura, 190, 12)
    c.drawString(40,480-altura,'FECHA DE REGRESO A LABORES:')
    c.drawCentredString(450,480-altura,str(fin))
    c.rect(35,478-altura, 195, 12)
    c.rect(360,478-altura, 190, 12)
    if cuenta >= 3: ########### Para formatos largos
        c.drawCentredString(300,440,'Entrega-Recepción')
        c.setFont('Helvetica',11)
        c.drawString(40,400,'DATOS DE QUIEN RECIBE:')
        c.drawString(40,380,'Nombre:')
        if solicitud.recibe_nombre:
            c.drawString(100,380,solicitud.recibe_nombre)
        c.line(90,378,375,378)
        c.drawString(385,380,'Area:')
        if solicitud.recibe_area:
            c.drawString(435,380,solicitud.recibe_area)
        c.line(420,378,560,378)
        c.drawString(40,360,'Puesto:')
        if solicitud.recibe_puesto:
            c.drawString(100,360,solicitud.recibe_puesto)
        c.line(90,358,375,358)
        c.drawString(40,340,'Sector:')
        if solicitud.recibe_sector:
            c.drawString(100,340,solicitud.recibe_sector)
        c.line(90,338,375,338)
        c.setFont('Helvetica-Bold',14)
        c.drawString(40,300,'SITUACIÓN DE TRABAJOS ENCOMENDADOS:')
        c.setFillColor(black)
        c.setFont('Helvetica',11)

        # Estilo de párrafo para los datos en la tabla
        styleSheet = getSampleStyleSheet()
        paragraphStyle = styleSheet['Normal']
        paragraphStyle.fontSize = 8
        #Tabla y altura guia
        high = 130
        trabajos = Trabajos_encomendados.objects.filter(solicitud_vacaciones__id=solicitud.id)
        data = []
        data.append(['No.', 'DENOMINACIÓN ASUNTO', 'ESTADO'])

        numero = 1  # Inicializar el número

        for index, trabajo in enumerate(trabajos, start=1):
            trabajo_data = []
            for i in range(1, 11):
                asunto = getattr(trabajo, f'asunto{i}', '')
                estado = getattr(trabajo, f'estado{i}', '')
                trabajo_data.append((numero, asunto, estado))
                numero += 1  # Incrementar el número
            data.extend(trabajo_data)
        high = high - 20

        #Propiedades de la tabla
        width, height = landscape(letter)
        table_style = TableStyle([ #estilos de la tabla
            #ENCABEZADO
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('TEXTCOLOR',(0,0),(-1,0), colors.black),
            ('FONTSIZE',(0,0),(-1,0), 12),
            ('BACKGROUND',(0,0),(-1,0), colors.white),
            #CUERPO
            ('TEXTCOLOR',(0,1),(-1,-1), colors.black),
            ('FONTSIZE',(0,1),(-1,-1), 12),
            ('GRID',(0,0),(-1,-1),0.5,colors.grey),
        ])

        # Definir variables para el salto de página
        rows_per_page = 7
        total_rows = len(data) - 1  # Excluye el encabezado
        current_row = 1  # Comenzar desde la primera fila (excluyendo el encabezado)

        # Generar páginas con el contenido restante
        while current_row <= total_rows:
            # Calcular el espacio disponible en la página actual
            available_height = height - 70 - 20  # Ajustar según el espaciado
            
            # Calcular cuántas filas caben en la página actual
            if current_row == 1:
                rows_on_page = min(rows_per_page, math.floor((available_height - 20) / 20))  # Para la primera página
            else:
                rows_on_page = min(20, math.floor((available_height - 20) / 20))  # Para las páginas restantes

            # Obtener los datos para la página actual
            end_row = int(current_row + rows_on_page) if current_row + rows_on_page <= total_rows else total_rows + 1
            page_data = data[current_row:end_row]

            # Reemplazar valores None con un espacio en blanco
            page_data = [[cell if cell is not None else " " for cell in row] for row in page_data]

            # Calcular la altura para dibujar la tabla
            table_height = len(page_data) * 20
            
            # Calcular la posición vertical para la tabla
            if current_row == 1:
                # Ajustar el margen superior para la primera tabla
                table_y = height - 130 - table_height - 275  # Usar la altura específica para la primera tabla
            else:
                # Calcular el margen desde la parte superior de la página
                margin_top = 40
                table_y = height - table_height - margin_top

            # Dentro del bucle para crear la tabla de cada página
            table_data = []
            for row in page_data:
                table_row = []
                for cell_data in row:
                    if isinstance(cell_data, str) and len(cell_data) > 30:
                        # Aplicar estilo CSS para dividir palabras largas
                        cell_data = cell_data.replace(' ', '<br/>')
                        cell_data = f'<font size="12">{cell_data}</font>'
                        cell = Paragraph(cell_data, paragraphStyle)
                    else:
                        cell = cell_data
                    table_row.append(cell)
                table_data.append(table_row)

            # Crear la tabla para la página actual
            table = Table([data[0]] + table_data, colWidths=[1.5 * cm, 8 * cm, 10 * cm], repeatRows=1)
            table.setStyle(table_style)

            # Dibujar la tabla en la página actual
            table.wrapOn(c, width, height)
            table.drawOn(c, 25, table_y)

            # Avanzar al siguiente conjunto de filas
            current_row += rows_on_page
            
            # Cambiar la cantidad de filas por página después de la primera página
            if current_row == 1:
                rows_per_page = 20
            
            # Agregar una nueva página si quedan más filas por dibujar
            if current_row <= total_rows:
                c.showPage()
        c.showPage()
        c.setFont('Helvetica-Bold',12)
        #Parrafo con salto de linea automatica si el texto es muy largo
        text = " "
        if solicitud.informacion_adicional:
            text = solicitud.informacion_adicional
        x = 40
        y = 750
        c.setFillColor(black)
        c.setFont('Helvetica', 12)
        c.drawString(x + 5, y - 15, 'INFORMACIÓN ADICIONAL:')
        c.setFont('Helvetica', 9)
        lines = textwrap.wrap(text, width=100)
        for line in lines:
            c.drawString(x + 10, y - 35, line)
            y -= 15
      
        # Estilo de párrafo para los comentarios
        styleSheet = getSampleStyleSheet()
        commentStyle = styleSheet['Normal']
        commentStyle.fontSize = 8

        def format_comment(comment):
            if comment is None:
                return ""
            return Paragraph(comment, commentStyle)
        
        # Datos y ajustes de la tabla
        data2 = []
        high = 465
        data2.append(['No.', 'TEMAS', 'COMENTARIO'])
        data2.append(["1", "Información sobre personal a su cargo", format_comment(solicitud.temas.comentario1)])
        data2.append(["2", "Documentos", format_comment(solicitud.temas.comentario2)])
        data2.append(["3", Paragraph("Arqueo de caja o cuenta bancaria a su cargo (cuando aplique)"), format_comment(solicitud.temas.comentario3)])
        data2.append(["4", "Proyectos pendientes", format_comment(solicitud.temas.comentario4)])
        data2.append(["5", "Estado de las operaciones a su cargo", format_comment(solicitud.temas.comentario5)])
        data2.append(["6", "Deudas con la empresa", format_comment(solicitud.temas.comentario6)])
        data2.append(["7", "Saldos por comprobar a contabilidad", format_comment(solicitud.temas.comentario7)])
        data2.append(["8", "Activos asignados", format_comment(solicitud.temas.comentario8)])
        data2.append(["9", "Otros", format_comment(solicitud.temas.comentario9)])

        table = Table(data2, colWidths=[1.5 * cm, 8 * cm, 11 * cm,], repeatRows=1)
        table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('TEXTCOLOR',(0,0),(-1,0), colors.black),
            ('FONTSIZE',(0,0),(-1,0), 13),
            ('BACKGROUND',(0,0),(-1,0), colors.white),
            ('GRID',(0,0),(-1,-1),0.5,colors.grey),
        ]))

        # Dibujar la tabla en el lienzo
        table.wrapOn(c, width, height)
        table.drawOn(c, 25, high)

        #c.drawString(40,375,'ANEXOS:')
        text = " "
        if solicitud.anexos:
            text = solicitud.anexos
        x = 40
        y = 380
        c.setFillColor(black)
        c.setFont('Helvetica', 12)
        c.drawString(x + 5, y - 15, 'Anexos:')
        c.setFont('Helvetica', 9)
        lines = textwrap.wrap(text, width=100)
        for line in lines:
            c.drawString(x + 10, y - 30, line)
            y -= 25
        c.line(40,345,570,345)
        c.line(40,320,570,320)
        c.line(40,293,570,293)
        c.line(40,270,570,270)
        c.drawCentredString(200,170,'ENTREGUE (NOMBRE Y FIRMA)')
        c.drawCentredString(200,190,nombre_completo)
        c.line(105,185,295,185)
        c.drawCentredString(400,170,'RECIBI (NOMBRE Y FIRMA)')
        c.drawCentredString(400,190,solicitud.recibe_nombre)
        c.line(320,185,480,185)

    c.drawCentredString(200,70,'FIRMA DE GERENCIA')
    c.rect(120,68, 160, 70)
    c.line(120,80,280,80)
    c.drawCentredString(400,70,'FIRMA DE JEFE INMEDIATO')
    c.rect(300,68, 195, 70)
    c.line(300,80,495,80)
    c.save()
    c.showPage()
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename='Formato_Vacaciones.pdf')

@login_required(login_url='user-login')
def FormatoEconomicos(request):
    usuario = UserDatos.objects.get(user__id=request.user.id)
    dato = Economicos.objects.filter(status__perfil__numero_de_trabajador=usuario.numero_de_trabajador).last()
    datos = Economicos.objects.filter(status__perfil__numero_de_trabajador=usuario.numero_de_trabajador)
    if request.method =='POST' and 'Pdf' in request.POST:
        return PdfFormatoEconomicos(usuario)
    context= {
        'usuario':usuario,
        'dato':dato,
        'datos':datos,
        }

    return render(request, 'proyecto/Formato_economicos.html',context)

@login_required(login_url='user-login')
def SolicitudEconomicos(request):
    usuario = UserDatos.objects.get(user__id=request.user.id)
    status = Status.objects.get(perfil__numero_de_trabajador=usuario.numero_de_trabajador, perfil__distrito=usuario.distrito)

    #Se reinician las vacaciones para los empleados que ya cumplan otro año de antiguedad con su planta anterior o actual
    fecha_actual = date.today()
    año_actual = str(fecha_actual.year)
    fecha_hace_un_año = fecha_actual - relativedelta(years=1)
    #Busca todos los status que no tengan vacaciones del año actual (periodo)
    status_filtrados = Status.objects.exclude(Q(fecha_planta_anterior__isnull=True, fecha_planta__isnull=True) |Q(economicos__periodo=año_actual)) 
    fecha_hace_un_año = fecha_actual - relativedelta(years=1)
    #Filtra todos aquellos con un año o mas de dias con respecto a la fecha actual
    reinicio = status_filtrados.filter(complete=True,perfil__baja=False,fecha_planta_anterior__lte=fecha_hace_un_año)
    #Busco el fecha de planta en los que no tengan fecha de planta anterior
    reinicio2 = status_filtrados.filter(complete=True, perfil__baja=False, fecha_planta_anterior=None, fecha_planta__lte=fecha_hace_un_año,)
    reinicio = reinicio | reinicio2 #Junto los datos de los empleados que ya tienen mas de 1 año de antiguedad
    if reinicio:
        for empleado in reinicio:                        
            economicos = Economicos.objects.create(complete=True, status=empleado, periodo=año_actual, dias_disfrutados=0, dias_pendientes=3, fecha=None, 
                                                comentario="Generado autom. al cumplir otro año de antigüedad", editado="Sistema")
            empleado.complete_economicos = True #Para confirmar que ya tiene economico actual
            empleado.save()


    solicitud, created = Solicitud_economicos.objects.get_or_create(complete=False)
    form = SolicitudEconomicosForm()
    now = date.today()
    periodo = str(now.year)
    valido = True
    datos= Economicos.objects.get(complete=True,status=status,periodo=periodo)
    if request.method == 'POST' and 'btnSend' in request.POST:

        form = SolicitudEconomicosForm(request.POST, instance=solicitud)
        form.save(commit=False)


        if datos.dias_disfrutados < 3:
            if Solicitud_economicos.objects.filter(status=status):
                verificar = Solicitud_economicos.objects.filter(status=status,periodo=periodo).last()
                if verificar.autorizar == None:
                    messages.error(request,'Tiene una solicitud generada sin revisar')
                    valido = False
        else:
            messages.error(request,'Usted ya a utilizado sus 3 días económicos')
            valido = False

        fecha_form = request.POST.get('fecha')
        year, month, day = map(int, fecha_form.split('-'))
        fecha_form = date(year, month, day)
        if datos.fecha == fecha_form - timedelta(days=1):
            messages.error(request,'Los días económicos no pueden ser seguidos')
            valido = False
        if valido and form.is_valid():
            messages.success(request, 'Solicitud enviada a RH')
            now = date.today()
            solicitud.periodo = str(now.year)
            solicitud.status = status
            solicitud.complete=True
            form.save()
            return redirect('index')

    context= {
        'usuario':usuario,
        'form':form,
        'datos':datos,
        'status':status,
    }

    return render(request, 'proyecto/Formato_EconomicosForm.html',context)

@login_required(login_url='user-login')
def solicitud_economico_verificar(request, pk):
    user_filter = UserDatos.objects.get(user=request.user)
    solicitud = Solicitud_economicos.objects.get(id=pk)

    if request.method == 'POST' and 'btnSend' in request.POST:
        form = SolicitudEconomicosUpdateForm(request.POST, instance=solicitud)
        solicitud = form.save(commit=False)

        if form.is_valid():
            solicitud = form.save(commit=False)
            solicitud.comentario = request.POST.get('observaciones')
            solicitud.save()

            if solicitud.autorizar == True:
                observaciones = request.POST.get('observaciones')
                # Buscamos o creamos una instancia de Economicos
                economico, created = Economicos.objects.get_or_create(complete=True,status=solicitud.status,periodo=solicitud.periodo)

                if not created:
                    anterior = Economicos.objects.get(complete=True,status=solicitud.status,periodo=solicitud.periodo)
                    # Si no se creó una nueva instancia, editamos los campos necesarios
                    economico.comentario = observaciones
                    economico.dias_disfrutados = anterior.dias_disfrutados + 1
                    economico.dias_pendientes = anterior.dias_pendientes - 1
                    economico.fecha = solicitud.fecha
                else:
                    economico.comentario = observaciones
                    economico.dias_disfrutados = 1
                    economico.dias_pendientes = 2
                    economico.fecha = solicitud.fecha
                # Actualizamos el objeto status
                status = Status.objects.get(id=economico.status.id)
                status.complete_economicos = True
                if economico.dias_pendientes == 0:
                    economico.complete_dias =True
                # Guardamos los cambios en la base de datos
                nombre = Perfil.objects.get(numero_de_trabajador = user_filter.numero_de_trabajador, distrito = user_filter.distrito)
                economico.editado = str("A:"+nombre.nombres+" "+nombre.apellidos)
                economico.save()
                status.save()
                messages.success(request, 'Solicitud autorizada y días economicos agregados')
            else:
                messages.success(request, 'Solicitud guardada como no autorizado')
            return redirect('Solicitudes_economicos')
    else:
        form = SolicitudEconomicosUpdateForm(instance=solicitud)

    context = {'form':form,'solicitud':solicitud}

    return render(request,'proyecto/solicitud_economicos_update.html',context)

def PdfFormatoEconomicos(request, pk):
    solicitud= Solicitud_economicos.objects.get(id=pk)
    now = date.today()
    fecha = solicitud.fecha
    periodo = str(fecha.year)
    economico = 0
    if not Economicos.objects.filter(status=solicitud.status):
        economico = 0
    else:
        last_economico = Economicos.objects.filter(status=solicitud.status).last()
        economico = last_economico.dias_disfrutados
    #Para ubicar el dia de regreso en un dia habil (Puede caer en día festivo)
    #if status.regimen.regimen == 'L-V':
    #    inhabil1 = 6
    #    inhabil2 = 7
    #elif status.regimen.regimen == 'L-S':
    #    inhabil1 = 7
    #    inhabil2 = None
    regreso = fecha + timedelta(days=1)
    #if regreso.isoweekday() == inhabil1:
    #    regreso = regreso + timedelta(days=1)
    #if regreso.isoweekday() == inhabil2:
    #    regreso = regreso + timedelta(days=1)


    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)

    #Colores utilizados
    azul = Color(0.16015625,0.5,0.72265625)
    rojo = Color(0.59375, 0.05859375, 0.05859375)

    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',16)
    c.drawCentredString(305,765,'GRUPO VORCAB SA DE CV')
    c.setFont('Helvetica-Bold',11)
    c.drawCentredString(305,750,'SOLICITUD DE DIA ECONOMICO')
    if solicitud.autorizar == False:
        c.setFillColor(rojo)
        c.setFont('Helvetica-Bold',16)
        c.drawCentredString(305,725,'SOLICITUD NO AUTORIZADA')
    c.setFillColor(black)
    c.setFont('Helvetica-Bold',11)
    c.drawString(40,690,'NOMBRE:')
    c.line(95,688,325,688)
    espacio = ' '
    nombre_completo = str(solicitud.status.perfil.nombres + espacio + solicitud.status.perfil.apellidos)
    c.drawString(100,690,nombre_completo)
    c.drawString(40,670,'PUESTO:')
    c.line(95,668,325,668)
    c.drawString(100,670,solicitud.status.puesto.puesto)
    c.drawString(335,670,'TELEFONO PARTICULAR:')
    c.line(475,668,580,668)
    c.drawString(485,670,solicitud.status.telefono)
    c.drawString(40,620,'FECHA DE PLANTA:')
    if solicitud.status.fecha_planta != None:
        dia = str(solicitud.status.fecha_planta.day)
        mes = str(solicitud.status.fecha_planta.month)
        año = str(solicitud.status.fecha_planta.year)
    else:
        dia = "NR"
        mes = "NR"
        año = "NR"
    #rect(x, y, alto, ancho, stroke=1, fill=0)
    c.rect(185,600, 150, 50)
    c.line(185,618,335,618)
    c.line(185,638,335,638)
    c.line(230,650,230,600)
    c.line(290,650,290,600)
    c.drawCentredString(210,620,dia)
    c.drawCentredString(260,620,mes)
    c.drawCentredString(310,620,año)
    c.drawString(40,600,'FECHA DE SOLICITUD:')
    c.drawCentredString(210,605,str(now.day))
    c.drawCentredString(260,605,str(now.month))
    c.drawCentredString(310,605,str(now.year))
    c.drawString(200,640,'DIA')
    c.drawString(250,640,'MES')
    c.drawString(300,640,'AÑO')
    c.drawString(400,600,'FIRMA DEL SOLICITANTE')
    c.rect(390,598, 155, 50)
    c.line(390,610,545,610)
    c.drawString(40,540,'PERIODO CORRESPONDIENTE:')
    c.drawCentredString(450,540, periodo)
    c.rect(35,538, 255, 12)
    c.rect(360,538, 190, 12)
    c.drawCentredString(385,520,'1')
    c.drawCentredString(435,520,'2')
    c.drawCentredString(485,520,'3')

    c.drawString(40,500,'NO. DE DIA ECONOMICO:')
    c.rect(35,498, 175, 12)
    c.rect(360,498, 150, 12)
    c.line(410,510,410,498)
    c.line(460,510,460,498)
    c.setFillColorRGB(0.8, 0.8, 0.8)  # Color de relleno
    if economico == 1:
        c.rect(360,498, 50, 12, stroke = 1, fill = 1)
    elif economico == 2:
        c.rect(410,498, 50, 12, stroke = 1, fill = 1)
    elif economico == 3:
        c.rect(460,498, 50, 12, stroke = 1, fill = 1)
    c.setFillColor(black)
    c.drawString(40,480,'CON GOCE DE SUELDO:')
    c.rect(35,478, 140, 12)
    c.drawString(380,480,'SI')
    c.rect(360,478, 50, 12)
    c.drawString(425,480,'NO')
    c.rect(410,478, 50, 12)
    c.drawString(40,460,'FECHA QUE DESEA SALIR DEL PERMISO:')
    c.drawCentredString(450,460,str(fecha))
    c.rect(35,458, 250, 12)
    c.rect(360,458, 190, 12)
    c.drawString(40,440,'FECHA DE REGRESO A LABORES:')
    c.drawCentredString(450,440,str(regreso))
    c.rect(35,438, 195, 12)
    c.rect(360,438, 190, 12)
    #c.drawCentredString(305,370,'OBSERVACIONES')
    text = solicitud.comentario
    x = 40
    y = 385
    c.setFillColor(black)
    c.setFont('Helvetica-Bold',12)
    c.drawCentredString(310, y - 15, 'OBSERVACIONES')
    c.setFont('Helvetica', 9)
    lines = textwrap.wrap(text, width=100)
    for line in lines:
        c.drawString(x + 10, y - 30, line)
        y -= 25
    c.rect(40,368, 530, 12)
    c.rect(40,300, 530, 68)
    c.drawCentredString(170,125,'FIRMA GERENCIA')
    c.rect(70,123, 200, 12)
    c.rect(70,135, 200, 50)
    c.drawCentredString(440,125,'FIRMA DE JEFE INMEDIATO')
    c.rect(330,123, 210, 12)
    c.rect(330,135, 210, 50)
    c.save()
    c.showPage()
    buf.seek(0)

    return FileResponse(buf, as_attachment=True, filename='Formato_Economico.pdf')

    #Reportes generales
def excel_reporte_general(perfil,status,bancarios,costo,bonos,vacaciones,economicos,):

    
    fecha_actual = date.today()
    año_actual = str(fecha_actual.year)
    fecha_hace_un_año = fecha_actual - relativedelta(years=1)
    status= Status.objects.filter(complete=True, perfil__baja=False)
    reinicio = status.filter(Q(fecha_planta_anterior__lte=fecha_hace_un_año) |Q(fecha_planta__lte=fecha_hace_un_año))
    reinicio = reinicio.count()

    matriz= perfil.filter(distrito__distrito = 'Matriz')
    matriz = matriz.count()
    altamira= perfil.filter(distrito__distrito = 'Altamira')
    altamira = altamira.count()
    planta= perfil.filter(distrito__distrito = 'Planta Veracruz')
    planta = planta.count()
    poza = perfil.filter(distrito__distrito = 'Poza Rica')
    poza = poza.count()
    villa = perfil.filter(distrito__distrito = 'Villahermosa')
    villa = villa.count()
    veracruz= perfil.filter(distrito__distrito = 'Veracruz')
    veracruz = veracruz.count()
    hombres = status.filter(sexo__sexo = 'Masculino')
    hombres = hombres.count()
    mujeres = status.filter(sexo__sexo = 'Femenino')
    mujeres = mujeres.count()
    perfil = perfil.count()
    status = status.count()
    bancarios = bancarios.count()
    costo = costo.count()
    bonos = bonos.count()
    vacaciones = vacaciones.count()
    economicos = economicos.count()
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Reporte_general_' + str(datetime.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Reporte')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)


    (ws.cell(column = 10, row = 1, value='{Reporte Creado Automáticamente por Savia RH. UH}')).style = messages_style
    (ws.cell(column = 10, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}')).style = messages_style
    my_png = openpyxl.drawing.image.Image('static/images/logo/SAVIA_Logo1.png')
    ws.add_image(my_png, 'K4')
    #(ws.cell(column = 1, row = 1, value='REPORTE GENERAL SAVIA RH')).style = head_style
    (ws.cell(column = 1, row = 3, value='Empleados:')).style = messages_style
    (ws.cell(column = 2, row = 3, value=perfil)).style = body_style
    (ws.cell(column = 1, row = 4, value='Status:')).style = messages_style
    (ws.cell(column = 2, row = 4, value=status)).style = body_style
    (ws.cell(column = 1, row = 5, value='Bancarios:')).style = messages_style
    (ws.cell(column = 2, row = 5, value=bancarios)).style = body_style
    (ws.cell(column = 1, row = 6, value='Costo:')).style = messages_style
    (ws.cell(column = 2, row = 6, value=costo)).style = body_style
    (ws.cell(column = 1, row = 7, value='Bonos:')).style = messages_style
    (ws.cell(column = 2, row = 7, value=bonos)).style = body_style
    (ws.cell(column = 1, row = 8, value='Vacaciones:')).style = messages_style
    (ws.cell(column = 2, row = 8, value=vacaciones)).style = body_style
    (ws.cell(column = 1, row = 9, value='Economicos:')).style = messages_style
    (ws.cell(column = 2, row = 9, value=economicos)).style = body_style
    (ws.cell(column = 1, row = 10, value='Hombres:')).style = messages_style
    (ws.cell(column = 2, row = 10, value=hombres)).style = body_style
    (ws.cell(column = 1, row = 11, value='Mujeres:')).style = messages_style
    (ws.cell(column = 2, row = 11, value=mujeres)).style = body_style
    (ws.cell(column = 1, row = 12, value='Empleados con 1 año antigüedad:')).style = messages_style
    (ws.cell(column = 2, row = 12, value=reinicio)).style = body_style

    (ws.cell(column = 1, row = 15, value='Matriz')).style = messages_style
    (ws.cell(column = 2, row = 15, value=matriz)).style = body_style
    (ws.cell(column = 1, row = 16, value='Altamira')).style = messages_style
    (ws.cell(column = 2, row = 16, value=altamira)).style = body_style
    (ws.cell(column = 1, row = 17, value='Planta Veracruz')).style = messages_style
    (ws.cell(column = 2, row = 17, value=planta)).style = body_style
    (ws.cell(column = 1, row = 18, value='Poza Rica')).style = messages_style
    (ws.cell(column = 2, row = 18, value=poza)).style = body_style
    (ws.cell(column = 1, row = 19, value='Villa Hermosa')).style = messages_style
    (ws.cell(column = 2, row = 19, value=villa)).style = body_style
    (ws.cell(column = 1, row = 20, value='Veracruz')).style = messages_style
    (ws.cell(column = 2, row = 20, value=veracruz)).style = body_style
    pie_chart = PieChart()
    labels = Reference(ws, min_col=1, min_row=15, max_row=20)
    data = Reference(ws, min_col=2, min_row=15, max_row=20)
    pie_chart.add_data(data, titles_from_data=True)
    pie_chart.set_categories(labels)
    # Agrega porcentajes a cada sector de la gráfica
    pie_chart.dataLabels = DataLabelList(showPercent=True, separator=' ')
    # Cambia el título de la gráfica
    pie_chart.title = 'Empleados por distrito'
    # Cambia el tamaño del gráfico
    pie_chart.width = 8
    pie_chart.height = 8
    ws.add_chart(pie_chart, "H12")
    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)

def reporte_pdf_general(perfil,status,bancarios,costo,bonos,vacaciones,economicos,):
    #Configuration of the PDF object
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)

    now = datetime.date.today()
    fecha = str(now)
    hombres = status.filter(sexo__sexo = 'Masculino')
    hombres = str(hombres.count())
    mujeres = status.filter(sexo__sexo = 'Femenino')
    mujeres = str(mujeres.count())
    perfil = str(perfil.count())
    status = str(status.count())
    bancarios = str(bancarios.count())
    costo = str(costo.count())
    bonos = str(bonos.count())
    vacaciones = str(vacaciones.count())
    economicos = str(economicos.count())
    #Colores utilizados
    azul = Color(0.16015625,0.5,0.72265625)
    rojo = Color(0.59375, 0.05859375, 0.05859375)
    #Encabezado
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',10)
    c.drawString(440,735,'Fecha:')
    c.drawString(480,735,fecha)

    c.setFillColor(azul)
    c.setStrokeColor(azul)
    c.setLineWidth(20)
    c.line(20,760,585,760) #Linea azul superior
    c.setLineWidth(0.2)
    c.line(20,727.5,585,727.5) #Linea posterior horizontal
    c.line(250,727.5,250,590) #Linea vertical

    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica',10)
    c.drawCentredString(295,755,'REPORTE GENERAL SAVIA RH')

    c.drawInlineImage('static/images/logo/SAVIA_Logo1.png',65,580, 5 * cm, 5 * cm) #Imagen Savia
    #Primera columna
    c.setFillColor(black)
    c.setFont('Helvetica',10)
    c.drawString(260,710,'Empleados:')
    c.drawString(335,710,perfil)
    c.drawString(260,690,'Status:',)
    c.drawString(335,690,status)
    c.drawString(260,670,'Bancarios:')
    c.drawString(335,670,bancarios)
    c.drawString(260,650,'Costos:')
    c.drawString(335,650,costo)
    c.drawString(260,630,'Bonos:')
    c.drawString(335,630,bonos)
    c.drawString(260,610,'Vacaciones:')
    c.drawString(335,610,vacaciones)
    c.drawString(260,590,'Economicos:')
    c.drawString(335,590,economicos)

    #Segunda columna
    c.drawString(420,710,'Hombres:')
    c.drawString(495,710, hombres)
    c.drawString(420,690, 'Mujeres:')
    c.drawString(495,690, mujeres)
    #c.drawString(420,630, 'Fecha Emisión:')
    #c.drawString(420,610,'28-06-2022 11:16:21')
    c.setFillColor(rojo) ## NUMERO DEL FOLIO

    #Tabla y altura guia
    #data =[]
    high = 550
    #data.append(['''Orden #''','''Producto''','''Cantidad''', '''Talla''',])
    #for uniforme in uniformes: #Salen todos los datos
    #    data.append([uniforme.id,uniforme.ropa,uniforme.cantidad,uniforme.talla,])
    #    high = high - 18

    #Observaciones
    #c.setFillColor(azul)
    #c.setLineWidth(20)
    #c.line(20,high-35,585,high-35) #Linea posterior horizontal
    #c.setFillColor(white)
    #c.setLineWidth(.1)
    #c.setFont('Helvetica-Bold',10)
    #c.drawCentredString(295,high-40,'Observaciones')
    #c.setFillColor(black)
    #c.setFont('Helvetica',8)
    #c.drawCentredString(295,high-60,'                                                                                                                ')
    #c.drawCentredString(295,high-70,'                                                                                                                ')

    #Autorizacion parte
    #c.setFillColor(azul)
    #c.setFont('Helvetica',8)
    #c.setLineWidth(1)
    #c.line(150,high-150,275,high-150) #Linea posterior horizontal
    #c.line(350,high-150,475,high-150) #Linea posterior horizontal
    #c.setFillColor(black)
    #c.drawCentredString(212.5,high-160,'Empleado')
    #c.drawCentredString(412.5,high-160,'Aprobación')


    #c.drawCentredString(412.5,high-145,'Nombre aprobador')

    #Pie de pagina
    c.setFillColor(azul)
    c.setLineWidth(40)
    c.line(20,50,585,50) #Linea posterior horizontal
    c.setFillColor(white)
    #c.drawCentredString(70,53,'Clasificación:')
    #c.drawCentredString(140,53,'Nivel:')
    #c.drawCentredString(240,53,'Preparado por:')
    #c.drawCentredString(350,53,'Aprobado:')
    #c.drawCentredString(450,53,'Fecha emisión:')
    #c.drawCentredString(550,53,'Rev:')
    #Parte de abajo
    #c.drawCentredString(70,39,'Controlado')
    #c.drawCentredString(140,39,'N5')
    #c.drawCentredString(240,39,'SEOV-ALM-N4-01-01')
    #c.drawCentredString(350,39,'SUB ADM')
    #c.drawCentredString(450,39,'24/Oct/2018')
    #c.drawCentredString(550,39,'001')

    #Propiedades de la tabla
    #width, height = letter
    #table = Table(data, colWidths=[2.6 * cm, 2.6 * cm, 11.8 * cm, 2.6 * cm], repeatRows=1)
    #table.setStyle(TableStyle([ #estilos de la tabla
        #ENCABEZADO
    #    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
    #    ('TEXTCOLOR',(0,0),(-1,0), white),
    #    ('FONTSIZE',(0,0),(-1,0), 13),
    #    ('BACKGROUND',(0,0),(-1,0), azul),
        #CUERPO
    #    ('TEXTCOLOR',(0,1),(-1,-1), colors.black),
    #    ('FONTSIZE',(0,1),(-1,-1), 10),
    #    ]))
    #table.wrapOn(c, width, height)
    #table.drawOn(c, 25, high)
    c.save()
    c.showPage()
    buf.seek(0)

    return FileResponse(buf, as_attachment=True, filename='Reporte_general.pdf')

def excel_reporte_especifico(distrito_seleccionado,perfill,statuss,bancarioss,costoo,bonoss,vacacioness,economicoss,):
    hombres = statuss.filter(sexo__sexo = 'Masculino')
    hombres = hombres.count()
    mujeres = statuss.filter(sexo__sexo = 'Femenino')
    mujeres = mujeres.count()
    perfil = perfill.count()
    status = statuss.count()
    bancarios = bancarioss.count()
    costo = costoo.count()
    bonos = bonoss.count()
    vacaciones = vacacioness.count()
    economicos = economicoss.count()
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Reporte_general_' + distrito_seleccionado+'_'+ str(datetime.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Reporte')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)


    (ws.cell(column = 10, row = 1, value='{Reporte Creado Automáticamente por Savia RH. UH}')).style = messages_style
    (ws.cell(column = 10, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}')).style = messages_style
    my_png = openpyxl.drawing.image.Image('static/images/logo/SAVIA_Logo1.png')
    ws.add_image(my_png, 'K4')
    #(ws.cell(column = 1, row = 1, value='REPORTE GENERAL SAVIA RH')).style = head_style
    (ws.cell(column = 1, row = 3, value='Empleados:')).style = messages_style
    (ws.cell(column = 2, row = 3, value=perfil)).style = body_style
    (ws.cell(column = 1, row = 4, value='Status:')).style = messages_style
    (ws.cell(column = 2, row = 4, value=status)).style = body_style
    (ws.cell(column = 1, row = 5, value='Bancarios:')).style = messages_style
    (ws.cell(column = 2, row = 5, value=bancarios)).style = body_style
    (ws.cell(column = 1, row = 6, value='Costo:')).style = messages_style
    (ws.cell(column = 2, row = 6, value=costo)).style = body_style
    (ws.cell(column = 1, row = 7, value='Bonos:')).style = messages_style
    (ws.cell(column = 2, row = 7, value=bonos)).style = body_style
    (ws.cell(column = 1, row = 8, value='Vacaciones:')).style = messages_style
    (ws.cell(column = 2, row = 8, value=vacaciones)).style = body_style
    (ws.cell(column = 1, row = 9, value='Economicos:')).style = messages_style
    (ws.cell(column = 2, row = 9, value=economicos)).style = body_style
    (ws.cell(column = 1, row = 10, value='Hombres:')).style = messages_style
    (ws.cell(column = 2, row = 10, value=hombres)).style = body_style
    (ws.cell(column = 1, row = 11, value='Mujeres:')).style = messages_style
    (ws.cell(column = 2, row = 11, value=mujeres)).style = body_style

    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)
"""
def excel_reporte_puestos():
    puesto = Puesto.objects.all()
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Puestos'+'_'+ str(datetime.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Reporte')
    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)

    # Encabezado de las columnas
    columns = ['Puesto', 'Complete']
    ws.append(columns)
    row_num = 1
    for row in puesto:
        row_num += 1
        ws.cell(row=row_num, column=1, value=row.puesto).style = body_style
        ws.cell(row=row_num, column=2, value=row.complete).style = body_style
        
    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)
"""
def reporte_pdf_especifico(distrito_seleccionado,perfill,statuss,bancarioss,costoo,bonoss,vacacioness,economicoss,):
    #Configuration of the PDF object
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)

    now = datetime.date.today()
    fecha = str(now)
    hombres = statuss.filter(sexo__sexo = 'Masculino')
    hombres = str(hombres.count())
    mujeres = statuss.filter(sexo__sexo = 'Femenino')
    mujeres = str(mujeres.count())
    perfil = str(perfill.count())
    status = str(statuss.count())
    bancarios = str(bancarioss.count())
    costo = str(costoo.count())
    bonos = str(bonoss.count())
    vacaciones = str(vacacioness.count())
    economicos = str(economicoss.count())
    #Colores utilizados
    azul = Color(0.16015625,0.5,0.72265625)
    rojo = Color(0.59375, 0.05859375, 0.05859375)
    #Encabezado
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',10)
    c.drawString(440,735,'Fecha:')
    c.drawString(480,735,fecha)

    c.setFillColor(azul)
    c.setStrokeColor(azul)
    c.setLineWidth(20)
    c.line(20,760,585,760) #Linea azul superior
    c.setLineWidth(0.2)
    c.line(20,727.5,585,727.5) #Linea posterior horizontal
    c.line(250,727.5,250,590) #Linea vertical

    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica',10)
    c.drawCentredString(295,755,'Reporte '+distrito_seleccionado+' SAVIA RH')

    c.drawInlineImage('static/images/logo/SAVIA_Logo1.png',65,580, 5 * cm, 5 * cm) #Imagen Savia
    #Primera columna
    c.setFillColor(black)
    c.setFont('Helvetica',10)
    c.drawString(260,710,'Empleados:')
    c.drawString(335,710,perfil)
    c.drawString(260,690,'Status:',)
    c.drawString(335,690,status)
    c.drawString(260,670,'Bancarios:')
    c.drawString(335,670,bancarios)
    c.drawString(260,650,'Costos:')
    c.drawString(335,650,costo)
    c.drawString(260,630,'Bonos:')
    c.drawString(335,630,bonos)
    c.drawString(260,610,'Vacaciones:')
    c.drawString(335,610,vacaciones)
    c.drawString(260,590,'Economicos:')
    c.drawString(335,590,economicos)

    #Segunda columna
    c.drawString(420,710,'Hombres:')
    c.drawString(495,710, hombres)
    c.drawString(420,690, 'Mujeres:')
    c.drawString(495,690, mujeres)
    #c.drawString(420,630, 'Fecha Emisión:')
    #c.drawString(420,610,'28-06-2022 11:16:21')
    c.setFillColor(rojo) ## NUMERO DEL FOLIO

    #Tabla y altura guia
    #data =[]
    high = 550
    #data.append(['''Orden #''','''Producto''','''Cantidad''', '''Talla''',])
    #for uniforme in uniformes: #Salen todos los datos
    #    data.append([uniforme.id,uniforme.ropa,uniforme.cantidad,uniforme.talla,])
    #    high = high - 18

    #Observaciones
    #c.setFillColor(azul)
    #c.setLineWidth(20)
    #c.line(20,high-35,585,high-35) #Linea posterior horizontal
    #c.setFillColor(white)
    #c.setLineWidth(.1)
    #c.setFont('Helvetica-Bold',10)
    #c.drawCentredString(295,high-40,'Observaciones')
    #c.setFillColor(black)
    #c.setFont('Helvetica',8)
    #c.drawCentredString(295,high-60,'                                                                                                                ')
    #c.drawCentredString(295,high-70,'                                                                                                                ')

    #Autorizacion parte
    #c.setFillColor(azul)
    #c.setFont('Helvetica',8)
    #c.setLineWidth(1)
    #c.line(150,high-150,275,high-150) #Linea posterior horizontal
    #c.line(350,high-150,475,high-150) #Linea posterior horizontal
    #c.setFillColor(black)
    #c.drawCentredString(212.5,high-160,'Empleado')
    #c.drawCentredString(412.5,high-160,'Aprobación')


    #c.drawCentredString(412.5,high-145,'Nombre aprobador')

    #Pie de pagina
    c.setFillColor(azul)
    c.setLineWidth(40)
    c.line(20,50,585,50) #Linea posterior horizontal
    c.setFillColor(white)
    #c.drawCentredString(70,53,'Clasificación:')
    #c.drawCentredString(140,53,'Nivel:')
    #c.drawCentredString(240,53,'Preparado por:')
    #c.drawCentredString(350,53,'Aprobado:')
    #c.drawCentredString(450,53,'Fecha emisión:')
    #c.drawCentredString(550,53,'Rev:')
    #Parte de abajo
    #c.drawCentredString(70,39,'Controlado')
    #c.drawCentredString(140,39,'N5')
    #c.drawCentredString(240,39,'SEOV-ALM-N4-01-01')
    #c.drawCentredString(350,39,'SUB ADM')
    #c.drawCentredString(450,39,'24/Oct/2018')
    #c.drawCentredString(550,39,'001')

    #Propiedades de la tabla
    #width, height = letter
    #table = Table(data, colWidths=[2.6 * cm, 2.6 * cm, 11.8 * cm, 2.6 * cm], repeatRows=1)
    #table.setStyle(TableStyle([ #estilos de la tabla
        #ENCABEZADO
    #    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
    #    ('TEXTCOLOR',(0,0),(-1,0), white),
    #    ('FONTSIZE',(0,0),(-1,0), 13),
    #    ('BACKGROUND',(0,0),(-1,0), azul),
        #CUERPO
    #    ('TEXTCOLOR',(0,1),(-1,-1), colors.black),
    #    ('FONTSIZE',(0,1),(-1,-1), 10),
    #    ]))
    #table.wrapOn(c, width, height)
    #table.drawOn(c, 25, high)
    c.save()
    c.showPage()
    buf.seek(0)

    return FileResponse(buf, as_attachment=True, filename='Reporte_'+distrito_seleccionado+'.pdf')

def Tabla_solicitud_vacaciones(request):
    user_filter = UserDatos.objects.get(user=request.user)

    if user_filter.distrito.distrito == 'Matriz':
        perfiles= Perfil.objects.filter(complete=True, baja=False).order_by("numero_de_trabajador")
    else:
        perfiles= Perfil.objects.filter(distrito=user_filter.distrito,complete=True, baja=False).order_by("numero_de_trabajador")

    solicitudes = Solicitud_vacaciones.objects.filter(status__perfil__in=perfiles, complete=True, autorizar=None)
    solicitudes_revisadas = Solicitud_vacaciones.objects.filter(status__perfil__in=perfiles, complete=True).exclude(Q(autorizar=None)).order_by("-id")

    solicitud_filter = SolicitudesVacacionesFilter(request.GET, queryset=solicitudes)
    solicitudes = solicitud_filter.qs
    solicitud2_filter = SolicitudesVacacionesFilter(request.GET, queryset=solicitudes_revisadas)
    solicitudes_revisadas = solicitud2_filter.qs

    context= {
        'perfiles':perfiles,
        'solicitud_filter':solicitud_filter,
        'solicitudes':solicitudes,
        'solicitudes_revisadas':solicitudes_revisadas,
        }

    return render(request, 'proyecto/Solicitudes_vacaciones_tabla.html',context)

def Tabla_solicitud_economicos(request):
    user_filter = UserDatos.objects.get(user=request.user)

    if user_filter.distrito.distrito == 'Matriz':
        perfiles= Perfil.objects.filter(complete=True,baja=False).order_by("numero_de_trabajador")
    else:
        perfiles= Perfil.objects.filter(distrito=user_filter.distrito,complete=True,baja=False).order_by("numero_de_trabajador")

    solicitudes = Solicitud_economicos.objects.filter(status__perfil__in=perfiles, complete=True, autorizar=None)
    solicitudes_revisadas = Solicitud_economicos.objects.filter(status__perfil__in=perfiles, complete=True).exclude(Q(autorizar=None)).order_by("-id")

    solicitud_filter = SolicitudesEconomicosFilter(request.GET, queryset=solicitudes)
    solicitudes = solicitud_filter.qs
    solicitud2_filter = SolicitudesEconomicosFilter(request.GET, queryset=solicitudes_revisadas)
    solicitudes_revisadas = solicitud2_filter.qs

    context= {
        'perfiles':perfiles,
        'solicitud_filter':solicitud_filter,
        'solicitudes':solicitudes,
        'solicitudes_revisadas':solicitudes_revisadas,
        }

    return render(request, 'proyecto/Solicitudes_economicos_tabla.html',context)

@login_required(login_url='user-login')
def upload_batch_vacaciones_anteriores(request):

    form = Vacaciones_anteriores_BatchForm(request.POST or None, request.FILES or None)

    if form.is_valid():
        form.save()
        form = Vacaciones_anteriores_BatchForm()
        empleados_list = Vacaciones_anteriores_Batch.objects.get(activated = False)
        f = open(empleados_list.file_name.path, 'r', encoding='latin1')
        reader = csv.reader(f)
        next(reader) # Advance past the reader
        fecha_str = '2019-12-01 01:00:00'
        fecha1 = datetime.datetime.strptime(fecha_str, '%Y-%m-%d %H:%M:%S')
        fecha_str = '2020-12-01 01:00:00'
        fecha2 = datetime.datetime.strptime(fecha_str, '%Y-%m-%d %H:%M:%S')
        fecha_str = '2021-12-01 01:00:00'
        fecha3 = datetime.datetime.strptime(fecha_str, '%Y-%m-%d %H:%M:%S')
        fecha_str = '2022-12-01 01:00:00'
        fecha4 = datetime.datetime.strptime(fecha_str, '%Y-%m-%d %H:%M:%S')
        for row in reader:
            if Perfil.objects.filter(numero_de_trabajador=row[0], distrito__distrito=row[1]):
                status = Status.objects.get(perfil__numero_de_trabajador=row[0], perfil__distrito__distrito=row[1])
                dia = Dia_vacacion.objects.get(nombre='Domingo')
                if row[2] != 0:
                    vacacion = Vacaciones(status=status, periodo='2019', dias_de_vacaciones=None, dia_inhabil=dia,
                                          fecha_inicio=None, fecha_fin=None, dias_disfrutados=None,
                                          total_pendiente=row[2], comentario=str('Vacaciones años previos'),
                                          complete=True)

                    vacacion._meta.get_field('created_at').auto_now = False
                    vacacion.created_at = fecha1
                    vacacion.save()
                    vacacion._meta.get_field('created_at').auto_now = True

                if row[3] != 0:
                    vacacion1 = Vacaciones(status=status, periodo='2020', dias_de_vacaciones=None, dia_inhabil=dia,
                                           fecha_inicio=None, fecha_fin=None, dias_disfrutados=None,
                                           total_pendiente=row[3], comentario=str('Vacaciones años previos'),
                                           complete=True)

                    vacacion1._meta.get_field('created_at').auto_now = False
                    vacacion1.created_at = fecha2
                    vacacion1.save()
                    vacacion1._meta.get_field('created_at').auto_now = True

                if row[4] != 0:
                    vacacion2 = Vacaciones(status=status, periodo='2021', dias_de_vacaciones=None, dia_inhabil=dia,
                                           fecha_inicio=None, fecha_fin=None, dias_disfrutados=None,
                                           total_pendiente=row[4], comentario=str('Vacaciones años previos'),
                                           complete=True)

                    vacacion2._meta.get_field('created_at').auto_now = False
                    vacacion2.created_at = fecha3
                    vacacion2.save()
                    vacacion2._meta.get_field('created_at').auto_now = True

                if row[5] != 0:
                    vacacion3 = Vacaciones(status=status, periodo='2022', dias_de_vacaciones=None, dia_inhabil=dia,
                                           fecha_inicio=None, fecha_fin=None, dias_disfrutados=None,
                                           total_pendiente=row[5], comentario=str('Vacaciones años previos'),
                                           complete=True)

                    vacacion3._meta.get_field('created_at').auto_now = False
                    vacacion3.created_at = fecha4
                    vacacion3.save()
                    vacacion3._meta.get_field('created_at').auto_now = True

                if row[6] != 0:
                    economico = Economicos(status=status, periodo='2023', dias_disfrutados=(3 - int(row[6])),
                                           dias_pendientes=row[6], fecha=None,
                                           comentario="Economicos pendientes cargados", complete=True)

                    if economico.dias_disfrutados == 3:
                        economico.complete_dias = True

                    economico.save()
                    status.complete_economicos = True
                    status.save()
            else:
                messages.error(request, f'No se ha encontrado el empleado, empleado #{row[0]}')

        empleados_list.activated = True
        empleados_list.save()

    context = {
        'form': form,
    }

    return render(request, 'proyecto/upload_batch_vacaciones_anteriores.html', context)

@login_required(login_url='user-login')
def Baja_empleado(request, pk):
    user_filter = UserDatos.objects.get(user=request.user)
    empleado = Perfil.objects.get(id=pk)
    form = BajaEmpleadoForm()
    if request.method == 'POST':
        form = BajaEmpleadoForm(request.POST)
        if form.is_valid():
            form.instance.perfil = empleado 
            form.instance.complete = True
            form.save()
            empleado.baja = True
            nombre = Perfil.objects.get(numero_de_trabajador = user_filter.numero_de_trabajador, distrito = user_filter.distrito)
            empleado.editado = str("B:"+nombre.nombres+" "+nombre.apellidos)
            empleado.save()
            messages.success(request, f'El empleado {empleado.nombres} {empleado.apellidos} se ha dado de baja en el Sistema')
            return redirect('Perfil')
            
        else:
            messages.error(request, 'Error al llenar el formulario')


    context = {
        'empleado': empleado,
        'form':form,
    }

    return render(request, 'proyecto/Baja_empleado.html', context)

@login_required(login_url='user-login')
def Baja_update(request, pk):
    user_filter = UserDatos.objects.get(user=request.user)

    item = Datos_baja.objects.get(id=pk)
    form = BajaEmpleadoForm(instance=item)  # Aquí inicializamos el formulario con los datos del item

    if request.method == 'POST':
        form = BajaEmpleadoForm(request.POST, instance=item)  # Aquí también debemos usar 'instance'
        if form.is_valid():
            nombre = Perfil.objects.get(numero_de_trabajador=user_filter.numero_de_trabajador, distrito=user_filter.distrito)
            item.editado = str("B:" + nombre.nombres + " " + nombre.apellidos)
            form.save()
            messages.success(request, f'Cambios guardados con éxito')
            return redirect('Perfil_baja')
            
        else:
            messages.error(request, 'Error al llenar el formulario')

    context = {
        'item': item,
        'form': form,
    }

    return render(request, 'proyecto/Baja_empleado_update.html', context)

@login_required(login_url='user-login')
def Antiguedad(request, pk): #Comprobar que el empleado si puede entrar a los link de solcitud si tiene la antiguedad y evitar error
    user_filter = UserDatos.objects.get(user=request.user)
    perfil = Perfil.objects.get(numero_de_trabajador=user_filter.numero_de_trabajador, distrito=user_filter.distrito)
    status = Status.objects.get(perfil=perfil)
    fecha_actual = date.today()
    fecha_hace_un_año = fecha_actual - relativedelta(years=1)
    
    if status.fecha_planta_anterior is not None:
        if status.fecha_planta_anterior <= fecha_hace_un_año:
            if pk == 1:
                return redirect('Solicitar_vacacion')
            elif pk == 2:
                return redirect('Solicitar_economico')
            elif pk == 3:
                return redirect('Solicitudes_vista_empleado')
            else:
                messages.error(request, 'Opción inválida para pk')
        else:
            messages.error(request, f'El empleado aún no tiene un año de antigüedad para la solicitud, fecha planta: {status.fecha_planta_anterior}')
    elif status.fecha_planta is not None:
        if status.fecha_planta <= fecha_hace_un_año:
            if pk == 1:
                return redirect('Solicitar_vacacion')
            elif pk == 2:
                return redirect('Solicitar_economico')
            elif pk == 3:
                return redirect('Solicitudes_vista_empleado')
            else:
                messages.error(request, 'Opción inválida para pk')
        else:
            messages.error(request, f'El empleado aún no tiene un año de antigüedad para la solicitud, fecha planta: {status.fecha_planta}')
    else:
        messages.error(request, 'El usuario no cuenta con una fecha de planta para poder realizar la solicitud')
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

@login_required(login_url='user-login')
def Cv_datos(request, pk):

    status = Status.objects.get(id=pk)
    datos = Empleado_cv.objects.filter(status=status).order_by("fecha_fin")

    if request.method =='POST' and 'PDF' in request.POST:
        return generar_curp_pdf(datos,status)
    
    context = {
        'status':status,
        'datos':datos,
        }

    return render(request, 'proyecto/Cv_datos.html',context)

def generar_curp_pdf(datos,status):
    #Configuration of the PDF object
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    
    now = datetime.date.today()
    fecha = str(now)
    #Colores utilizados
    azul = Color(0.16015625,0.5,0.72265625)
    rojo = Color(0.59375, 0.05859375, 0.05859375)
    naranja = Color(1, 0.5, 0)
    gris = Color(0.75, 0.75, 0.75)
    c.setFillColor(gris)
    c.rect(0, 0, 900, 20, fill=True)
    c.rect(0, 0, 20, 900, fill=True)
    #c.setFillColor(naranja)
    #c.rect(550, 750, 100, 100, fill=True)




    #Encabezado
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',10)
    c.drawString(460,760,'Fecha:')
    c.drawString(500,760,fecha)

    #c.setFillColor(white)
    #c.setLineWidth(.2)
    #c.setFont('Helvetica',10)
    #c.drawCentredString(295,755,'Curriculum vitae')
    c.setFillColor(azul)
    c.setFont('Helvetica-Bold',23)
    c.drawString(40,710,status.perfil.nombres)
    c.drawString(40,680,status.perfil.apellidos)
    c.setFillColor(black)
    c.setFont('Helvetica',10)
    c.drawString(40,660,'Lugar de nacimiento:')
    if status.lugar_nacimiento:
        c.drawString(140,660,status.lugar_nacimiento)
    c.drawString(40,640,'Fecha de nacimiento:')
    if status.perfil.fecha_nacimiento:
        edad = now.year - status.perfil.fecha_nacimiento.year - ((now.month, now.day) < (status.perfil.fecha_nacimiento.month, status.perfil.fecha_nacimiento.day))
        c.drawString(140,640,str(status.perfil.fecha_nacimiento))
        c.drawString(250,640,'Edad:')
        c.drawString(280,640,str(edad))
    c.drawString(40,620,'Número de INE:')
    if status.numero_ine:
        c.drawString(120,620,"status.numero_ine")
    c.drawString(40,600,'RFC:')
    if status.rfc:
        c.drawString(120,600,"status.rfc")
    c.drawString(40,580,'Curp:')
    if status.curp:
        c.drawString(120,580,"status.curp")

    #icono = status.perfil.fotoURL
    #if icono:
    #    c.drawInlineImage(icono, 400, 600, 5 * cm, 5 * cm)
    #else:
    if status.sexo.sexo == "Masculino":
        c.drawInlineImage('static/images/perfil/Masculino.png',400,580, 5 * cm, 5 * cm) 
    elif status.sexo.sexo == "Femenino":
        c.drawInlineImage('static/images/perfil/Femenino.png',400,580, 5 * cm, 5 * cm)
    c.line(40,550,300,550) #Linea 1
    c.setFont('Helvetica-Bold',14)
    c.setFillColor(azul)
    c.drawString(40,530,'Experiencia')
    c.setFillColor(black)
    c.setFont('Helvetica',10)
    c.drawString(65,515,'Fecha')
    data =[]
    high = 460
    data.append(['''Inicio''','''Fin''','''Cargo''', '''Puesto''','''Empresa''',])
    for dato in datos: #Salen todos los datos
        fecha_inicio_str = dato.fecha_inicio.strftime('%Y-%m-%d')
        fecha_fin_str = dato.fecha_fin.strftime('%Y-%m-%d')
        puesto = str(dato.puesto)
        empresa = str(dato.empresa)
        styles = getSampleStyleSheet()
        fila = [
            Paragraph(fecha_inicio_str, styles['Normal']),
            Paragraph(fecha_fin_str, styles['Normal']),
            Paragraph(dato.comentario, styles['Normal']),
            Paragraph(puesto, styles['Normal']),
            Paragraph(empresa, styles['Normal'])
        ]
        data.append(fila)
        high = high - 18

    # Añadir un espacio en blanco (un espacio en blanco del tamaño de una línea)
    data.append([Spacer(1, 18)] * 5)
    #Propiedades de la tabla
    width, height = letter
    table = Table(data, colWidths=[2.5 * cm, 2.5 * cm, 7 * cm, 3.5 * cm, 2.5 * cm], repeatRows=1)
    table.setStyle(TableStyle([ #estilos de la tabla
        #ENCABEZADO
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('TEXTCOLOR',(0,0),(-1,0), gris),
        ('FONTSIZE',(0,0),(-1,0), 10),
        ('BACKGROUND',(0,0),(-1,0), white),
        #CUERPO
        ('TEXTCOLOR',(0,1),(-1,-1), colors.black),
        ('FONTSIZE',(0,1),(-1,-1), 10),
        ]))
    table.wrapOn(c, width, height)
    table.drawOn(c, 30, high)
    
    c.line(40,280,300,280) #Linea 2
    c.setFont('Helvetica-Bold',14)
    c.setFillColor(azul)
    c.drawString(40,260,'Formación')
    c.setFillColor(black)
    c.setFont('Helvetica',10)
    c.drawString(40,240,'Profesión:')
    if status.profesion:
        c.drawString(90,240,status.profesion)
    c.drawString(40,220,'No. Cedula:')
    if status.no_cedula:
        c.drawString(95,220,status.no_cedula)
    c.drawString(40,200,'fecha_cedula:')
    if status.fecha_cedula:
        c.drawString(95,200,status.fecha_cedula)
    c.drawString(40,180,'Escuelas donde egreso:')
    if status.escuela:
        c.drawString(150,180,status.escuela)
    """
    c.line(310,300,580,300) #Linea 3
    c.setFont('Helvetica-Bold',14)
    c.setFillColor(azul)
    c.drawString(310,280,'Comunicación')
    c.setFillColor(black)
    c.setFont('Helvetica',10)
    c.line(40,200,300,200) #Linea 4
    c.setFont('Helvetica-Bold',14)
    c.setFillColor(azul)
    c.drawString(40,180,'Liderazgo')
    c.setFillColor(black)
    c.setFont('Helvetica',10)
    c.line(310,200,580,200) #Linea 5
    c.setFont('Helvetica-Bold',14)
    c.setFillColor(azul)
    c.drawString(310,180,'Referencias')
    c.setFillColor(black)
    c.setFont('Helvetica',10)
    """
    c.line(40,100,40,60) #Linea 6
    c.setFont('Helvetica-Bold',14)
    c.setFillColor(azul)
    c.drawString(45,95,'Contacto')
    c.setFillColor(black)
    c.setFont('Helvetica',10)
    c.drawString(45,80,'Telefono:')
    c.drawString(90,80,status.telefono)
    c.drawString(45,70,'Dirección:')
    c.drawString(90,70,status.domicilio)
    c.drawString(45,60,'Correo electronico:')
    c.drawString(135,60,status.perfil.correo_electronico)

    c.save()
    c.showPage()
    buf.seek(0)

    return FileResponse(buf, as_attachment=True, filename='CV_empleado.pdf')

@login_required(login_url='user-login')
def Cv_agregar(request, pk):
    user_filter = UserDatos.objects.get(user=request.user)
    status = Status.objects.get(id=pk)
    puestos = Puesto.objects.all()
    form = CvAgregar()
    if request.method == 'POST':
        form = CvAgregar(request.POST)
        if form.is_valid():
            nombre = Perfil.objects.get(numero_de_trabajador=user_filter.numero_de_trabajador, distrito=user_filter.distrito)
            form.instance.editado = str("B:" + nombre.nombres + " " + nombre.apellidos)
            form.instance.status = status
            form.instance.complete = True
            form.save()
            messages.success(request, f'Se agrego el dato al empleado: {status.perfil.nombres} {status.perfil.apellidos}')
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
            
        else:
            messages.error(request, 'Error al llenar el formulario')


    context = {
        'status': status,
        'form':form,
        'puestos':puestos,
    }

    return render(request, 'proyecto/Cv_agregar.html', context)

@login_required(login_url='user-login')
def Reingreso(request, pk):
    empleado = Perfil.objects.get(id=pk)
    status = Status.objects.get(perfil=empleado)
    anterior = Datos_baja.objects.filter(perfil=empleado).last()
    ahora = datetime.date.today()
    puestos = Puesto.objects.all()
    subproyectos = SubProyecto.objects.all()
    form2 = StatusUpdateForm(instance=status)

    if request.method == 'POST':
        form = PerfilUpdateForm(request.POST, request.FILES, instance=empleado)
        form2 = StatusUpdateForm(request.POST, request.FILES, instance=status)
        if empleado.foto and empleado.foto.size > 2097152:
            messages.error(request,'El tamaño del archivo es mayor de 2 MB')
        elif empleado.fecha_nacimiento >= ahora:
            messages.error(request, 'La fecha de nacimiento no puede ser mayor o igual a hoy')
        elif form.is_valid() and form2.is_valid():
            empleado = form.save(commit=False)
            status = form2.save(commit=False)
            user_filter = UserDatos.objects.get(user=request.user)
            nombre = Perfil.objects.get(numero_de_trabajador = user_filter.numero_de_trabajador, distrito = user_filter.distrito)
            empleado.editado = str("Reingreso:"+nombre.nombres+" "+nombre.apellidos)
            status.editado = str("Reingreso:"+nombre.nombres+" "+nombre.apellidos)
            empleado.baja = False
            empleado.save()
            status.save()
            anterior.pasado=True
            anterior.save()
            messages.success(request, f'Empleado: {empleado.nombres} {empleado.apellidos}, dado de alta')
            return redirect('Perfil')
    else:
        form = PerfilUpdateForm(instance=empleado)
        form2 = StatusUpdateForm(instance=status)

    context = {
        'form':form,
        'form2':form2,
        'empleado':empleado,
        'status':status,
        'subproyectos':subproyectos,
        'puestos':puestos,
        }


    return render(request, 'proyecto/ReingresoForm.html',context)